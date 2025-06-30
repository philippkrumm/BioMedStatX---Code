import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import seaborn as sns
from scipy import stats
from scipy.stats import boxcox, boxcox_normmax
import xlsxwriter
import os
import warnings
from statsmodels.stats.multicomp import pairwise_tukeyhsd
from statsmodels.stats.multitest import multipletests
from itertools import combinations
from datetime import datetime
from matplotlib.ticker import ScalarFormatter
from PyQt5.QtWidgets import QDialog, QVBoxLayout, QLabel, QRadioButton, QDialogButtonBox
from decisiontreevisualizer import DecisionTreeVisualizer
# DISABLED: Nonparametric fallbacks are not yet supported
# from nonparametricanovas import NonParametricFactory, NonParametricRMANOVA

# --------------------------------------------------------------
#  Fallback QApplication to prevent dialogs blocking when script
#  is run purely via CLI
# --------------------------------------------------------------
from PyQt5.QtWidgets import QApplication
import sys as _sys

_QT_APP = QApplication.instance()
if _QT_APP is None:
    _QT_APP = QApplication(_sys.argv)

warnings.simplefilter(action='ignore', category=FutureWarning)

def safe_format(val, fmt="{:.4f}", none_text="N/A"):
    """
    Format a value safely: if numeric, use the given format;
    if None, return the none_text; otherwise, cast to string.
    """
    if isinstance(val, (float, int)):
        try:
            return fmt.format(val)
        except Exception:
            return str(val)
    elif val is None:
        return none_text
    else:
        return str(val)
    
class PostHocAnalyzer:
    """Base class for all post-hoc tests with uniform methods."""
    
    @staticmethod
    def create_result_template(test_name):
        """Creates a standard dictionary for post-hoc results."""
        return {
            "posthoc_test": test_name,
            "pairwise_comparisons": [],
            "error": None
        }
    
    @staticmethod
    def add_comparison(results, group1, group2, test, p_value, statistic=None, 
                       corrected=True, correction_method=None, effect_size=None, 
                       effect_size_type=None, confidence_interval=(None, None), 
                       alpha=0.05):
        """Adds a standardized pairwise comparison to the results."""
        significant = p_value < alpha if isinstance(p_value, (float, int)) else False
        
        comparison = {
            "group1": str(group1),
            "group2": str(group2),
            "test": test,
            "p_value": float(p_value) if isinstance(p_value, (float, int)) else p_value,
            "statistic": float(statistic) if isinstance(statistic, (float, int)) else statistic,
            "significant": significant,
            "corrected": corrected,
            "effect_size": float(effect_size) if isinstance(effect_size, (float, int)) else effect_size,
            "effect_size_type": effect_size_type,
            "confidence_interval": confidence_interval
        }
        
        if correction_method:
            comparison["correction"] = correction_method
            
        results["pairwise_comparisons"].append(comparison)
        return comparison
        
    @staticmethod
    def _holm_sidak_correction(p_values):
        """Applies Holm-Sidak correction to a list of p-values."""
        if not p_values:
            return []
        
        # Use statsmodels implementation instead of custom one
        reject, corrected_p, _, _ = multipletests(p_values, method='holm-sidak')
        return corrected_p.tolist()

class TwoWayPostHocAnalyzer(PostHocAnalyzer):
    """Post-hoc tests for Two-Way ANOVA with a uniform interface."""
    
    @staticmethod
    def perform_test(df, dv, factors, alpha=0.05):
        """
        Performs post-hoc tests for Two-Way ANOVA.
        
        Parameters:
        -----------
        df : pandas.DataFrame
            Data in long format
        dv : str
            Dependent variable
        factors : list
            List of the two factors [factor_a, factor_b]
        alpha : float
            Significance level
            
        Returns:
        --------
        dict
            Standardized post-hoc results
        """
        result = PostHocAnalyzer.create_result_template("Two-Way ANOVA Post-hoc Tests")
        try:
            import pingouin as pg
            has_pingouin = True
        except ImportError:
            has_pingouin = False
            
        try:
            if has_pingouin:
                # Get raw p-values first without adjustment
                ph = pg.pairwise_tests(data=df, dv=dv, between=factors, padjust=None)
                
                # Extract p-values for all comparisons to apply Holm-Sidak correction
                p_values = ph['p-unc'].tolist()
                
                # Apply Holm-Sidak correction
                from statsmodels.stats.multitest import multipletests
                reject, p_corrected, _, _ = multipletests(p_values, alpha=alpha, method='holm-sidak')
                
                # Add corrected p-values back to dataframe
                ph['p-corr'] = p_corrected
                ph['significant'] = reject
                
                for _, row in ph.iterrows():
                    # Standardized result formatting via add_comparison
                    g1 = f"{factors[0]}={row['A']} {factors[1]}={row.get('A2', '')}"
                    g2 = f"{factors[0]}={row['B']} {factors[1]}={row.get('B2', '')}"
                    
                    PostHocAnalyzer.add_comparison(
                        result,
                        group1=g1.strip(),
                        group2=g2.strip(), 
                        test="Pairwise t-test",
                        p_value=row.get('p-corr', row.get('p-unc')),
                        statistic=row.get('T', None),
                        corrected=True,
                        correction_method="Holm-Sidak",
                        effect_size=row.get('hedges', None),
                        effect_size_type="hedges_g",
                        confidence_interval=tuple(row.get('CI95%', (None, None))),
                        alpha=alpha
                    )
            else:
                # Statsmodels fallback implementation - already using Holm-Sidak
                from statsmodels.stats.multicomp import pairwise_tukeyhsd
                
                # Create interaction group for Tukey HSD
                df['interaction_group'] = df[factors[0]].astype(str) + "_" + df[factors[1]].astype(str)
                
                # Run Tukey HSD on the interaction groups
                tukey = pairwise_tukeyhsd(df[dv], df['interaction_group'], alpha=alpha)
                
                # For the Tukey HSD test in the fallback, we'll need to manually apply Holm-Sidak
                # First collect all pairwise comparisons and p-values
                comparisons = []
                for i in range(len(tukey.pvalues)):
                    group1 = tukey.groupsunique[tukey.pairindices[i, 0]]
                    group2 = tukey.groupsunique[tukey.pairindices[i, 1]]
                    p_val = tukey.pvalues[i]
                    conf_int = tukey.confint[i]
                    
                    comparisons.append({
                        'group1': group1,
                        'group2': group2,
                        'p_value': p_val,
                        'conf_int': conf_int
                    })
                
                # Apply Holm-Sidak correction
                p_values = [comp['p_value'] for comp in comparisons]
                # Use statsmodels implementation instead of custom _holm_sidak_correction
                from statsmodels.stats.multitest import multipletests
                reject, corrected_p_values, _, _ = multipletests(p_values, alpha=alpha, method='holm-sidak')
                
                # Convert results into standardized format with corrected p-values
                for i, comp in enumerate(comparisons):
                    is_significant = corrected_p_values[i] < alpha
                    
                    PostHocAnalyzer.add_comparison(
                        result,
                        group1=comp['group1'],
                        group2=comp['group2'],
                        test="Pairwise t-test",
                        p_value=corrected_p_values[i],
                        statistic=None,
                        corrected=True,
                        correction_method="Holm-Sidak",
                        confidence_interval=tuple(comp['conf_int']),
                        alpha=alpha
                    )
                
            return result
        except Exception as e:
            result["error"] = f"Error in Two-Way ANOVA post-hoc tests: {str(e)}"
            return result
        
class MixedAnovaPostHocAnalyzer(PostHocAnalyzer):
    """Post-hoc tests for Mixed ANOVA with a uniform interface."""
    
    @staticmethod
    def perform_test(df, dv, subject, between, within, alpha=0.05):
        """
        Performs post-hoc tests for Mixed ANOVA.
        
        Parameters:
        -----------
        df : pandas.DataFrame
            Data in long format
        dv : str
            Dependent variable
        subject : str
            Column with subject ID
        between : list
            List with the between-factor [between_factor]
        within : list
            List with the within-factor [within_factor]
        alpha : float
            Significance level
            
        Returns:
        --------
        dict
            Standardized post-hoc results
        """
        result = PostHocAnalyzer.create_result_template("Mixed ANOVA Post-hoc Tests")
        
        try:
            import pingouin as pg
            between_factor = between[0]
            within_factor = within[0]
            
            # Get raw p-values first without adjustment
            interaction_ph = pg.pairwise_tests(
                data=df, 
                dv=dv, 
                between=between_factor,
                within=within_factor,
                subject=subject,
                padjust=None  # Get raw uncorrected p-values
            )
            
            # Apply Holm-Sidak correction
            from statsmodels.stats.multitest import multipletests
            p_values = interaction_ph['p-unc'].tolist()
            reject, p_corrected, _, _ = multipletests(p_values, alpha=alpha, method='holm-sidak')
            
            # Add corrected p-values back to dataframe
            interaction_ph['p-corr'] = p_corrected
            interaction_ph['significant'] = reject
            
            for _, row in interaction_ph.iterrows():
                # Create group labels
                group1 = f"{between_factor}={row['A']}, {within_factor}={row['Time']}"
                group2 = f"{between_factor}={row['B']}, {within_factor}={row['Time']}"
                
                # Add standardized comparison
                PostHocAnalyzer.add_comparison(
                    result,
                    group1=group1,
                    group2=group2,
                    test="Paired t-test" if row['Type'] == 'within' else "Independent t-test",
                    p_value=row['p-corr'],
                    statistic=row['T'],
                    corrected=True,
                    correction_method="Holm-Sidak",
                    effect_size=row.get('hedges', None),
                    effect_size_type="hedges_g",
                    confidence_interval=tuple(row.get('CI95%', (None, None))),
                    alpha=alpha
                )
            
            return result
        except Exception as e:
            result["error"] = f"Error in Mixed ANOVA post-hoc tests: {str(e)}"
            return result
        
class RMAnovaPostHocAnalyzer(PostHocAnalyzer):
        
    @staticmethod
    def perform_test(df, dv, subject, within, alpha=0.05):
        result = PostHocAnalyzer.create_result_template("RM ANOVA Post-hoc Tests")
        try:
            from itertools import combinations
            import numpy as np
            import scipy.stats as stats

            within_factor = within[0]
            within_levels = sorted(df[within_factor].unique())
            n_tests = len(list(combinations(within_levels, 2)))

            # Collect all comparisons first
            comparisons = []
            for level1, level2 in combinations(within_levels, 2):
                data1 = df[df[within_factor] == level1].sort_values(by=subject)[dv].values
                data2 = df[df[within_factor] == level2].sort_values(by=subject)[dv].values
                min_len = min(len(data1), len(data2))
                data1 = data1[:min_len]
                data2 = data2[:min_len]

                t_stat, p_val = stats.ttest_rel(data1, data2)
                
                diff = data1 - data2
                mean_diff = np.mean(diff)
                std_diff = np.std(diff, ddof=1)
                effect_size = mean_diff / std_diff if std_diff > 0 else 0

                n = len(diff)
                stderr = std_diff / np.sqrt(n)
                from scipy.stats import t
                # Sidak-adjusted alpha for family-wise confidence intervals
                alpha_sidak = 1 - (1-alpha)**(1/n_tests) 
                t_crit = t.ppf(1 - alpha_sidak/2, n-1)
                ci_lower = mean_diff - t_crit * stderr
                ci_upper = mean_diff + t_crit * stderr
                
                comparisons.append({
                    "level1": level1,
                    "level2": level2,
                    "t_stat": t_stat,
                    "p_val": p_val,
                    "effect_size": effect_size,
                    "ci_lower": ci_lower,
                    "ci_upper": ci_upper
                })
            
            # Apply Holm-Sidak correction
            p_values = [comp["p_val"] for comp in comparisons]
            corrected_p_values = PostHocAnalyzer._holm_sidak_correction(p_values)
            
            # Add each pairwise comparison result with corrected p-values
            for i, comp in enumerate(comparisons):
                PostHocAnalyzer.add_comparison(
                    result,
                    group1=str(comp["level1"]),
                    group2=str(comp["level2"]),
                    test="Paired t-test (Holm-Sidak)",  # Changed from "Paired t-test (Bonferroni)"
                    p_value=corrected_p_values[i],
                    statistic=comp["t_stat"],
                    corrected=True,
                    correction_method="Holm-Sidak",  # Changed from "Bonferroni"
                    effect_size=comp["effect_size"],
                    effect_size_type="cohen_d",
                    confidence_interval=(comp["ci_lower"], comp["ci_upper"]),
                    alpha=alpha
                )

            # If no comparisons were added, add a placeholder
            if not result["pairwise_comparisons"]:
                PostHocAnalyzer.add_comparison(
                    result,
                    group1="No comparison available",
                    group2="",
                    test="RM ANOVA Post-hoc",
                    p_value=None,
                    statistic=None,
                    corrected=False,
                    correction_method=None,
                    effect_size=None,
                    effect_size_type=None,
                    confidence_interval=(None, None),
                    alpha=alpha
                )

            return result
        except Exception as e:
            result["error"] = f"Error in RM ANOVA post-hoc tests: {str(e)}"
            return result
        
class PostHocStatistics:
    """Statistical calculations for various post-hoc tests."""
    
    @staticmethod
    def calculate_cohens_d(group1_data, group2_data, paired=False):
        """Calculates Cohen's d effect size with appropriate adjustments."""
        import numpy as np
        
        if paired:
            diff = np.array(group1_data) - np.array(group2_data)
            return np.mean(diff) / np.std(diff, ddof=1) if np.std(diff, ddof=1) > 0 else 0
        else:
            n1, n2 = len(group1_data), len(group2_data)
            s1, s2 = np.var(group1_data, ddof=1), np.var(group2_data, ddof=1)
            s_pooled = np.sqrt(((n1-1)*s1 + (n2-1)*s2) / (n1+n2-2))
            return (np.mean(group1_data) - np.mean(group2_data)) / s_pooled if s_pooled > 0 else 0
    
    @staticmethod
    def calculate_ci_mean_diff(group1_data, group2_data, alpha=0.05, paired=False):
        """Calculates confidence intervals for the mean difference."""
        import numpy as np
        from scipy.stats import t
        
        try:
            if paired:
                diff = np.array(group1_data) - np.array(group2_data)
                n = len(diff)
                mean_diff = np.mean(diff)
                se = np.std(diff, ddof=1) / np.sqrt(n)
                df = n - 1
            else:
                n1, n2 = len(group1_data), len(group2_data)
                mean_diff = np.mean(group1_data) - np.mean(group2_data)
                s1, s2 = np.var(group1_data, ddof=1), np.var(group2_data, ddof=1)
                se = np.sqrt(s1/n1 + s2/n2)
                df = (s1/n1 + s2/n2)**2 / ((s1/n1)**2/(n1-1) + (s2/n2)**2/(n2-1))
                
            t_crit = t.ppf(1 - alpha/2, df)
            ci_lower = mean_diff - t_crit * se
            ci_upper = mean_diff + t_crit * se
            
            return (float(ci_lower), float(ci_upper))
        except Exception:
            return (None, None)
        
class TukeyHSD(PostHocAnalyzer):
    
    @staticmethod
    def perform_test(valid_groups, samples, alpha=0.05):
        """Performs the Tukey HSD test."""
        import numpy as np
        from statsmodels.stats.multicomp import pairwise_tukeyhsd
        
        result = PostHocAnalyzer.create_result_template("Tukey HSD Test")
        
        try:
            all_data = []
            group_labels = []
            
            for group in valid_groups:
                values = samples[group]
                all_data.extend(values)
                group_labels.extend([str(group)] * len(values))
            
            if len(set(group_labels)) < 2:
                result["error"] = "Tukey HSD requires at least two groups."
                return result

            tukey_result = pairwise_tukeyhsd(endog=all_data, groups=group_labels, alpha=alpha)
            
            # Check if tukey_result has a summary() attribute
            if hasattr(tukey_result, 'summary'):
                summary = tukey_result.summary()
                
                # Extract data from the summary table
                for i in range(len(tukey_result.meandiffs)):
                    group1, group2 = summary.data[i+1][0:2]  # First two columns are the groups
                    p_val = summary.data[i+1][3]  # Fourth column is the p-value
                    lower, upper = summary.data[i+1][4:6]  # Fifth and sixth columns are the confidence intervals
                    reject = summary.data[i+1][6]  # Seventh column is the reject info
                    is_significant = reject  # Diese Variable wird nicht mehr als Parameter übergeben!

                    # Calculate Cohen's d effect size
                    group1_data = samples[group1]
                    group2_data = samples[group2]
                    effect_size = PostHocStatistics.calculate_cohens_d(group1_data, group2_data)

                    # Use the common method to add a comparison
                    PostHocAnalyzer.add_comparison(
                        result,
                        group1=group1,
                        group2=group2,
                        test="Tukey HSD",
                        p_value=p_val,
                        statistic=tukey_result.meandiffs[i],
                        corrected=True,
                        correction_method="Tukey HSD",
                        effect_size=effect_size,
                        effect_size_type="cohen_d",
                        confidence_interval=(float(lower), float(upper)),
                        alpha=alpha
                        # Der Parameter significant=is_significant wurde entfernt
                    )
            else:
                result["error"] = "TukeyHSDResults object has no summary() attribute"
                return result
            
            return result
        except Exception as e:
            result["error"] = f"Error in Tukey HSD test: {str(e)}"
            return result
        
class DunnettTest(PostHocAnalyzer):
    """Implementation of the Dunnett test for comparing multiple groups to a control group."""

    @staticmethod
    def perform_test(valid_groups, samples, control_group, alpha=0.05):
        """
        Performs the Dunnett test (compares each group to the control group).
        """
        result = PostHocAnalyzer.create_result_template(f"Dunnett Test (Control group: {control_group})")
        result["control_group"] = control_group

        try:
            import scikit_posthocs as sp
            all_data = []
            group_labels = []
            for group in valid_groups:
                values = samples[group]
                all_data.extend(values)
                group_labels.extend([str(group)] * len(values))
            df = pd.DataFrame({"value": all_data, "group": group_labels})

            dunnett_result = sp.posthoc_dunnett(df, val_col="value", group_col="group", control=str(control_group))

            group_means = {str(g): np.mean(samples[g]) for g in valid_groups}
            group_n = {str(g): len(samples[g]) for g in valid_groups}
            group_std = {str(g): np.std(samples[g], ddof=1) for g in valid_groups}

            p_vals = []
            group_pairs = []
            mean_diffs = []
            ci_lowers = []
            ci_uppers = []
            effect_sizes = []

            k = len(valid_groups) - 1  # Number of comparisons with control
            alpha_sidak = 1 - (1 - alpha) ** (1 / k) if k > 0 else alpha

            for group in valid_groups:
                if str(group) == str(control_group):
                    continue
                # Robust matrix lookup (row/col swap if needed)
                try:
                    p_val = float(dunnett_result.loc[str(group), str(control_group)])
                except KeyError:
                    try:
                        p_val = float(dunnett_result.loc[str(control_group), str(group)])
                    except Exception:
                        p_val = 1.0
                p_vals.append(p_val)
                group_pairs.append(group)

                mean_diff = group_means[str(group)] - group_means[str(control_group)]
                n1 = group_n[str(group)]
                n2 = group_n[str(control_group)]
                s1 = group_std[str(group)]
                s2 = group_std[str(control_group)]
                pooled_std = np.sqrt(((n1 - 1) * (s1 ** 2) + (n2 - 1) * (s2 ** 2)) / (n1 + n2 - 2))
                se_diff = pooled_std * np.sqrt(1 / n1 + 1 / n2)
                df_val = n1 + n2 - 2
                from scipy.stats import t
                t_crit = t.ppf(1 - alpha_sidak / 2, df_val)
                ci_lower = mean_diff - t_crit * se_diff
                ci_upper = mean_diff + t_crit * se_diff
                effect_size = mean_diff / pooled_std if pooled_std > 0 else 0

                mean_diffs.append(mean_diff)
                ci_lowers.append(ci_lower)
                ci_uppers.append(ci_upper)
                effect_sizes.append(effect_size)

            from statsmodels.stats.multitest import multipletests
            reject, p_adj, _, _ = multipletests(p_vals, alpha=alpha, method='holm-sidak')

            for i, group in enumerate(group_pairs):
                PostHocAnalyzer.add_comparison(
                    result,
                    group1=group,
                    group2=control_group,
                    test="Dunnett",
                    p_value=p_adj[i],
                    statistic=None,
                    corrected=True,
                    correction_method="Holm-Sidak",
                    effect_size=effect_sizes[i],
                    effect_size_type="cohen_d",
                    confidence_interval=(float(ci_lowers[i]), float(ci_uppers[i])),
                    alpha=alpha
                )

            return result
        except Exception as e:
            import traceback
            result["error"] = f"Error in Dunnett test: {str(e)}"
            traceback.print_exc()
            return result
        
import numpy as np
from scipy.stats import mannwhitneyu
from statsmodels.stats.multitest import multipletests

class DunnTest(PostHocAnalyzer):
    @staticmethod
    def perform_test(valid_groups, samples, alpha=0.05, n_boot=1000):
        result = PostHocAnalyzer.create_result_template("Dunn-Test")

        try:
            import scikit_posthocs as sp
        except ImportError:
            result["error"] = "scikit-posthocs is not installed."
            return result

        # 1) Get raw p-values matrix
        data_array = [samples[g] for g in valid_groups]
        raw_p = sp.posthoc_dunn(data_array, p_adjust=None)  # no internal correction

        # 2) Flatten into list and correct with Holm-Sidak
        pairs = []
        pvals = []
        for i, g1 in enumerate(valid_groups):
            for j, g2 in enumerate(valid_groups):
                if i < j:
                    pairs.append((g1, g2))
                    pvals.append(raw_p.iloc[i, j])
        reject, p_adj, _, _ = multipletests(pvals, alpha=alpha, method='holm-sidak')

        # 3) Loop over pairs and compute effect & CI
        for (g1, g2), pval_adj, sig in zip(pairs, p_adj, reject):
            x, y = samples[g1], samples[g2]
            # Mann–Whitney U for effect‐size r
            U, _ = mannwhitneyu(x, y, alternative='two-sided')
            n1, n2 = len(x), len(y)
            z = (U - n1 * n2 / 2) / np.sqrt(n1 * n2 * (n1 + n2 + 1) / 12)
            effect_r = abs(z) / np.sqrt(n1 + n2)

            # Hodges-Lehmann
            diffs = [xi - yi for xi in x for yi in y]
            hl = np.median(diffs)

            # Bootstrap CI on HL
            boots = []
            for _ in range(n_boot):
                b1 = np.random.choice(x, n1, replace=True)
                b2 = np.random.choice(y, n2, replace=True)
                boots.append(np.median([u - v for u in b1 for v in b2]))
            ci_low, ci_high = np.percentile(boots, [100*alpha/2, 100*(1-alpha/2)])

            # Median difference
            med_diff = np.median(x) - np.median(y)

            PostHocAnalyzer.add_comparison(
                result,
                group1=g1,
                group2=g2,
                test="Dunn",
                p_value=pval_adj,
                statistic=None,
                corrected=True,
                correction_method="Holm-Sidak",
                effect_size=effect_r,
                effect_size_type="r",
                confidence_interval=(float(ci_low), float(ci_high)),
                alpha=alpha
            )

        return result

class DependentPostHoc(PostHocAnalyzer):
    @staticmethod
    def perform_test(valid_groups, samples, alpha=0.05, parametric=True):
        name = "Parametric paired t-tests" if parametric else "Wilcoxon signed-rank tests"
        result = PostHocAnalyzer.create_result_template(name)

        # 1) check equal lengths
        sizes = [len(samples[g]) for g in valid_groups]
        if len(set(sizes)) != 1:
            result["error"] = "All groups must have same length for dependent tests."
            return result

        # 2) collect stats
        pvals, stats_list, pairs = [], [], []
        for g1, g2 in combinations(valid_groups, 2):
            x, y = np.array(samples[g1]), np.array(samples[g2])
            if parametric:
                tstat, p = stats.ttest_rel(x, y)
                stats_list.append(tstat)
            else:
                wstat, p = stats.wilcoxon(x, y)
                stats_list.append(wstat)
            pvals.append(p)
            pairs.append((g1, g2, x, y))

        # 3) Holm-Sidak correction
        reject, p_adj, _, _ = multipletests(pvals, alpha=alpha, method='holm-sidak')

        # 4) add comparisons
        for i, (g1, g2, x, y) in enumerate(pairs):
            if parametric:
                # paired CI and d
                ci = PostHocStatistics.calculate_ci_mean_diff(x, y, alpha=alpha, paired=True)
                d = PostHocStatistics.calculate_cohens_d(x, y, paired=True)
                test = "Paired t-test"
                stat = stats_list[i]
                es, estype = d, "cohen_d"
            else:
                # r from Wilcoxon
                n = len(x)
                W = stats_list[i]
                mu = n*(n+1)/4
                sigma = np.sqrt(n*(n+1)*(2*n+1)/24)
                z = (W - mu)/sigma
                r = abs(z)/np.sqrt(n)
                ci = (None, None)
                test = "Wilcoxon signed-rank"
                stat = W
                es, estype = r, "r"

            PostHocAnalyzer.add_comparison(
                result,
                group1=g1,
                group2=g2,
                test=test,
                p_value=p_adj[i],
                statistic=stat,
                corrected=True,
                correction_method="Holm-Sidak",
                effect_size=es,
                effect_size_type=estype,
                confidence_interval=ci,
                alpha=alpha
            )

        return result
            
class PostHocFactory:
    @staticmethod
    def create_test(test_type, is_parametric=True, is_dependent=False):
        """Creates the correct post-hoc test implementation based on parameters."""
        if is_dependent:
            return DependentPostHoc()
        
        if is_parametric:
            if test_type == "tukey":
                return TukeyHSD()
            elif test_type == "dunnett":
                return DunnettTest()
        else:
            if test_type == "dunn":
                return DunnTest()
            elif test_type == "conover":
                # Return None or a message since ConoverPostHoc is removed
                return None
            elif test_type == "nemenyi":
                # Return None or a message since NemenyiPostHoc is removed
                return None
        
        return None
    
    @staticmethod
    def create_anova_posthoc(anova_type, **kwargs):
        """Creates specialized post-hoc tests for different ANOVA types."""
        if anova_type == "two_way":
            return TwoWayPostHocAnalyzer()
        elif anova_type == "mixed":
            return MixedAnovaPostHocAnalyzer()
        elif anova_type == "rm":
            return RMAnovaPostHocAnalyzer()
        return None
    
    @staticmethod
    def perform_posthoc_for_anova(anova_type, df, dv, subject=None, between=None, within=None, alpha=0.05):
        """
        Performs post-hoc tests for an ANOVA type and returns standardized results.
        
        Parameters:
        -----------
        anova_type : str
            Type of ANOVA ('two_way', 'mixed', 'rm')
        df : pandas.DataFrame
            Dataset in long format
        dv : str
            Name of the dependent variable
        subject : str, optional
            Name of the subject variable (for Mixed and RM ANOVA)
        between : list, optional
            List of between factors
        within : list, optional
            List of within factors
        alpha : float, optional
            Significance level (default: 0.05)
            
        Returns:
        --------
        dict
            Standardized post-hoc results
        """
        analyzer = PostHocFactory.create_anova_posthoc(anova_type)
        if analyzer is None:
            return {"error": f"No post-hoc test available for ANOVA type '{anova_type}'"}
        
        if anova_type == "two_way":
            if not between or len(between) != 2:
                return {"error": "Two-Way ANOVA requires two between factors"}
            return analyzer.perform_test(df=df, dv=dv, factors=between, alpha=alpha)
        
        elif anova_type == "mixed":
            # Full implementation for Mixed ANOVA
            if not subject:
                return {"error": "Mixed ANOVA requires a subject variable"}
            if not between or len(between) != 1:
                return {"error": "Mixed ANOVA requires exactly one between factor"}
            if not within or len(within) != 1:
                return {"error": "Mixed ANOVA requires exactly one within factor"}
            
            return analyzer.perform_test(df=df, dv=dv, subject=subject, between=between, within=within, alpha=alpha)
        
        elif anova_type == "rm":
            # Full implementation for RM-ANOVA
            if not subject:
                return {"error": "RM-ANOVA requires a subject variable"}
            if not within or len(within) < 1:
                return {"error": "RM-ANOVA requires at least one within factor"}
            
            # Get post-hoc results from analyzer
            posthoc = analyzer.perform_test(df=df, dv=dv, subject=subject, within=within, alpha=alpha)
            
            # Add validation to ensure we're getting valid results
            if posthoc and 'pairwise_comparisons' in posthoc:
                print(f"DEBUG: Found {len(posthoc['pairwise_comparisons'])} rm-anova post-hoc comparisons")
            else:
                print("DEBUG: No valid rm-anova post-hoc results found!")
                
            # Explicitly pass through the posthoc results without modification
            return posthoc
        
        return {"error": f"Unknown ANOVA type: {anova_type}"}
    
class DataImporter:
    @staticmethod
    def import_data(file_path, sheet_name=0, group_col="Group", value_cols=None, combine_columns=False):
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"The file {file_path} was not found.")
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
        if group_col not in df.columns:
            raise ValueError(f"The group column '{group_col}' was not found. Available columns: {', '.join(df.columns)}")
        if value_cols is None:
            numeric_cols = [col for col in df.columns if pd.api.types.is_numeric_dtype(df[col]) and col != group_col]
            if len(numeric_cols) == 0:
                raise ValueError("No numeric columns found that can be used as measurements.")
            value_cols = numeric_cols
        for col in value_cols:
            if col not in df.columns:
                raise ValueError(f"The value column '{col}' was not found. Available columns: {', '.join(df.columns)}")
        groups = sorted(df[group_col].unique())
        samples = {}
        if combine_columns:
            for group in groups:
                combined_values = []
                for col in value_cols:
                    values = df[df[group_col] == group][col].dropna().tolist()
                    combined_values.extend(values)
                samples[group] = combined_values
        else:  # if combine_columns=False
            if len(value_cols) > 1:
                print("Warning: Multiple value columns specified, but combine_columns=False. Only the first column will be used.")
            
            for group in groups:
                values = df[df[group_col] == group][value_cols[0]].dropna().tolist()
                samples[group] = values
        return samples, df

import numpy as np
import pandas as pd
import scipy.stats as stats
try:
    import scikit_posthocs as sp
    HAS_SCPH = True
except ImportError:
    HAS_SCPH = False      
    
class StatisticalTester:
    @staticmethod
    def _standardize_results(results):
        """
        Ensures that all results dictionaries have the same structure with consistent keys.
        Fills in any missing keys with None or appropriate default values.
        
        Parameters:
        -----------
        results : dict
            The results dictionary to standardize
            
        Returns:
        --------
        dict
            The standardized results dictionary
        """
        standard_keys = {
            "test": None,
            "p_value": None,
            "statistic": None,
            "posthoc_test": None,
            "pairwise_comparisons": [],
            "descriptive": {},
            "descriptive_transformed": {},
            "alpha": 0.05,
            "null_hypothesis": "The means/medians of all groups are equal.",
            "alternative_hypothesis": "At least one mean/median differs from the others.",
            "confidence_interval": None,
            "ci_level": 0.95,
            "power": None,
            "effect_size": None,
            "effect_size_type": None,
            "error": None,
            "df1": None,
            "df2": None
        }
        
        # Debug logging
        if 'pairwise_comparisons' in results:
            print(f"DEBUG: Standardizing results with {len(results['pairwise_comparisons'])} pairwise comparisons")
        
        # Copy all existing values first
        standardized = dict(results)
        
        # Fill in any missing top-level keys with default values (without overwriting existing ones)
        for key, default_value in standard_keys.items():
            if key not in standardized:
                standardized[key] = default_value

        # Define the standard keys for each post-hoc comparison entry
        comparison_keys = {
            "group1": "",
            "group2": "",
            "test": "",
            "p_value": None,
            "statistic": None,
            "significant": False,
            "corrected": False,
            "effect_size": None,
            "effect_size_type": None,
            "confidence_interval": (None, None),
            "correction": None  # optional key if a correction method was applied
        }
        
        # If the standardized results contain pairwise comparisons,
        # ensure each comparison has all the expected keys.
        if isinstance(standardized.get("pairwise_comparisons"), list):
            for comp in standardized.get("pairwise_comparisons", []):
                for key, default in comparison_keys.items():
                    if key not in comp:
                        comp[key] = default

        return standardized

    @staticmethod
    def check_normality_and_variance(groups, samples, dataset_name=None, progress_text=None, column_name=None, already_transformed=False):
        """
        Checks normality and homogeneity of variance of the data.
        Performs transformations if necessary.
        Returns transformed data, test recommendation, and test info.
        
        Parameters:
        -----------
        groups : list
            List of groups to analyze
        samples : dict
            Dictionary with group names as keys and lists of measurements as values
        dataset_name : str, optional
            Name of the dataset for logging purposes
        progress_text : str, optional
            Progress text for dialog title
        column_name : str, optional
            Column name for dialog title
        already_transformed : bool, optional
            Indicates if the data has already been transformed (prevents second transformation)
        """
        from scipy.stats import boxcox, boxcox_normmax

        test_info = {"normality_tests": {}, "variance_test": {}, "transformation": None}
        valid_groups = [g for g in groups if g in samples and len(samples[g]) > 0]
        transformed_samples = {g: samples[g].copy() for g in valid_groups}
        
        # Initialize test_recommendation with default value
        test_recommendation = "parametric"  # Default to parametric if all tests pass

        # --- Normality: Shapiro-Wilk-Test für jede Gruppe einzeln
        test_info["normality_tests"] = {}
        all_normal = True
        for group in valid_groups:
            values = samples[group]
            if len(values) >= 3 and len(set(values)) > 1:
                try:
                    stat, pval = stats.shapiro(values)
                    is_normal = pval > 0.05
                    test_info["normality_tests"][group] = {
                        "statistic": stat, "p_value": pval, "is_normal": is_normal
                    }
                    if not is_normal:
                        all_normal = False
                except Exception as e:
                    test_info["normality_tests"][group] = {"statistic": None, "p_value": None, "error": str(e)}
                    all_normal = False
            else:
                test_info["normality_tests"][group] = {"statistic": None, "p_value": None, "note": "Too few values"}
                all_normal = False

        # --- Levene test for homogeneity of variance
        data_for_levene = [samples[g] for g in valid_groups]
        has_equal_variance = True
        if len(valid_groups) >= 2 and all(len(v) >= 3 for v in data_for_levene):
            try:
                stat, pval = stats.levene(*data_for_levene)
                has_equal_variance = pval > 0.05
                test_info["variance_test"] = {
                    "statistic": stat, 
                    "p_value": pval, 
                    "equal_variance": has_equal_variance
                }
            except Exception as e:
                test_info["variance_test"] = {
                    "statistic": None, 
                    "p_value": None, 
                    "error": str(e),
                    "equal_variance": False  # Explicitly set to False on error
                }
                has_equal_variance = False
        else:
            test_info["variance_test"] = {
                "statistic": None, 
                "p_value": None, 
                "note": "Too few groups or values",
                "equal_variance": False  # Explicitly set to False if too little data
            }

        # --- Transformation if not normal
        if not all_normal or not has_equal_variance:
            # If already transformed, immediately recommend non-parametric test
            if already_transformed:
                test_info["transformation"] = "No further"
                return transformed_samples, "non_parametric", test_info
            
            transformation_type = None
            try:
                # Only show dialog if not already specified
                if transformation_type is None:
                    # Use the progress_text and column_name directly
                    # UIDialogManager now handles the normalization internally
                    transformation_type = UIDialogManager.select_transformation_dialog(
                        parent=None, progress_text=progress_text, column_name=column_name
                    )
            except Exception:
                transformation_type = "none"
                
            # Only store the transformation selection if it's valid
            if transformation_type and transformation_type != "none":
                test_info["transformation"] = transformation_type

                # Apply the selected transformation
                if transformation_type == "log10":
                    for group in valid_groups:
                        values = samples[group]
                        min_val = min(values)
                        shift = -min_val + 1 if min_val <= 0 else 0
                        transformed_samples[group] = [np.log10(v + shift) for v in values]
                elif transformation_type == "boxcox":
                    for group in valid_groups:
                        values = samples[group]
                        min_val = min(values)
                        shift = -min_val + 1 if min_val <= 0 else 0
                        shifted = [v + shift for v in values]
                        try:
                            lambda_val = boxcox_normmax(shifted)
                            transformed_samples[group] = list(boxcox(shifted, lambda_val))
                            test_info["boxcox_lambda"] = lambda_val
                        except Exception as e:
                            test_info["transformation_error"] = str(e)
                            transformed_samples[group] = [np.log10(v + shift) for v in values]
                elif transformation_type == "arcsin_sqrt":
                    # Add arcsin_sqrt transformation
                    for group in valid_groups:
                        values = samples[group]
                        min_val = min(values)
                        max_val = max(values)
                        # Scale values to 0-1 range if needed
                        if min_val < 0 or max_val > 1:
                            scaled = [(v - min_val) / (max_val - min_val) for v in values]
                        else:
                            scaled = values
                        transformed_samples[group] = [np.arcsin(np.sqrt(v)) for v in scaled]
                    test_info["transformation"] = "arcsin_sqrt"
            else:
                # No transformation selected
                test_info["transformation"] = "None"
                return transformed_samples, "non_parametric", test_info

            # Post-test after transformation
            try:
                transformed_data = []
                for group in valid_groups:
                    for value in transformed_samples[group]:
                        transformed_data.append({'Group': group, 'Value': value})
                df_t = pd.DataFrame(transformed_data)
                from statsmodels.formula.api import ols
                model_t = ols('Value ~ C(Group)', data=df_t).fit()
                residuals_t = model_t.resid
                if len(residuals_t) >= 3 and len(set(residuals_t)) > 1:
                    stat, pval = stats.shapiro(residuals_t)
                    test_info["normality_tests"]["transformed_data"] = {
                        "statistic": stat, "p_value": pval, "is_normal": pval > 0.05
                    }
            except Exception as e:
                test_info["normality_tests"]["transformed_data"] = {
                    "statistic": None, "p_value": None, "error": str(e)
                }
            try:
                data_for_levene_trans = [transformed_samples[g] for g in valid_groups]
                if len(valid_groups) >= 2 and all(len(v) >= 3 for v in data_for_levene_trans):
                    stat, pval = stats.levene(*data_for_levene_trans)
                    test_info["variance_test"]["transformed"] = {
                        "statistic": stat, "p_value": pval, "equal_variance": pval > 0.05
                    }
            except Exception as e:
                test_info["variance_test"]["transformed"] = {
                    "statistic": None, "p_value": None, "error": str(e)
                }

            # --- Test recommendation after transformation (ONLY if transformation was done)
            normal_after = test_info["normality_tests"].get("transformed_data", {}).get("is_normal", False)
            varhom_after = test_info["variance_test"].get("transformed", {}).get("equal_variance", False)

            if normal_after and varhom_after:
                test_recommendation = "parametric"
            elif normal_after and not varhom_after:
                # Key change: For t-tests with unequal variances, use Welch's t-test
                if len(valid_groups) == 2:
                    test_recommendation = "parametric"  # Still parametric, but will use Welch's t-test
                    test_info["note"] = "Normal distribution but unequal variances – Welch's t-test will be used."
                else:
                    # For ANOVA with unequal variances
                    test_recommendation = "parametric"  # Could be "welch" for Welch's ANOVA
                    test_info["note"] = "Normal distribution but unequal variances – Welch's ANOVA will be used."
            else:
                test_recommendation = "non_parametric"

            # Add debug log
            print(f"DEBUG: Final recommendation after transformation: {test_recommendation}")
            print(f"DEBUG: normal_after={normal_after}, varhom_after={varhom_after}")
            print(f"DEBUG: recommendation = {test_recommendation}")
            print("DEBUG TREE: Taking parametric path" if test_recommendation == "parametric" else "DEBUG TREE: Taking non-parametric path")

        return transformed_samples, test_recommendation, test_info
    
    @staticmethod
    def perform_statistical_test(
    groups, transformed_samples, original_samples, 
    dependent=False, test_recommendation="parametric", alpha=0.05, test_info=None
    ):
        """
        Robustly performs the statistical test (t-test, ANOVA, Mann-Whitney, Kruskal-Wallis, posthoc, etc.).
        Catches all typical sources of error and always returns a meaningful result dict.
        Important: For parametric tests, transformed data is used,
        for non-parametric tests, the original data is used.
        """
        results = {
            "test": None,
            "p_value": None,
            "statistic": None,
            "posthoc_test": None,
            "pairwise_comparisons": [],
            "descriptive": {},
            "descriptive_transformed": {},
            "alpha": alpha,
            "null_hypothesis": "The means/medians of all groups are equal.",
            "alternative_hypothesis": "At least one mean/median differs from the others.",
            "confidence_interval": None,
            "ci_level": 0.95,
            "power": None
        }

        valid_groups = [g for g in groups if g in transformed_samples and len(transformed_samples[g]) > 0]

        # Always compute descriptive stats for both original and transformed data
        results["descriptive"] = {g: StatisticalTester._compute_descriptive_stats(original_samples[g]) for g in valid_groups}
        
        # IMPORTANT: Always include transformed descriptive stats when transformation was performed
        if any(original_samples[g] != transformed_samples[g] for g in valid_groups):
            results["descriptive_transformed"] = {g: StatisticalTester._compute_descriptive_stats(transformed_samples[g]) for g in valid_groups}
        
        # Store raw data for both original and transformed values
        results["raw_data"] = {g: original_samples[g].copy() for g in valid_groups}
        if original_samples != transformed_samples:
            results["raw_data_transformed"] = {g: transformed_samples[g].copy() for g in valid_groups}

        if len(valid_groups) == 0:
            return StatisticalTester._stat_test_no_valid_groups(results)

        if len(valid_groups) == 1:
            return StatisticalTester._stat_test_one_group(results, valid_groups, original_samples, transformed_samples)

        # Descriptive statistics for all groups
        results["descriptive"] = {g: StatisticalTester._compute_descriptive_stats(original_samples[g]) for g in valid_groups}
        if original_samples != transformed_samples:
            results["descriptive_transformed"] = {g: StatisticalTester._compute_descriptive_stats(transformed_samples[g]) for g in valid_groups}

        samples_to_use = transformed_samples if test_recommendation == "parametric" else original_samples

        if len(valid_groups) == 2:
            return StatisticalTester._stat_test_two_groups(
                results, valid_groups, samples_to_use, original_samples, dependent, test_recommendation, alpha
            )

        return StatisticalTester._stat_test_multi_groups(
        results, valid_groups, samples_to_use, dependent, test_recommendation, alpha, test_info=test_info
        )

    @staticmethod
    def _stat_test_no_valid_groups(results):
        results["test"] = "No test possible"
        results["error"] = "No valid groups found"
        return StatisticalTester._standardize_results(results)

    @staticmethod
    def _stat_test_one_group(results, valid_groups, original_samples, transformed_samples):
        g = valid_groups[0]
        results["descriptive"] = {
            g: StatisticalTester._compute_descriptive_stats(original_samples[g])
        }
        if original_samples[g] != transformed_samples[g]:
            results["descriptive_transformed"] = {
                g: StatisticalTester._compute_descriptive_stats(transformed_samples[g])
            }
        results["test"] = "Descriptive statistics only"
        return StatisticalTester._standardize_results(results)

    @staticmethod
    def _stat_test_two_groups(results, valid_groups, samples_to_use, original_samples, dependent, test_recommendation, alpha):
        g1, g2 = valid_groups
        data1, data2 = samples_to_use[g1], samples_to_use[g2]
        try:
            if dependent:
                if test_recommendation == "parametric":
                    return StatisticalTester._paired_ttest(results, g1, g2, data1, data2, alpha)
                else:
                    return StatisticalTester._wilcoxon_test(results, g1, g2, data1, data2, alpha)
            else:
                if test_recommendation == "parametric":
                    return StatisticalTester._independent_ttest(results, g1, g2, data1, data2, alpha)
                else:
                    return StatisticalTester._mannwhitney_test(results, g1, g2, data1, data2, alpha)
        except Exception as e:
            results["test"] = "Error during test"
            results["error"] = str(e)
            results["posthoc_test"] = "Not performed (error in main test)"
            results["pairwise_comparisons"] = []
            return StatisticalTester._standardize_results(results)

    @staticmethod
    def _paired_ttest(results, g1, g2, data1, data2, alpha):
        statistic, p_value = stats.ttest_rel(data1, data2)
        test_name = "Paired t-test"
        diff = np.array(data1) - np.array(data2)
        cohen_d = np.mean(diff) / np.std(diff, ddof=1)
        results["effect_size"] = cohen_d
        results["effect_size_type"] = "cohen_d"
        n = len(diff)
        stderr = np.std(diff, ddof=1) / np.sqrt(n)
        from scipy.stats import t
        ci = t.interval(0.95, n-1, loc=np.mean(diff), scale=stderr)
        results["confidence_interval"] = ci
        try:
            from statsmodels.stats.power import TTestPower
            effect_size = abs(cohen_d)
            power_analysis = TTestPower()
            results["power"] = float(power_analysis.power(effect_size=effect_size, nobs=n, alpha=alpha))
        except Exception as e:
            results["power"] = None
        results["test"] = test_name
        results["statistic"] = statistic
        results["p_value"] = p_value
        results["pairwise_comparisons"] = [{
            "group1": g1, "group2": g2, "test": test_name,
            "p_value": p_value, "statistic": statistic,
            "significant": p_value < alpha, "corrected": False,
            "effect_size": results.get("effect_size"),
            "effect_size_type": results.get("effect_size_type"),
            "confidence_interval": results.get("confidence_interval"),
            "power": results.get("power")
        }]
        return StatisticalTester._standardize_results(results)

    @staticmethod
    def _wilcoxon_test(results, g1, g2, data1, data2, alpha):
        statistic, p_value = stats.wilcoxon(data1, data2)
        test_name = "Wilcoxon test"
        n = len(data1)
        r = statistic / (n * (n + 1) / 2)
        results["effect_size"] = r
        results["effect_size_type"] = "r"
        results["confidence_interval"] = (None, None)
        try:
            from statsmodels.stats.power import TTestPower
            effect_size_corrected = r * 0.955
            power_analysis = TTestPower()
            results["power"] = float(power_analysis.power(effect_size=effect_size_corrected, nobs=n, alpha=alpha))
        except Exception as e:
            results["power"] = None
        results["test"] = test_name
        results["statistic"] = statistic
        results["p_value"] = p_value
        results["pairwise_comparisons"] = [{
            "group1": g1, "group2": g2, "test": test_name,
            "p_value": p_value, "statistic": statistic,
            "significant": p_value < alpha, "corrected": False,
            "effect_size": results.get("effect_size"),
            "effect_size_type": results.get("effect_size_type"),
            "confidence_interval": results.get("confidence_interval"),
            "power": results.get("power")
        }]
        return StatisticalTester._standardize_results(results)
    
    @staticmethod
    def _independent_ttest(results, g1, g2, data1, data2, alpha):
        statistic, p_value = stats.ttest_ind(data1, data2, equal_var=True)
        test_name = "t-test (independent)"
        n1, n2 = len(data1), len(data2)
        s1, s2 = np.var(data1, ddof=1), np.var(data2, ddof=1)
        s_pooled = np.sqrt(((n1-1)*s1 + (n2-1)*s2) / (n1+n2-2))
        cohen_d = (np.mean(data1) - np.mean(data2)) / s_pooled
        results["effect_size"] = cohen_d
        results["effect_size_type"] = "cohen_d"
        mean_diff = np.mean(data1) - np.mean(data2)
        stderr_diff = np.sqrt(s1/n1 + s2/n2)
        from scipy.stats import t
        ci = t.interval(0.95, n1+n2-2, loc=mean_diff, scale=stderr_diff)
        results["confidence_interval"] = ci
        try:
            from statsmodels.stats.power import TTestIndPower
            effect_size = abs(cohen_d)
            power_analysis = TTestIndPower()
            results["power"] = float(power_analysis.power(effect_size=effect_size, nobs1=n1, ratio=n2/n1, alpha=alpha))
        except Exception as e:
            results["power"] = None
        results["test"] = test_name
        results["statistic"] = statistic
        results["p_value"] = p_value
        results["pairwise_comparisons"] = [{
            "group1": g1, "group2": g2, "test": test_name,
            "p_value": p_value, "statistic": statistic,
            "significant": p_value < alpha, "corrected": False,
            "effect_size": results.get("effect_size"),
            "effect_size_type": results.get("effect_size_type"),
            "confidence_interval": results.get("confidence_interval"),
            "power": results.get("power")
        }]
        return StatisticalTester._standardize_results(results)

    @staticmethod
    def _mannwhitney_test(results, g1, g2, data1, data2, alpha):
        statistic, p_value = stats.mannwhitneyu(data1, data2, alternative='two-sided')
        test_name = "Mann-Whitney-U"
        n1, n2 = len(data1), len(data2)
        u = statistic
        mean_u = n1 * n2 / 2
        std_u = np.sqrt(n1 * n2 * (n1 + n2 + 1) / 12)
        z = (u - mean_u) / std_u if std_u > 0 else 0
        r = abs(z) / np.sqrt(n1 + n2)
        results["effect_size"] = r
        results["effect_size_type"] = "r"
        results["confidence_interval"] = (None, None)
        try:
            from statsmodels.stats.power import TTestIndPower
            effect_size_corrected = r * 0.955
            power_analysis = TTestIndPower()
            results["power"] = float(power_analysis.power(effect_size=effect_size_corrected, nobs1=n1, ratio=n2/n1, alpha=alpha))
        except Exception as e:
            results["power"] = None
        results["test"] = test_name
        results["statistic"] = statistic
        results["p_value"] = p_value
        results["pairwise_comparisons"] = [{
            "group1": g1, "group2": g2, "test": test_name,
            "p_value": p_value, "statistic": statistic,
            "significant": p_value < alpha, "corrected": False,
            "effect_size": results.get("effect_size"),
            "effect_size_type": results.get("effect_size_type"),
            "confidence_interval": results.get("confidence_interval"),
            "power": results.get("power")
        }]
        return StatisticalTester._standardize_results(results)

    @staticmethod
    def _stat_test_multi_groups(results, valid_groups, samples_to_use, dependent, test_recommendation, alpha, test_info=None, df=None, dv=None, subject=None, within=None):
        # Welch ANOVA: If explicitly requested OR if parametric but variances are unequal
        welch_condition = (
            test_recommendation == "welch"  # Direct request for Welch ANOVA
            or (
                test_recommendation == "parametric" 
                and test_info is not None
                and (
                    # Check if transformed data has unequal variance
                    (test_info.get("transformation") and not test_info.get("variance_test", {}).get("transformed", {}).get("equal_variance", True))
                    # Or if original data has unequal variance (when no transformation)
                    or (not test_info.get("transformation") and not test_info.get("variance_test", {}).get("equal_variance", True))
                )
                # Ensure we're still normally distributed
                and (
                    (test_info.get("transformation") and test_info.get("normality_tests", {}).get("transformed_data", {}).get("is_normal", False))
                    or (not test_info.get("transformation") and test_info.get("normality_tests", {}).get("all_data", {}).get("is_normal", False))
                )
            )
        )
        
        print(f"DEBUG WELCH CHECK: test_recommendation = {test_recommendation}")
        if test_info and test_info.get("transformation"):
            print(f"DEBUG WELCH CHECK: transformed data normality = {test_info.get('normality_tests', {}).get('transformed_data', {}).get('is_normal', False)}")
            print(f"DEBUG WELCH CHECK: transformed variance equal = {test_info.get('variance_test', {}).get('transformed', {}).get('equal_variance', True)}")
        elif test_info:
            print(f"DEBUG WELCH CHECK: original data normality = {test_info.get('normality_tests', {}).get('all_data', {}).get('is_normal', False)}")
            print(f"DEBUG WELCH CHECK: original variance equal = {test_info.get('variance_test', {}).get('equal_variance', True)}")
        else:
            print(f"DEBUG WELCH CHECK: test_info is None")
        print(f"DEBUG WELCH CHECK: welch_condition = {welch_condition}")
        
        if welch_condition:
            # Use Welch ANOVA on transformed data
            print("DEBUG WELCH CHECK: Using Welch ANOVA!")
            welch_result = StatisticalTester._welch_anova_test(results, valid_groups, samples_to_use, alpha)
            if welch_result is not None:
                return welch_result
            else:
                print("DEBUG WELCH CHECK: Welch ANOVA returned None, falling back to regular ANOVA")
                # Fall through to regular ANOVA
        
        try:
            if dependent and len(valid_groups) > 2:
                results["test"] = "Repeated Measures ANOVA not supported in simple pipeline"
                results["error"] = "Please use perform_advanced_test for RM-ANOVA."
                return StatisticalTester._standardize_results(results)
            else:
                if test_recommendation == "parametric":
                    try:
                        print("DEBUG: Attempting to use Pingouin for ANOVA...")
                        import pingouin as pg
                        import pandas as pd
                        import numpy as np
                        
                        # Create a completely fresh DataFrame with proper numeric types
                        data_for_anova = []
                        
                        # Use standard column names that won't confuse Pingouin
                        for i, group in enumerate(valid_groups):
                            values = samples_to_use[group]
                            for val in values:
                                data_for_anova.append({
                                    'dependent_var': float(val),
                                    'group_var': i  # Integer group identifier
                                })
                        
                        # Convert to DataFrame with explicit dtypes
                        df_pg = pd.DataFrame(data_for_anova)
                        df_pg['dependent_var'] = df_pg['dependent_var'].astype(float)
                        df_pg['group_var'] = df_pg['group_var'].astype('category')
                        
                        print(f"DEBUG: New df_pg shape = {df_pg.shape}")
                        print(f"DEBUG: New groups = {df_pg['group_var'].unique()}")
                        print(f"DEBUG: New dependent_var min={df_pg['dependent_var'].min()}, max={df_pg['dependent_var'].max()}")
                        
                        # Run ANOVA with robust column names
                        aov = pg.anova(data=df_pg, dv='dependent_var', between='group_var', detailed=True)
                        results["anova_table"] = aov.copy()
                        
                        # Extract results more carefully
                        print(f"DEBUG: ANOVA table columns: {list(aov.columns)}")
                        print(f"DEBUG: ANOVA table index/rows: {list(aov.index)}")
                        
                        # Get the between-groups row (should be the first row)
                        row_between = aov.iloc[0]
                        
                        # Get residuals row (should be the second row)
                        if len(aov) > 1:
                            row_residual = aov.iloc[1]
                            df2 = row_residual['DF']  # This should be numeric already
                            results["df2"] = int(df2) if pd.notnull(df2) else None
                        else:
                            # Fallback: calculate residual df from total observations - groups
                            total_observations = sum(len(samples_to_use[g]) for g in valid_groups)
                            results["df2"] = total_observations - len(valid_groups)
                        
                        # Extract the main ANOVA results
                        results["test"] = "One-way ANOVA (Pingouin)"
                        results["statistic"] = float(row_between["F"])
                        results["p_value"] = float(row_between["p-unc"])
                        results["effect_size"] = float(row_between["np2"])
                        results["effect_size_type"] = "partial_eta_squared"
                        results["confidence_interval"] = (None, None)
                        results["power"] = None
                        
                        # Map group identifiers back to names for easier reading
                        group_map = {i: group for i, group in enumerate(valid_groups)}
                        
                        print("DEBUG: Successfully used Pingouin for ANOVA!")
                        
                    except Exception as e:
                        print(f"DEBUG: Pingouin failed with error: {str(e)}")
                        # Detailed traceback for better debugging
                        import traceback
                        print(f"DEBUG: Full traceback: {traceback.format_exc()}")
                        # Fallback to scipy if Pingouin fails
                        teststat, pval = stats.f_oneway(*[samples_to_use[g] for g in valid_groups])
                        results["test"] = "One-way ANOVA (scipy fallback)"
                        results["p_value"] = pval
                        results["statistic"] = teststat
                        all_data = np.concatenate([samples_to_use[g] for g in valid_groups])
                        grand_mean = np.mean(all_data)
                        ss_between = sum(len(samples_to_use[g]) * (np.mean(samples_to_use[g]) - grand_mean) ** 2 for g in valid_groups)
                        ss_total = sum((x - grand_mean) ** 2 for x in all_data)
                        eta_sq = ss_between / ss_total if ss_total > 0 else None
                        results["effect_size"] = eta_sq
                        results["effect_size_type"] = "eta_squared"
                        results["anova_table"] = None
                        try:
                            from statsmodels.stats.power import FTestAnovaPower
                            k = len(valid_groups)
                            n = sum(len(samples_to_use[g]) for g in valid_groups)
                            eta_sq = results.get("effect_size", 0)
                            f2 = eta_sq / (1 - eta_sq) if eta_sq < 1 else 0
                            power_analysis = FTestAnovaPower()
                            results["power"] = float(power_analysis.power(
                                effect_size=f2, k_groups=k, nobs=n, alpha=alpha
                            ))
                        except Exception:
                            results["power"] = None
                        results["confidence_interval"] = (None, None)
                else:
                    teststat, pval = stats.kruskal(*[samples_to_use[g] for g in valid_groups])
                    results["test"] = "Kruskal-Wallis test"
                    results["p_value"] = pval
                    results["statistic"] = teststat
                    n = sum(len(samples_to_use[g]) for g in valid_groups)
                    h = teststat
                    k = len(valid_groups)
                    epsilon_sq = (h - k + 1) / (n - k) if n > k else None
                    results["effect_size"] = epsilon_sq
                    results["effect_size_type"] = "epsilon_squared"
                    results["anova_table"] = None
                    try:
                        from statsmodels.stats.power import FTestAnovaPower
                        k = len(valid_groups)
                        n = sum(len(samples_to_use[g]) for g in valid_groups)
                        epsilon_sq = results.get("effect_size", 0)
                        f2_approx = (epsilon_sq / (1 - epsilon_sq)) * 0.955 if epsilon_sq < 1 else 0
                        power_analysis = FTestAnovaPower()
                        results["power"] = float(power_analysis.power(
                            effect_size=f2_approx, k_groups=k, nobs=n, alpha=alpha
                        ))
                    except Exception:
                        results["power"] = None
                    results["confidence_interval"] = (None, None)
            if test_recommendation == "parametric" and results.get("p_value") is not None and results["p_value"] < alpha:
                posthoc_choice = UIDialogManager.select_posthoc_test_dialog(
                    parent=None, progress_text=None, column_name=None
                )
                if posthoc_choice and posthoc_choice != "none":
                    control_group = None
                    if posthoc_choice == "dunnett":
                        control_group = UIDialogManager.select_control_group_dialog(valid_groups)
                    posthoc_results = StatisticalTester.perform_refactored_posthoc_testing(
                        valid_groups, samples_to_use, test_recommendation="parametric", alpha=alpha,
                        posthoc_choice=posthoc_choice, control_group=control_group
                    )
                    results["posthoc_test"] = posthoc_results.get("posthoc_test")
                    results["pairwise_comparisons"] = posthoc_results.get("pairwise_comparisons", [])
            elif test_recommendation == "non_parametric" and results.get("p_value") is not None and results["p_value"] < alpha:
                # Default to Dunn test for non-parametric post-hoc tests
                posthoc_choice = "dunn"
                analysis_log = results.get("analysis_log", [])
                analysis_log += "\nAutomatically selected Dunn test as post-hoc for non-parametric test.\n"
                results["analysis_log"] = analysis_log
                
                posthoc_results = StatisticalTester.perform_refactored_posthoc_testing(
                    valid_groups, samples_to_use, test_recommendation,
                    alpha=0.05, posthoc_choice=posthoc_choice
                )
                
                if posthoc_results:
                    results["posthoc_test"] = posthoc_results.get("posthoc_test")
                    results["pairwise_comparisons"] = posthoc_results.get("pairwise_comparisons", [])
                else:
                    results["posthoc_test"] = "No post-hoc tests performed (error or unsupported test)"
                    results["pairwise_comparisons"] = []
        except Exception as e:
            results["test"] = "Error during test"
            results["p_value"] = None
            results["statistic"] = None
            results["effect_size"] = None
            results["effect_size_type"] = None
            results["confidence_interval"] = None
            results["power"] = None
            results["posthoc_test"] = "Not performed (error in main test)"
            results["pairwise_comparisons"] = []
            results["error"] = str(e)
            for key in ["test", "p_value", "statistic", "effect_size", "effect_size_type"]:
                if key not in results or results[key] is None:
                    results[key] = None
            return results
        return StatisticalTester._standardize_results(results)

    @staticmethod
    def _welch_anova_test(results, valid_groups, samples_to_use, alpha):
        """
        Perform Welch's ANOVA (for unequal variances).
        
        Parameters:
        -----------
        results : dict
            Results dictionary to store the output
        valid_groups : list
            List of group identifiers
        samples_to_use : list or dict
            Data samples for each group
        alpha : float
            Significance level
            
        Returns:
        --------
        dict
            Updated results dictionary with Welch ANOVA results
        """
        try:
            import pingouin as pg
            import pandas as pd
            import numpy as np
            
            # Prepare data for pingouin
            data_for_pingouin = []
            
            # Handle both list and dict format for samples_to_use
            if isinstance(samples_to_use, dict):
                for group in valid_groups:
                    if group in samples_to_use:
                        for value in samples_to_use[group]:
                            data_for_pingouin.append({
                                'value': float(value),
                                'group': group
                            })
            else:  # Assume it's a list of lists, matching valid_groups order
                for i, (group, samples) in enumerate(zip(valid_groups, samples_to_use)):
                    for value in samples:
                        data_for_pingouin.append({
                            'value': float(value),
                            'group': group
                        })
            
            # Create DataFrame for pingouin
            df_pg = pd.DataFrame(data_for_pingouin)
            
            # Skip test if not enough groups or data
            if len(df_pg['group'].unique()) < 2:
                results["test"] = "Welch's ANOVA (failed)"
                results["error"] = "At least two groups with data are required for Welch's ANOVA"
                results["effects"] = []
                return results
                
            print(f"DEBUG: Welch ANOVA - df_pg shape = {df_pg.shape}")
            print(f"DEBUG: Welch ANOVA - groups = {df_pg['group'].unique()}")
            
            # Perform Welch's ANOVA using pingouin
            welch_results = pg.welch_anova(data=df_pg, dv='value', between='group')
            
            # Extract results
            F_value = float(welch_results['F'].iloc[0])
            p_value = float(welch_results['p-unc'].iloc[0])
            df1 = float(welch_results['ddof1'].iloc[0])
            df2 = float(welch_results['ddof2'].iloc[0])
            
            # Store results in the format expected by the calling code
            results["test"] = "Welch's ANOVA"
            results["test_type"] = "parametric"
            results["error"] = None
            results["effects"] = [{
                "name": "between",
                "F": F_value,
                "p": p_value,
                "significant": p_value < alpha,
                "df_num": df1,
                "df_den": df2,
                "effect_size": None,  # Could calculate eta-squared if needed
                "ci_lower": None,
                "ci_upper": None,
                "posthoc_tests": None  # We'll add post-hoc separately if needed
            }]
            
            # Add test-level results 
            results["p_value"] = p_value
            results["statistic"] = F_value
            results["df1"] = df1
            results["df2"] = df2
            results["anova_table"] = welch_results
            
            # Post-hoc: Dunnett's T3 if significant
            if p_value < alpha:
                try:
                    print("DEBUG: Performing Dunnett's T3 post-hoc tests")
                    pairwise = StatisticalTester._perform_dunnett_t3_posthoc(
                        valid_groups, samples_to_use, alpha=alpha
                    )
                    results["posthoc_test"] = "Dunnett's T3"
                    results["pairwise_comparisons"] = pairwise
                except Exception as ph_err:
                    print(f"DEBUG: Post-hoc Dunnett's T3 failed: {str(ph_err)}")
                    results["posthoc_test"] = "No post-hoc tests performed (error)"
                    results["pairwise_comparisons"] = []
            else:
                results["posthoc_test"] = "No post-hoc tests performed (not significant)"
                results["pairwise_comparisons"] = []
            
            return results
        except Exception as e:
            import traceback
            print(f"DEBUG: Welch ANOVA failed with error: {str(e)}")
            print(f"DEBUG: Traceback: {traceback.format_exc()}")
            results["test"] = "Welch's ANOVA (failed)"
            results["test_type"] = "parametric"
            results["error"] = str(e)
            results["effects"] = []
            results["p_value"] = None
            results["statistic"] = None
            results["df1"] = None
            results["df2"] = None
            return results

    @staticmethod
    def _perform_dunnett_t3_posthoc(valid_groups, samples_to_use, alpha=0.05):
        """
        Performs Dunnett's T3 post-hoc test for unequal variances.
        
        Parameters:
        -----------
        valid_groups : list
            List of group names
        samples_to_use : dict
            Dictionary with group names as keys and lists of values as values
        alpha : float
            Significance level
            
        Returns:
        --------
        list
            List of pairwise comparison results
        """
        try:
            import numpy as np
            from scipy import stats
            from itertools import combinations
            
            pairwise_results = []
            
            # Create all possible pairs of groups
            pairs = list(combinations(valid_groups, 2))
            
            for group1, group2 in pairs:
                # Extract data for the two groups
                data1 = np.array(samples_to_use[group1])
                data2 = np.array(samples_to_use[group2])
                
                # Sample sizes
                n1 = len(data1)
                n2 = len(data2)
                
                # Calculate means
                mean1 = np.mean(data1)
                mean2 = np.mean(data2)
                
                # Calculate variances (with correction for sample size)
                var1 = np.var(data1, ddof=1)
                var2 = np.var(data2, ddof=1)
                
                # Standard error of the difference
                se = np.sqrt(var1/n1 + var2/n2)
                
                # T-statistic
                t_stat = (mean1 - mean2) / se
                
                # Degrees of freedom using Welch-Satterthwaite equation
                df_num = (var1/n1 + var2/n2)**2
                df_den = (var1/n1)**2/(n1-1) + (var2/n2)**2/(n2-1)
                df = df_num / df_den
                
                # Number of groups for the critical value calculation
                k = len(valid_groups)
                
                # Get critical value from studentized range distribution and transform for SMM
                try:
                    # For p-value: use studentized range distribution
                    # First, convert t-statistic to q-statistic format
                    q_stat = abs(t_stat) * np.sqrt(2)
                    
                    # Calculate p-value from studentized range distribution
                    # We use 1-cdf because we want P(q > |q_stat|)
                    p_value = 1 - stats.studentized_range.cdf(q_stat, k, df)
                    
                    # Get critical value
                    q_crit = stats.studentized_range.ppf(1-alpha, k, df)
                    crit_value = q_crit / np.sqrt(2)
                    
                    # Determine significance
                    significant = abs(t_stat) > crit_value
                    
                    # Calculate effect size (Cohen's d with pooled SD)
                    cohens_d = (mean1 - mean2) / np.sqrt((var1 + var2) / 2)
                    
                    # Calculate confidence interval
                    # For the CI, we use the critical value from the studentized range distribution
                    ci_lower = (mean1 - mean2) - crit_value * se
                    ci_upper = (mean1 - mean2) + crit_value * se
                    
                    # Add the result to our list
                    pairwise_results.append({
                        "group1": group1,
                        "group2": group2,
                        "test": "Dunnett's T3",
                        "statistic": float(t_stat),
                        "p_value": float(p_value),
                        "significant": significant,
                        "corrected": True,
                        "effect_size": float(cohens_d),
                        "effect_size_type": "cohen_d",
                        "confidence_interval": (float(ci_lower), float(ci_upper))
                    })
                    
                except Exception as err:
                    print(f"WARNING: Error calculating critical value: {str(err)}")
                    # Fallback: use t-distribution (conservative)
                    p_value = 2 * (1 - stats.t.cdf(abs(t_stat), df))
                    significant = p_value < (alpha / len(pairs))  # Bonferroni correction
                    
                    pairwise_results.append({
                        "group1": group1,
                        "group2": group2,
                        "test": "Dunnett's T3 (t approximation)",
                        "statistic": float(t_stat),
                        "p_value": float(p_value),
                        "significant": significant,
                        "corrected": True,
                        "effect_size": None,
                        "effect_size_type": None,
                        "confidence_interval": (None, None)
                    })
            
            return pairwise_results
        
        except Exception as e:
            print(f"ERROR in Dunnett's T3 post-hoc: {str(e)}")
            import traceback
            traceback.print_exc()
            return []

    @staticmethod
    def _compute_descriptive_stats(values):
        if not values or len(values) == 0:
            return {"n": 0, "mean": None, "median": None, "std": None, "stderr": None, "min": None, "max": None}
        arr = np.array(values)
        n = arr.size
        mean = np.mean(arr)
        std = np.std(arr, ddof=1) if n > 1 else 0
        stderr = std / np.sqrt(n) if n > 1 else 0
        # 95% confidence interval
        from scipy.stats import t
        if n > 1:
            ci = t.interval(0.95, n-1, loc=mean, scale=stderr)
        else:
            ci = (None, None)
        return {
            "n": n,
            "mean": mean,
            "median": np.median(arr),
            "std": std,
            "stderr": stderr,
            "ci_lower": ci[0],
            "ci_upper": ci[1],
            "min": np.min(arr),
            "max": np.max(arr)
        }
    
    @staticmethod
    def validate_dependent_data(samples, groups):
        """
        Checks whether the data are suitable for dependent tests.
        
        Parameters:
        -----------
        samples : dict
            Dictionary with group names as keys and lists of measurements as values
        groups : list
            List of groups to analyze
            
        Returns:
        --------
        dict
            Validation results with status and error messages if applicable
        """
        validation = {"valid": True, "messages": []}
        
        # Check if all specified groups exist
        for group in groups:
            if group not in samples:
                validation["valid"] = False
                validation["messages"].append(f"Group '{group}' not found in the data.")
        
        if not validation["valid"]:
            return validation
        
        # Check sample sizes
        sample_sizes = [len(samples[g]) for g in groups]
        if len(set(sample_sizes)) > 1:
            validation["valid"] = False
            validation["messages"].append(
                f"Unequal sample sizes: {', '.join([f'{g}: {len(samples[g])}' for g in groups])}"
            )
            validation["messages"].append(
                "For dependent tests, all groups must have the same number of measurements."
            )
        
        # Check for empty groups
        for group in groups:
            if len(samples[group]) == 0:
                validation["valid"] = False
                validation["messages"].append(f"Group '{group}' contains no data.")
        
        # Check minimum number of measurements
        min_samples = min(sample_sizes) if sample_sizes else 0
        if min_samples < 3:
            validation["valid"] = False
            validation["messages"].append(
                f"Too few measurements ({min_samples}). At least 3 measurements per group are recommended."
            )
        
        return validation
    
    @staticmethod
    def prepare_advanced_test(df, test, dv, subject, between=None, within=None):
        """
        Prepares an advanced statistical test by checking the assumptions.
        Returns test_info, recommendation, samples and groups.
        """
        recommendation = 'parametric'
        
        try:
            # 1. Extract group data
            samples = {}
            groups = []
    
            if test == 'mixed_anova':
                if not between or not within:
                    return {"error": "Mixed ANOVA requires between and within factor"}
                b_factor, w_factor = between[0], within[0]
                for b_val in df[b_factor].unique():
                    for w_val in df[w_factor].unique():
                        group_label = f"{b_factor}={b_val}, {w_factor}={w_val}"
                        subset = df[(df[b_factor] == b_val) & (df[w_factor] == w_val)]
                        samples[group_label] = subset[dv].tolist()
                groups = list(samples.keys())
    
            elif test == 'repeated_measures_anova':
                if not within:
                    return {"error": "RM-ANOVA requires within factor"}
                w_factor = within[0]
                for lvl in df[w_factor].unique():
                    samples[lvl] = df[df[w_factor] == lvl][dv].tolist()
                groups = list(samples.keys())
    
            elif test == 'two_way_anova':
                if not between or len(between) != 2:
                    return {"error": "Two-Way ANOVA requires two between factors"}
                fA, fB = between
                for a_val in df[fA].unique():
                    for b_val in df[fB].unique():
                        group_label = f"{fA}={a_val}, {fB}={b_val}"
                        subset = df[(df[fA] == a_val) & (df[fB] == b_val)]
                        samples[group_label] = subset[dv].tolist()
                groups = list(samples.keys())
    
            else:
                return {"error": f"Unknown test type: {test}"}
    
            transformed_samples, recommendation, test_info = StatisticalTester.check_normality_and_variance(
                groups, samples, dataset_name=dv,
                progress_text=f"{test}",  # Add this parameter for consistent dialog caching
                column_name=dv            # Add this parameter for consistent dialog caching
            )

            return {
                "test_info": test_info,
                "recommendation": recommendation,
                "transformed_samples": transformed_samples,
                "samples": samples,
                "groups": groups
            }

        except Exception as e:
            return {"error": str(e)} 
    
    @staticmethod
    def perform_advanced_test(
        df, test, dv, subject, between=None, within=None, alpha=0.05,
        transformed_samples=None, recommendation=None, test_info=None,
        transform_fn=None, force_parametric=False, skip_excel=False, file_name=None, manual_transform=None,
        analysis_log=None  # Add this parameter
    ):
        # Initialize analysis_log if None
        if analysis_log is None:
            analysis_log = []
        """
        Performs advanced statistical tests (Mixed ANOVA, RM-ANOVA, Two-Way ANOVA).
        Checks assumptions, applies transformation if necessary, performs main and post-hoc tests.
        Returns a complete result dict.
        """
        from scipy import stats
        from datetime import datetime

        try:
            # 1. Extract group data
            samples = {}
            groups = []
            df_original = df.copy()

            if test == 'mixed_anova':
                # Mixed ANOVA extraction code...
                if not between or not within:
                    return {"error": "Mixed ANOVA requires between and within factor", "test": test}
                b_factor, w_factor = between[0], within[0]
                for b_val in df[b_factor].unique():
                    for w_val in df[w_factor].unique():
                        group_label = f"{b_factor}={b_val}, {w_factor}={w_val}"
                        subset = df[(df[b_factor] == b_val) & (df[w_factor] == w_val)]
                        samples[group_label] = subset[dv].tolist()
                groups = list(samples.keys())

            elif test == 'repeated_measures_anova':
                if not within or not subject:
                    return {"error": "RM-ANOVA requires within factor and subject", "test": test}
                w_factor = within[0]
                for lvl in df[w_factor].unique():
                    samples[lvl] = df[df[w_factor] == lvl][dv].tolist()
                groups = list(samples.keys())
                group_sizes = [len(samples[g]) for g in groups]
                if len(set(group_sizes)) > 1:
                    return {
                        "error": "For dependent samples, all groups must have the same number of measurements.",
                        "test": "Repeated Measures ANOVA (failed)"
                    }

            elif test == 'two_way_anova':
                # Two-way ANOVA extraction code...
                if not between or len(between) != 2:
                    return {"error": "Two-Way ANOVA requires two between factors", "test": test}
                fA, fB = between
                for a_val in df[fA].unique():
                    for b_val in df[fB].unique():
                        group_label = f"{fA}={a_val}, {fB}={b_val}"
                        subset = df[(df[fA] == a_val) & (df[fB] == b_val)]
                        samples[group_label] = subset[dv].tolist()
                groups = list(samples.keys())

            else:
                return {"error": f"Invalid test type: {test}", "test": test}

            # CRITICAL FIX: Store the original samples BEFORE any transformation
            original_samples = {k: v.copy() for k, v in samples.items()}

            # 2. Check assumptions - BUT ONLY IF NOT ALREADY DONE
            # If transformation was manually specified or already applied, skip the dialog
            transformation_already_applied = manual_transform is not None or transformed_samples is not None
            
            if transformed_samples is None or recommendation is None:
                # We haven't performed checks yet, need to do them
                
                # If manual_transform is provided, we should bypass showing the dialog
                # by directly setting up the test_info and returning the appropriate values
                if manual_transform is not None:
                    print(f"DEBUG: Using manually specified transformation '{manual_transform}'")
                    
                    # Initialize test_info if it doesn't exist
                    if test_info is None:
                        test_info = {
                            "normality_tests": {"all_data": {}, "transformed_data": {}},
                            "variance_test": {},
                            "transformation": manual_transform
                        }
                    else:
                        # Just ensure the transformation is correctly set
                        test_info["transformation"] = manual_transform
                    
                    # Apply the manually selected transformation to generate transformed_samples
                    transformed_samples = {k: v.copy() for k, v in samples.items()}
                    
                    # FIRST: Run tests on ORIGINAL data
                    try:
                        # Combine all raw data for overall normality test
                        all_values = [value for group in groups for value in samples[group]]
                        if len(all_values) >= 3 and len(set(all_values)) > 1:
                            stat, pval = stats.shapiro(all_values)
                            is_normal = pval > 0.05
                            test_info["normality_tests"]["all_data"] = {
                                "statistic": stat, "p_value": pval, "is_normal": is_normal
                            }
                            
                        # Run variance test on raw data
                        if len(groups) >= 2:
                            data_for_levene = [samples[g] for g in groups]
                            if all(len(v) >= 3 for v in data_for_levene):
                                stat, pval = stats.levene(*data_for_levene)
                                has_equal_variance = pval > 0.05
                                test_info["variance_test"] = {
                                    "statistic": stat, 
                                    "p_value": pval, 
                                    "equal_variance": has_equal_variance
                                }
                    except Exception as e:
                        print(f"DEBUG: Error in original data tests: {str(e)}")
                    
                    # SECOND: Apply the transformation
                    if manual_transform == "log10":
                        # Apply log transformation to each group
                        min_val = min([min(samples[g]) for g in groups])
                        shift = -min_val + 1 if min_val <= 0 else 0
                        for group in groups:
                            transformed_samples[group] = [np.log10(v + shift) for v in samples[group]]
                        print(f"DEBUG: Applied log10 transformation with shift {shift}")
                    elif manual_transform == "boxcox":
                        # Apply Box-Cox transformation
                        min_val = min([min(samples[g]) for g in groups])
                        shift = -min_val + 1 if min_val <= 0 else 0
                        for group in groups:
                            shifted = [v + shift for v in samples[group]]
                            try:
                                lambda_val = boxcox_normmax(shifted)
                                transformed_samples[group] = list(boxcox(shifted, lambda_val))
                                test_info["boxcox_lambda"] = lambda_val
                            except Exception as e:
                                print(f"DEBUG: Error in Box-Cox, falling back to log10: {str(e)}")
                                transformed_samples[group] = [np.log10(v + shift) for v in samples[group]]
                    elif manual_transform == "arcsin_sqrt":
                        # Apply arcsin square root transformation
                        for group in groups:
                            values = samples[group]
                            min_val = min(values)
                            max_val = max(values)
                            # Scale values to 0-1 range if needed
                            if min_val < 0 or max_val > 1:
                                scaled = [(v - min_val) / (max_val - min_val) for v in values]
                            else:
                                scaled = values
                            transformed_samples[group] = [np.arcsin(np.sqrt(v)) for v in scaled]
                    
                    # THIRD: Test the transformed data
                    try:
                        # Combine all transformed data for overall normality test
                        all_values_trans = [value for group in groups for value in transformed_samples[group]]
                        if len(all_values_trans) >= 3 and len(set(all_values_trans)) > 1:
                            stat, pval = stats.shapiro(all_values_trans)
                            is_normal = pval > 0.05
                            test_info["normality_tests"]["transformed_data"] = {
                                "statistic": stat, "p_value": pval, "is_normal": is_normal
                            }
                            
                        # Run variance test on transformed data
                        if len(groups) >= 2:
                            data_for_levene = [transformed_samples[g] for g in groups]
                            if all(len(v) >= 3 for v in data_for_levene):
                                stat, pval = stats.levene(*data_for_levene)
                                has_equal_variance = pval > 0.05
                                test_info["variance_test"]["transformed"] = {
                                    "statistic": stat, 
                                    "p_value": pval, 
                                    "equal_variance": has_equal_variance
                                }
                    except Exception as e:
                        print(f"DEBUG: Error in transformed data tests: {str(e)}")
                    
                    recommendation = "parametric" if manual_transform != "none" else "non_parametric"
                    # But still check normality after transformation to be sure
                    if "normality_tests" in test_info and "transformed_data" in test_info["normality_tests"]:
                        is_normal = test_info["normality_tests"]["transformed_data"].get("is_normal", False)
                        if not is_normal:
                            print(f"DEBUG: Data still not normal after manual transformation, using non_parametric")
                            recommendation = "non_parametric"
                    
                    # Cache this transformation choice in the dialog manager to prevent showing the dialog again
                    if not hasattr(UIDialogManager.select_transformation_dialog, '_shown_dialogs'):
                        UIDialogManager.select_transformation_dialog._shown_dialogs = {}
                        
                    dialog_key = f"for Advanced Test_{dv}"
                    UIDialogManager.select_transformation_dialog._shown_dialogs[dialog_key] = manual_transform
                    print(f"DEBUG: Cached transformation choice '{manual_transform}' with key '{dialog_key}'")
                else:
                    # Only call check_normality_and_variance if no manual transform is provided
                    transformed_samples, recommendation, test_info = StatisticalTester.check_normality_and_variance(
                        groups, samples, dataset_name=dv, 
                        already_transformed=False,
                        progress_text=f"for Advanced Test",  # Use consistent key format that matches open_advanced_tests
                        column_name=dv
                    )
            
            print("DEBUG: transformed_samples =", transformed_samples)
            valid_groups = [g for g in groups if g in transformed_samples and len(transformed_samples[g]) > 0]
            print("DEBUG: valid_groups =", valid_groups)
            print("DEBUG: recommendation =", recommendation)

            df_transformed = df.copy()

            # 3. Apply transformation (automatically or manually, but never twice!)
            transformation_type = None
            if manual_transform is not None:
                transformation_type = manual_transform
                # Ensure this is also set in test_info for consistent reporting
                if test_info:
                    test_info["transformation"] = manual_transform
            elif test_info and test_info.get("transformation"):
                transformation_type = test_info["transformation"]

            # IMPORTANT: Only apply transformation if we have a valid transformation type
            # and it's not "none" or "None"
            if transformation_type and transformation_type not in ["none", "None", "Keine"]:
                if transformation_type == "log10":
                    min_val = df[dv].min()
                    shift = -min_val + 1 if min_val <= 0 else 0
                    df_transformed[dv] = np.log10(df[dv] + shift)
                    print(f"DEBUG: Applied log10 transformation with shift {shift}")
                elif transformation_type == "boxcox":
                    min_val = df[dv].min()
                    shift = -min_val + 1 if min_val <= 0 else 0
                    if shift > 0:
                        df_transformed[dv] = df[dv] + shift
                    lambda_val = test_info.get("boxcox_lambda")
                    if lambda_val is None:
                        from scipy.stats import boxcox_normmax
                        lambda_val = boxcox_normmax(df_transformed[dv])
                    from scipy.stats import boxcox
                    df_transformed[dv] = boxcox(df_transformed[dv], lambda_val)
                elif transformation_type == "arcsin_sqrt":
                    min_val = df[dv].min()
                    max_val = df[dv].max()
                    if min_val < 0 or max_val > 1:
                        df_transformed[dv] = (df[dv] - min_val) / (max_val - min_val)
                    df_transformed[dv] = np.arcsin(np.sqrt(df_transformed[dv]))

                transformed_values = df_transformed[dv].values
                print("DEBUG: Transformed data statistics:")
                print(f"- Min: {np.min(transformed_values)}")
                print(f"- Max: {np.max(transformed_values)}")
                print(f"- Contains NaN: {np.isnan(transformed_values).any()}")
                print(f"- Contains Inf: {np.isinf(transformed_values).any()}")

                # Extract transformed samples again
                samples_for_transform = {}
                if test == 'mixed_anova':
                    # Extract for mixed ANOVA...
                    b_factor, w_factor = between[0], within[0]
                    for b_val in df_transformed[b_factor].unique():
                        for w_val in df_transformed[w_factor].unique():
                            group_label = f"{b_factor}={b_val}, {w_factor}={w_val}"
                            subset = df_transformed[(df_transformed[b_factor] == b_val) & (df_transformed[w_factor] == w_val)]
                            samples_for_transform[group_label] = subset[dv].tolist()
                elif test == 'repeated_measures_anova':
                    # Extract for RM ANOVA...
                    w_factor = within[0]
                    for lvl in df_transformed[w_factor].unique():
                        samples_for_transform[lvl] = df_transformed[df_transformed[w_factor] == lvl][dv].tolist()
                elif test == 'two_way_anova':
                    # Extract for two-way ANOVA...
                    fA, fB = between
                    for a_val in df_transformed[fA].unique():
                        for b_val in df_transformed[fB].unique():
                            group_label = f"{fA}={a_val}, {fB}={b_val}"
                            subset = df_transformed[(df_transformed[fA] == a_val) & (df_transformed[fB] == b_val)]
                            samples_for_transform[group_label] = subset[dv].tolist()
                transformed_samples = samples_for_transform
            else:
                # No transformation applied, use original data
                print("DEBUG: No transformation applied, using original data")

            # 4. Perform test
            result = {"test_info": test_info, "recommendation": recommendation}
            # To:
            if force_parametric:
                print(f"DEBUG: User explicitly forced parametric test, overriding recommendation '{recommendation}'")
                recommendation = 'parametric'
            else:
                # Honor recommendation from normality tests
                print(f"DEBUG: Using recommendation from normality tests: '{recommendation}'")
                # Double-check normality for extra safety
                if test_info and "normality_tests" in test_info:
                    if "transformed_data" in test_info["normality_tests"]:
                        is_normal = test_info["normality_tests"]["transformed_data"].get("is_normal", False)
                        if not is_normal:
                            print(f"DEBUG: Data is NOT normal after transformation, forcing non_parametric")
                            recommendation = "non_parametric"

            if recommendation == 'parametric':
                if test == 'mixed_anova':
                    res = StatisticalTester._run_mixed_anova_logged(df_transformed, dv, subject, between, within, alpha)
                elif test == 'repeated_measures_anova':
                    res = StatisticalTester._run_repeated_measures_anova_logged(
                        df_transformed, dv, subject, within, alpha,
                        test_info=test_info  # Pass test_info here
                    )
                else:
                    res = StatisticalTester._run_two_way_anova_logged(df_transformed, dv, between, alpha)
                res.update(result)
                # Set test info at top level
                if test_info:
                    if "test_info" not in res:
                        res["test_info"] = test_info
                    else:
                        # Merge test_info into existing test_info
                        for key, value in test_info.items():
                            if key not in res["test_info"]:
                                res["test_info"][key] = value

                # Ensure normality and variance tests are directly accessible 
                if "test_info" in res:
                    ti = res["test_info"]
                    if "normality_tests" in ti and "normality_tests" not in res:
                        res["normality_tests"] = ti["normality_tests"]
                    if "variance_test" in ti and "variance_test" not in res:
                        res["variance_test"] = ti["variance_test"]
                    if "transformation" in ti and "transformation" not in res:
                        res["transformation"] = ti["transformation"]
                    if "boxcox_lambda" in ti and "boxcox_lambda" not in res:
                        res["boxcox_lambda"] = ti["boxcox_lambda"]

                # --- POST-HOC for all advanced tests ---
                if res.get("p_value") is not None and res["p_value"] < alpha:
                    if test == "two_way_anova":
                        posthoc = PostHocFactory.perform_posthoc_for_anova(
                            "two_way", df=df_transformed, dv=dv, between=between, alpha=alpha
                        )
                    elif test == "mixed_anova":
                        posthoc = PostHocFactory.perform_posthoc_for_anova(
                            "mixed", df=df_transformed, dv=dv, subject=subject, between=between, within=within, alpha=alpha
                        )
                    elif test == "repeated_measures_anova":
                        posthoc = PostHocFactory.perform_posthoc_for_anova(
                            "rm", df=df_transformed, dv=dv, subject=subject, within=within, alpha=alpha
                        )
                    else:
                        posthoc = None
                    if posthoc and "pairwise_comparisons" in posthoc:
                        res["pairwise_comparisons"] = posthoc["pairwise_comparisons"]
                        res["posthoc_test"] = posthoc.get("posthoc_test")

                # CRITICAL FIX: Always store the original and transformed samples as separate entities
                res["raw_data"] = original_samples
                if transformation_type and transformation_type not in ["none", "None", "Keine"]:
                    res["raw_data_transformed"] = transformed_samples

                # Excel export
                if not skip_excel:
                    excel_file = file_name if file_name else get_output_path(f"{test}_{datetime.now().strftime('%Y%m%d_%H%M%S')}", "xlsx")
                    print("DEBUG: Results dict before Excel export:", res)  # <--- Add this line
                    ResultsExporter.export_results_to_excel(res, excel_file, res.get("analysis_log", None))
                    res["excel_file"] = excel_file
                return res

            else:  # non-parametric path
                # DISABLED: Nonparametric fallbacks are not yet supported
                # When parametric assumptions are violated, provide informational message
                # instead of automatically running nonparametric alternatives
                test_names = {
                    'two_way_anova': 'NonParametricTwoWayANOVA (available in nonparametricanovas.py)',
                    'repeated_measures_anova': 'NonParametricRMANOVA (available in nonparametricanovas.py)',
                    'mixed_anova': 'NonParametricMixedANOVA (available in nonparametricanovas.py)'
                }
                
                suggested_test = test_names.get(test, 'Non-parametric alternative')
                
                result = {
                    "test": f"{test} (parametric assumptions violated)",
                    "test_info": test_info,
                    "recommendation": "non_parametric",
                    "error": f"Nonparametric alternatives are currently disabled. The {suggested_test} class is available in nonparametricanovas.py but nonparametric fallbacks are disabled for this analysis.",
                    "parametric_violated": True,
                    "suggested_alternative": suggested_test
                }
                
                # Generate analysis log for the information message
                if not skip_excel:
                    analysis_log = []
                    analysis_log.append(f"Advanced Test Analysis: {test}")
                    analysis_log.append(f"Dataset: {dv}")
                    analysis_log.append(f"Test recommendation: {recommendation}")
                    if transformation_type:
                        analysis_log.append(f"Applied transformation: {transformation_type}")
                    analysis_log.append(f"Result: Parametric assumptions violated")
                    analysis_log.append(f"Suggested alternative: {suggested_test} class (available in nonparametricanovas.py)")
                    analysis_log.append("Note: Nonparametric alternatives are currently disabled for this analysis")
                    
                    analysis_log_text = "\n".join(analysis_log)
                    
                    # Create output file name
                    output_file = f"{file_name}_results.xlsx" if file_name else get_output_path(
                        f"{test}_{datetime.now().strftime('%Y%m%d_%H%M%S')}", 
                        "xlsx"
                    )
                    
                    try:
                        ResultsExporter.export_results_to_excel(result, output_file, analysis_log_text)
                        result["excel_file"] = output_file
                        print(f"Results exported to: {output_file}")
                    except Exception as export_e:
                        print(f"Excel export could not be performed: {export_e}")
                        result["excel_export_error"] = str(export_e)
                
                return result
                
                # COMMENTED OUT: Original nonparametric fallback code
                # print("DEBUG TREE: Taking non-parametric path - using NonParametricFactory")
                # try:
                #     # Map ANOVA test types to non-parametric design types
                #     np_design_type = {
                #         'two_way_anova': 'two_way',
                #         'repeated_measures_anova': 'rm_anova',
                #         'mixed_anova': 'mixed_anova'
                #     }.get(test)
                #     print(f"DEBUG: Using non-parametric analysis with design type: {np_design_type}")
                #     print(f"DEBUG: Creating non-parametric test from factory...")
                #     
                #     if np_design_type:
                #         # Create the appropriate non-parametric test
                #         np_test = NonParametricFactory.create_nonparametric_test(np_design_type)
                #         
                #         # Run the test with appropriate parameters based on type
                #         if test == 'two_way_anova':
                #             result = np_test.run(df=df_transformed, dv=dv, factors=between)
                #         elif test == 'repeated_measures_anova':
                #             result = np_test.run(df=df_transformed, dv=dv, subject=subject, within=within)
                #         elif test == 'mixed_anova':
                #             result = np_test.run(df=df_transformed, dv=dv, subject=subject, between=between, within=within)
                #             
                #         # Standardize the results for export
                #         result = NonParametricFactory.standardize_results_for_export(result)
                #         
                #         # Add info about the transformation if one was applied
                #         if transformation_type:
                #             result["transformation"] = transformation_type
                #             
                #         # Add test info
                #         result["test_info"] = test_info
                #         result["recommendation"] = "non_parametric"
                #         
                #     else:
                #         # Fallback if unknown test type
                #         test_names = {
                #             'two_way_anova': 'Scheirer-Ray-Hare test',
                #             'repeated_measures_anova': 'Friedman test',
                #             'mixed_anova': 'Aligned Rank Transform ANOVA'
                #         }
                #         
                #         result = {
                #             "test": f"{test_names.get(test, 'Non-parametric alternative')} (not performed)",
                #             "test_info": test_info,
                #             "recommendation": "non_parametric",
                #             "error": "Unknown test type for non-parametric analysis"
                #         }
                #         
                #     # FIXED: Excel export moved inside the try block
                #     if not skip_excel:
                #         # Generate analysis log
                #         analysis_log = []
                #         analysis_log.append(f"Advanced Test Analysis: {test}")
                #         analysis_log.append(f"Dataset: {dv}")
                #         analysis_log.append(f"Test recommendation: {recommendation}")
                #         if transformation_type:
                #             analysis_log.append(f"Applied transformation: {transformation_type}")
                #         
                #         test_name = result.get("test", test)
                #         p_value = result.get("p_value")
                #         if p_value is not None:
                #             if p_value < alpha:
                #                 analysis_log.append(f"Result: Significant difference found (p = {p_value:.4f})")
                #             else:
                #                 analysis_log.append(f"Result: No significant difference (p = {p_value:.4f})")
                #         else:
                #             analysis_log.append("Result: Test could not be completed")
                #         
                #         analysis_log_text = "\n".join(analysis_log)
                #         
                #         # Create output file name
                #         output_file = f"{file_name}_results.xlsx" if file_name else get_output_path(
                #             f"{test}_{datetime.now().strftime('%Y%m%d_%H%M%S')}", 
                #             "xlsx"
                #         )
                #         
                #         try:
                #             ResultsExporter.export_results_to_excel(result, output_file, analysis_log_text)
                #             result["excel_file"] = output_file
                #             print(f"Results exported to: {output_file}")
                #         except Exception as export_e:
                #             print(f"Excel export could not be performed: {export_e}")
                #             result["excel_export_error"] = str(export_e)
                #     
                #     return result
                #     
                # except Exception as e:
                #     import traceback
                #     print(f"ERROR in non-parametric test execution: {str(e)}")
                #     traceback.print_exc()
                #     
                #     # FIXED: Error handling moved inside the except block where 'e' is defined
                #     result = {
                #         "test": f"Non-parametric {test} (failed)",
                #         "test_info": test_info,
                #         "recommendation": "non_parametric",
                #         "error": f"Error running non-parametric test: {str(e)}"
                #     }
                #     
                #     # Initialize an analysis log for the error case
                #     analysis_log = [
                #         f"Error in non-parametric test execution: {str(e)}",
                #         "The analysis could not be completed successfully.",
                #         f"Test type: {test}",
                #         f"Recommendation: {recommendation}"
                #     ]
                #     analysis_log = "\n".join(analysis_log)
                #     
                #     if not skip_excel:
                #         output_file = f"{file_name}_results.xlsx" if file_name else get_output_path(f"{test}_{datetime.now().strftime('%Y%m%d_%H%M%S')}", "xlsx")
                #         try:
                #             ResultsExporter.export_results_to_excel(result, output_file, analysis_log)
                #             result["excel_file"] = output_file
                #         except Exception as ee:
                #             print(f"Excel export could not be performed: {ee}")
                #     
                #     return result
                
        except Exception as e:
            import traceback
            print(f"ERROR in perform_advanced_test: {str(e)}")
            traceback.print_exc()
            return {
                "error": f"Error performing the test: {str(e)}",
                "test": f"{test} (failed)",
                "p_value": None,
                "statistic": None
            }
            
    @staticmethod
    def _run_any_parametric_test(
        df, dv, subject=None, between=None, within=None,
        alpha=0.05, test_func=None, extract_raw=None, test_info=None, **kwargs
    ):
        # 1. Initialize logger
        log_messages = []
        def log_step(msg):
            log_messages.append(msg)

        # Add logging about assumptions from test_info if available
        if test_info:
            log_step("Starting with test_info from preceding normality and variance tests.")
            
            # Log normality test results
            if "normality_tests" in test_info:
                norm_tests = test_info["normality_tests"]
                if "all_data" in norm_tests and "p_value" in norm_tests["all_data"]:
                    p_val = norm_tests["all_data"]["p_value"]
                    is_normal = norm_tests["all_data"].get("is_normal", False)
                    log_step(f"Original normality test: p = {p_val:.4f} - {'Normal' if is_normal else 'Not normal'}")
                    
                if "transformed_data" in norm_tests and "p_value" in norm_tests["transformed_data"]:
                    p_val = norm_tests["transformed_data"]["p_value"]
                    is_normal = norm_tests["transformed_data"].get("is_normal", False)
                    log_step(f"Transformed normality test: p = {p_val:.4f} - {'Normal' if is_normal else 'Not normal'}")
            
            # Log variance test results
            if "variance_test" in test_info:
                var_test = test_info["variance_test"]
                if "p_value" in var_test:
                    p_val = var_test["p_value"]
                    is_equal = var_test.get("equal_variance", False)
                    log_step(f"Original variance test: p = {p_val:.4f} - {'Equal' if is_equal else 'Not equal'}")
                    
                if "transformed" in var_test and "p_value" in var_test["transformed"]:
                    p_val = var_test["transformed"]["p_value"] 
                    is_equal = var_test["transformed"].get("equal_variance", False)
                    log_step(f"Transformed variance test: p = {p_val:.4f} - {'Equal' if is_equal else 'Not equal'}")

        log_step(f"Start parametric test: {test_func.__name__}")
        log_step("Check test assumptions (normality, variance)...")
        log_step("Performing test...")

        # Determine which parameters the function actually accepts
        import inspect
        sig = inspect.signature(test_func)
        valid_params = sig.parameters.keys()
        
        # Create a dictionary with the required parameters
        test_params = {}
        if 'df' in valid_params: test_params['df'] = df
        if 'dv' in valid_params: test_params['dv'] = dv
        if 'subject' in valid_params and subject is not None: test_params['subject'] = subject
        if 'between' in valid_params and between is not None: test_params['between'] = between
        if 'within' in valid_params and within is not None: test_params['within'] = within
        if 'alpha' in valid_params: test_params['alpha'] = alpha
        if 'test_info' in valid_params and test_info is not None: test_params['test_info'] = test_info
        
        # Filter kwargs to only include parameters accepted by the test function
        filtered_kwargs = {k: v for k, v in kwargs.items() if k in valid_params}
        test_params.update(filtered_kwargs)  # Add filtered additional parameters
        
        # 3. Run test with only the required parameters
        results = test_func(**test_params)

        # 4. Post-hoc tests if significant
        p = results.get("p_value")
        if p is not None and p < alpha:
            log_step("Significant result (p={:.3f}). Performing post-hoc tests...".format(p))

        # 5. Extract raw data
        if extract_raw:
            log_step("Extracting raw data for DV and factors...")
            results["raw_data"] = extract_raw(df, dv, between, within, subject)

        # 6. Add main ANOVA results to log
        # Main and interaction effects
        if "factors" in results:
            for factor in results["factors"]:
                log_step(f"Main effect {factor['factor']}: F({factor['df1']}, {factor['df2']}) = {factor['F']:.3f}, p = {factor['p_value']:.4f}, effect size: {factor.get('effect_size', 'N/A')}")
        if "interactions" in results:
            for inter in results["interactions"]:
                log_step(f"Interaction {inter['factors'][0]} x {inter['factors'][1]}: F({inter['df1']}, {inter['df2']}) = {inter['F']:.3f}, p = {inter['p_value']:.4f}, effect size: {inter.get('effect_size', 'N/A')}")

        # Post-hoc comparisons
        if "pairwise_comparisons" in results and results["pairwise_comparisons"]:
            log_step("Post-hoc pairwise comparisons:")
            for comp in results["pairwise_comparisons"]:
                g1 = comp.get("group1", "")
                g2 = comp.get("group2", "")
                pval = comp.get("p_value", "")
                signif = "significant" if comp.get("significant", False) else "not significant"
                log_step(f"  {g1} vs {g2}: p = {pval:.4f} ({signif})")

        results["analysis_log"] = log_messages
        
        # Add test_info to results if not already there
        if test_info is not None and "test_info" not in results:
            results["test_info"] = test_info
        
        return StatisticalTester._standardize_results(results)
    
    @staticmethod
    def _run_mixed_anova_logged(df, dv, subject, between, within, alpha=0.05):
        # 'extract_raw' can be a function that extracts raw data
        return StatisticalTester._run_any_parametric_test(
            df=df,
            dv=dv,
            subject=subject,
            between=between,
            within=within,
            alpha=alpha,
            test_func=StatisticalTester._run_mixed_anova,
            extract_raw=StatisticalTester._extract_raw_data_mixed_anova
        )
    
    @staticmethod
    def _extract_raw_data_mixed_anova(df, dv, between, within, subject):
        # Example implementation: return all individual values per group
        raw = {}
        b, w = between[0], within[0]
        for b_val in df[b].unique():
            for w_val in df[w].unique():
                key = f"{b}={b_val}, {w}={w_val}"
                raw[key] = df[(df[b] == b_val) & (df[w] == w_val)][dv].tolist()
        return raw
    
    @staticmethod
    def _run_repeated_measures_anova_logged(df, dv, subject, within, alpha=0.05, force_posthoc=False, custom_posthoc_alpha=None, **kwargs):
        """Wrapper function that includes logging for repeated measures ANOVA"""
        
        # Capture the test_info parameter and pass it through
        test_info = None
        if 'test_info' in kwargs:
            test_info = kwargs.pop('test_info')
            
        results = StatisticalTester._run_any_parametric_test(
            df=df,
            dv=dv,
            subject=subject,
            between=None,
            within=within,
            alpha=alpha,
            force_posthoc=force_posthoc,
            custom_posthoc_alpha=custom_posthoc_alpha,
            test_func=StatisticalTester._run_repeated_measures_anova,
            extract_raw=StatisticalTester._extract_raw_data_rm_anova,
            test_info=test_info  # Pass the test_info parameter
        )
        
        # Ensure test_info is added to results
        if test_info is not None and "test_info" not in results:
            results["test_info"] = test_info
            
        return results
    
    @staticmethod
    def _extract_raw_data_rm_anova(df, dv, between, within, subject):
        raw = {}
        w = within[0]
        for lvl in df[w].unique():
            raw[f"{w}={lvl}"] = df[df[w] == lvl][dv].tolist()
        return raw
    
    @staticmethod
    def _run_two_way_anova_logged(df, dv, between, alpha=0.05):
        return StatisticalTester._run_any_parametric_test(
            df=df,
            dv=dv,
            subject=None,
            between=between,
            within=None,
            alpha=alpha,
            test_func=StatisticalTester._run_two_way_anova,
            extract_raw=StatisticalTester._extract_raw_data_two_way_anova
        )
    
    @staticmethod
    def _extract_raw_data_two_way_anova(df, dv, between, within, subject):
        raw = {}
        a, b = between
        for a_val in df[a].unique():
            for b_val in df[b].unique():
                key = f"{a}={a_val}, {b}={b_val}"
                raw[key] = df[(df[a] == a_val) & (df[b] == b_val)][dv].tolist()
        return raw
        
    @staticmethod
    def _run_mixed_anova(df, dv, subject, between, within, alpha=0.05):
        """
        Performs a Mixed ANOVA. Prefers pingouin, fallback to statsmodels.
        """
        import numpy as np
        results = {
            "test": "Mixed ANOVA",
            "p_value": None,
            "statistic": None,
            "effect_size": None,
            "effect_size_type": None,
            "descriptive": {},
            "factors": [],
            "interactions": [],
            "error": None
        }

        if not between or not within:
            results["error"] = "Mixed ANOVA requires both between and within factors"
            return StatisticalTester._standardize_results(results)

        between_factor = between[0]
        rm_factor = within[0]

        try:
            import pingouin as pg
            has_pingouin = True
        except ImportError:
            has_pingouin = False
            results["warning"] = "Pingouin not installed, using statsmodels"

        try:
            if has_pingouin:
                print("DEBUG: DataFrame columns:", df.columns)
                print("DEBUG: Unique values for within factor:", df[within[0]].unique())
                print("DEBUG: Unique values for subject:", df[subject].unique())
                print("DEBUG: First few rows of df:\n", df.head())
                print("DEBUG: Using Pingouin for Mixed ANOVA")
                aov = pg.mixed_anova(data=df, dv=dv, within=rm_factor, between=between_factor, subject=subject)
                results["anova_table"] = aov.copy()
                for factor in [rm_factor, between_factor]:
                    mask = aov["Source"] == factor
                    if mask.any():
                        row = aov.loc[mask].iloc[0]
                        results["factors"].append({
                            "factor": factor,
                            "type": "within" if factor == rm_factor else "between",
                            "F": float(row["F"]),
                            "p_value": float(row["p-unc"]),
                            "df1": int(row["DF1"]),
                            "df2": int(row["DF2"]),
                            "effect_size": float(row["np2"]),
                            "effect_size_type": "partial_eta_squared"
                        })
                    else:
                        results.setdefault("warnings", []).append(f"No result for factor '{factor}' found in Mixed-ANOVA.")
                
                interaction_name = f"{rm_factor} * {between_factor}"
                mask_int = aov["Source"] == interaction_name
                if mask_int.any():
                    row = aov.loc[mask_int].iloc[0]
                    interaction = {
                        "factors": [rm_factor, between_factor],
                        "F": float(row["F"]),
                        "p_value": float(row["p-unc"]),
                        "df1": int(row["DF1"]),
                        "df2": int(row["DF2"]),
                        "effect_size": float(row["np2"]),
                        "effect_size_type": "partial_eta_squared"
                    }
                    results["interactions"].append(interaction)
                    # Set top-level fields
                    results.update({
                        "p_value": interaction["p_value"],
                        "statistic": interaction["F"],
                        "effect_size": interaction["effect_size"],
                        "effect_size_type": interaction["effect_size_type"],
                        "df1": interaction["df1"],
                        "df2": interaction["df2"],
                    })
                else:
                    results.setdefault("warnings", []).append(f"No interaction '{interaction_name}' found in Mixed-ANOVA.")
                #  POST-HOC: pairwise t-tests (Bonferroni) if interaction is significant
                try:
                    # 1. Check if interaction is significant
                    int_row = aov.loc[aov["Source"] == interaction_name]
                    if not int_row.empty and float(int_row["p-unc"].iloc[0]) < alpha:
                        # Interaction is significant: t-tests for all combinations
                        ph = pg.pairwise_tests(
                            data=df,
                            dv=dv,
                            between=between_factor,
                            within=rm_factor,
                            subject=subject,
                            padjust="holm"  # Changed from "bonf" to "holm"
                        )
                        results["posthoc_test"] = "Pairwise t-tests for interaction (Holm-Sidak)"  # Changed from "Bonferroni" to "Holm-Sidak"
                        for _, r in ph.iterrows():
                            results.setdefault("pairwise_comparisons", []).append({
                                "group1": f"{between_factor}={r['A']}, {rm_factor}={r['Time']}",
                                "group2": f"{between_factor}={r['B']}, {rm_factor}={r['Time']}",
                                "test": "Paired t-test" if r['Type'] == 'within' else "Independent t-test",
                                "statistic": float(r["T"]),
                                "p_value": float(r["p-corr"]),
                                "significant": bool(r["significant"]),
                                "corrected": True,
                                "effect_size": float(r["hedges"]) if "hedges" in r else None,
                                "effect_size_type": "hedges_g"
                            })
                    else:
                        # 2. Interaction not significant, check main effects
                        # Between-factor post-hoc with Tukey (if significant)
                        between_row = aov.loc[aov["Source"] == between_factor]
                        if not between_row.empty and float(between_row["p-unc"].iloc[0]) < alpha:
                            # Tukey HSD for between-factor
                            from statsmodels.stats.multicomp import pairwise_tukeyhsd
                            between_groups = df[between_factor].unique()
                            if len(between_groups) > 1:
                                tukey = pairwise_tukeyhsd(
                                    endog=df[dv],
                                    groups=df[between_factor],
                                    alpha=alpha
                                )
                                results["between_posthoc_test"] = "Tukey HSD"
                
                                # More robust way to handle various versions of statsmodels
                                try:
                                    # First try with the pairindices attribute
                                    for i in range(len(tukey.pvalues)):
                                        group1 = tukey.groupsunique[tukey.pairindices[i, 0]]
                                        group2 = tukey.groupsunique[tukey.pairindices[i, 1]]
                                        p_val = tukey.pvalues[i]
                                        is_significant = tukey.reject[i]
                                        
                                        results.setdefault("between_pairwise_comparisons", []).append({
                                            "group1": f"{between_factor}={group1}",
                                            "group2": f"{between_factor}={group2}",
                                            "test": "Tukey HSD",
                                            "p_value": float(p_val),
                                            "significant": bool(is_significant),
                                            "corrected": True
                                        })
                                except (AttributeError, IndexError):
                                    # Fall back to using summary() method, which works in newer versions
                                    summary = tukey.summary()
                                    for i in range(len(summary.data) - 1):  # Skip header row
                                        row = summary.data[i+1]
                                        group1, group2 = row[0], row[1]
                                        p_val = row[3]
                                        is_significant = row[6]  # reject column
                                        
                                        results.setdefault("between_pairwise_comparisons", []).append({
                                            "group1": f"{between_factor}={group1}",
                                            "group2": f"{between_factor}={group2}",
                                            "test": "Tukey HSD",
                                            "p_value": float(p_val),
                                            "significant": bool(is_significant),
                                            "corrected": True
                                        })

                        # Within-factor post-hoc with paired t-tests (if significant)
                        within_row = aov.loc[aov["Source"] == rm_factor]
                        if not within_row.empty and float(within_row["p-unc"].iloc[0]) < alpha:
                            # Paired t-tests for within-factor with Bonferroni
                            from itertools import combinations
                            within_groups = df[rm_factor].unique()
                            results["within_posthoc_test"] = "Paired t-tests (Holm-Sidak)"  # Changed from "Bonferroni" to "Holm-Sidak"
                            
                            # Create comparison groups
                            n_comparisons = len(list(combinations(within_groups, 2)))
                            
                            # Perform paired t-tests and store p-values for Holm-Sidak correction
                            p_values = []
                            t_stats = []
                            data_pairs = []
                            for group1, group2 in combinations(within_groups, 2):
                                # Prepare data for paired t-tests
                                data1 = df[df[rm_factor] == group1][dv].values
                                data2 = df[df[rm_factor] == group2][dv].values
                                
                                # Store data pairs for later calculations
                                data_pairs.append((group1, group2, data1, data2))
                                
                                # Calculate t-statistic and p-value
                                t_stat, p_val = stats.ttest_rel(data1, data2)
                                p_values.append(p_val)
                                t_stats.append(t_stat)

                            # Apply Holm-Sidak correction to all p-values at once
                            corrected_p_values = PostHocAnalyzer._holm_sidak_correction(p_values)

                            # Create comparison results using corrected p-values
                            for i, (group1, group2, data1, data2) in enumerate(data_pairs):
                                t_stat = t_stats[i]
                                p_val = p_values[i]  # Original p-value
                                corrected_p = corrected_p_values[i]  # Holm-Sidak corrected p-value
                                
                                # Calculate effect size (Cohen's d)
                                d = (np.mean(data1) - np.mean(data2)) / np.std(np.array(data1) - np.array(data2))
                                
                                results.setdefault("within_pairwise_comparisons", []).append({
                                    "group1": f"{rm_factor}={group1}",
                                    "group2": f"{rm_factor}={group2}",
                                    "test": "Paired t-test (Holm-Sidak)",  # Changed from "Bonferroni" to "Holm-Sidak"
                                    "statistic": float(t_stat),
                                    "p_value": float(corrected_p),
                                    "original_p": float(p_val),
                                    "significant": corrected_p < alpha,
                                    "corrected": True,
                                    "effect_size": float(d),
                                    "effect_size_type": "cohen_d"
                                })
                except Exception as ph_err:
                    results["warnings"] = results.get("warnings", []) + [f"Post-hoc failed: {ph_err}"]
                    
                try:
                    rm_factor = within[0]
                    within_levels = df[rm_factor].unique()
                    if len(within_levels) > 2:
                        try:
                            sphericity_result = pg.sphericity(df, dv=dv, subject=subject, within=rm_factor)
                            if isinstance(sphericity_result, tuple) and len(sphericity_result) == 3:
                                W, pval, spher = sphericity_result
                            elif isinstance(sphericity_result, (list, tuple)) and len(sphericity_result) == 1 and isinstance(sphericity_result[0], tuple):
                                W, pval, spher = sphericity_result[0]
                            else:
                                raise ValueError("Unexpected output from pg.sphericity")
                            has_sphericity = bool(spher)
                            results["sphericity_test"] = {
                                "W": float(W),
                                "p_value": float(pval),
                                "has_sphericity": has_sphericity
                            }
                        except Exception as e:
                            results["sphericity_test"] = {
                                "W": None,
                                "p_value": None,
                                "has_sphericity": None,
                                "note": f"Sphericity test failed: {str(e)}"
                            }
                    else:
                        results["sphericity_test"] = {
                            "W": None,
                            "p_value": None,
                            "has_sphericity": True,
                            "note": "Sphericity not relevant (only 2 levels)"
                        }
                except Exception as e:
                    # DEFINE has_sphericity HERE to ensure it exists for the later code
                    has_sphericity = True  # Default to assuming sphericity when test fails
                    # Fallback: Try to extract from ANOVA table
                    try:
                        if aov is not None and 'sphericity' in aov.columns:
                            has_sphericity = bool(aov['sphericity'].iloc[0])
                            W = float(aov['W-spher'].iloc[0]) if 'W-spher' in aov.columns and pd.notnull(aov['W-spher'].iloc[0]) else None
                            pval = float(aov['p-spher'].iloc[0]) if 'p-spher' in aov.columns and pd.notnull(aov['p-spher'].iloc[0]) else None
                            results["sphericity_test"] = {
                                "W": W, 
                                "p_value": pval,
                                "has_sphericity": has_sphericity,
                                "note": f"Extracted from ANOVA table (sphericity test failed: {str(e)})"
                            }
                        else:
                            results["sphericity_test"] = {
                                "W": None, 
                                "p_value": None,
                                "has_sphericity": has_sphericity,  # Use the default we set above
                                "note": f"Sphericity test failed: {str(e)}"
                            }
                    except Exception as inner_e:
                        results["sphericity_test"] = {
                            "W": None, 
                            "p_value": None,
                            "has_sphericity": has_sphericity,  # Use the default we set above
                            "note": f"Sphericity test failed with multiple errors"
                        }
                    
                    # Add the following section to compute and apply corrections
                    try:
                        # Only apply corrections if sphericity violation is detected
                        if has_sphericity is False:
                            interaction_name = f"{rm_factor} * {between_factor}"
                            row = aov.loc[aov["Source"] == interaction_name].iloc[0] if interaction_name in aov["Source"].values else None
                            
                            if row is not None and 'GG-eps' in row and 'HF-eps' in row:
                                gg_epsilon = float(row["GG-eps"])
                                results["sphericity_corrections"] = {
                                    "greenhouse_geisser": {
                                        "epsilon": gg_epsilon,
                                        "df1": float(row["DF1"]) * gg_epsilon,
                                        "df2": float(row["DF2"]) * gg_epsilon,
                                        "p_value": float(row["p-GG"]),
                                        "note": "Greenhouse-Geisser correction"
                                    },
                                    "huynh_feldt": {
                                        "epsilon": float(row["HF-eps"]),
                                        "df1": float(row["DF1"]) * float(row["HF-eps"]),
                                        "df2": float(row["DF2"]) * float(row["HF-eps"]),
                                        "p_value": float(row["p-HF"]),
                                        "note": "Huynh-Feldt correction"
                                    }
                                }

                                if gg_epsilon > 0.75:
                                    # For epsilon > 0.75 use Huynh-Feldt
                                    results["corrected_p_value"] = float(row["p-HF"])
                                    results["correction_used"] = "Huynh-Feldt (ε > 0.75)"
                                else:
                                    # For epsilon <= 0.75 use Greenhouse-Geisser
                                    results["corrected_p_value"] = float(row["p-GG"])
                                    results["correction_used"] = "Greenhouse-Geisser (ε ≤ 0.75)"
                            else:
                                results["sphericity_corrections"] = {
                                    "note": "Correction information not available in ANOVA table"
                                }
                        else:
                            # Sphericity is met or couldn't be determined
                            results["corrected_p_value"] = results.get("p_value")
                            results["correction_used"] = "None (sphericity met or undetermined)"
                    except Exception as corr_error:
                        results["sphericity_corrections_error"] = str(corr_error)
            else:
                print("DEBUG: Using statsmodels for Mixed ANOVA")
                # Fallback with statsmodels
                import statsmodels.api as sm
                from statsmodels.formula.api import ols
                formula = f"{dv} ~ C({between_factor}) + C({rm_factor}) + C({between_factor}):C({rm_factor})"
                model = ols(formula, data=df).fit()
                anova = sm.stats.anova_lm(model, typ=2)

                # Effect sizes not available in fallback
                for factor in [rm_factor, between_factor]:
                    row = anova.loc[f"C({factor})"]
                    results["factors"].append({
                        "factor": factor,
                        "type": "within" if factor == rm_factor else "between",
                        "F": float(row["F"]),
                        "p_value": float(row["PR(>F)"]),
                        "df1": int(row["df"]),
                        "df2": int(anova.loc["Residual", "df"]),
                        "effect_size": None,
                        "effect_size_type": None
                    })

                interaction_key = f"C({between_factor}):C({rm_factor})"
                row = anova.loc[interaction_key]
                interaction = {
                    "factors": [rm_factor, between_factor],
                    "F": float(row["F"]),
                    "p_value": float(row["PR(>F)"]),
                    "df1": int(row["df"]),
                    "df2": int(anova.loc["Residual", "df"]),
                    "effect_size": None,
                    "effect_size_type": None
                }
                results["interactions"].append(interaction)
                results.update({
                    "p_value": interaction["p_value"],
                    "statistic": interaction["F"],
                    "df1": interaction["df1"],
                    "df2": interaction["df2"],
                    "effect_size": interaction["effect_size"],
                    "effect_size_type": interaction["effect_size_type"],
                    "test": f"Mixed ANOVA ({rm_factor} * {between_factor}) [statsmodels]"
                })
        except Exception as e:
            results["error"] = str(e)

        # Descriptive statistics
        for b in df[between_factor].unique():
            for w in df[rm_factor].unique():
                subset = df[(df[between_factor] == b) & (df[rm_factor] == w)][dv]
                key = f"{between_factor}={b}, {rm_factor}={w}"
                if len(subset) > 0:
                    n = len(subset)
                    mean = float(np.mean(subset))
                    std = float(np.std(subset, ddof=1))
                    stderr = std / np.sqrt(n) if n > 0 else 0
                    
                    # Calculate confidence interval
                    ci_lower = None
                    ci_upper = None
                    if n > 1:
                        try:
                            from scipy.stats import t
                            ci_lower, ci_upper = t.interval(0.95, n-1, loc=mean, scale=stderr)
                        except Exception:
                            pass
                    
                    results["descriptive"][key] = {
                        "n": n,
                        "mean": mean,
                        "sd": std,
                        "stderr": stderr,  # Changed from 'se' to 'stderr' for consistency
                        "ci_lower": ci_lower,  # Added confidence interval
                        "ci_upper": ci_upper,  # Added confidence interval
                        "min": float(np.min(subset)),
                        "max": float(np.max(subset)),
                        "median": float(np.median(subset))
                    }
        
        # Ensure a top-level result if nothing was set above
        if results.get("p_value", None) is None:
            if results["interactions"]:
                iv = results["interactions"][0]
                results["p_value"]         = iv["p_value"]
                results["statistic"]       = iv["F"]
                results["effect_size"]     = iv.get("effect_size")
                results["effect_size_type"]= iv.get("effect_size_type")
                results["df1"]             = iv.get("df1")
                results["df2"]             = iv.get("df2")
            elif results["factors"]:
                fv = results["factors"][0]
                results["p_value"]         = fv["p_value"]
                results["statistic"]       = fv["F"]
                results["effect_size"]     = fv.get("effect_size")
                results["effect_size_type"]= fv.get("effect_size_type")
                results["df1"]             = fv.get("df1")
                results["df2"]             = fv.get("df2")
        # Ensure pairwise_comparisons exists
        if "pairwise_comparisons" not in results:
            results["pairwise_comparisons"] = []

        # Consolidate all post-hoc results into the main pairwise_comparisons array
        if "between_pairwise_comparisons" in results and results["between_pairwise_comparisons"]:
            results["pairwise_comparisons"].extend(results["between_pairwise_comparisons"])
            
        if "within_pairwise_comparisons" in results and results["within_pairwise_comparisons"]:
            results["pairwise_comparisons"].extend(results["within_pairwise_comparisons"])
            
        return StatisticalTester._standardize_results(results)          
    
    @staticmethod
    def _run_repeated_measures_anova(df, dv, subject, within, alpha=0.05):
        """
        Performs a Repeated Measures ANOVA (one or more within factors).
        Prefers pingouin, fallback to statsmodels.
        """
        import numpy as np
        results = {
            "test": "Repeated Measures ANOVA",
            "p_value": None,
            "statistic": None,
            "effect_size": None,
            "effect_size_type": None,
            "factors": [],
            "interactions": [],
            "descriptive": {},
            "error": None
        }

        try:
            import pingouin as pg
            has_pingouin = True
        except ImportError:
            has_pingouin = False
            results["warning"] = "Pingouin not installed, using statsmodels"

        try:
            if has_pingouin:
                print("DEBUG: DataFrame columns:", df.columns)
                print("DEBUG: Unique values for within factor:", df[within[0]].unique())
                print("DEBUG: Unique values for subject:", df[subject].unique())
                print("DEBUG: First few rows of df:\n", df.head())
                print("DEBUG: Using Pingouin for RM ANOVA")    
                if len(within) == 1:
                    factor = within[0]
                    # Add correction=True to apply corrections for sphericity violation
                    aov = pg.rm_anova(data=df, dv=dv, within=factor, subject=subject, detailed=True, correction=True)
                    print("DEBUG: ANOVA result:", aov)
                    print("DEBUG: Results structure:", results)
                    results["anova_table"] = aov.copy()
                    row = aov.iloc[0]
                    error_row = aov[aov["Source"] == "Error"].iloc[0]
                    results["factors"].append({
                        "factor": factor,
                        "type": "within",
                        "F": float(row["F"]),
                        "p_value": float(row["p-unc"]),
                        "df1": int(row["DF"]),
                        "df2": int(error_row["DF"]),
                        "effect_size": float(row["ng2"]) if "ng2" in row else None,
                        "effect_size_type": "partial_eta_squared"
                    })
                    results.update({
                        "p_value": float(row["p-unc"]),
                        "statistic": float(row["F"]),
                        "effect_size": float(row["ng2"]) if "ng2" in row else None,
                        "effect_size_type": "partial_eta_squared",
                        "df1": int(row["DF"]),
                        "df2": int(error_row["DF"]),
                        "test": f"Repeated Measures ANOVA ({factor})"
                    })

                    try:
                        factor = within[0]
                        within_levels = df[factor].unique()
                        if len(within_levels) > 2:
                            try:
                                sphericity_result = pg.sphericity(df, dv=dv, subject=subject, within=factor)
                                if isinstance(sphericity_result, tuple) and len(sphericity_result) == 3:
                                    W, pval, spher = sphericity_result
                                elif isinstance(sphericity_result, (list, tuple)) and len(sphericity_result) == 1 and isinstance(sphericity_result[0], tuple):
                                    W, pval, spher = sphericity_result[0]
                                else:
                                    raise ValueError("Unexpected output from pg.sphericity")
                                has_sphericity = bool(spher)
                                results["sphericity_test"] = {
                                    "W": float(W) if W is not None else None,
                                    "p_value": float(pval) if pval is not None else None,
                                    "has_sphericity": has_sphericity
                                }
                            except Exception as e:
                                results["sphericity_test"] = {
                                    "W": None,
                                    "p_value": None,
                                    "has_sphericity": None,
                                    "note": f"Sphericity test failed: {str(e)}"
                                }
                        else:
                            # Only 2 levels: sphericity is always met
                            results["sphericity_test"] = {
                                "W": None,
                                "p_value": None,
                                "has_sphericity": True,
                                "note": "Sphericity not relevant (only 2 levels)"
                            }
                    except Exception as e:
                        # Fallback: Try to extract from ANOVA table
                        try:
                            has_sphericity = True  # Default assumption
                            if aov is not None and 'sphericity' in aov.columns:
                                has_sphericity = bool(aov['sphericity'].iloc[0])
                                W = float(aov['W-spher'].iloc[0]) if 'W-spher' in aov.columns and pd.notnull(aov['W-spher'].iloc[0]) else None
                                pval = float(aov['p-spher'].iloc[0]) if 'p-spher' in aov.columns and pd.notnull(aov['p-spher'].iloc[0]) else None
                                results["sphericity_test"] = {
                                    "W": W, 
                                    "p_value": pval,
                                    "has_sphericity": has_sphericity,
                                    "note": f"Extracted from ANOVA table (sphericity test failed: {str(e)})"
                                }
                            else:
                                results["sphericity_test"] = {
                                    "W": None, 
                                    "p_value": None,
                                    "has_sphericity": has_sphericity,
                                    "note": f"Sphericity test failed: {str(e)}"
                                }
                        except Exception as inner_e:
                            results["sphericity_test"] = {
                                "W": None, 
                                "p_value": None,
                                "has_sphericity": True,  # Default to True when test fails
                                "note": f"Sphericity test failed with multiple errors"
                            }
                        
                        # Add the following section to compute and apply corrections
                        try:
                            # Only apply corrections if sphericity violation is detected
                            if not has_sphericity and 'GG-eps' in row and 'p-GG' in row:
                                gg_epsilon = float(row["GG-eps"])
                                results["sphericity_corrections"] = {
                                    "greenhouse_geisser": {
                                        "epsilon": gg_epsilon,
                                        "df1": float(row["DF"]) * gg_epsilon,
                                        "df2": float(error_row["DF"]) * gg_epsilon,
                                        "p_value": float(row["p-GG"]),
                                        "note": "Greenhouse-Geisser correction"
                                    }
                                }
                                
                                if 'HF-eps' in row and 'p-HF' in row:
                                    results["sphericity_corrections"]["huynh_feldt"] = {
                                        "epsilon": float(row["HF-eps"]),
                                        "df1": float(row["DF"]) * float(row["HF-eps"]),
                                        "df2": float(error_row["DF"]) * float(row["HF-eps"]),
                                        "p_value": float(row["p-HF"]),
                                        "note": "Huynh-Feldt correction"
                                    }

                                if gg_epsilon > 0.75 and 'p-HF' in row:
                                    # For epsilon > 0.75 use Huynh-Feldt
                                    results["corrected_p_value"] = float(row["p-HF"])
                                    results["correction_used"] = "Huynh-Feldt (ε > 0.75)"
                                else:
                                    # For epsilon <= 0.75 use Greenhouse-Geisser
                                    results["corrected_p_value"] = float(row["p-GG"])
                                    results["correction_used"] = "Greenhouse-Geisser (ε ≤ 0.75)"
                            else:
                                results["corrected_p_value"] = float(row["p-unc"])
                                results["correction_used"] = "None (sphericity met or indeterminable)"
                        except Exception as corr_error:
                            print(f"DEBUG: Error applying sphericity corrections: {str(corr_error)}")
                                
                    # Automatic post-hoc tests for significant main effect
                    if results["p_value"] is not None and results["p_value"] < alpha:
                        try:
                            # Extract data for post-hoc tests
                            factor_levels = df[factor].unique()
                            factor_data = {}
                            for level in factor_levels:
                                factor_data[level] = df[df[factor] == level][dv].tolist()
                            
                            # Perform paired t-tests with Holm-Sidak correction
                            posthoc_results = StatisticalTester.perform_dependent_posthoc_tests(
                                factor_data, list(factor_levels), alpha=alpha, parametric=True
                            )
                            print(f"DEBUG: Post-hoc for RM-ANOVA created with {len(posthoc_results.get('pairwise_comparisons', []))} comparisons")
                            results["posthoc_test"] = posthoc_results.get("posthoc_test", "Paired t-tests (Holm-Sidak)")

                            # Initialize with empty list as default
                            results["pairwise_comparisons"] = []

                            # If we got valid posthoc results, use them
                            if posthoc_results and 'pairwise_comparisons' in posthoc_results and posthoc_results['pairwise_comparisons']:
                                # Deep copy to ensure data isn't lost
                                import copy
                                results["pairwise_comparisons"] = copy.deepcopy(posthoc_results['pairwise_comparisons'])
                                print(f"DEBUG: Added {len(results['pairwise_comparisons'])} pairwise comparisons to RM-ANOVA results")
                        except Exception as ph_err:
                            results["warnings"] = results.get("warnings", []) + [f"Post-hoc failed: {ph_err}"]
                            print(f"DEBUG: Post-hoc test error: {str(ph_err)}")
                else:
                    aov = pg.rm_anova(data=df, dv=dv, within=within, subject=subject, detailed=True)
                    for _, row in aov.iterrows():
                        if "*" in row["Source"]:
                            results["interactions"].append({
                                "factors": row["Source"].split("*"),
                                "F": float(row["F"]),
                                "p_value": float(row["p-unc"]),
                                "df1": int(row["DF1"]),
                                "df2": int(row["DF2"]),
                                "effect_size": float(row["np2"]),
                                "effect_size_type": "partial_eta_squared"
                            })
                        else:
                            results["factors"].append({
                                "factor": row["Source"],
                                "type": "within",
                                "F": float(row["F"]),
                                "p_value": float(row["p-unc"]),
                                "df1": int(row["DF1"]),
                                "df2": int(row["DF2"]),
                                "effect_size": float(row["np2"]),
                                "effect_size_type": "partial_eta_squared"
                            })

                    # Best result for main output
                    best_row = aov.iloc[aov["F"].argmax()]
                    results.update({
                        "p_value": float(best_row["p-unc"]),
                        "statistic": float(best_row["F"]),
                        "effect_size": float(best_row["np2"]),
                        "effect_size_type": "partial_eta_squared",
                        "df1": int(best_row["DF1"]),
                        "df2": int(best_row["DF2"]),
                        "test": "Repeated Measures ANOVA (multiple factors)"
                    })
            else:
                print("DEBUG: Using statsmodels for RM ANOVA")
                # Only simple fallback for one factor
                if len(within) != 1:
                    results["error"] = "Multiple within factors only possible with pingouin"
                    return StatisticalTester._standardize_results(results)
                    
                factor = within[0]
                from statsmodels.formula.api import ols
                import statsmodels.api as sm
                formula = f"{dv} ~ C({factor}) + C({subject})"
                model = ols(formula, data=df).fit()
                anova = sm.stats.anova_lm(model, typ=2)
                row = anova.loc[f"C({factor})"]
                results["factors"].append({
                    "factor": factor,
                    "type": "within",
                    "F": float(row["F"]),
                    "p_value": float(row["PR(>F)"]),
                    "df1": int(row["df"]),
                    "df2": int(anova.loc['Residual', 'df']),
                    "effect_size": None,
                    "effect_size_type": None
                })
                results.update({
                    "p_value": float(row["PR(>F)"]),
                    "statistic": float(row["F"]),
                    "df1": int(row["df"]),
                    "df2": int(anova.loc['Residual', 'df']),
                    "test": f"Repeated Measures ANOVA ({factor}) [statsmodels]"
                })
        except Exception as e:
            results["error"] = str(e)
            print(f"DEBUG: Error in RM-ANOVA: {str(e)}")
            import traceback
            traceback.print_exc()

        # Descriptive statistics
        for factor in within:
            for val in df[factor].unique():
                subset = df[df[factor] == val][dv]
                n = len(subset)
                mean = float(np.mean(subset))
                std = float(np.std(subset, ddof=1))
                stderr = std / np.sqrt(n) if n > 0 else 0
                
                # Calculate confidence interval
                ci_lower = None
                ci_upper = None
                if n > 1:
                    try:
                        from scipy.stats import t
                        ci_lower, ci_upper = t.interval(0.95, n-1, loc=mean, scale=stderr)
                    except Exception:
                        pass
                
                results["descriptive"][f"{factor}={val}"] = {
                    "n": n,
                    "mean": mean,
                    "sd": std,
                    "stderr": stderr,
                    "ci_lower": ci_lower,
                    "ci_upper": ci_upper,
                    "min": float(np.min(subset)),
                    "max": float(np.max(subset)),
                    "median": float(np.median(subset))
                }

        # Ensure a top-level result if nothing was set above
        if results.get("p_value", None) is None:
            if results["interactions"]:
                iv = results["interactions"][0]
                results["p_value"]         = iv["p_value"]
                results["statistic"]       = iv["F"]
                results["effect_size"]     = iv.get("effect_size")
                results["effect_size_type"]= iv.get("effect_size_type")
                results["df1"]             = iv.get("df1")
                results["df2"]             = iv.get("df2")
            elif results["factors"]:
                fv = results["factors"][0]
                results["p_value"]         = fv["p_value"]
                results["statistic"]       = fv["F"]
                results["effect_size"]     = fv.get("effect_size")
                results["effect_size_type"]= fv.get("effect_size_type")
                results["df1"]             = fv.get("df1")
                results["df2"]             = fv.get("df2")

        return StatisticalTester._standardize_results(results)
    
    @staticmethod
    def _run_two_way_anova(df, dv, between, alpha=0.05):
        """
        Performs a Two-Way ANOVA (two between factors).

        Parameters:
        -----------
        df : pandas.DataFrame
            Data in long format
        dv : str
            Dependent variable
        between : list
            Two between factors
        alpha : float
            Significance level

        Returns:
        --------
        dict
            Results including main effects, interaction, effect sizes
        """
        import numpy as np
        import pandas as pd
        results = {
            "test": f"Two-Way ANOVA ({between[0]} * {between[1]})",
            "factors": [],
            "interactions": [],
            "p_value": None,
            "statistic": None,
            "df1": None,
            "df2": None,
            "effect_size": None,
            "effect_size_type": "partial_eta_squared",
            "descriptive": {},
            "error": None,
            "pairwise_comparisons": []
        }

        try:
            try:
                import pingouin as pg
                has_pingouin = True
            except ImportError:
                has_pingouin = False

            factor_a, factor_b = between[0], between[1]

            if has_pingouin:
                aov = pg.anova(data=df, dv=dv, between=between, detailed=True)
                results["anova_table"] = aov.copy()
                if "Residual" not in aov["Source"].values:
                    results["error"] = "Residuals not found in Pingouin ANOVA output. Cannot determine df2."
                    return StatisticalTester._standardize_results(results)

                residual_df_series = aov.loc[aov["Source"] == "Residual", "DF"]
                if residual_df_series.empty:
                    results["error"] = "Residual DF not found in Pingouin ANOVA output."
                    return StatisticalTester._standardize_results(results)
                residual_df = int(residual_df_series.iloc[0])

                # Main effects
                for factor in between:
                    if factor not in aov["Source"].values:
                        results.setdefault("warnings", []).append(f"Factor '{factor}' not found in Pingouin ANOVA output.")
                        continue
                    row = aov.loc[aov["Source"] == factor].iloc[0]
                    results["factors"].append({
                        "factor": factor,
                        "type": "between",
                        "F": float(row["F"]),
                        "p_value": float(row["p-unc"]),
                        "df1": int(row["DF"]),
                        "df2": residual_df,
                        "effect_size": float(row["np2"]),
                        "effect_size_type": "partial_eta_squared"
                    })

                # Interaction
                interaction_label = f"{factor_a} * {factor_b}"
                possible_interaction_labels = [interaction_label, f"{factor_b} * {factor_a}"]
                actual_interaction_label = None
                for label in possible_interaction_labels:
                    if label in aov["Source"].values:
                        actual_interaction_label = label
                        break

                if actual_interaction_label:
                    row = aov.loc[aov["Source"] == actual_interaction_label].iloc[0]
                    interaction_result = {
                        "factors": [factor_a, factor_b],
                        "F": float(row["F"]),
                        "p_value": float(row["p-unc"]),
                        "df1": int(row["DF"]),
                        "df2": residual_df,
                        "effect_size": float(row["np2"]),
                        "effect_size_type": "partial_eta_squared"
                    }
                    results["interactions"].append(interaction_result)
                    results["p_value"] = float(row["p-unc"])
                    results["statistic"] = float(row["F"])
                    results["df1"] = int(row["DF"])
                    results["df2"] = residual_df
                    results["effect_size"] = float(row["np2"])
                else:
                    results.setdefault("warnings", []).append(f"Interaction term for '{factor_a}' and '{factor_b}' not found in Pingouin ANOVA output.")
                    if not results["p_value"] and results["factors"]:
                        results["p_value"] = results["factors"][0]["p_value"]
                        results["statistic"] = results["factors"][0]["F"]
                        results["df1"] = results["factors"][0]["df1"]
                        results["df2"] = results["factors"][0]["df2"]
                        results["effect_size"] = results["factors"][0]["effect_size"]

                if actual_interaction_label and results["p_value"] is not None and results["p_value"] < alpha:
                    try:
                        posthoc_df = pg.pairwise_tests(data=df, dv=dv, between=between, padjust='tukey', subject=None)
                        if not posthoc_df.empty:
                            results["posthoc_test"] = "Tukey HSD for Interaction (Pingouin)"
                            for _, ph_row in posthoc_df.iterrows():
                                g1_label = str(ph_row.get('A', 'Group1'))
                                g2_label = str(ph_row.get('B', 'Group2'))
                                if 'Contrast' in ph_row and isinstance(ph_row['Contrast'], list) and len(ph_row['Contrast']) == 2:
                                    g1_label = str(ph_row['Contrast'][0])
                                    g2_label = str(ph_row['Contrast'][1])
                                elif 'Contrast' in ph_row and isinstance(ph_row['Contrast'], str) and 'vs.' in ph_row['Contrast']:
                                    parts = ph_row['Contrast'].split(' vs. ')
                                    if len(parts) == 2:
                                        g1_label = parts[0].strip()
                                        g2_label = parts[1].strip()
                                pval_col = 'p-tukey' if 'p-tukey' in ph_row else 'p-corr' if 'p-corr' in ph_row else 'p-unc'
                                confidence_interval = (None, None)
                                if 'CI95%' in ph_row and isinstance(ph_row['CI95%'], (list, np.ndarray)) and len(ph_row['CI95%']) == 2:
                                    confidence_interval = tuple(ph_row['CI95%'])
                                elif 'CI95' in ph_row and isinstance(ph_row['CI95'], (list, np.ndarray)) and len(ph_row['CI95']) == 2:
                                    confidence_interval = tuple(ph_row['CI95'])
                                elif 'CLES' in ph_row and isinstance(ph_row['CLES'], (list, np.ndarray)) and len(ph_row['CLES']) == 2:
                                    confidence_interval = tuple(ph_row['CLES'])
                                elif 'ci' in ph_row and isinstance(ph_row['ci'], (list, np.ndarray)) and len(ph_row['ci']) == 2:
                                    confidence_interval = tuple(ph_row['ci'])
                                results["pairwise_comparisons"].append({
                                    "group1": g1_label,
                                    "group2": g2_label,
                                    "test": "Tukey HSD (Pingouin)",
                                    "p_value": float(ph_row[pval_col]),
                                    "statistic": float(ph_row["T"]) if "T" in ph_row else None,
                                    "significant": float(ph_row[pval_col]) < alpha,
                                    "corrected": "tukey" in pval_col or "corr" in pval_col,
                                    "confidence_interval": confidence_interval
                                })
                        else:
                            results.setdefault("warnings", []).append("Pingouin pairwise_tests for interaction returned empty.")
                    except Exception as e_ph:
                        results.setdefault("warnings", []).append(f"Post-hoc tests (Pingouin) for interaction failed: {str(e_ph)}")

            else: # Fallback to statsmodels
                import statsmodels.api as sm
                from statsmodels.formula.api import ols
                print("DEBUG: WARNING! Two-Way ANOVA uses statsmodels fallback!")
                print("DEBUG: Pingouin not installed or import failed.")

                formula = f"`{dv}` ~ C(`{factor_a}`) * C(`{factor_b}`)"
                model = ols(formula, data=df).fit()
                aov = sm.stats.anova_lm(model, typ=2)
                if "Residual" not in aov.index:
                    results["error"] = "Residuals not found in statsmodels ANOVA output."
                    return StatisticalTester._standardize_results(results)
                residual_df = int(aov.loc["Residual", "df"])

                # Main effects
                for factor in [factor_a, factor_b]:
                    factor_term = f"C(`{factor}`)"
                    if factor_term not in aov.index:
                        results.setdefault("warnings", []).append(f"Factor term '{factor_term}' not found in statsmodels ANOVA output.")
                        continue
                    row = aov.loc[factor_term]
                    results["factors"].append({
                        "factor": factor,
                        "type": "between",
                        "F": float(row["F"]),
                        "p_value": float(row["PR(>F)"]),
                        "df1": int(row["df"]),
                        "df2": residual_df,
                        "effect_size": None,
                        "effect_size_type": None
                    })

                # Interaction
                interaction_term = f"C(`{factor_a}`):C(`{factor_b}`)"
                if interaction_term in aov.index:
                    row = aov.loc[interaction_term]
                    interaction_result = {
                        "factors": [factor_a, factor_b],
                        "F": float(row["F"]),
                        "p_value": float(row["PR(>F)"]),
                        "df1": int(row["df"]),
                        "df2": residual_df,
                        "effect_size": None,
                        "effect_size_type": None
                    }
                    results["interactions"].append(interaction_result)
                    results["p_value"] = float(row["PR(>F)"])
                    results["statistic"] = float(row["F"])
                    results["df1"] = int(row["df"])
                    results["df2"] = residual_df
                    results["effect_size"] = None
                    results["test"] += " [statsmodels]"
                else:
                    results.setdefault("warnings", []).append(f"Interaction term '{interaction_term}' not found in statsmodels ANOVA output.")
                    if not results["p_value"] and results["factors"]:
                        results["p_value"] = results["factors"][0]["p_value"]
                        results["statistic"] = results["factors"][0]["F"]
                        results["df1"] = results["factors"][0]["df1"]
                        results["df2"] = results["factors"][0]["df2"]

                if (results["p_value"] is not None and results["p_value"] < alpha) or any(factor["p_value"] < alpha for factor in results["factors"]):
                    try:
                        factor_a, factor_b = between[0], between[1]
                        df['interaction_group'] = df[factor_a].astype(str) + "_" + df[factor_b].astype(str)
                        from statsmodels.stats.multicomp import pairwise_tukeyhsd
                        tukey = pairwise_tukeyhsd(df[dv], df['interaction_group'], alpha=alpha)
                        results["posthoc_test"] = "Tukey HSD für Interaktionseffekt"
                        if "pairwise_comparisons" not in results:
                            results["pairwise_comparisons"] = []
                        for i in range(len(tukey.pvalues)):
                            group1 = tukey.groupsunique[tukey.pairindices[i, 0]]
                            group2 = tukey.groupsunique[tukey.pairindices[i, 1]]
                            p_val = tukey.pvalues[i]
                            is_significant = tukey.reject[i]
                            conf_int = tukey.confint[i]
                            group1_values = df[df['interaction_group'] == group1][dv].values
                            group2_values = df[df['interaction_group'] == group2][dv].values
                            effect_size = None
                            try:
                                from scipy import stats
                                n1, n2 = len(group1_values), len(group2_values)
                                s1, s2 = np.var(group1_values, ddof=1), np.var(group2_values, ddof=1)
                                s_pooled = np.sqrt(((n1-1)*s1 + (n2-1)*s2) / (n1+n2-2))
                                effect_size = (np.mean(group1_values) - np.mean(group2_values)) / s_pooled if s_pooled > 0 else 0
                            except Exception:
                                effect_size = None
                            results["pairwise_comparisons"].append({
                                "group1": str(group1),
                                "group2": str(group2),
                                "test": "Tukey HSD",
                                "p_value": float(p_val),
                                "significant": bool(is_significant),
                                "corrected": True,
                                "effect_size": effect_size,
                                "effect_size_type": "cohen_d",
                                "confidence_interval": tuple(conf_int)
                            })
                    except Exception as e_ph:
                        results.setdefault("warnings", []).append(f"Post-hoc Tests failed: {str(e_ph)}")

            # Descriptive statistics
            for a_val in df[factor_a].unique():
                for b_val in df[factor_b].unique():
                    group_key = f"{factor_a}={a_val}, {factor_b}={b_val}"
                    subset = df[(df[factor_a] == a_val) & (df[factor_b] == b_val)][dv]
                    if len(subset) > 0:
                        mean_val = float(np.mean(subset))
                        std_val = float(np.std(subset, ddof=1))
                        n_val = len(subset)
                        se_val = std_val / np.sqrt(n_val) if n_val > 0 else 0
                        ci_desc = (None, None)
                        if n_val > 1:
                            try:
                                from scipy.stats import t
                                ci_desc = t.interval(0.95, n_val - 1, loc=mean_val, scale=se_val)
                            except:
                                pass
                        results["descriptive"][group_key] = {
                            "n": n_val,
                            "mean": mean_val,
                            "sd": std_val,
                            "se": se_val,
                            "ci_lower": ci_desc[0],
                            "ci_upper": ci_desc[1],
                            "min": float(np.min(subset)),
                            "max": float(np.max(subset)),
                            "median": float(np.median(subset))
                        }
                    else:
                        results["descriptive"][group_key] = {
                            "n": 0, "mean": None, "sd": None, "se": None,
                            "ci_lower": None, "ci_upper": None,
                            "min": None, "max": None, "median": None
                        }

        except Exception as e:
            import traceback
            results["error"] = f"Error in Two-Way ANOVA: {str(e)}. Trace: {traceback.format_exc()}"
            results["test"] += " (failed)"

        # Ensure top-level results are set, prioritizing interaction, then first main effect.
        if results.get("p_value") is None:
            if results["interactions"] and results["interactions"][0].get("p_value") is not None:
                iv = results["interactions"][0]
                results["p_value"] = iv["p_value"]
                results["statistic"] = iv["F"]
                results["effect_size"] = iv.get("effect_size")
                results["effect_size_type"] = iv.get("effect_size_type")
                results["df1"] = iv.get("df1")
                results["df2"] = iv.get("df2")
            elif results["factors"] and results["factors"][0].get("p_value") is not None:
                fv = results["factors"][0]
                results["p_value"] = fv["p_value"]
                results["statistic"] = fv["F"]
                results["effect_size"] = fv.get("effect_size")
                results["effect_size_type"] = fv.get("effect_size_type")
                results["df1"] = fv.get("df1")
                results["df2"] = fv.get("df2")
        return StatisticalTester._standardize_results(results)
    
    @staticmethod
    def perform_dependent_posthoc_tests(data_dict, groups, alpha=0.05, parametric=True):
        """
        Performs post-hoc tests for dependent samples.
    
        Parameters:
        -----------
        data_dict : dict
            Dictionary with group names as keys and lists of values as values
        groups : list
            List of groups to analyze
        alpha : float
            Significance level
        parametric : bool
            Whether to perform parametric tests
    
        Returns:
        --------
        dict
            Results of the post-hoc tests
        """
        # In this transition phase, use the new implementation
        return StatisticalTester.perform_refactored_posthoc_testing(
            groups,
            data_dict,
            "parametric" if parametric else "non_parametric",
            alpha=alpha,
            posthoc_choice="dependent",
            is_dependent=True
        )
    @staticmethod
    def perform_refactored_posthoc_testing(valid_groups, samples, test_recommendation, alpha=0.05, posthoc_choice=None, control_group=None, is_dependent=False):
        """
        Central function for performing post-hoc tests with the new framework.
        Can be used as a replacement for the existing perform_posthoc_testing.

        Parameters:
        -----------
        valid_groups : list
            List of groups to analyze
        samples : dict
            Dictionary with group names as keys and lists of values
        test_recommendation : str
            "parametric" or "non_parametric", based on normality tests
        alpha : float, optional
            Significance level (default: 0.05)
        posthoc_choice : str, optional
            Specific post-hoc test: "tukey", "dunnett", "dunn" or None
        control_group : str, optional
            Control group for Dunnett test
        is_dependent : bool, optional
            Indicates whether samples are dependent

        Returns:
        --------
        dict
            Dictionary with post-hoc test results and pairwise comparisons
        """
        # Initialize default result
        result = {
            "posthoc_test": None,
            "pairwise_comparisons": [],
            "error": None
        }

        # Check validity
        if len(valid_groups) <= 1:
            result["error"] = "At least two groups are required for post-hoc tests."
            return result

        # Automatic test selection if not explicitly specified
        if posthoc_choice is None:
            if is_dependent:
                posthoc_choice = "dependent"
            elif test_recommendation == "parametric":
                posthoc_choice = "tukey"
            else:
                posthoc_choice = "dunn"

        try:
            is_parametric = test_recommendation == "parametric"

            # Create the appropriate test
            if posthoc_choice == "dependent":
                test_instance = PostHocFactory.create_test(None, is_parametric=is_parametric, is_dependent=True)
                if test_instance:
                    return test_instance.perform_test(valid_groups, samples, alpha=alpha, parametric=is_parametric)
            elif posthoc_choice == "dunnett":
                if control_group is None or control_group not in valid_groups:
                    result["error"] = "A valid control group must be specified for the Dunnett test."
                    return result
                test_instance = PostHocFactory.create_test("dunnett", is_parametric=is_parametric, is_dependent=False)
                if test_instance:
                    return test_instance.perform_test(valid_groups, samples, control_group=control_group, alpha=alpha)
            else:
                # tukey or dunn
                test_instance = PostHocFactory.create_test(posthoc_choice, is_parametric=is_parametric, is_dependent=False)
                if test_instance:
                    return test_instance.perform_test(valid_groups, samples, alpha=alpha)

            # If no suitable implementation was found
            result["error"] = f"No suitable test available for {posthoc_choice} (parametric: {is_parametric}, dependent: {is_dependent})"
            return result
        except Exception as e:
            import traceback
            result["error"] = f"Error performing post-hoc test: {str(e)}"
            traceback.print_exc()
            return result

    def process_results(results):
        print("Processing results:")
        print(f"Keys in results: {list(results.keys())}")
        if 'interactions' in results:
            print(f"Interaction effect p-value: {results['interactions'][0]['p_value'] if results['interactions'] else 'No interaction found'}")
        if 'pairwise_comparisons' in results and results['pairwise_comparisons']:
            print("Pairwise comparisons found:")
            for comparison in results['pairwise_comparisons']:
                print(f"{comparison['group1']} vs {comparison['group2']}: p = {comparison['p_value']}")
        else:
            print("No pairwise comparisons found.")
            
class UIDialogManager:
    @staticmethod
    def select_posthoc_test_dialog(parent=None, progress_text=None, column_name=None):
        """
        Opens a dialog to select the post-hoc test.
        The window title optionally contains the column name and progress.
        Returns the name of the post-hoc test (str), or None if cancelled.
        """
        dialog = QDialog(parent)
        layout = QVBoxLayout(dialog)

        # Set window title
        title = "Select Post-hoc Test"
        if column_name:
            title += f" for '{column_name}'"
        if progress_text:
            title += f" {progress_text}"
        dialog.setWindowTitle(title)

        info = QLabel("The ANOVA has revealed significant differences. Please select a post-hoc test:")
        layout.addWidget(info)

        # RadioButtons for post-hoc tests
        options = [
            ("Tukey-HSD Test", "tukey"),
            ("Dunnett Test", "dunnett"),
            ("Do not perform a post-hoc test", "none")
        ]
        radio_buttons = []
        for label, value in options:
            rb = QRadioButton(label)
            layout.addWidget(rb)
            radio_buttons.append((rb, value))
        radio_buttons[0][0].setChecked(True)  # Default: Tukey

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        layout.addWidget(buttons)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)

        if dialog.exec_() == QDialog.Accepted:
            for rb, value in radio_buttons:
                if rb.isChecked():
                    return value
        return None

    @staticmethod
    def select_control_group_dialog(groups):
        """Opens a dialog window to select the control group"""
        from PyQt5.QtWidgets import QDialog, QVBoxLayout, QLabel, QRadioButton, QDialogButtonBox

        class ControlGroupDialog(QDialog):
            def __init__(self, available_groups, parent=None):
                super().__init__(parent)
                self.setWindowTitle("Select Control Group")
                self.selected_group = None
                self.groups = available_groups

                layout = QVBoxLayout(self)
                label = QLabel("Please select the control group for the Dunnett test:")
                layout.addWidget(label)

                self.group_buttons = []
                for group in self.groups:
                    rb = QRadioButton(str(group))
                    self.group_buttons.append(rb)
                    layout.addWidget(rb)

                # Select first group by default
                if self.group_buttons:
                    self.group_buttons[0].setChecked(True)

                button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                button_box.accepted.connect(self.accept)
                button_box.rejected.connect(self.reject)
                layout.addWidget(button_box)

            def accept(self):
                for i, button in enumerate(self.group_buttons):
                    if button.isChecked():
                        self.selected_group = self.groups[i]
                        break
                super().accept()

        # Always create a new dialog
        dialog = ControlGroupDialog(groups)
        if dialog.exec_() == QDialog.Accepted:
            return dialog.selected_group
        return groups[0]  # Default: first group
    
    @staticmethod
    def reset_dialog_cache():
        """Resets all cached dialog selections to ensure fresh prompts for new analyses."""
        if hasattr(UIDialogManager.select_transformation_dialog, '_shown_dialogs'):
            UIDialogManager.select_transformation_dialog._shown_dialogs = {}
            print("DEBUG: Dialog cache reset - transformation selection will be prompted again")

    @staticmethod
    def select_transformation_dialog(parent=None, progress_text=None, column_name=None):
        """
        Opens a dialog to select the transformation.
        Uses consistent wording in all cases.
        Returns the name of the transformation (str), or None if cancelled.
        """
        # Add memory to prevent showing the dialog multiple times for the same dataset
        if not hasattr(UIDialogManager.select_transformation_dialog, '_shown_dialogs'):
            UIDialogManager.select_transformation_dialog._shown_dialogs = {}
        
        # Create a unique key for this specific dialog with a consistent format
        # Normalize the key so the same dialog is only shown once
        # The "for Advanced Test" format is used in StatisticalAnalyzerApp.open_advanced_tests
        normalized_progress_text = progress_text
        if progress_text and ("Advanced Test" in progress_text or 
                             any(test in progress_text for test in ["mixed_anova", "two_way_anova", "repeated_measures_anova"])):
            normalized_progress_text = "for Advanced Test"
            
        dialog_key = f"{normalized_progress_text}_{column_name}"
        
        # Check if we've already shown this dialog
        if dialog_key in UIDialogManager.select_transformation_dialog._shown_dialogs:
            # Return the previously selected value without showing dialog again
            return UIDialogManager.select_transformation_dialog._shown_dialogs[dialog_key]
        
        dialog = QDialog(parent)
        layout = QVBoxLayout(dialog)

        # Set window title
        title = "Select Transformation"
        if column_name:
            title += f" for '{column_name}'"
        if progress_text:
            title += f" {progress_text}"
        dialog.setWindowTitle(title)

        # Info text - CONSISTENT WORDING
        info = QLabel("Please select the desired transformation:")
        layout.addWidget(info)

        # RadioButtons for transformations
        options = [
            ("Log10 transformation (for positive, right-skewed data)", "log10"),
            ("Box-Cox transformation (automatic lambda optimization)", "boxcox"),
            ("Arcsin square root transformation (for percentages/proportions)", "arcsin_sqrt"),
            ("No transformation (use non-parametric test)", "none")
        ]
        radio_buttons = []
        for label, value in options:
            rb = QRadioButton(label)
            layout.addWidget(rb)
            radio_buttons.append((rb, value))
        radio_buttons[0][0].setChecked(True)  # Default: Log10

        # OK/Cancel buttons
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        layout.addWidget(buttons)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)

        # Show dialog
        if dialog.exec_() == QDialog.Accepted:
            for rb, value in radio_buttons:
                if rb.isChecked():
                    # Store the selected value to avoid showing dialog again
                    UIDialogManager.select_transformation_dialog._shown_dialogs[dialog_key] = value
                    return value
        
        # If canceled, store None to avoid showing dialog again
        UIDialogManager.select_transformation_dialog._shown_dialogs[dialog_key] = None
        return None
    
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
import pandas as pd
from matplotlib.ticker import ScalarFormatter, FuncFormatter
from matplotlib.patches import Rectangle
import matplotlib.patches as mpatches
from scipy import stats

class DataVisualizer:
    """Advanced data visualization class with extensive customization options"""
    
    # Default themes/styles
    THEMES = {
        'default': {
            'colors': ['#3357FF', '#FF5733', '#33FF57', '#F033FF', '#FF3366', '#33FFEC'],
            'hatches': ['/', '\\', '|', '-', '+', 'x', 'o', '.'],
            'style': 'whitegrid'
        },
        'nature': {
            'colors': ['#2E8B57', '#CD853F', '#4682B4', '#DAA520', '#DC143C', '#9370DB'],
            'hatches': ['...', '|||', '---', '+++', '\\\\\\', '///'],
            'style': 'white'
        },
        'academic': {
            'colors': ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', '#8c564b'],
            'hatches': [None] * 6,
            'style': 'ticks'
        },
        'colorblind': {
            'colors': ['#0173B2', '#DE8F05', '#029E73', '#CC78BC', '#CA9161', '#FBAFE4'],
            'hatches': ['////', '\\\\\\\\', '||||', '----', '++++', 'xxxx'],
            'style': 'whitegrid'
        }
    }
    
    @staticmethod
    def plot_bar(groups, samples, 
                 # Basic plot settings
                 width=8, height=6, dpi=300,
                 
                 # Styling and theme
                 theme='default', colors=None, hatches=None, 
                 color_palette='husl', alpha=0.8,
                 
                 # Bar customization
                 bar_width=0.8, bar_edge_color='black', bar_edge_width=0.5,
                 capsize=0.05, error_type="sd", show_error_bars=True,
                 
                 # Data points
                 show_points=True, point_style='jitter', max_points_per_group=None,
                 point_size=40, point_alpha=0.8, point_edge_width=0.5,
                 jitter_strength=0.3, strip_dodge=False,
                 
                 # Statistical annotations
                 show_significance_letters=True, show_pairwise_comparisons=True,
                 significance_height_offset=0.05, comparison_line_height=0.1,
                 significance_font_size=12, comparison_font_size=14,
                 
                 # Axes and labels
                 x_label=None, y_label=None, title=None,
                 x_label_size=12, y_label_size=12, title_size=14,
                 tick_label_size=10, rotate_x_labels=0,
                 
                 # Axis formatting
                 y_axis_format='auto', y_limits=None, x_limits=None,
                 grid_style='major', grid_alpha=0.3,
                 
                 # Legend
                 show_legend=True, legend_position='upper right',
                 legend_bbox=(1.15, 1), legend_fontsize=9,
                 legend_title="Samples", legend_title_size=12,
                 
                 # Advanced styling
                 spine_style='default', background_color='white',
                 figure_face_color='white',
                 
                 # Output options
                 save_plot=True, file_formats=['pdf', 'svg'], 
                 file_name=None, group_order=None,
                 
                 # Statistical data
                 compare=None, test_recommendation="parametric",
                 pairwise_results=None,
                 
                 # Advanced customization
                 custom_annotations=None, watermark=None,
                 subplot_margins=None, tight_layout=True):
       
        # Apply theme
        if theme in DataVisualizer.THEMES:
            theme_config = DataVisualizer.THEMES[theme]
            if colors is None:
                colors = theme_config['colors']
            if hatches is None:
                hatches = theme_config['hatches']
            sns.set_style(theme_config['style'])
        
        # Prepare data and groups
        if group_order is not None:
            groups = [g for g in group_order if g in samples and len(samples[g]) > 0]
        else:
            groups = [g for g in groups if len(samples.get(g, [])) > 0]
        
        # Color and hatch preparation
        if colors is None:
            if color_palette:
                colors = sns.color_palette(color_palette, len(groups))
            else:
                colors = DataVisualizer.THEMES['default']['colors']
        
        colors = DataVisualizer._extend_list(colors, len(groups))
        hatches = DataVisualizer._extend_list(hatches or [''] * len(groups), len(groups))
        
        # Create figure
        fig, ax = plt.subplots(figsize=(width, height), dpi=dpi)
        fig.patch.set_facecolor(figure_face_color)
        ax.set_facecolor(background_color)
        
        # Prepare data for plotting
        plot_data = DataVisualizer._prepare_plot_data(groups, samples, colors)
        df = pd.DataFrame(plot_data)
        
        # Create bar plot
        if show_error_bars:
            bars = sns.barplot(
                x='Group', y='Value', data=df, ax=ax,
                errorbar=error_type, palette=colors, 
                capsize=capsize, alpha=alpha,
                order=groups, width=bar_width,
                edgecolor=bar_edge_color, linewidth=bar_edge_width
            )
        else:
            bars = sns.barplot(
                x='Group', y='Value', data=df, ax=ax,
                errorbar=None, palette=colors,
                order=groups, width=bar_width,
                edgecolor=bar_edge_color, linewidth=bar_edge_width,
                alpha=alpha
            )
        
        # Apply hatches
        if any(h for h in hatches):
            for i, patch in enumerate(bars.patches):
                if i < len(hatches) and hatches[i]:
                    patch.set_hatch(hatches[i])
        
        # Add data points
        if show_points:
            DataVisualizer._add_data_points(
                ax, groups, samples, point_style, point_size, 
                point_alpha, jitter_strength, max_points_per_group,
                point_edge_width, strip_dodge
            )
        
        # Format axes
        DataVisualizer._format_axes(
            ax, y_axis_format, y_limits, x_limits, 
            grid_style, grid_alpha, spine_style
        )
        
        # Add labels and title
        DataVisualizer._add_labels(
            ax, x_label, y_label, title,
            x_label_size, y_label_size, title_size,
            tick_label_size, rotate_x_labels
        )
        
        # Add statistical annotations
        if show_significance_letters:
            DataVisualizer._add_significance_letters(
                ax, df, groups, samples, test_recommendation,
                significance_height_offset, significance_font_size,
                error_type
            )
        
        if show_pairwise_comparisons and compare and pairwise_results:
            DataVisualizer._add_pairwise_comparisons(
                ax, groups, compare, pairwise_results, df,
                comparison_line_height, comparison_font_size
            )
        
        # Add legend
        if show_legend and show_points:
            DataVisualizer._add_legend(
                ax, samples, groups, legend_position, legend_bbox,
                legend_fontsize, legend_title, legend_title_size
            )
        
        # Add custom annotations
        if custom_annotations:
            DataVisualizer._add_custom_annotations(ax, custom_annotations)
        
        # Add watermark
        if watermark:
            DataVisualizer._add_watermark(fig, watermark)
        
        # Adjust layout
        if tight_layout:
            if subplot_margins:
                plt.subplots_adjust(**subplot_margins)
            else:
                fig.tight_layout()
        
        # Save plot
        if save_plot:
            DataVisualizer._save_plot(fig, file_name, groups, file_formats, dpi)
        
        return fig, ax
    
    @staticmethod
    def plot_violin(
        groups, samples,
        width=8, height=6, dpi=300,
        theme='default', colors=None, hatches=None, color_palette='husl', alpha=0.8,
        violin_width=0.8, edge_color='black', edge_width=0.5,
        show_points=True, point_style='jitter', max_points_per_group=None,
        point_size=40, point_alpha=0.8, point_edge_width=0.5,
        jitter_strength=0.3, strip_dodge=False,
        show_significance_letters=True, significance_height_offset=0.05, significance_font_size=12,
        x_label=None, y_label=None, title=None,
        x_label_size=12, y_label_size=12, title_size=14,
        tick_label_size=10, rotate_x_labels=0,
        y_axis_format='auto', y_limits=None, x_limits=None,
        grid_style='major', grid_alpha=0.3,
        show_legend=True, legend_position='upper right',
        legend_bbox=(1.15, 1), legend_fontsize=9,
        legend_title="Samples", legend_title_size=12,
        spine_style='default', background_color='white',
        figure_face_color='white',
        save_plot=True, file_formats=['pdf', 'svg'],
        file_name=None, group_order=None,
        test_recommendation="parametric",
        custom_annotations=None, watermark=None,
        subplot_margins=None, tight_layout=True
    ):
        """Creates a violin plot with extensive customization options."""
        import seaborn as sns
        import matplotlib.pyplot as plt
        import pandas as pd
        import numpy as np

        if theme in DataVisualizer.THEMES:
            theme_config = DataVisualizer.THEMES[theme]
            if colors is None:
                colors = theme_config['colors']
            if hatches is None:
                hatches = theme_config['hatches']
            sns.set_style(theme_config['style'])
        if colors is None:
            colors = sns.color_palette(color_palette, len(groups))
        colors = DataVisualizer._extend_list(colors, len(groups))

        if group_order is not None:
            groups = [g for g in group_order if g in samples and len(samples[g]) > 0]
        else:
            groups = [g for g in groups if len(samples.get(g, [])) > 0]

        plot_data = DataVisualizer._prepare_plot_data(groups, samples, colors)
        df = pd.DataFrame(plot_data)

        fig, ax = plt.subplots(figsize=(width, height), dpi=dpi)
        fig.patch.set_facecolor(figure_face_color)
        ax.set_facecolor(background_color)

        sns.violinplot(
            x='Group', y='Value', data=df, ax=ax,
            palette=colors, order=groups, width=violin_width,
            linewidth=edge_width, edgecolor=edge_color, alpha=alpha
        )

        if show_points:
            DataVisualizer._add_data_points(
                ax, groups, samples, point_style, point_size,
                point_alpha, jitter_strength, max_points_per_group,
                point_edge_width, strip_dodge
            )

        DataVisualizer._format_axes(
            ax, y_axis_format, y_limits, x_limits,
            grid_style, grid_alpha, spine_style
        )
        DataVisualizer._add_labels(
            ax, x_label, y_label, title,
            x_label_size, y_label_size, title_size,
            tick_label_size, rotate_x_labels
        )
        if show_significance_letters:
            DataVisualizer._add_significance_letters(
                ax, df, groups, samples, test_recommendation,
                significance_height_offset, significance_font_size,
                "sd"
            )
        if show_legend and show_points:
            DataVisualizer._add_legend(
                ax, samples, groups, legend_position, legend_bbox,
                legend_fontsize, legend_title, legend_title_size
            )
        if custom_annotations:
            DataVisualizer._add_custom_annotations(ax, custom_annotations)
        if watermark:
            DataVisualizer._add_watermark(fig, watermark)
        if tight_layout:
            if subplot_margins:
                plt.subplots_adjust(**subplot_margins)
            else:
                fig.tight_layout()
        if save_plot:
            DataVisualizer._save_plot(fig, file_name, groups, file_formats, dpi)
        return fig, ax

    @staticmethod
    def plot_box(
        groups, samples,
        width=8, height=6, dpi=300,
        theme='default', colors=None, hatches=None, color_palette='husl', alpha=0.8,
        box_width=0.8, edge_color='black', edge_width=0.5,
        show_points=True, point_style='jitter', max_points_per_group=None,
        point_size=40, point_alpha=0.8, point_edge_width=0.5,
        jitter_strength=0.3, strip_dodge=False,
        show_significance_letters=True, significance_height_offset=0.05, significance_font_size=12,
        x_label=None, y_label=None, title=None,
        x_label_size=12, y_label_size=12, title_size=14,
        tick_label_size=10, rotate_x_labels=0,
        y_axis_format='auto', y_limits=None, x_limits=None,
        grid_style='major', grid_alpha=0.3,
        show_legend=True, legend_position='upper right',
        legend_bbox=(1.15, 1), legend_fontsize=9,
        legend_title="Samples", legend_title_size=12,
        spine_style='default', background_color='white',
        figure_face_color='white',
        save_plot=True, file_formats=['pdf', 'svg'],
        file_name=None, group_order=None,
        test_recommendation="parametric",
        custom_annotations=None, watermark=None,
        subplot_margins=None, tight_layout=True
    ):
        """Creates a box plot with extensive customization options."""
        import seaborn as sns
        import matplotlib.pyplot as plt
        import pandas as pd
        import numpy as np

        if theme in DataVisualizer.THEMES:
            theme_config = DataVisualizer.THEMES[theme]
            if colors is None:
                colors = theme_config['colors']
            if hatches is None:
                hatches = theme_config['hatches']
            sns.set_style(theme_config['style'])
        if colors is None:
            colors = sns.color_palette(color_palette, len(groups))
        colors = DataVisualizer._extend_list(colors, len(groups))

        if group_order is not None:
            groups = [g for g in group_order if g in samples and len(samples[g]) > 0]
        else:
            groups = [g for g in groups if len(samples.get(g, [])) > 0]

        plot_data = DataVisualizer._prepare_plot_data(groups, samples, colors)
        df = pd.DataFrame(plot_data)

        fig, ax = plt.subplots(figsize=(width, height), dpi=dpi)
        fig.patch.set_facecolor(figure_face_color)
        ax.set_facecolor(background_color)

        sns.boxplot(
            x='Group', y='Value', data=df, ax=ax,
            palette=colors, order=groups, width=box_width,
            linewidth=edge_width, fliersize=0, boxprops=dict(alpha=alpha)
        )

        if show_points:
            DataVisualizer._add_data_points(
                ax, groups, samples, point_style, point_size,
                point_alpha, jitter_strength, max_points_per_group,
                point_edge_width, strip_dodge
            )

        DataVisualizer._format_axes(
            ax, y_axis_format, y_limits, x_limits,
            grid_style, grid_alpha, spine_style
        )
        DataVisualizer._add_labels(
            ax, x_label, y_label, title,
            x_label_size, y_label_size, title_size,
            tick_label_size, rotate_x_labels
        )
        if show_significance_letters:
            DataVisualizer._add_significance_letters(
                ax, df, groups, samples, test_recommendation,
                significance_height_offset, significance_font_size,
                "sd"
            )
        if show_legend and show_points:
            DataVisualizer._add_legend(
                ax, samples, groups, legend_position, legend_bbox,
                legend_fontsize, legend_title, legend_title_size
            )
        if custom_annotations:
            DataVisualizer._add_custom_annotations(ax, custom_annotations)
        if watermark:
            DataVisualizer._add_watermark(fig, watermark)
        if tight_layout:
            if subplot_margins:
                plt.subplots_adjust(**subplot_margins)
            else:
                fig.tight_layout()
        if save_plot:
            DataVisualizer._save_plot(fig, file_name, groups, file_formats, dpi)
        return fig, ax

    @staticmethod
    def plot_raincloud(
        groups, samples,
        width=8, height=6, dpi=300,
        theme='default', colors=None, hatches=None, color_palette='husl', alpha=0.8,
        violin_width=0.8, box_width=0.2, edge_color='black', edge_width=0.5,
        show_points=True, point_style='jitter', max_points_per_group=None,
        point_size=40, point_alpha=0.8, point_edge_width=0.5,
        jitter_strength=0.3, strip_dodge=False,
        show_significance_letters=True, significance_height_offset=0.05, significance_font_size=12,
        x_label=None, y_label=None, title=None,
        x_label_size=12, y_label_size=12, title_size=14,
        tick_label_size=10, rotate_x_labels=0,
        y_axis_format='auto', y_limits=None, x_limits=None,
        grid_style='major', grid_alpha=0.3,
        show_legend=True, legend_position='upper right',
        legend_bbox=(1.15, 1), legend_fontsize=9,
        legend_title="Samples", legend_title_size=12,
        spine_style='default', background_color='white',
        figure_face_color='white',
        save_plot=True, file_formats=['pdf', 'svg'],
        file_name=None, group_order=None,
        test_recommendation="parametric",
        custom_annotations=None, watermark=None,
        subplot_margins=None, tight_layout=True
    ):
        """Creates a raincloud plot (violin + box + points)."""
        import seaborn as sns
        import matplotlib.pyplot as plt
        import pandas as pd
        import numpy as np

        if theme in DataVisualizer.THEMES:
            theme_config = DataVisualizer.THEMES[theme]
            if colors is None:
                colors = theme_config['colors']
            if hatches is None:
                hatches = theme_config['hatches']
            sns.set_style(theme_config['style'])
        if colors is None:
            colors = sns.color_palette(color_palette, len(groups))
        colors = DataVisualizer._extend_list(colors, len(groups))

        if group_order is not None:
            groups = [g for g in group_order if g in samples and len(samples[g]) > 0]
        else:
            groups = [g for g in groups if len(samples.get(g, [])) > 0]

        plot_data = DataVisualizer._prepare_plot_data(groups, samples, colors)
        df = pd.DataFrame(plot_data)

        fig, ax = plt.subplots(figsize=(width, height), dpi=dpi)
        fig.patch.set_facecolor(figure_face_color)
        ax.set_facecolor(background_color)

        # Violin (half)
        for i, group in enumerate(groups):
            vals = samples[group]
            color = colors[i % len(colors)]
            sns.violinplot(
                x=[group]*len(vals), y=vals, ax=ax, palette=[color],
                width=violin_width, linewidth=edge_width, cut=0, bw=0.2,
                inner=None, order=[group], alpha=alpha
            )
            # Overlay boxplot (narrow)
            sns.boxplot(
                x=[group]*len(vals), y=vals, ax=ax, palette=[color],
                width=box_width, linewidth=edge_width, fliersize=0,
                order=[group], boxprops=dict(alpha=0.7)
            )

        if show_points:
            DataVisualizer._add_data_points(
                ax, groups, samples, point_style, point_size,
                point_alpha, jitter_strength, max_points_per_group,
                point_edge_width, strip_dodge
            )

        DataVisualizer._format_axes(
            ax, y_axis_format, y_limits, x_limits,
            grid_style, grid_alpha, spine_style
        )
        DataVisualizer._add_labels(
            ax, x_label, y_label, title,
            x_label_size, y_label_size, title_size,
            tick_label_size, rotate_x_labels
        )
        if show_significance_letters:
            DataVisualizer._add_significance_letters(
                ax, df, groups, samples, test_recommendation,
                significance_height_offset, significance_font_size,
                "sd"
            )
        if show_legend and show_points:
            DataVisualizer._add_legend(
                ax, samples, groups, legend_position, legend_bbox,
                legend_fontsize, legend_title, legend_title_size
            )
        if custom_annotations:
            DataVisualizer._add_custom_annotations(ax, custom_annotations)
        if watermark:
            DataVisualizer._add_watermark(fig, watermark)
        if tight_layout:
            if subplot_margins:
                plt.subplots_adjust(**subplot_margins)
            else:
                fig.tight_layout()
        if save_plot:
            DataVisualizer._save_plot(fig, file_name, groups, file_formats, dpi)
        return fig, ax
    
    # Helper methods
    @staticmethod
    def _extend_list(lst, target_length):
        """Extend a list to target length by repeating elements"""
        if not lst:
            return [''] * target_length
        return (lst * (target_length // len(lst) + 1))[:target_length]
    
    @staticmethod
    def _prepare_plot_data(groups, samples, colors):
        """Prepare data for plotting"""
        data = []
        for i, group in enumerate(groups):
            values = samples.get(group, [])
            for val in values:
                data.append({
                    'Group': group, 
                    'Value': val, 
                    'Color': colors[i % len(colors)]
                })
        return data
    
    @staticmethod
    def _add_data_points(ax, groups, samples, style, size, alpha, 
                        jitter_strength, max_points, edge_width, dodge):
        """Add individual data points to the plot"""
        markers = ['o', 's', 'D', '^', 'v', '<', '>', 'p', '*', 'h', '+', 'x']
        
        for i, group in enumerate(groups):
            values = samples.get(group, [])
            if max_points and len(values) > max_points:
                values = np.random.choice(values, size=max_points, replace=False)
            
            x_pos = i
            
            if style == 'jitter':
                # Custom jitter implementation
                if len(values) > 1:
                    jitter = np.random.uniform(-jitter_strength/2, jitter_strength/2, len(values))
                else:
                    jitter = [0]
                
                for j, (val, jit) in enumerate(zip(values, jitter)):
                    marker_idx = j % len(markers)
                    ax.scatter(x_pos + jit, val, color='black', 
                             marker=markers[marker_idx], s=size, 
                             zorder=3, alpha=alpha, edgecolors='white',
                             linewidth=edge_width)
            
            elif style == 'strip':
                # Use seaborn stripplot
                df_group = pd.DataFrame({'x': [i] * len(values), 'y': values})
                sns.stripplot(data=df_group, x='x', y='y', ax=ax, 
                            size=size/10, alpha=alpha, color='black',
                            jitter=jitter_strength, dodge=dodge)
            
            elif style == 'swarm':
                # Use seaborn swarmplot
                df_group = pd.DataFrame({'x': [group] * len(values), 'y': values})
                sns.swarmplot(data=df_group, x='x', y='y', ax=ax,
                            size=size/10, alpha=alpha, color='black')
    
    @staticmethod
    def _format_axes(ax, y_format, y_limits, x_limits, grid_style, grid_alpha, spine_style):
        """Format axes according to specifications"""
        # Y-axis formatting
        if y_format == 'scientific':
            formatter = ScalarFormatter(useOffset=False, useMathText=True)
            formatter.set_scientific(True)
            formatter.set_powerlimits((0, 0))
            ax.yaxis.set_major_formatter(formatter)
        elif y_format == 'percentage':
            ax.yaxis.set_major_formatter(FuncFormatter(lambda x, p: f'{x:.1%}'))
        elif y_format == 'decimal':
            ax.yaxis.set_major_formatter(FuncFormatter(lambda x, p: f'{x:.2f}'))
        
        # Set limits
        if y_limits:
            ax.set_ylim(y_limits)
        if x_limits:
            ax.set_xlim(x_limits)
        
        # Grid
        if grid_style != 'none':
            ax.grid(True, which=grid_style, alpha=grid_alpha, linestyle='-', linewidth=0.5)
        
        # Spine styling
        if spine_style == 'minimal':
            sns.despine(ax=ax, top=True, right=True)
        elif spine_style == 'none':
            sns.despine(ax=ax, top=True, right=True, left=True, bottom=True)
        elif spine_style == 'box':
            # Keep all spines
            pass
    
    @staticmethod
    def _add_labels(ax, x_label, y_label, title, x_size, y_size, title_size, tick_size, rotation):
        """Add and format labels"""
        if x_label:
            ax.set_xlabel(x_label, fontsize=x_size, fontweight='bold')
        if y_label:
            ax.set_ylabel(y_label, fontsize=y_size, fontweight='bold')
        if title:
            ax.set_title(title, fontsize=title_size, fontweight='bold', pad=20)
        
        ax.tick_params(axis='both', which='major', labelsize=tick_size)
        
        if rotation != 0:
            plt.setp(ax.get_xticklabels(), rotation=rotation, ha='right')
    
    @staticmethod
    def _add_custom_annotations(ax, annotations):
        """Add custom text annotations"""
        for annotation in annotations:
            ax.annotate(
                annotation.get('text', ''),
                xy=annotation.get('xy', (0, 0)),
                xytext=annotation.get('xytext', (0, 0)),
                fontsize=annotation.get('fontsize', 10),
                color=annotation.get('color', 'black'),
                ha=annotation.get('ha', 'center'),
                va=annotation.get('va', 'center'),
                arrowprops=annotation.get('arrowprops', None)
            )
    
    @staticmethod
    def _add_watermark(fig, watermark_text):
        """Add watermark to the figure"""
        fig.text(0.95, 0.05, watermark_text, 
                fontsize=8, color='gray', alpha=0.5,
                ha='right', va='bottom', rotation=0)
    
    @staticmethod
    def _save_plot(fig, file_name, groups, formats, dpi):
        """Save plot in multiple formats"""
        if file_name is None:
            file_name = "_".join(map(str, groups))
        
        for fmt in formats:
            if fmt == 'pdf':
                pdf_path = get_output_path(file_name, "pdf")
                fig.savefig(pdf_path, dpi=dpi, bbox_inches='tight', format='pdf')
            elif fmt == 'svg':
                svg_path = get_output_path(file_name, "svg")
                fig.savefig(svg_path, bbox_inches='tight', format='svg')
            elif fmt == 'png':
                png_path = get_output_path(file_name, "png")
                fig.savefig(png_path, dpi=dpi, bbox_inches='tight', format='png')
            elif fmt == 'eps':
                eps_path = get_output_path(file_name, "eps")
                fig.savefig(eps_path, dpi=dpi, bbox_inches='tight', format='eps')
    
    @staticmethod
    def get_significance_letters(samples, groups, test_recommendation="parametric", alpha=0.05):
        """
        Calculate significance letters for groups based on statistical comparisons.
        Groups that share a letter are not significantly different.

        Parameters:
        -----------
        samples : dict
            Dictionary with group names as keys and measurement values as lists
        groups : list
            List of groups to compare
        test_recommendation : str
            Type of test to perform ("parametric" or "non_parametric")
        alpha : float
            Significance level

        Returns:
        --------
        dict
            Dictionary with groups as keys and significance letters as values
        """
        import string
        from scipy.stats import ttest_ind, mannwhitneyu

        # Initialize all groups with 'a'
        letters = {group: 'a' for group in groups}

        # If we have only one group, return immediately
        if len(groups) <= 1:
            return letters

        # Create matrix of p-values
        p_values = {}
        for i, group1 in enumerate(groups):
            for j, group2 in enumerate(groups):
                if i < j:  # Only compute once per pair
                    if test_recommendation == "parametric":
                        _, p_val = ttest_ind(samples[group1], samples[group2])
                    else:
                        _, p_val = mannwhitneyu(samples[group1], samples[group2], alternative='two-sided')
                    p_values[(group1, group2)] = p_val

        # Assign letters based on significant differences
        available_letters = list(string.ascii_lowercase)
        current_letter_idx = 0

        # Start assigning from the first group
        for i, group1 in enumerate(groups):
            # If this is the first group or it's significantly different from all previous,
            # we might need a new letter
            needs_new_letter = True

            # Check all previous groups
            for j, group2 in enumerate(groups[:i]):
                # If current group is NOT significantly different from a previous group
                pair = (group2, group1) if (group2, group1) in p_values else (group1, group2)
                if p_values[pair] >= alpha:  # Not significant
                    # Use the same letter as that group
                    if letters[group2] not in letters[group1]:
                        letters[group1] += letters[group2]
                    needs_new_letter = False

            # If we need a new letter and haven't exhausted the alphabet
            if needs_new_letter and current_letter_idx < len(available_letters):
                current_letter_idx += 1
                if current_letter_idx < len(available_letters):
                    letters[group1] = available_letters[current_letter_idx]

        return letters
    
    @staticmethod
    def _add_significance_letters(ax, df, groups, samples, test_recommendation, 
                                height_offset, font_size, error_type):
        """Add significance letters with enhanced formatting"""
        try:
            letters = DataVisualizer.get_significance_letters(
                samples, groups, test_recommendation=test_recommendation
            )
            
            y_max = df['Value'].max()
            y_offset = height_offset * y_max
            
            # Calculate bar heights with error bars
            bar_heights = []
            for group in groups:
                values = samples[group]
                mean_val = np.mean(values)
                if error_type == 'sd':
                    error = np.std(values, ddof=1)
                elif error_type == 'se':
                    error = np.std(values, ddof=1) / np.sqrt(len(values))
                else:  # ci
                    error = 1.96 * np.std(values, ddof=1) / np.sqrt(len(values))
                bar_heights.append(mean_val + error)
            
            # Place letters with enhanced styling
            for i, group in enumerate(groups):
                letter = letters[group]
                ax.text(i, bar_heights[i] + y_offset, letter,
                       horizontalalignment='center', 
                       verticalalignment='bottom',
                       color='black', fontweight='bold',
                       fontsize=font_size,
                       bbox=dict(boxstyle="round,pad=0.3", 
                                facecolor="white", 
                                edgecolor="gray",
                                alpha=0.8))
        except Exception as e:
            print(f"Error adding significance letters: {str(e)}")    
    
class ResultsExporter:
    _temp_files = set()
    @staticmethod
    def export_results_to_excel(results, output_file, analysis_log=None):
        import xlsxwriter
        import copy
        
        # Create a deep copy to prevent modifications during processing
        results_copy = copy.deepcopy(results)
        
        # Ensure pairwise_comparisons exists and is a list
        if 'pairwise_comparisons' not in results_copy:
            print("WARNING: No pairwise comparisons found, initializing empty list")
            results_copy['pairwise_comparisons'] = []
        elif not isinstance(results_copy['pairwise_comparisons'], list):
            print(f"WARNING: pairwise_comparisons is not a list, type: {type(results_copy['pairwise_comparisons'])}")
            results_copy['pairwise_comparisons'] = []
        
        print(f"DEBUG: Before Excel export - number of pairwise comparisons: {len(results_copy.get('pairwise_comparisons', []))}")
        
        # Initialize dataset_tree_paths for single dataset export
        dataset_tree_paths = {}
        
        workbook = xlsxwriter.Workbook(output_file, {'nan_inf_to_errors': True})
        fmt = ResultsExporter._get_excel_formats(workbook)

        ResultsExporter._write_summary_sheet(workbook, results, fmt)
        ResultsExporter._write_assumptions_sheet(workbook, results, fmt)
        ResultsExporter._write_results_sheet(workbook, results, fmt)
        ResultsExporter._write_descriptive_sheet(workbook, results, fmt)
        ResultsExporter._write_decision_tree_sheet(workbook, results, fmt)
        ResultsExporter._write_rawdata_sheet(workbook, results, fmt)
        ResultsExporter._write_pairwise_sheet(workbook, results, fmt)
        if analysis_log:
            ResultsExporter._write_analysislog_sheet(workbook, analysis_log, fmt)
            
        workbook.close()

        # Clean up all temporary decision tree files
        for dataset_name, tree_path in dataset_tree_paths.items():
            if tree_path and os.path.exists(tree_path):
                try:
                    os.remove(tree_path)
                    print(f"DEBUG MULTI: Cleaned up decision tree file for {dataset_name}: {tree_path}")
                except Exception as e:
                    print(f"DEBUG MULTI: Could not clean up {tree_path}: {str(e)}")

        # Clean up any other tracked temporary files
        if ResultsExporter._temp_files:
            print(f"DEBUG MULTI: Cleaning up {len(ResultsExporter._temp_files)} tracked temporary files")
            for temp_file in ResultsExporter._temp_files:
                if temp_file and os.path.exists(temp_file):
                    try:
                        os.remove(temp_file)
                        print(f"DEBUG MULTI: Removed tracked temp file: {temp_file}")
                    except Exception as e:
                        print(f"DEBUG MULTI: Failed to remove temp file: {str(e)}")
            ResultsExporter._temp_files.clear()

    @staticmethod
    def export_multi_dataset_results(all_results, excel_path):
        print(f"DEBUG MULTI: export_multi_dataset_results called with excel_path='{excel_path}'")
        print("DEBUG MULTI: Received all_results with contents:")   
        for ds_name, results in all_results.items():
            print(f"  Dataset: {ds_name} → Keys in results: {list(results.keys())}")
            print(f"    p_value: {results.get('p_value')} | pairwise_comparisons: {len(results.get('pairwise_comparisons', []))}")
        
        """Exports the results of all dataset analyses into a shared Excel file."""
        import os
        import time
        import xlsxwriter
        from decisiontreevisualizer import DecisionTreeVisualizer
        
        # Create a dictionary to track all decision tree images for this multi-dataset export
        dataset_tree_paths = {}
        
        # Generate all decision trees first, before creating the workbook
        for dataset_name, results in all_results.items():
            print(f"DEBUG MULTI: Pre-generating decision tree for {dataset_name}...")
            # Generate decision tree and track the file path
            tree_path = DecisionTreeVisualizer.generate_and_save_for_excel(results)
            if tree_path and os.path.exists(tree_path):
                print(f"DEBUG MULTI: Generated decision tree for {dataset_name}: {tree_path}")
                # Store in dictionary mapping dataset to file path
                dataset_tree_paths[dataset_name] = tree_path
            else:
                print(f"DEBUG MULTI: Warning - Failed to generate decision tree for {dataset_name}")
        
        # Now create the Excel workbook with all necessary formats
        workbook = xlsxwriter.Workbook(excel_path, {'nan_inf_to_errors': True})
        fmt = ResultsExporter._get_excel_formats(workbook)
        
        # DEBUG: Print available format keys
        print(f"DEBUG MULTI: Available format keys: {list(fmt.keys())}")
        
        # Create an overview sheet
        overview_sheet = workbook.add_worksheet("Overview")
        overview_sheet.set_column('A:A', 30)
        overview_sheet.set_column('B:E', 15)
        
        # Write overview headers
        overview_sheet.write(0, 0, "Dataset", fmt["header"])
        overview_sheet.write(0, 1, "Test", fmt["header"])
        overview_sheet.write(0, 2, "p-value", fmt["header"])
        overview_sheet.write(0, 3, "Significant", fmt["header"])
        overview_sheet.write(0, 4, "Transformation", fmt["header"])
        
        # For each dataset: write overview row with basic info
        row = 1
        for dataset_name, results in all_results.items():
            overview_sheet.write(row, 0, str(dataset_name), fmt["header"])
            overview_sheet.write(row, 1, str(results.get("test", "N/A")), fmt["cell"])
            
            p_value = results.get("p_value", None)
            if p_value is not None and isinstance(p_value, (float, int)):
                if p_value < 0.001:
                    overview_sheet.write(row, 2, "<0.001", fmt["cell"])
                else:
                    overview_sheet.write(row, 2, f"{p_value:.4f}", fmt["cell"])
            else:
                overview_sheet.write(row, 2, "N/A", fmt["cell"])
            
            is_significant = p_value is not None and isinstance(p_value, (float, int)) and p_value < 0.05
            sig_fmt = fmt["significant"] if is_significant else fmt["cell"]
            overview_sheet.write(row, 3, "Yes" if is_significant else "No", sig_fmt)
            
            transformation = results.get("transformation", "None")
            overview_sheet.write(row, 4, str(transformation), fmt["cell"])
            
            row += 1
            
        # Add detailed information for each dataset
        row += 2  # Add some space
        for dataset_name, results in all_results.items():
            # Dataset header
            overview_sheet.merge_range(f'A{row}:E{row}', f"DATASET: {dataset_name}", fmt["title"])
            row += 1
            
            # RAW DATA section
            overview_sheet.merge_range(f'A{row}:E{row}', "RAW DATA", fmt["section_header"])
            row += 1
            overview_sheet.write(row, 0, "These data are the basis of all calculations.", fmt["explanation"])
            row += 1  # FIX: Changed from row += 2 to row += 1 to prevent misalignment
            
            # Get raw data for this dataset
            raw_data = results.get("raw_data", results.get("original_data", {})) or {}
            print("DEBUG: raw_data keys:", list(raw_data.keys()))

            # Filtere evtl. "Group"-Key raus
            data_to_write = {k: v for k, v in raw_data.items() if k.lower() not in ["group", "sample", ""]}          

            row += 1  # die Zeile, in der gleich Group & Values stehen sollen

            overview_sheet.write(row, 0, "Group", fmt["header"])
            overview_sheet.write(row, 1, "Values", fmt["header"])
            row += 1
            for group_name, values in data_to_write.items():
                # 4) Gruppe in Spalte A
                overview_sheet.write(row, 0, group_name, fmt["cell"])
                # 5) Werte-String in Spalte B
                values_str = ", ".join([
                    f"{v:.4f}" if isinstance(v, (int, float)) else str(v)
                    for v in values
                ])
                overview_sheet.write(row, 1, values_str, fmt["cell"])
                row += 1

            # Get raw data for this dataset and apply new alignment function
            raw_data = results.get("raw_data", results.get("original_data", {})) or {}

            print(f"DEBUG: Processing raw data for {dataset_name}")
            print(f"DEBUG: Raw data keys: {list(raw_data.keys())}")

            # TRANSFORMED DATA section for this dataset
            transformed_data = results.get("raw_data_transformed", results.get("transformed_data", {})) or {}
            transformation = results.get("transformation", "None")
            print("DEBUG: transformed_data keys:", list(transformed_data.keys()))
            # Only show transformed data if a transformation was performed
            if transformed_data and transformation and transformation.lower() != "none":
                print(f"DEBUG: Processing transformed data for {dataset_name}")
                print(f"DEBUG: Transformed data keys: {list(transformed_data.keys())}")
                
                # Use the same alignment function for transformed data
                transformed_to_write = transformed_data
                
                # Check if transformed data actually differ from raw data
                is_different = False
                if data_to_write and transformed_to_write:

                    if set(data_to_write.keys()) != set(transformed_to_write.keys()):
                        is_different = True
                    else:
                        for group in data_to_write:
                            if group in transformed_to_write:
 
                                raw_vals = data_to_write[group]
                                trans_vals = transformed_to_write[group]
                                if len(raw_vals) != len(trans_vals):
                                    is_different = True
                                    break

                                for r, t in zip(raw_vals, trans_vals):
                                    if abs(r - t) > 1e-10:
                                        is_different = True
                                        break
                                if is_different:
                                    break
                
                if is_different:
                    row += 1
                    transformed_to_write = {k: v for k, v in transformed_data.items() if k.lower() not in ["group", "sample", ""]}
                    overview_sheet.merge_range(f'A{row}:E{row}', "TRANSFORMED DATA", fmt["section_header"])
                    row += 1

                    overview_sheet.write(row, 0, "Group", fmt["header"])
                    overview_sheet.write(row, 1, "Values", fmt["header"])

                    row += 1
                    for group_name, values in transformed_to_write.items():
                        overview_sheet.write(row, 0, group_name, fmt["cell"])
                        values_str = ", ".join([
                            f"{v:.4f}" if isinstance(v, (int, float)) else str(v)
                            for v in values
                        ])
                        overview_sheet.write(row, 1, values_str, fmt["cell"])
                        row += 1

            # PAIRWISE COMPARISONS section
            row += 2
            overview_sheet.merge_range(f'A{row}:E{row}', "PAIRWISE COMPARISONS", fmt["section_header"])
            row += 1
            
            # Headers for pairwise comparisons
            headers = ["Group 1", "Group 2", "Test", "p-Value", "Corrected"]
            for i, header in enumerate(headers):
                overview_sheet.write(row, i, header, fmt["header"])
            row += 1
            
            comps = results.get("pairwise_comparisons", [])
            if comps and len(comps) > 0:
                for comp in comps[:5]:  # Limit to first 5 comparisons to save space
                    group1 = str(comp.get('group1', 'N/A'))
                    group2 = str(comp.get('group2', 'N/A'))
                    test_name = comp.get('test', 'N/A')
                    pval = comp.get('p_value', None)
                    pval_str = "<0.001" if isinstance(pval, (float, int)) and pval < 0.001 else f"{pval:.4f}" if isinstance(pval, (float, int)) else "N/A"
                    corrected = "Yes" if comp.get('corrected', False) else "No"
                    
                    overview_sheet.write(row, 0, group1, fmt["cell"])
                    overview_sheet.write(row, 1, group2, fmt["cell"])
                    overview_sheet.write(row, 2, test_name, fmt["cell"])
                    overview_sheet.write(row, 3, pval_str, fmt["cell"])
                    overview_sheet.write(row, 4, corrected, fmt["cell"])
                    row += 1
                    
                if len(comps) > 5:
                    overview_sheet.merge_range(f'A{row}:E{row}', f"... and {len(comps) - 5} more comparisons (see {dataset_name}_Pairwise sheet)", fmt["explanation"])
                    row += 1
            else:
                message = "No pairwise comparisons performed or available."
                if p_value is not None and p_value >= results.get("alpha", 0.05) and len(results.get("groups", [])) > 2:
                    message = "No pairwise comparisons performed because the main test was not significant."
                
                overview_sheet.merge_range(f'A{row}:E{row}', message, fmt["cell"])
                row += 1
            
            # Add separator between datasets
            row += 3
        
        # For each dataset: create all detail sheets as in single analysis
        for dataset_name, results in all_results.items():
            # Use the pre-generated decision tree path
            pre_generated_tree = dataset_tree_paths.get(dataset_name)
            
            try:
                # Create all the detailed sheets for this dataset
                ResultsExporter._write_summary_sheet(workbook, results, fmt, f"{dataset_name}_Summary")
                ResultsExporter._write_assumptions_sheet(workbook, results, fmt, f"{dataset_name}_Assumptions")
                ResultsExporter._write_results_sheet(workbook, results, fmt, f"{dataset_name}_Results")
                ResultsExporter._write_descriptive_sheet(workbook, results, fmt, f"{dataset_name}_Descriptive")
                ResultsExporter._write_decision_tree_sheet(workbook, results, fmt, f"{dataset_name}_DecisionTree", pre_generated_tree)
                ResultsExporter._write_rawdata_sheet(workbook, results, fmt, f"{dataset_name}_RawData")
                ResultsExporter._write_pairwise_sheet(workbook, results, fmt, f"{dataset_name}_Pairwise")
                
                # Add analysis log if available
                if results.get("analysis_log"):
                    ResultsExporter._write_analysislog_sheet(workbook, results["analysis_log"], fmt, f"{dataset_name}_Log")
                    
            except Exception as e:
                print(f"DEBUG MULTI: Error creating sheets for {dataset_name}: {str(e)}")
                import traceback
                traceback.print_exc()
        
        # Close the workbook to save changes
        workbook.close()
        print(f"DEBUG MULTI: Excel file created at {excel_path}")
        
        # Clean up all temporary decision tree files
        for dataset_name, tree_path in dataset_tree_paths.items():
            if tree_path and os.path.exists(tree_path):
                try:
                    os.remove(tree_path)
                    print(f"DEBUG MULTI: Cleaned up decision tree file for {dataset_name}")
                except Exception as e:
                    print(f"DEBUG MULTI: Could not clean up {tree_path}: {str(e)}")
        
        # Clean up any other tracked temporary files
        if hasattr(ResultsExporter, '_temp_files'):
            for temp_file in ResultsExporter._temp_files:
                if os.path.exists(temp_file):
                    try:
                        os.remove(temp_file)
                    except Exception:
                        pass
                ResultsExporter._temp_files.clear()
        
        return excel_path

    @staticmethod
    def _get_excel_formats(workbook):
        return {
            "title": workbook.add_format({'bold': True, 'font_size': 14, 'align': 'center', 'valign': 'vcenter'}),
            "header": workbook.add_format({'bold': True, 'font_size': 12, 'align': 'center', 'bottom': 2}),
            "cell": workbook.add_format({'align': 'center', 'text_wrap': True}),
            "significant": workbook.add_format({'align': 'center', 'color': 'red', 'bold': True, 'text_wrap': True}),
            "explanation": workbook.add_format({'text_wrap': True, 'valign': 'top', 'font_color': '#1F4E78'}),
            "section_header": workbook.add_format({'bold': True, 'bg_color': '#B4C6E7', 'border': 1}),
            "effect_strong": workbook.add_format({'align': 'center', 'color': '#006400', 'bold': True, 'text_wrap': True}),
            "effect_medium": workbook.add_format({'align': 'center', 'color': '#FFA500', 'bold': True, 'text_wrap': True}),
            "effect_weak": workbook.add_format({'align': 'center', 'color': '#A52A2A', 'bold': True, 'text_wrap': True}),
            "key": workbook.add_format({'bold': True, 'align': 'right'}),
            "bold": workbook.add_format({'bold': True})
        }

    @staticmethod
    def _write_anova_table(ws, anova_table, fmt, start_row=0):
        """
        Writes an ANOVA table (as DataFrame or dict) to the worksheet at the given row.
        Returns the next empty row after the table.
        """
        import pandas as pd
        if isinstance(anova_table, dict):
            anova_table = pd.DataFrame(anova_table)
        elif not isinstance(anova_table, pd.DataFrame):
            return start_row  # Nothing to write

        # Write header
        for col, colname in enumerate(anova_table.columns):
            ws.write(start_row, col, str(colname), fmt["header"])
        # Write rows
        for row_idx, (_, row) in enumerate(anova_table.iterrows()):
            for col, val in enumerate(row):
                ws.write(start_row + 1 + row_idx, col, val, fmt["cell"])
        return start_row + 1 + len(anova_table)

    @staticmethod
    def _write_summary_sheet(workbook, results, fmt, sheet_name="Summary"):
        ws = workbook.add_worksheet(sheet_name)
        ws.set_column(0, 5, 28)
        ws.set_row(0, 30)

        test_info = results.get("test", "Not specified")
        p_value = results.get("p_value", None)
        is_significant = p_value is not None and isinstance(p_value, (float, int)) and p_value < results.get("alpha", 0.05)
        significant_text = "Yes" if is_significant else "No"
        title = f"SUMMARY OF ANALYSIS - {test_info}"
        ws.merge_range('A1:F1', title, fmt["title"])

        # Key statement
        ws.merge_range('A3:F3', "KEY STATEMENT", fmt["section_header"])
        if is_significant:
            effect_size_text = ""
            if "effect_size" in results and results["effect_size"] is not None:
                effect_size = results["effect_size"]
                effect_magnitude = ""
                if "effect_size_type" in results:
                    effect_type = results["effect_size_type"]
                    # Define magnitude based on effect size type
                    if effect_type.lower() == "cohen_d":
                        if abs(effect_size) < 0.2: effect_magnitude = "very small"
                        elif abs(effect_size) < 0.5: effect_magnitude = "small"
                        elif abs(effect_size) < 0.8: effect_magnitude = "medium"
                        else: effect_magnitude = "large"
                    elif effect_type.lower() in ["eta_squared", "partial_eta_squared", "epsilon_squared", "kendall_w", "r"]:
                        # Simplified thresholds for other effect sizes
                        if abs(effect_size) < 0.1: effect_magnitude = "very small"
                        elif abs(effect_size) < 0.3: effect_magnitude = "small"
                        elif abs(effect_size) < 0.5: effect_magnitude = "medium"
                        else: effect_magnitude = "large"
                effect_size_text = f" with a {effect_magnitude} effect (effect size: {effect_size:.3f})"
            p_val_text = "<0.001" if isinstance(p_value, (float, int)) and p_value < 0.001 else f"={p_value:.4f}" if isinstance(p_value, (float, int)) else "not available"
            conclusion = (
                f"The performed test ({test_info}) shows SIGNIFICANT differences "
                f"between the groups under investigation (p{p_val_text})"
                f"{effect_size_text}."
            )
        else:
            p_val_text = f"={p_value:.4f}" if isinstance(p_value, (float, int)) else "not available"
            conclusion = (
                f"The performed test ({test_info}) shows NO significant differences "
                f"between the groups under investigation (p{p_val_text})."
            )
        ws.merge_range('A4:F4', conclusion, fmt["cell"])
        ws.set_row(3, ResultsExporter.get_text_height(conclusion, 28))

        # Key information
        row = 6
        ws.merge_range(f'A{row}:F{row}', "KEY INFORMATION", fmt["section_header"])
        row += 1

        key_value_pairs = [
            ("Test:", test_info),
            ("Significant:", significant_text),
            ("p-Value:", f"{'<0.001' if p_value and isinstance(p_value, (float,int)) and p_value < 0.001 else f'={p_value:.4f}' if isinstance(p_value, (float,int)) else 'Not available'}")
        ]

        if "df1" in results and results["df1"] is not None and "df2" in results and results["df2"] is not None:
            key_value_pairs.append(("Degrees of freedom (numerator, denominator):", f"{results['df1']}, {results['df2']}"))
        elif "df" in results and results["df"] is not None: # For chi-square etc.
                key_value_pairs.append(("Degrees of freedom (df):", f"{results['df']}"))


        if "sphericity_test" in results:
            sphericity = results["sphericity_test"]
            if sphericity and sphericity.get("has_sphericity") is not None:
                sphericity_text = "Yes" if sphericity["has_sphericity"] else "No"
                key_value_pairs.append(("Sphericity (Mauchly's Test):", sphericity_text))
                if sphericity.get("p_value") is not None:
                    p_val_text = f"{sphericity['p_value']:.4f}" if sphericity["p_value"] >= 0.001 else "<0.001"
                    key_value_pairs.append(("  p-Value Sphericity:", p_val_text))
                if not sphericity["has_sphericity"] and "correction_used" in results:
                    key_value_pairs.append(("  Correction applied:", results["correction_used"]))


        stat_value = results.get("statistic")
        if stat_value is not None:
            stat_name = "Statistic"
            if "t-Test" in test_info: stat_name = "t-Statistic"
            elif "ANOVA" in test_info or "Welch" in test_info: stat_name = "F-Statistic"
            elif "Mann-Whitney" in test_info: stat_name = "U-Statistic"
            elif "Kruskal-Wallis" in test_info: stat_name = "H-Statistic"
            elif "Wilcoxon" in test_info: stat_name = "W-Statistic"
            elif "Friedman" in test_info: stat_name = "Chi²-Statistic"
            key_value_pairs.append((f"{stat_name}:", f"{stat_value:.4f}" if isinstance(stat_value, (float,int)) else str(stat_value)))


        if "effect_size" in results and results["effect_size"] is not None:
            effect_size = results["effect_size"]
            effect_type = results.get("effect_size_type", "")
            effect_desc = ""
            magnitude = ""
            format_to_use = fmt["cell"] # Default format

            if effect_type.lower() == "cohen_d":
                effect_desc = "Cohen's d"
                if abs(effect_size) < 0.2: magnitude = "very small"; format_to_use = fmt["effect_weak"]
                elif abs(effect_size) < 0.5: magnitude = "small"; format_to_use = fmt["effect_weak"]
                elif abs(effect_size) < 0.8: magnitude = "medium"; format_to_use = fmt["effect_medium"]
                else: magnitude = "large"; format_to_use = fmt["effect_strong"]
            elif effect_type.lower() in ["eta_squared", "partial_eta_squared", "omega_squared"]:
                effect_desc = effect_type.replace("_", " ").title()
                if abs(effect_size) < 0.01: magnitude = "very small"; format_to_use = fmt["effect_weak"]
                elif abs(effect_size) < 0.06: magnitude = "small"; format_to_use = fmt["effect_weak"]
                elif abs(effect_size) < 0.14: magnitude = "medium"; format_to_use = fmt["effect_medium"]
                else: magnitude = "large"; format_to_use = fmt["effect_strong"]
            elif effect_type.lower() == "epsilon_squared":
                effect_desc = "Epsilon²"
                if abs(effect_size) < 0.01: magnitude = "very small"; format_to_use = fmt["effect_weak"]
                elif abs(effect_size) < 0.08: magnitude = "small"; format_to_use = fmt["effect_weak"]
                elif abs(effect_size) < 0.26: magnitude = "medium"; format_to_use = fmt["effect_medium"]
                else: magnitude = "large"; format_to_use = fmt["effect_strong"]
            elif effect_type.lower() == "kendall_w":
                effect_desc = "Kendall's W"
                if abs(effect_size) < 0.1: magnitude = "very small"; format_to_use = fmt["effect_weak"]
                elif abs(effect_size) < 0.3: magnitude = "small"; format_to_use = fmt["effect_weak"]
                elif abs(effect_size) < 0.5: magnitude = "medium"; format_to_use = fmt["effect_medium"]
                else: magnitude = "large"; format_to_use = fmt["effect_strong"]
            elif effect_type.lower() == "r": # For Wilcoxon, Mann-Whitney U
                effect_desc = "r (rank correlation)"
                if abs(effect_size) < 0.1: magnitude = "very small"; format_to_use = fmt["effect_weak"]
                elif abs(effect_size) < 0.3: magnitude = "small"; format_to_use = fmt["effect_weak"]
                elif abs(effect_size) < 0.5: magnitude = "medium"; format_to_use = fmt["effect_medium"]
                else: magnitude = "large"; format_to_use = fmt["effect_strong"]
            else:
                effect_desc = effect_type if effect_type else "Effect size"
                magnitude = "not classified"

            key_value_pairs.append((f"{effect_desc}:", f"{effect_size:.4f} ({magnitude})"))
        else:
            format_to_use = fmt["cell"] # Ensure format_to_use is defined

        ci = results.get("confidence_interval", (None, None))
        ci_level = results.get("ci_level", 0.95) * 100

        ci_text = "Not calculated; see confidence intervals of pairwise comparisons (if available)."
        if ci is not None and isinstance(ci, (list, tuple)) and len(ci) == 2 and ci[0] is not None and ci[1] is not None:
            ci_text = f"[{ci[0]:.4f}, {ci[1]:.4f}]"
        elif test_info and ("ANOVA" in test_info or "Kruskal-Wallis" in test_info or "Friedman" in test_info) and len(results.get("groups",[])) > 2 :
                # For ANOVA-like tests with >2 groups, the main CI is often less informative than post-hoc CIs
            pass # ci_text remains the default message
        elif ci == (None, None) or ci is None : # Explicitly (None,None) or just None
            pass # ci_text remains the default message

        key_value_pairs.append((f"{ci_level:.0f}% Confidence interval:", ci_text))


        if "power" in results:
            power = results["power"]
            if power is not None:
                power_desc = "low" if power < 0.5 else "moderate" if power < 0.8 else "high"
                key_value_pairs.append(("Statistical power:", f"{power:.2f} ({power_desc})"))
            else:
                key_value_pairs.append(("Statistical power:", "Not calculated/available"))

        for key, value in key_value_pairs:
            ws.write(row, 0, key, fmt["key"])
            current_format = fmt["cell"]
            if key == "Significant:" and value == "Yes":
                current_format = fmt["significant"]
            elif key == "p-Value:" and is_significant:
                current_format = fmt["significant"]
            elif "Effect size" in key or "Cohen's d" in key or "Eta²" in key or "Epsilon²" in key or "Kendall's W" in key or "r (" in key:
                    # Use the format_to_use determined during effect size magnitude check
                current_format = format_to_use
            ws.write(row, 1, value, current_format)
            row += 1

        # Navigation
        row += 2
        ws.merge_range(f'A{row}:F{row}', "NAVIGATION TO DETAILED RESULTS", fmt["section_header"])
        row += 1
        nav_text = (
            "• Statistical results: Details on test and significance\n"
            "• Assumptions check: Tests for normality and variance homogeneity\n"
            "• Descriptive statistics: Metrics with confidence intervals for each group\n"
            "• Pairwise comparisons: Details on individual group differences with effect sizes and CIs\n"
            "• Raw data: The original measured values\n"
            "• Analysis log: Chronological sequence of the analysis\n"
            "• Hypotheses: Tested null and alternative hypotheses\n")
        ws.merge_range(f'A{row}:F{row+7}', nav_text, fmt["explanation"])
        ws.set_row(row, ResultsExporter.get_text_height(nav_text, 28*6))

        # Update row position to be AFTER the merged navigation text
        row += 8  # This ensures we're beyond the previous merge range

        anova_table = results.get("anova_table")
        if anova_table is not None:
            row += 2
            ws.merge_range(f'A{row}:F{row}', "ANOVA TABLE", fmt["section_header"])
            row += 1
            next_row = ResultsExporter._write_anova_table(ws, anova_table, fmt, start_row=row)
            row = next_row + 2  # Add space after the table
            
            # Add explanation section for ANOVA table
            ws.merge_range(f'A{row}:F{row}', "UNDERSTANDING THE ANOVA TABLE", fmt["section_header"])
            row += 1
            
            # Determine which type of ANOVA is being used
            test_name = results.get("test", "").lower()
            is_welch = "welch" in test_name
            is_rm = "repeated measures" in test_name or "rm anova" in test_name
            is_mixed = "mixed" in test_name
            is_two_way = "two-way" in test_name or "two way" in test_name
            
            # Introduction text based on ANOVA type
            if is_welch:
                intro_text = (
                    "This is a Welch's ANOVA table, which does not assume equal variances between groups. "
                    "Welch's ANOVA is more robust when the homogeneity of variance assumption is violated."
                )
            elif is_rm:
                intro_text = (
                    "This is a Repeated Measures ANOVA table, which analyzes differences between repeated measurements "
                    "on the same subjects, accounting for the dependency between observations."
                )
            elif is_mixed:
                intro_text = (
                    "This is a Mixed ANOVA table, which combines between-subjects factors (different groups) "
                    "and within-subjects factors (repeated measures) in the same analysis."
                )
            elif is_two_way:
                intro_text = (
                    "This is a Two-Way ANOVA table, which analyzes the effect of two independent variables (factors) "
                    "on one dependent variable, including their potential interaction."
                )
            else:
                intro_text = (
                    "This is a standard One-Way ANOVA table, which compares means across multiple groups "
                    "to determine if there are significant differences between them."
                )
            
            ws.merge_range(f'A{row}:F{row}', intro_text, fmt["explanation"])
            ws.set_row(row, ResultsExporter.get_text_height(intro_text, 28*6))
            row += 2
            
            # Column explanations
            ws.merge_range(f'A{row}:F{row}', "EXPLANATION OF ANOVA TABLE COLUMNS:", fmt["key"])
            row += 1
            
            explanations = [
                ("Source", "Identifies the component being analyzed:\n• Group/factor names indicate variation between groups/factors\n• Residual (or Error/Within) represents unexplained variation within groups\n• Interaction terms (in factorial designs) show how factors work together"),
                
                ("SS (Sum of Squares)", "Measures the total variation attributed to each source:\n• Higher values indicate more variation explained by that source\n• SS between groups shows variation due to group differences\n• SS within groups (residual) shows variation due to individual differences\n• The ratio of SS(between) to SS(total) provides an estimate of effect size"),
                
                ("DF (Degrees of Freedom)", "The number of values that are free to vary:\n• For between-groups: DF = number of groups - 1\n• For residuals (simple design): DF = total observations - number of groups\n• For residuals (factorial design): DF = total sample size - number of estimated parameters\n• For factors: DF = number of levels - 1\n• For interactions: DF = product of the individual factors' DFs"),
                
                ("MS (Mean Square)", "Average variation per degree of freedom (MS = SS/DF):\n• MS between represents average between-group variation\n• MS within (residual) represents average within-group variation (error variance)\n• The ratio MS(between)/MS(within) forms the F-statistic"),
                
                ("F", "The F-statistic (MS between / MS within):\n• Compares between-group variation to within-group variation\n• Larger F values suggest stronger group differences\n• F = 1 would indicate no difference between groups\n• The critical F value depends on the degrees of freedom and alpha level"),
                
                ("p-unc", "Probability value (significance level):\n• p < 0.05 is commonly used to indicate statistical significance, though this threshold is arbitrary\n• 'unc' indicates these are uncorrected p-values (not adjusted for multiple comparisons)\n• Small p-values suggest the observed differences are unlikely under the null hypothesis\n• For multiple tests, consider using corrected p-values to control error rates"),
                
                ("np2 (Partial Eta Squared)", "Effect size measure (proportion of variance explained):\n• 0.01 = small effect\n• 0.06 = medium effect\n• 0.14 = large effect\n• Higher values indicate stronger effects\n• Interpretation is context-dependent and varies between research fields\n• Consider field-specific benchmarks when interpreting effect sizes")
            ]
            
            for term, explanation in explanations:
                ws.write(row, 0, term, fmt["key"])
                ws.write(row, 1, explanation, fmt["explanation"])
                ws.set_row(row, ResultsExporter.get_text_height(explanation, 28*5))
                row += 1
            
            row += 1
            
            # How to interpret section
            ws.merge_range(f'A{row}:F{row}', "HOW TO INTERPRET THE RESULTS", fmt["key"])
            row += 1
            
            if is_two_way or is_mixed:
                interpret_text = (
                    "1. Check main effects: Look at p-values for each factor. If p < 0.05, that factor has a significant effect.\n\n"
                    "2. Check interaction: If the interaction p-value is < 0.05, the effect of one factor depends on the level of the other factor. "
                    "In this case, interpret main effects with caution and focus on pairwise comparisons.\n\n"
                    "3. Effect size (np2): Indicates the practical significance - how much variance is explained by each factor.\n\n"
                    "4. Post-hoc tests: For significant effects, examine post-hoc tests to identify which specific groups differ."
                )
            else:
                interpret_text = (
                    "1. Statistical significance: If the p-value for the between-groups factor is < 0.05, there are significant differences between at least some groups.\n\n"
                    "2. Effect size (np2): Indicates the practical significance - how much variance is explained by group differences.\n\n"
                    "3. F-statistic: Higher values indicate stronger evidence against the null hypothesis.\n\n"
                    "4. Post-hoc tests: For significant results, examine post-hoc tests to identify which specific groups differ from each other."
                )

            
            ws.merge_range(f'A{row}:F{row}', interpret_text, fmt["explanation"])
            ws.set_row(row, ResultsExporter.get_text_height(interpret_text, 28*6))
            row += 1
    
    @staticmethod
    def _write_assumptions_sheet(workbook, results, fmt, sheet_name="Assumptions"):
        ws = workbook.add_worksheet(sheet_name)
        ws.set_column(0, 5, 30)
        ws.set_row(0, 30)
        ws.merge_range('A1:F1', 'TEST ASSUMPTIONS CHECK', fmt["title"])

        # Introduction
        introduction = (
            "This sheet documents the tests for checking the assumptions for the statistical analysis. "
            "Depending on the type of test (parametric or non-parametric), different assumptions must be met."
        )
        ws.merge_range('A2:F2', introduction, fmt["explanation"])
        ws.set_row(1, ResultsExporter.get_text_height(introduction, 30*6))

        row = 4

        # Test type and general notes
        test_type = results.get("recommendation", results.get("test_type", "Not specified"))
        if test_type == "parametric":
            assumptions_overview = (
                "For parametric tests (such as t-test, ANOVA), the following assumptions apply:\n"
                "  • Normal distribution of the data in each group\n"
                "  • Homogeneity of variances between groups\n"
                "  • Independence of observations\n"
                "  • Interval scale of the dependent variable\n\n"
                "Points 1 and 2 are tested statistically. Points 3 and 4 are ensured by the study design and data collection."
            )
        else:
            assumptions_overview = (
                "For non-parametric tests (such as Mann-Whitney U, Kruskal-Wallis), the following assumptions apply:\n"
                "  • Similar distribution shape in all groups (not necessarily normal distribution)\n"
                "  • Independence of observations\n"
                "  • At least ordinal scale of the dependent variable\n\n"
                "Assumption 1 is assessed visually. Points 2 and 3 are ensured by the study design and data collection."
            )
        ws.merge_range(f'A{row}:F{row}', "OVERVIEW OF ASSUMPTIONS", fmt["section_header"])
        row += 1
        ws.merge_range(f'A{row}:F{row}', assumptions_overview, fmt["explanation"])
        ws.set_row(row, ResultsExporter.get_text_height(assumptions_overview, 30*6))
        row += 2

        # Sphericity (only for Repeated Measures ANOVA)
        if "sphericity_test" in results:
            row += 1
            ws.write(row, 0, "Sphericity (Mauchly test):", fmt["section_header"])
            row += 1
            
            # Check if we have a two-level within factor (where sphericity is always met)
            has_two_levels = False
            
            # Method 1: Check ANOVA table for epsilon=1.0
            if "anova_table" in results and isinstance(results["anova_table"], pd.DataFrame):
                if "eps" in results["anova_table"].columns:
                    eps_values = results["anova_table"]["eps"].dropna()
                    if not eps_values.empty and (eps_values == 1.0).all():
                        has_two_levels = True
            
            # Method 2: Check within factor directly if available
            if "factors" in results:
                for factor in results["factors"]:
                    if factor.get("type") == "within":
                        within_factor = factor.get("factor")
                        if within_factor and "df1" in factor and factor["df1"] == 1:
                            has_two_levels = True
                            break
            
            if has_two_levels:
                # Special explanation for two-level within factors
                explanation = (
                    "With only two levels of the within-factor, sphericity is automatically satisfied mathematically.\n\n"
                    "Sphericity concerns the equality of variances of differences between all combinations of within-subject levels. "
                    "When there are only two levels, there is only one possible difference (level 1 - level 2), "
                    "so no comparison of different variances is possible and sphericity is perfectly met by definition.\n\n"
                    "Therefore, no sphericity test is necessary, and no corrections (Greenhouse-Geisser or Huynh-Feldt) are needed."
                )
                ws.merge_range(row, 0, row, 5, explanation, fmt["explanation"])
                ws.set_row(row, ResultsExporter.get_text_height(explanation, 30*6))
                row += 2
            else:
                # Regular sphericity test information
                sph_headers = ["Mauchly's W", "p-Value", "Sphericity assumed?", "Interpretation"]
                for i, h in enumerate(sph_headers):
                    ws.write(row, i, h, fmt["header"])
                row += 1
                
                sphericity = results["sphericity_test"]
                w_val = sphericity.get("W", "N/A")
                p_val = sphericity.get("p_value", "N/A")
                has_sphericity = sphericity.get("has_sphericity", None)
                sph_text = "Yes" if has_sphericity else "No" if has_sphericity is not None else "Indeterminable"
                
                interpretation = (
                    "No significant deviation from sphericity"
                    if has_sphericity else
                    "Significant deviation from sphericity, correction necessary"
                    if has_sphericity is not None else
                    "Sphericity could not be tested"
                )
                
                values = [
                    f"{w_val:.4f}" if isinstance(w_val, (float, int)) else w_val,
                    f"{p_val:.4f}" if isinstance(p_val, (float, int)) else p_val,
                    sph_text,
                    interpretation
                ]
                
                for col, val in enumerate(values):
                    ws.write(row, col, val, fmt["cell"])
                row += 1
    
        # Normality tests per group
        ws.write(row, 0, "Normality (Shapiro-Wilk test per group):", fmt["section_header"])
        row += 1
        norm_headers = ["Group", "Shapiro-Wilk statistic", "p-Value", "Normally distributed?", "Interpretation"]
        for i, h in enumerate(norm_headers):
            ws.write(row, i, h, fmt["header"])
        row += 1
        normality_results = results.get("normality_tests", {})
    
        for group, test_result in normality_results.items():
            if group == "all_data" or group == "transformed_data":
                continue  # Skip these special entries
            stat = test_result.get('statistic', 'N/A')
            p_val = test_result.get('p_value', 'N/A')
            is_normal = (isinstance(p_val, (float, int)) and p_val > 0.05)
            normal_text = "Yes" if is_normal else "No"
            interpretation = (
                "No significant deviation from normality"
                if is_normal else
                "Significant deviation from normality"
            )
            values = [
                str(group),
                f"{stat:.4f}" if isinstance(stat, (float, int)) else stat,
                f"{p_val:.4f}" if isinstance(p_val, (float, int)) else p_val,
                normal_text,
                interpretation
            ]
            for col, val in enumerate(values):
                ws.write(row, col, val, fmt["cell"])
            row += 1
    
        row += 1
    
        # Homogeneity of variances (Levene test)
        ws.write(row, 0, "Homogeneity of variances (Levene test):", fmt["section_header"])
        row += 1
        var_headers = ["Levene statistic", "p-Value", "Variances equal?", "Interpretation"]
        for i, h in enumerate(var_headers):
            ws.write(row, i, h, fmt["header"])
        row += 1
        
        var_test = results.get("variance_test", {})
        stat = var_test.get('statistic', 'N/A')
        p_val = var_test.get('p_value', 'N/A')
        var_equal = (isinstance(p_val, (float, int)) and p_val > 0.05)
        var_text = "Yes" if var_equal else "No"
        interpretation = (
            "No significant differences in variances"
            if var_equal else
            "Significant differences in variances"
        )
        values = [
            f"{stat:.4f}" if isinstance(stat, (float, int)) else stat,
            f"{p_val:.4f}" if isinstance(p_val, (float, int)) else p_val,
            var_text,
            interpretation
        ]
        for col, val in enumerate(values):
            ws.write(row, col, val, fmt["cell"])
        row += 2
    
        # Add a clear post-transformation section (after the transformation announcement)
        transformation = results.get("transformation", "None")
        if transformation and transformation != "None":
            ws.write(row, 0, f"Transformation applied: {transformation}", fmt["section_header"])
            trans_info = results.get("transformation_info", "")
            if trans_info:
                row += 1
                ws.merge_range(f'A{row}:F{row}', f"Details: {trans_info}", fmt["explanation"])
                ws.set_row(row, ResultsExporter.get_text_height(str(trans_info), 30*6))
            # Show lambda value if Box-Cox transformation
            if transformation == "boxcox" and "boxcox_lambda" in results:
                row += 1
                lambda_val = results["boxcox_lambda"]
                ws.merge_range(f'A{row}:F{row}', f"Box-Cox Lambda (MLE): {lambda_val:.4f}", fmt["cell"])
            
            # IMPORTANT ADDITION: Note about data usage
            row += 1
            test_type = results.get("test_type", "")
            data_usage_note = (
                f"Note: {'Transformed data WAS used for statistical tests' if test_type == 'parametric' else 'Original (untransformed) data was used for statistical tests'} "
                f"based on test recommendation: {test_type}."
            )
            ws.merge_range(f'A{row}:F{row}', data_usage_note, fmt["explanation"])
            
            # Add post-transformation test results section
            row += 2
            ws.merge_range(f'A{row}:F{row}', "TESTS AFTER TRANSFORMATION", fmt["section_header"])
            row += 1
            
            # Normality tests after transformation
            if "normality_tests" in results and "transformed_data" in results["normality_tests"]:
                ws.write(row, 0, "Normality after transformation:", fmt["key"])
                row += 1
                norm_headers = ["Test", "Statistic", "p-Value", "Normally distributed?", "Interpretation"]
                for i, h in enumerate(norm_headers):
                    ws.write(row, i, h, fmt["header"])
                row += 1
                
                norm_result = results["normality_tests"]["transformed_data"]
                stat = norm_result.get('statistic', 'N/A')
                p_val = norm_result.get('p_value', 'N/A')
                is_normal = (isinstance(p_val, (float, int)) and p_val > 0.05)
                normal_text = "Yes" if is_normal else "No"
                interpretation = (
                    "No significant deviation from normality"
                    if is_normal else
                    "Significant deviation from normality"
                )
                
                values = [
                    "Shapiro-Wilk",
                    f"{stat:.4f}" if isinstance(stat, (float, int)) else stat,
                    f"{p_val:.4f}" if isinstance(p_val, (float, int)) else p_val,
                    normal_text,
                    interpretation
                ]
                
                for col, val in enumerate(values):
                    ws.write(row, col, val, fmt["cell"])
                row += 2
            
            # Homogeneity of variances after transformation
            if "variance_test" in results and "transformed" in results["variance_test"]:
                ws.write(row, 0, "Variance homogeneity after transformation:", fmt["key"])
                row += 1
                var_headers = ["Test", "Statistic", "p-Value", "Variances equal?", "Interpretation"]
                for i, h in enumerate(var_headers):
                    ws.write(row, i, h, fmt["header"])
                row += 1
                
                var_result = results["variance_test"]["transformed"]
                stat = var_result.get('statistic', 'N/A')
                p_val = var_result.get('p_value', 'N/A')
                var_equal = (isinstance(p_val, (float, int)) and p_val > 0.05)
                var_text = "Yes" if var_equal else "No"
                interpretation = (
                    "No significant differences in variances"
                    if var_equal else
                    "Significant differences in variances"
                )
                
                values = [
                    "Levene",
                    f"{stat:.4f}" if isinstance(stat, (float, int)) else stat,
                    f"{p_val:.4f}" if isinstance(p_val, (float, int)) else p_val,
                    var_text,
                    interpretation
                ]
                
                for col, val in enumerate(values):
                    ws.write(row, col, val, fmt["cell"])
                row += 2

            row += 1
        
            # Summary text
            row += 1
            ws.merge_range(f'A{row}:F{row}', "SUMMARY OF ASSUMPTIONS CHECK", fmt["section_header"])
            row += 1
            summary = results.get("assumptions_summary", "")
            if not summary:
                summary = (
                    "The assumptions were checked as documented above. "
                    "See test results for each group for details."
                )
            ws.merge_range(f'A{row}:F{row}', summary, fmt["explanation"])
            ws.set_row(row, ResultsExporter.get_text_height(summary, 30*6))
            row += 2
        
            # Decision tree
            ws.merge_range(f'A{row}:F{row}', "DECISION TREE FOR TEST SELECTION", fmt["section_header"])
            row += 1
            decision_tree = results.get("decision_tree_text", None)
            if not decision_tree:
                decision_tree = (
                    "1. Are the data in all groups normally distributed?\n"
                    "2. Are the variances between the groups equal?\n"
                    "→ If yes: Parametric test (e.g., t-test, ANOVA)\n"
                    "→ If no: Try transformation or non-parametric test (e.g., Mann-Whitney U, Kruskal-Wallis)"
                )
            ws.merge_range(f'A{row}:F{row+4}', decision_tree, fmt["explanation"])
            ws.set_row(row, ResultsExporter.get_text_height(decision_tree, 30*6))
    
        # Set column widths
        for col in range(6):
            ws.set_column(col, col, 28)
            
    @staticmethod
    def _write_results_sheet(workbook, results, fmt, sheet_name="Statistical Results"):
        ws = workbook.add_worksheet(sheet_name)
        ws.set_column(0, 12, 22)
        ws.set_row(0, 30)
        ws.merge_range('A1:M1', 'STATISTICAL RESULTS', fmt["title"])

        # Introduction
        # Detect if this is a nonparametric permutation ANOVA
        is_perm = results.get("test_type", "").lower() == "non-parametric" or results.get("permutation_test", False)
        
        intro = (
            "This sheet contains the main results of the statistical analysis: "
            "test statistics, p-value, effect size, confidence interval, power, "
            "and – if relevant – alternative tests. "
        )
        
        # Add explanation about Freedman-Lane when permutation tests are used
        if is_perm:
            intro += (
                "For permutation-based nonparametric ANOVA, p-values are computed using the Freedman–Lane scheme."
            )
            
        ws.merge_range('A2:M2', intro, fmt["explanation"])
        ws.set_row(1, ResultsExporter.get_text_height(intro, 22*10))

        # Main result table
        row = 4
        
        # Define column headers (with permutation-specific headers when applicable)
        headers = [
            "Test", "Test statistic", 
            "Permutation p-value" if is_perm else "p-Value", 
            "Effect size", "Confidence interval", "Power", "Significant?",
            "Permutation Test" if is_perm else "",
            "Permutation Scheme" if is_perm else ""
        ]
        # Remove empty headers
        headers = [h for h in headers if h]
        
        for col, header in enumerate(headers):
            ws.write(row, col, header, fmt["header"])
        row += 1

        # Values
        test = results.get("test", "N/A")
        stat_val = (
            results.get("t_statistic") or results.get("u_statistic") or
            results.get("f_statistic") or results.get("h_statistic") or
            results.get("statistic", None)
        )
        p_val = results.get("p_value", None)
        # Special handling for non-parametric test effects
        if results.get("test_type") == "non-parametric" and "effects" in results:
            for effect in results.get("effects", []):
                if effect.get("name") and "within_effect" in effect.get("name", "").lower():
                    stat_val = effect.get("F")
                    p_val = effect.get("p")
                    print(f"DEBUG: Using effect data for non-parametric test: F={stat_val}, p={p_val}")
                    break
        effect_size = results.get("effect_size", None)
        ci = results.get("confidence_interval", None)
        power = results.get("power", None)
        is_significant = p_val is not None and p_val < 0.05

        stat_val_str = f"{stat_val:.4f}" if isinstance(stat_val, (float, int)) else (stat_val or "N/A")
        
        # Format p-value differently for permutation tests
        if is_perm:
            p_val_str = f"{p_val:.4f}" if isinstance(p_val, (float, int)) else (p_val or "N/A")
        else:
            p_val_str = (
                "p < 0.001" if isinstance(p_val, (float, int)) and p_val < 0.001 else
                f"p = {p_val:.4f}" if isinstance(p_val, (float, int)) else (p_val or "N/A")
            )
            
        effect_type = results.get("effect_size_type", "")
        if effect_size is not None:
            if effect_type == "cohen_d":
                if effect_size < 0.2: effect_str = f"{effect_size:.4f} (very small)"
                elif effect_size < 0.5: effect_str = f"{effect_size:.4f} (small)"
                elif effect_size < 0.8: effect_str = f"{effect_size:.4f} (medium)"
                else: effect_str = f"{effect_size:.4f} (large)"
            elif effect_type in ["eta_squared", "partial_eta_squared"]:
                if effect_size < 0.01: effect_str = f"{effect_size:.4f} (very small)"
                elif effect_size < 0.06: effect_str = f"{effect_size:.4f} (small)"
                elif effect_size < 0.14: effect_str = f"{effect_size:.4f} (medium)"
                else: effect_str = f"{effect_size:.4f} (large)"
            elif effect_type == "epsilon_squared":
                if effect_size < 0.01: effect_str = f"{effect_size:.4f} (very small)"
                elif effect_size < 0.08: effect_str = f"{effect_size:.4f} (small)"
                elif effect_size < 0.26: effect_str = f"{effect_size:.4f} (medium)"
                else: effect_str = f"{effect_size:.4f} (large)"
            else:
                effect_str = f"{effect_size:.4f}"
        else:
            effect_str = "N/A"
            
        if ci is not None and isinstance(ci, (tuple, list)) and len(ci) == 2 and ci[0] is not None and ci[1] is not None:
            ci_str = f"[{ci[0]:.4f}, {ci[1]:.4f}]"
        else:
            ci_str = "N/A"
            
        power_str = f"{power:.2f}" if isinstance(power, (float, int)) else "N/A"
        sig_str = "Yes" if is_significant else "No"

        # Create list of values to write
        values = [test, stat_val_str, p_val_str, effect_str, ci_str, power_str, sig_str]
        
        # Add permutation-specific columns if needed
        if is_perm:
            values.extend(["Yes", perm_scheme := results.get("permutation_scheme", "Freedman–Lane")])

        # Write all values
        for col, val in enumerate(values):
            fmtx = fmt["significant"] if (col == 2 and is_significant) or (col == 6 and is_significant) else fmt["cell"]
            ws.write(row, col, val, fmtx)
        row += 2

        # Show sphericity corrections if present
        if "sphericity_corrections" in results:
            ws.merge_range(f'A{row}:F{row}', "CORRECTIONS FOR SPHERICITY VIOLATION", fmt["section_header"])
            row += 1
            
            # Show which correction was used, based on Girden (1992)
            if "correction_used" in results:
                ws.merge_range(f'A{row}:F{row}', f"Correction used: {results['correction_used']}", fmt["explanation"])
                row += 1
            
            corr_headers = ["Correction type", "Epsilon", "Corrected df1", "Corrected df2", "Corrected p-Value", "Significant?"]
            for col, header in enumerate(corr_headers):
                ws.write(row, col, header, fmt["header"])
            row += 1
            
            # Greenhouse-Geisser correction
            gg_corr = results["sphericity_corrections"]["greenhouse_geisser"]
            gg_p = gg_corr["p_value"]
            gg_sig = gg_p < results.get("alpha", 0.05) if isinstance(gg_p, (float, int)) else False
            ws.write(row, 0, "Greenhouse-Geisser", fmt["cell"])
            ws.write(row, 1, f"{gg_corr['epsilon']:.4f}", fmt["cell"])
            ws.write(row, 2, f"{gg_corr['df1']:.4f}", fmt["cell"])
            ws.write(row, 3, f"{gg_corr['df2']:.4f}", fmt["cell"])
            ws.write(row, 4, f"{gg_p:.4f}" if isinstance(gg_p, (float, int)) else "N/A", 
                    fmt["significant"] if gg_sig else fmt["cell"])
            ws.write(row, 5, "Yes" if gg_sig else "No",
                    fmt["significant"] if gg_sig else fmt["cell"])
            row += 1
            
            # Huynh-Feldt correction
            hf_corr = results["sphericity_corrections"]["huynh_feldt"]
            hf_p = hf_corr["p_value"]
            hf_sig = hf_p < results.get("alpha", 0.05) if isinstance(hf_p, (float, int)) else False
            ws.write(row, 0, "Huynh-Feldt", fmt["cell"])
            ws.write(row, 1, f"{hf_corr['epsilon']:.4f}", fmt["cell"])
            ws.write(row, 2, f"{hf_corr['df1']:.4f}", fmt["cell"])
            ws.write(row, 3, f"{hf_corr['df2']:.4f}", fmt["cell"])
            ws.write(row, 4, f"{hf_p:.4f}" if isinstance(hf_p, (float, int)) else "N/A", 
                    fmt["significant"] if hf_sig else fmt["cell"])
            ws.write(row, 5, "Yes" if hf_sig else "No",
                    fmt["significant"] if hf_sig else fmt["cell"])
            row += 2

        # Alternative tests
        alt_tests = results.get("alternative_tests", [])
        if alt_tests:
            ws.merge_range(f'A{row}:F{row}', "RESULTS OF ALTERNATIVE TESTS", fmt["section_header"])
            row += 1
            alt_headers = [
                "Test", "Test statistic", "p-Value", "Significant?", "Effect size", "Effect interpretation"
            ]
            for col, header in enumerate(alt_headers):
                ws.write(row, col, header, fmt["header"])
            row += 1
            for alt in alt_tests:
                test = alt.get("test", "")
                stat = alt.get("statistic", "N/A")
                p = alt.get("p_value", "N/A")
                eff = alt.get("effect_size", "N/A")
                eff_type = alt.get("effect_size_type", "")
                sig = p < 0.05 if isinstance(p, (float, int)) else False
                if eff != "N/A" and eff is not None:
                    if eff_type == "cohen_d":
                        if eff < 0.2: effint = "very small"
                        elif eff < 0.5: effint = "small"
                        elif eff < 0.8: effint = "medium"
                        else: effint = "large"
                    elif eff_type in ["eta_squared", "partial_eta_squared"]:
                        if eff < 0.01: effint = "very small"
                        elif eff < 0.06: effint = "small"
                        elif eff < 0.14: effint = "medium"
                        else: effint = "large"
                    else:
                        effint = ""
                else:
                    effint = ""
                vals = [
                    test,
                    f"{stat:.4f}" if isinstance(stat, (float, int)) else stat,
                    f"{p:.4f}" if isinstance(p, (float, int)) else p,
                    "Yes" if sig else "No",
                    f"{eff:.4f}" if isinstance(eff, (float, int)) else eff,
                    effint
                ]
                for col, val in enumerate(vals):
                    fmtx = fmt["significant"] if (col == 2 and sig) or (col == 3 and sig) else fmt["cell"]
                    ws.write(row, col, val, fmtx)
                row += 1
            row += 1

        # Interpretation
        ws.merge_range(f'A{row}:F{row}', "INTERPRETATION", fmt["section_header"])
        row += 1
        interpretation = (
            "The analysis shows a statistically significant difference between the groups."
            if is_significant else
            "The analysis shows no statistically significant difference between the groups."
        )
        ws.merge_range(f'A{row}:F{row}', interpretation, fmt["explanation"])
        ws.set_row(row, ResultsExporter.get_text_height(interpretation, 22*6))
        row += 2

        # Add a permutation explanation if applicable
        if is_perm:
            ws.merge_range(f'A{row}:F{row}', "ABOUT PERMUTATION TESTS", fmt["section_header"])
            row += 1
            perm_explanation = (
                "This analysis used a permutation-based approach with the Freedman–Lane scheme. "
                "In permutation tests, the data is repeatedly shuffled (permuted) to create a "
                "distribution of test statistics under the null hypothesis. The p-value represents "
                "the proportion of permuted datasets that produce a test statistic as extreme as "
                "or more extreme than the observed one. This approach is more robust when parametric "
                "assumptions are violated."
            )
            ws.merge_range(f'A{row}:F{row}', perm_explanation, fmt["explanation"])
            ws.set_row(row, ResultsExporter.get_text_height(perm_explanation, 22*6))
    

    @staticmethod
    def _write_descriptive_sheet(workbook, results, fmt, sheet_name="Descriptive Statistics"):
        ws = workbook.add_worksheet(sheet_name)
        ws.set_column(0, 9, 20)
        ws.set_row(0, 28)
        ws.merge_range('A1:J1', 'DESCRIPTIVE STATISTICS', fmt["title"])

        # Introduction
        desc_explanation = (
            "This sheet contains summary statistics for each group:\n"
            "• n: Sample size of the group\n"
            "• Mean: Average of the values\n"
            "• 95% Confidence interval: Confidence interval for the mean\n"
            "• Median, standard deviation, standard error, minimum, maximum\n"
            "Transformed values are also shown if a transformation was performed."
        )
        ws.merge_range('A2:J2', desc_explanation, fmt["explanation"])
        ws.set_row(1, ResultsExporter.get_text_height(desc_explanation, 20*10))

        # Header
        headers = [
            "Group", "n", "Mean", "95% CI Lower", "95% CI Upper",
            "Median", "Std. Dev.", "Std. Error", "Min", "Max"
        ]
        ws.set_row(3, 22)
        for col, header in enumerate(headers):
            ws.write(3, col, header, fmt["header"])

        # Original data
        desc = results.get('descriptive', results.get('descriptive_stats', {}))
        row = 4
        for group, grp in desc.items():
            n = grp.get('n', None)
            mean = grp.get('mean', None)
            median = grp.get('median', None)
            std = grp.get('std', None)
            stderr = grp.get('stderr', None)
            minv = grp.get('min', None)
            maxv = grp.get('max', None)
            
            # Calculate confidence interval if needed
            ci_lower = grp.get('ci_lower', None)
            ci_upper = grp.get('ci_upper', None)
            
            if ci_lower is None or ci_upper is None:
                try:
                    from scipy.stats import t
                    if n and n > 1 and stderr is not None:
                        ci_lower, ci_upper = t.interval(0.95, n - 1, loc=mean, scale=stderr)
                    else:
                        ci_lower, ci_upper = None, None
                except Exception:
                    ci_lower, ci_upper = None, None

            ws.write(row, 0, group, fmt["cell"])
            ws.write(row, 1, n if n is not None else "", fmt["cell"])
            ws.write(row, 2, f"{mean:.4f}" if mean is not None else "", fmt["cell"])
            ws.write(row, 3, f"{ci_lower:.4f}" if ci_lower is not None else "", fmt["cell"])
            ws.write(row, 4, f"{ci_upper:.4f}" if ci_upper is not None else "", fmt["cell"])
            ws.write(row, 5, f"{median:.4f}" if median is not None else "", fmt["cell"])
            ws.write(row, 6, f"{std:.4f}" if std is not None else "", fmt["cell"])
            ws.write(row, 7, f"{stderr:.4f}" if stderr is not None else "", fmt["cell"])
            ws.write(row, 8, f"{minv:.4f}" if minv is not None else "", fmt["cell"])
            ws.write(row, 9, f"{maxv:.4f}" if maxv is not None else "", fmt["cell"])
            row += 1

        # Transformed data, if present - Enhanced section
        desc_t = results.get('descriptive_transformed', {})
        transformation = results.get('transformation', 'None')
        
        # Show transformed data even if it wasn't used for tests
        if desc_t and transformation and transformation != 'None':
            row += 2
            header_text = "Descriptive Statistics (after transformation)"
            if results.get("test_type") != "parametric":
                header_text += " - Not used for statistical test"
            
            ws.merge_range(f'A{row}:J{row}', header_text, fmt["section_header"])
            row += 1
            
            # Add transformation method info
            transform_info = f"Transformation method: {transformation.capitalize()}"
            if transformation == "boxcox" and "boxcox_lambda" in results:
                transform_info += f", λ = {results['boxcox_lambda']:.4f}"
            ws.merge_range(f'A{row}:J{row}', transform_info, fmt["explanation"])
            row += 1
            
            # Column headers for transformed data
            for col, header in enumerate(headers):
                ws.write(row, col, header, fmt["header"])
            row += 1
            
            # Write transformed data
            for group, grp in desc_t.items():
                n = grp.get('n', None)
                mean = grp.get('mean', None)
                median = grp.get('median', None)
                std = grp.get('std', None)
                stderr = grp.get('stderr', None)
                minv = grp.get('min', None)
                maxv = grp.get('max', None)
                
                # Calculate confidence interval if needed
                ci_lower = grp.get('ci_lower', None)
                ci_upper = grp.get('ci_upper', None)
                
                if ci_lower is None or ci_upper is None:
                    try:
                        from scipy.stats import t
                        if n and n > 1 and stderr is not None:
                            ci_lower, ci_upper = t.interval(0.95, n - 1, loc=mean, scale=stderr)
                        else:
                            ci_lower, ci_upper = None, None
                    except Exception:
                        ci_lower, ci_upper = None, None
                        
                ws.write(row, 0, group, fmt["cell"])
                ws.write(row, 1, n if n is not None else "", fmt["cell"])
                ws.write(row, 2, f"{mean:.4f}" if mean is not None else "", fmt["cell"])
                ws.write(row, 3, f"{ci_lower:.4f}" if ci_lower is not None else "", fmt["cell"])
                ws.write(row, 4, f"{ci_upper:.4f}" if ci_upper is not None else "", fmt["cell"])
                ws.write(row, 5, f"{median:.4f}" if median is not None else "", fmt["cell"])
                ws.write(row, 6, f"{std:.4f}" if std is not None else "", fmt["cell"])
                ws.write(row, 7, f"{stderr:.4f}" if stderr is not None else "", fmt["cell"])
                ws.write(row, 8, f"{minv:.4f}" if minv is not None else "", fmt["cell"])
                ws.write(row, 9, f"{maxv:.4f}" if maxv is not None else "", fmt["cell"])
                row += 1
        
    @staticmethod
    def _write_pairwise_sheet(workbook, results, fmt, sheet_name="Pairwise Comparisons"):
        # RECONSTRUCTION SAFETY: If main list is empty but component lists exist, rebuild it
        if (not results.get('pairwise_comparisons') or len(results.get('pairwise_comparisons', [])) == 0):
            # Try to reconstruct from between and within comparisons
            all_comparisons = []
            
            if "between_pairwise_comparisons" in results and results["between_pairwise_comparisons"]:
                all_comparisons.extend(results["between_pairwise_comparisons"])
                
            if "within_pairwise_comparisons" in results and results["within_pairwise_comparisons"]:
                all_comparisons.extend(results["within_pairwise_comparisons"])
                
            if all_comparisons:
                # Use the reconstructed comparisons
                results["pairwise_comparisons"] = all_comparisons
                print(f"DEBUG: Reconstructed {len(all_comparisons)} pairwise comparisons for Excel export")
        
        print(f"DEBUG POSTHOC EXCEL: Number of pairwise comparisons when writing: {len(results.get('pairwise_comparisons', []))}")
        ws = workbook.add_worksheet(sheet_name)
        ws.set_column(0, 7, 22)  # Increased column count for CI
        ws.set_row(0, 28)
        posthoc_test_name = results.get("posthoc_test", "")
        title_text = 'RESULTS OF PAIRWISE COMPARISONS'
        if posthoc_test_name:
            title_text += f' – {posthoc_test_name}'
        ws.merge_range('A1:H1', title_text, fmt["title"])  # Increased merge range
    
        # Introduction
        pw_explanation = (
            "This sheet shows the results of the pairwise comparisons between the groups.\n"
            "• Group 1 & Group 2: The compared groups\n"
            "• Test: Test performed for the comparison\n"
            "• p-Value: (Corrected) significance value of the comparison\n"
            "• Corrected: Indicates whether a correction for multiple testing was applied\n"
            "• Significant: 'Yes' if p < Alpha (usually 0.05)\n"
            "• Effect size: Magnitude of the difference (e.g., Cohen's d, Hedges' g)\n"
            "• 95% CI: Confidence interval for the difference between groups (if calculated)\n"
            "Interpretation of significance (typical): * p<0.05; ** p<0.01; *** p<0.001"
        )
        ws.merge_range('A2:H2', pw_explanation, fmt["explanation"])  # Increased merge range
        ws.set_row(1, ResultsExporter.get_text_height(pw_explanation, 22*8))  # Adjusted width factor
    
        # Header
        headers = ["Group 1", "Group 2", "Test", "p-Value", "Corrected", "Significant", "Effect size", "95% CI Difference"]
        for col, header in enumerate(headers):
            ws.write(3, col, header, fmt["header"])
    
        # Data
        comps = results.get("pairwise_comparisons", [])
        if comps is None:  # Extra safety check
            comps = []
            print("WARNING: pairwise_comparisons was None, converted to empty list")
        
        print(f"DEBUG: comps type = {type(comps)}, content = {str(comps[:3]) if comps else 'empty'}")
        print(f"DEBUG: comps type = {type(comps)}, content = {comps[:3]}...")
        row = 4
    
        if len(comps) == 0:
            message = "No pairwise comparisons performed or available."
            if results.get("p_value") is not None and results.get("p_value") >= results.get("alpha", 0.05) and len(results.get("groups", [])) > 2:
                message = "No pairwise comparisons performed because the main test was not significant."
            elif results.get("error") and "Post-hoc" in results.get("error"):
                message = f"Error in post-hoc tests: {results.get('error')}"
    
            ws.merge_range(row, 0, row, len(headers)-1, message, fmt["cell"])
            return
    
        for comp_idx, comp in enumerate(comps):
            group1 = str(comp.get('group1', 'N/A'))
            group2 = str(comp.get('group2', 'N/A'))
            test_name = comp.get('test', posthoc_test_name or 'N/A')
            pval = comp.get('p_value', None)
    
            # Correction info
            corrected_info = "N/A"
            if comp.get('corrected') is True:
                corrected_info = comp.get('correction', 'Yes') if comp.get('correction') else 'Yes'
            elif comp.get('corrected') is False:
                corrected_info = "No"
    
            is_sign = comp.get('significant', False)
            if pval is not None and not isinstance(is_sign, bool):  # Fallback if 'significant' field is missing
                is_sign = pval < results.get("alpha", 0.05)
    
            effect_size_val = comp.get('effect_size', None)
            effect_size_type = comp.get('effect_size_type', '')
    
            pval_str = "N/A"
            if isinstance(pval, (float, int)):
                if pval < 0.001:
                    pval_str = "<0.001"
                else:
                    pval_str = f"{pval:.4f}"
    
            sign_str = "Yes" if is_sign else "No"
    
            eff_text = "N/A"
            eff_fmt = fmt["cell"]
            if isinstance(effect_size_val, (float, int)):
                magnitude = ""
                # Simplified magnitude for pairwise comparisons
                if effect_size_type.lower() in ["cohen_d", "hedges_g", "r"]:
                    if abs(effect_size_val) < 0.2:
                        magnitude = "very small"
                        eff_fmt = fmt["effect_weak"]
                    elif abs(effect_size_val) < 0.5:
                        magnitude = "small"
                        eff_fmt = fmt["effect_weak"]
                    elif abs(effect_size_val) < 0.8:
                        magnitude = "medium"
                        eff_fmt = fmt["effect_medium"]
                    else:
                        magnitude = "large"
                        eff_fmt = fmt["effect_strong"]
                eff_text = f"{effect_size_val:.3f}"
                if magnitude:
                    eff_text += f" ({magnitude})"
    
            ci_val = comp.get('confidence_interval', None)
            ci_str = "N/A"
            if ci_val and isinstance(ci_val, (tuple, list)) and len(ci_val) == 2 and ci_val[0] is not None and ci_val[1] is not None:
                ci_str = f"[{ci_val[0]:.3f}, {ci_val[1]:.3f}]"
    
            current_row_data = [group1, group2, test_name, pval_str, corrected_info, sign_str, eff_text, ci_str]
    
            for col, val_to_write in enumerate(current_row_data):
                current_fmt = fmt["cell"]
                if headers[col] == "p-Value" and is_sign:
                    current_fmt = fmt["significant"]
                elif headers[col] == "Significant" and is_sign:
                    current_fmt = fmt["significant"]
                elif headers[col] == "Effect size" and isinstance(effect_size_val, (float, int)):
                    current_fmt = eff_fmt  # Use pre-determined format for effect size
                ws.write(row + comp_idx, col, val_to_write, current_fmt)
    
    @staticmethod
    def _write_decision_tree_sheet(workbook, results, fmt, sheet_name="Decision Tree", pre_generated_tree=None):
        """Write decision tree sheet with visualization."""
        from decisiontreevisualizer import DecisionTreeVisualizer
        
        sheet = workbook.add_worksheet(sheet_name)
        sheet.set_column('A:A', 120)  # Wide column for the image
        
        # Write header
        sheet.write(0, 0, "Decision Tree Visualization", fmt["title"])
        sheet.write(1, 0, "Test Methodology: This decision tree shows the hypothesis workflow and statistical decisions.", fmt["explanation"])
        sheet.write(2, 0, "Highlighted path: The red path shows the decisions made for this specific analysis.", fmt["explanation"])
        
        # Use pre-generated path if provided, otherwise generate a new one
        image_path = pre_generated_tree
        if not image_path or not os.path.exists(image_path):
            print(f"DEBUG: No valid pre-generated tree, generating new one...")
            image_path = DecisionTreeVisualizer.generate_and_save_for_excel(results)
            # Track the newly generated file
            ResultsExporter.track_temp_file(image_path)
        
        # Insert the image if it exists
        if image_path and os.path.exists(image_path):
            print(f"Inserting decision tree image: {image_path}")
            
            # Get image dimensions to scale appropriately
            try:
                from PIL import Image
                with Image.open(image_path) as img:
                    width, height = img.size
                    aspect_ratio = width / height
                    print(f"DEBUG: Image dimensions: {width}x{height}, aspect ratio: {aspect_ratio:.2f}")
                    
                    # Scale to fit within Excel cell constraints
                    scale_factor = 0.75 if width > 4000 else 1.0
                    new_width = int(width * scale_factor)
                    new_height = int(height * scale_factor)
                    print(f"DEBUG: Scale factor: {scale_factor}, resulting size: {new_width}x{new_height}")
                    
                    # Insert image at row 5
                    sheet.insert_image(5, 0, image_path, {'x_scale': scale_factor, 'y_scale': scale_factor})
                    print(f"Successfully inserted decision tree image at row 5")
                    
            except Exception as e:
                print(f"DEBUG: Error processing image dimensions: {e}")
                # Fallback: insert without scaling
                sheet.insert_image(5, 0, image_path)
            
            # Add image filename for reference
            sheet.write(3, 0, f"Image file: {os.path.basename(image_path)}", fmt["explanation"])
        else:
            sheet.write(5, 0, "Error: Failed to generate decision tree visualization.", fmt["explanation"])
        
        return image_path
    
    @staticmethod
    def _write_rawdata_sheet(workbook, results, fmt, sheet_name="Raw Data"):
        ws = workbook.add_worksheet(sheet_name)
        ws.set_column(0, 10, 15)
        
        # Title and description
        ws.merge_range('A1:K1', "RAW DATA", fmt["title"])
        ws.write('A3', "This sheet shows the original data and processing steps for each group.", fmt["explanation"])
        ws.write('A4', "These data are the basis of all calculations.", fmt["explanation"])
        
        # Check if this is a non-parametric test with special data storage
        if results.get("test_type") == "non-parametric":
            # Handle non-parametric test data
            original_data = results.get("original_data", {})
            aggregated_data = results.get("aggregated_data", {})
            ranked_data = results.get("ranked_data", {})
            
            if original_data or aggregated_data or ranked_data:
                row = 6
                
                # Original Data Section
                ws.merge_range(f'A{row}:K{row}', "ORIGINAL DATA (Before any processing)", fmt["section_header"])
                row += 1
                ws.write(row, 0, "Group", fmt["header"])
                ws.write(row, 1, "Original Values", fmt["header"])
                row += 1
                
                for group_name, values in original_data.items():
                    ws.write(row, 0, group_name, fmt["cell"])
                    if values:
                        values_str = ", ".join([f"{v:.4f}" if isinstance(v, (int, float)) else str(v) for v in values])
                        ws.write(row, 1, values_str, fmt["cell"])
                    else:
                        ws.write(row, 1, "No data", fmt["cell"])
                    row += 1
                
                row += 1
                
                # Aggregated Data Section (if different from original)
                if aggregated_data and any(original_data.get(k, []) != aggregated_data.get(k, []) for k in aggregated_data.keys()):
                    ws.merge_range(f'A{row}:K{row}', "AGGREGATED DATA (Means of replicates)", fmt["section_header"])
                    row += 1
                    ws.write(row, 0, "Group", fmt["header"])
                    ws.write(row, 1, "Aggregated Values", fmt["header"])
                    row += 1
                    
                    for group_name, values in aggregated_data.items():
                        ws.write(row, 0, group_name, fmt["cell"])
                        if values:
                            values_str = ", ".join([f"{v:.4f}" if isinstance(v, (int, float)) else str(v) for v in values])
                            ws.write(row, 1, values_str, fmt["cell"])
                        else:
                            ws.write(row, 1, "No data", fmt["cell"])
                        row += 1
                    
                    row += 1
                
                # Ranked Data Section
                if ranked_data:
                    ws.merge_range(f'A{row}:K{row}', "RANKED DATA (Used for statistical test)", fmt["section_header"])
                    row += 1
                    ws.write(row, 0, "Group", fmt["header"])
                    ws.write(row, 1, "Ranked Values", fmt["header"])
                    row += 1
                    
                    for group_name, values in ranked_data.items():
                        ws.write(row, 0, group_name, fmt["cell"])
                        if values:
                            values_str = ", ".join([f"{v:.2f}" if isinstance(v, (int, float)) else str(v) for v in values])
                            ws.write(row, 1, values_str, fmt["cell"])
                        else:
                            ws.write(row, 1, "No data", fmt["cell"])
                        row += 1
                    
                    # Add explanation
                    row += 2
                    explanation = results.get("data_explanation", {})
                    if explanation:
                        ws.merge_range(f'A{row}:K{row}', "DATA PROCESSING EXPLANATION", fmt["section_header"])
                        row += 1
                        for key, value in explanation.items():
                            ws.write(row, 0, key.replace("_", " ").title() + ":", fmt["key"])
                            ws.write(row, 1, str(value), fmt["explanation"])
                            row += 1
                    
                    return
        
        # Handle regular parametric test data or fallback
        raw_data = results.get("raw_data", {})
        transformed_data = results.get("raw_data_transformed", {})
        
        if not raw_data and not transformed_data:
            # Try to get data from descriptive statistics
            descriptive = results.get("descriptive", {})
            if descriptive:
                row = 6
                ws.write(row, 0, "Group", fmt["header"])
                ws.write(row, 1, "Sample Size", fmt["header"])
                ws.write(row, 2, "Mean", fmt["header"])
                ws.write(row, 3, "Std Dev", fmt["header"])
                row += 1
                
                for group_name, stats in descriptive.items():
                    ws.write(row, 0, group_name, fmt["cell"])
                    ws.write(row, 1, stats.get("n", "N/A"), fmt["cell"])
                    ws.write(row, 2, f"{stats.get('mean', 0):.4f}" if stats.get('mean') is not None else "N/A", fmt["cell"])
                    ws.write(row, 3, f"{stats.get('std', 0):.4f}" if stats.get('std') is not None else "N/A", fmt["cell"])
                    row += 1
            else:
                ws.write(6, 0, "Group", fmt["header"])
                ws.write(6, 1, "Original Value", fmt["header"])
                ws.write(7, 0, "No data available", fmt["cell"])
            return
        
        # Handle parametric test data
        row = 6
        if raw_data:
            ws.merge_range(f'A{row}:K{row}', "ORIGINAL DATA", fmt["section_header"])
            row += 1
            ws.write(row, 0, "Group", fmt["header"])
            ws.write(row, 1, "Original Values", fmt["header"])
            row += 1
            
            for group_name, values in raw_data.items():
                ws.write(row, 0, group_name, fmt["cell"])
                if values:
                    values_str = ", ".join([f"{v:.4f}" if isinstance(v, (int, float)) else str(v) for v in values])
                    ws.write(row, 1, values_str, fmt["cell"])
                else:
                    ws.write(row, 1, "No data", fmt["cell"])
                row += 1
            
            row += 1
        
        if transformed_data and transformed_data != raw_data:
            ws.merge_range(f'A{row}:K{row}', "TRANSFORMED DATA", fmt["section_header"])
            row += 1
            ws.write(row, 0, "Group", fmt["header"])
            ws.write(row, 1, "Transformed Values", fmt["header"])
            row += 1
            
            for group_name, values in transformed_data.items():
                ws.write(row, 0, group_name, fmt["cell"])
                if values:
                    values_str = ", ".join([f"{v:.4f}" if isinstance(v, (int, float)) else str(v) for v in values])
                    ws.write(row, 1, values_str, fmt["cell"])
                else:
                    ws.write(row, 1, "No data", fmt["cell"])
                row += 1
            
    @staticmethod
    def _write_analysislog_sheet(workbook, log, fmt, sheet_name="Analysis Log"):
        ws = workbook.add_worksheet(sheet_name)
        ws.set_column(0, 0, 80)
        ws.set_row(0, 28)
        ws.write('A1', 'ANALYSIS LOG', fmt["title"])

        # Introduction/Legend
        log_explanation = (
            "This sheet documents the course of the statistical analysis and the decisions made. "
            "The log provides a chronological overview of the individual analysis steps, "
            "methods used, transformations, test selection, and special notes.\n"
            "Each paragraph describes a key step or decision in the analysis process."
        )
        ws.write(1, 0, log_explanation, fmt["explanation"])
        ws.set_row(1, ResultsExporter.get_text_height(log_explanation, 80))

        row = 3

        # Apply structured formatting to all logs
        if isinstance(log, str):
            # Split the log into clear sections
            sections = {
                "header": [],
                "setup": [],
                "analysis": [],
                "results": [],
                "posthoc": []
            }

            current_section = "header"
            lines = log.split('\n')

            # Enhanced section detection for all log types
            for line in lines:
                line = line.strip()
                if not line:
                    continue

                # Section detection - handle both advanced and basic test logs
                if "--- ANALYSE ---" in line or "--- ANALYSIS ---" in line:
                    current_section = "analysis"
                    continue
                elif "Testempfehlung:" in line or "Test recommendation:" in line:
                    current_section = "analysis"
                elif "Durchgeführter Test:" in line or "Test performed:" in line or "Two-Way ANOVA" in line or "t-test" in line or "ANOVA" in line:
                    current_section = "results"
                elif "paarweise Vergleiche" in line or "Post-hoc" in line or "Pairwise comparisons:" in line:
                    current_section = "posthoc"
                elif line.startswith("Datei:") or line.startswith("Arbeitsblatt:") or line.startswith("File:") or line.startswith("Worksheet:") or line.startswith("Group column:") or line.startswith("Value column"):
                    current_section = "setup"

                sections[current_section].append(line)

            # Write each section with consistent formatting
            # Header section
            for line in sections["header"]:
                ws.write(row, 0, line, fmt["cell"])
                row += 1
            row += 1  # Empty line after header

            # Setup section - file info, columns, groups
            if sections["setup"]:
                ws.write(row, 0, "DATASET INFORMATION", fmt["section_header"])
                row += 1
                for line in sections["setup"]:
                    ws.write(row, 0, line, fmt["cell"])
                    row += 1
                row += 1  # Empty line after setup

            # Analysis section - tests performed
            if sections["analysis"]:
                ws.write(row, 0, "ANALYSIS PREPARATION AND ASSUMPTIONS", fmt["section_header"])
                row += 1
                for line in sections["analysis"]:
                    ws.write(row, 0, line, fmt["cell"])
                    row += 1
                row += 1  # Empty line after analysis

            # Results section - main results
            if sections["results"]:
                ws.write(row, 0, "MAIN RESULTS", fmt["section_header"])
                row += 1
                
                # Format all results with bullet points for improved readability
                for line in sections["results"]:
                    if ":" in line and ("p =" in line or "p=" in line or "p <" in line or "p<" in line):
                        # This looks like a result line - add bullet point and highlight if significant
                        ws.write(row, 0, f"• {line}", fmt["significant"] if "significant" in line.lower() else fmt["cell"])
                    else:
                        ws.write(row, 0, line, fmt["cell"])
                    row += 1
                    
                row += 1  # Empty line after results

            # Post-hoc section
            if sections["posthoc"]:
                ws.write(row, 0, "POST-HOC ANALYSES", fmt["section_header"])
                row += 1
                
                # Add heading for post-hoc tests
                ws.write(row, 0, "Pairwise Comparisons (Post-hoc):", fmt["key"])
                row += 1
                
                # Format post-hoc results with bullet points for all tests
                for line in sections["posthoc"]:
                    if " vs " in line and ("p =" in line or "p=" in line or "p <" in line or "p<" in line):
                        # This is a comparison line - add bullet point and highlight if significant
                        is_significant = "significant" in line.lower() and "not significant" not in line.lower()
                        ws.write(row, 0, f"• {line}", fmt["significant"] if is_significant else fmt["cell"])
                    else:
                        ws.write(row, 0, line, fmt["cell"])
                    row += 1
                    
                row += 1  # Empty line after post-hoc

        elif isinstance(log, list):
            for entry in log:
                ws.write(row, 0, str(entry), fmt["cell"])
                row += 1
        else:
            ws.write(row, 0, str(log), fmt["cell"])

        # Add summary section for all tests when significant results are found
        if isinstance(log, str) and "significant" in log.lower() and "p < 0.05" in log:
            ws.write(row, 0, "SUMMARY OF RESULTS", fmt["section_header"])
            row += 1

            summary_text = (
                "The statistical analysis revealed significant effects. "
                "Please refer to the detailed results for the specific findings. "
                "Post-hoc tests were performed for more precise group comparisons where applicable."
            )
            ws.write(row, 0, summary_text, fmt["explanation"])
            ws.set_row(row, ResultsExporter.get_text_height(summary_text, 80))
                
    @staticmethod
    def get_text_height(text, width):
        """
        Calculates the approximate row height based on text length and cell width,
        with caching to optimize repeated calls.
        """
        # Static cache for already calculated heights
        if not hasattr(ResultsExporter.get_text_height, "_cache"):
            ResultsExporter.get_text_height._cache = {}
    
        # Build cache key (text length + width as approximation)
        cache_key = (len(text), width)
    
        # Return from cache if already calculated
        if cache_key in ResultsExporter.get_text_height._cache:
            return ResultsExporter.get_text_height._cache[cache_key]
    
        # Default height for empty text
        if not text:
            return 15
    
        # Count actual lines in the text
        lines = text.count('\n') + 1
    
        # Estimate additional lines based on text length and cell width
        avg_chars_per_line = width * 0.8
        text_length = len(text)
        estimated_lines = text_length / avg_chars_per_line
    
        # Take the higher value
        total_lines = max(lines, estimated_lines)
    
        # 15 points per line plus some spacing
        result = max(15, int(total_lines * 15 + 5))
    
        # Store result in cache
        ResultsExporter.get_text_height._cache[cache_key] = result
    
        return result
    
    @staticmethod
    def track_temp_file(filepath):
        """Track a temporary file for later cleanup"""
        if filepath and os.path.exists(filepath):
            ResultsExporter._temp_files.add(filepath)
            print(f"DEBUG: Tracking temporary file: {filepath}")
        return filepath
    
    @staticmethod
    def cleanup_old_tree_files(max_age_hours=24):
        """Clean up old decision tree files."""
        import glob
        import time
        import os
        import tempfile
        
        # First check temp directory for pattern-matching files
        temp_dir = tempfile.gettempdir()
        current_time = time.time()
        cleaned_count = 0
        
        # Find and delete old files
        for pattern in ["decision_tree_*.png", "tree_visualization_*.png"]:
            for file_path in glob.glob(os.path.join(temp_dir, pattern)):
                try:
                    file_age = current_time - os.path.getctime(file_path)
                    if file_age > max_age_hours * 3600:
                        os.remove(file_path)
                        cleaned_count += 1
                        print(f"Removed old decision tree: {file_path}")
                except Exception as e:
                    print(f"Error cleaning up file {file_path}: {str(e)}")
        
        # Also check legacy location (Documents/StatisticsTemp)
        docs_dir = os.path.join(os.path.expanduser("~"), "Documents", "StatisticsTemp")
        if os.path.exists(docs_dir):
            for pattern in ["decision_tree_*.png", "tree_visualization_*.png"]:
                for file_path in glob.glob(os.path.join(docs_dir, pattern)):
                    try:
                        file_age = current_time - os.path.getctime(file_path)
                        if file_age > max_age_hours * 3600:
                            os.remove(file_path)
                            cleaned_count += 1
                            print(f"Removed old decision tree from legacy location: {file_path}")
                    except Exception as e:
                        print(f"Error cleaning up legacy file {file_path}: {str(e)}")
        
        return cleaned_count

class DatasetSelector:
    """Helper class to manage dataset selection in the UI"""
    
    @staticmethod
    def get_available_datasets(file_path, sheet_name=None):
        """
        Get all available datasets (sheets) from an Excel file
        
        Returns:
        --------
        dict: {sheet_name: preview_info}
        """
        try:
            if file_path.endswith('.xlsx') or file_path.endswith('.xls'):
                # Get all sheet names
                xl_file = pd.ExcelFile(file_path)
                datasets = {}
                
                for sheet in xl_file.sheet_names:
                    try:
                        # Get a preview of each sheet
                        df_preview = pd.read_excel(file_path, sheet_name=sheet, nrows=5)
                        datasets[sheet] = {
                            'columns': df_preview.columns.tolist(),
                            'shape': f"{len(pd.read_excel(file_path, sheet_name=sheet))} rows",
                            'preview': df_preview.head(3).to_dict('records')
                        }
                    except Exception as e:
                        datasets[sheet] = {'error': str(e)}
                
                return datasets
            else:
                # For CSV files, return single dataset
                df_preview = pd.read_csv(file_path, nrows=5)
                return {
                    'CSV Data': {
                        'columns': df_preview.columns.tolist(),
                        'shape': f"{len(pd.read_csv(file_path))} rows",
                        'preview': df_preview.head(3).to_dict('records')
                    }
                }
        except Exception as e:
            return {'Error': {'error': str(e)}}

# Modified AnalysisManager.analyze function
class AnalysisManager:
    @staticmethod
    def analyze(file_path, group_col, groups, sheet_name=0, value_cols=None, 
                selected_datasets=None, combine_columns=False, width=12, height=10, 
                dependent=False, compare=None, colors=None, hatches=None,
                title=None, x_label=None, y_label=None, file_name=None, 
                save_plot=True, skip_plots=False, error_type="sd", skip_excel=False, 
                dataset_name=None, additional_factors=None, show_individual_lines=True, 
                **kwargs):
        """
        Performs analysis on one or multiple datasets.

        New Parameters:
        ---------------
        selected_datasets : list or None
            List of sheet names to analyze. If None, analyzes single dataset.
            If provided, analyzes multiple datasets and creates combined Excel output.
        
        All other parameters remain the same as before.
        
        Returns:
        --------
        dict or list
            If single dataset: returns analysis results dict
            If multiple datasets: returns list of results dicts with combined Excel file
        """
        
        # Single dataset analysis (existing functionality)
        if selected_datasets is None or len(selected_datasets) <= 1:
            # Use existing single dataset logic
            actual_sheet = selected_datasets[0] if selected_datasets else sheet_name
            return AnalysisManager._analyze_single_dataset(
                file_path, group_col, groups, actual_sheet, value_cols, 
                combine_columns, width, height, dependent, compare, colors, hatches,
                title, x_label, y_label, file_name, save_plot, skip_plots, 
                error_type, skip_excel, dataset_name, additional_factors, 
                show_individual_lines, **kwargs
            )
        
        # Multiple dataset analysis
        else:
            return AnalysisManager._analyze_multiple_datasets(
                file_path, group_col, groups, selected_datasets, value_cols,
                combine_columns, width, height, dependent, compare, colors, hatches,
                title, x_label, y_label, file_name, save_plot, skip_plots,
                error_type, skip_excel, additional_factors, show_individual_lines, **kwargs
            )
            
    @staticmethod
    def _analyze_multiple_datasets(file_path, group_col, groups, selected_datasets, value_cols,
                                  combine_columns, width, height, dependent, compare, colors, hatches,
                                  title, x_label, y_label, file_name, save_plot, skip_plots,
                                  error_type, skip_excel, additional_factors, show_individual_lines, **kwargs):
        """
        Multiple dataset analysis with unified Excel output
        """
        from datetime import datetime
        
        all_results = {}
        failed_datasets = {}
        
        print(f"Starting analysis of {len(selected_datasets)} datasets...")
        
        # Analyze each selected dataset
        for i, dataset_name in enumerate(selected_datasets):
            print(f"Analyzing dataset {i+1}/{len(selected_datasets)}: {dataset_name}")
            
            try:
                # Reset dialog cache for each dataset to ensure fresh prompts
                UIDialogManager.reset_dialog_cache()
                
                # Analyze single dataset
                result = AnalysisManager._analyze_single_dataset(
                    file_path=file_path,
                    group_col=group_col,
                    groups=groups,
                    sheet_name=dataset_name,
                    value_cols=value_cols,
                    combine_columns=combine_columns,
                    width=width,
                    height=height,
                    dependent=dependent,
                    compare=compare,
                    colors=colors,
                    hatches=hatches,
                    title=f"{title} - {dataset_name}" if title else dataset_name,
                    x_label=x_label,
                    y_label=y_label,
                    file_name=f"{file_name}_{dataset_name}" if file_name else dataset_name,
                    save_plot=save_plot,
                    skip_plots=skip_plots,
                    error_type=error_type,
                    skip_excel=True,  # Skip individual Excel files
                    dataset_name=dataset_name,
                    additional_factors=additional_factors,
                    show_individual_lines=show_individual_lines,
                    dialog_progress=f"({i+1}/{len(selected_datasets)})",
                    dialog_column=dataset_name,
                    **kwargs
                )
                
                if "error" in result:
                    failed_datasets[dataset_name] = result["error"]
                    print(f"ERROR analyzing {dataset_name}: {result['error']}")
                else:
                    all_results[dataset_name] = result
                    print(f"Successfully analyzed {dataset_name}")
                    
            except Exception as e:
                error_msg = f"Exception during analysis: {str(e)}"
                failed_datasets[dataset_name] = error_msg
                print(f"ERROR analyzing {dataset_name}: {error_msg}")
        
        # Create combined Excel output
        if all_results:
            base_name = file_name if file_name else "multi_dataset_analysis"
            excel_path = f"{base_name}_combined_results.xlsx"
            
            try:
                ResultsExporter.export_multi_dataset_results(all_results, excel_path)
                print(f"Combined results saved to: {excel_path}")
            except Exception as e:
                print(f"Error creating combined Excel file: {str(e)}")
        
        # Return summary
        return {
            "type": "multi_dataset_analysis",
            "successful_datasets": list(all_results.keys()),
            "failed_datasets": failed_datasets,
            "results": all_results,
            "combined_excel": excel_path if all_results else None,
            "summary": {
                "total_datasets": len(selected_datasets),
                "successful": len(all_results),
                "failed": len(failed_datasets),
                "success_rate": f"{len(all_results)/len(selected_datasets)*100:.1f}%"
            }
        }
            
    @staticmethod
    def _analyze_single_dataset(file_path, group_col, groups, sheet_name, value_cols, 
                               combine_columns, width, height, dependent, compare, colors, hatches,
                               title, x_label, y_label, file_name, save_plot, skip_plots, 
                               error_type, skip_excel, dataset_name, additional_factors, 
                               show_individual_lines, **kwargs):
        from stats_functions import ResultsExporter
        # Basic parameter validation
        if not file_path or not os.path.exists(file_path):
            raise ValueError("Please specify a valid file")
        if not groups:
            raise ValueError("Please specify at least one group")
        if not group_col:
            raise ValueError("Please specify a valid group column")
        if error_type not in ["sd", "se"]:
            raise ValueError("Error bar type must be 'sd' or 'se'")

        from datetime import datetime
        analysis_log = f"Analysis Report\n"
        analysis_log += f"Date and time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
        analysis_log += f"File: {file_path}\n"
        analysis_log += f"Worksheet: {sheet_name}\n"
        analysis_log += f"Group column: {group_col}\n"
        analysis_log += f"Value column(s): {', '.join(value_cols) if value_cols else 'All numeric columns'}\n"
        analysis_log += f"Groups to analyze: {', '.join(map(str, groups))}\n"
        analysis_log += f"Dependent samples: {'Yes' if dependent else 'No'}\n"
        analysis_log += f"Error bar type: {'SEM (standard error)' if error_type == 'se' else 'SD (standard deviation)'}\n"

        if compare:
            compare_str = ", ".join([f"{g1} vs {g2}" for g1, g2 in compare])
            analysis_log += f"Specific comparisons: {compare_str}\n"

        analysis_log += "\n--- ANALYSIS ---\n\n"

        try:
            # Import and filter data
            samples, df = DataImporter.import_data(file_path, sheet_name=sheet_name, group_col=group_col, 
                                                  value_cols=value_cols, combine_columns=combine_columns)
            filtered_samples = {g: samples[g] for g in groups if g in samples}

            # Validations and logging
            if not filtered_samples:
                raise ValueError(f"None of the specified groups were found in the data. Available groups: {list(samples.keys())}")

            for group, values in filtered_samples.items():
                if len(values) < 1:
                    raise ValueError(f"Group '{group}' contains no data.")

            analysis_log += f"Data imported successfully.\n"
            analysis_log += "Number of data points per group:\n"
            for group, values in filtered_samples.items():
                analysis_log += f"  {group}: {len(values)} data points\n"

            # Initialize the result dictionary (important: before first assignments!)
            results = {}

            # Normality and variance check with dataset name
            transformed_samples, test_recommendation, test_info = StatisticalTester.check_normality_and_variance(
                groups,
                filtered_samples,
                dataset_name=dataset_name,
                progress_text=kwargs.get('dialog_progress', None),
                column_name=kwargs.get('dialog_column', None)
            )
            print(f"DEBUG: Test recommendation is '{test_recommendation}'")
            print(f"DEBUG: Test info transformation: '{test_info.get('transformation')}'")

            # Write test recommendation to log
            analysis_log += f"\nTest recommendation: {test_recommendation}\n"

            # For dependent samples, perform additional validation
            if dependent:
                validation = StatisticalTester.validate_dependent_data(filtered_samples, groups)
                if not validation["valid"]:
                    error_message = "Error validating dependent data:\n" + "\n".join(validation["messages"])
                    analysis_log += f"\n{error_message}\n"
                    if not kwargs.get('force_continue', False):
                        print(f"WARNING: {error_message}")
                        analysis_log += "\nAnalysis continues with warning, results may be unreliable."
                        
            # Adjust test type based on recommendation if auto-selection is enabled
            if kwargs.get('auto_nonparametric', True) and test_recommendation == 'non_parametric':
                if kwargs.get('test') == 'two_way_anova':
                    kwargs['test'] = 'nonparametric_two_way_anova'
                    analysis_log += "\nSwitching to nonparametric two-way ANOVA based on assumption check.\n"
                elif kwargs.get('test') == 'repeated_measures_anova':
                    kwargs['test'] = 'nonparametric_rm_anova'
                    analysis_log += "\nSwitching to nonparametric repeated measures ANOVA based on assumption check.\n"
                elif kwargs.get('test') == 'mixed_anova':
                    kwargs['test'] = 'nonparametric_mixed_anova'
                    analysis_log += "\nSwitching to nonparametric mixed ANOVA based on assumption check.\n"            

            # Perform the appropriate statistical test - only call ONCE
            if kwargs.get('test') == 'mixed_anova':
                between_factor, within_factor = kwargs['additional_factors']
                # Step 3: Call prepare_advanced_test first
                prep = StatisticalTester.prepare_advanced_test(
                    df, 'mixed_anova', value_cols[0], 'Subject', [between_factor], [within_factor]
                )
                if "error" in prep:
                    return prep  # or handle error

                # Step 4: Pass outputs to perform_advanced_test
                results = StatisticalTester.perform_advanced_test(
                    df=df,
                    test='mixed_anova',
                    dv=value_cols[0],
                    subject='Subject',
                    between=[between_factor],
                    within=[within_factor],
                    alpha=0.05,
                    transformed_samples=prep["transformed_samples"],
                    recommendation=prep["recommendation"],
                    test_info=prep["test_info"],
                    transform_fn=None,
                    force_parametric=kwargs.get('force_parametric', False)
                )
                # Get the transformation type from the test_info
                requested_transform = prep["test_info"].get("transformation", "None")
                print(f"DEBUG: Requested transformation: {requested_transform}")
                print(f"DEBUG: Applied transformation: {results.get('transformation')}")
            else:
                # Standard path for simple tests
                test_results = StatisticalTester.perform_statistical_test(
                    groups, transformed_samples, filtered_samples,
                    dependent=dependent, test_recommendation=test_recommendation, test_info=test_info
                )

                
            if kwargs.get('test') == 'nonparametric_two_way_anova':
                # DISABLED: Nonparametric fallbacks are not yet supported
                # from nonparametricanovas import NonParametricTwoWayANOVA, NonParametricFactory
                
                # Return informational message instead of running nonparametric test
                test_results = {
                    "test": "Two-Way ANOVA (parametric assumptions violated)",
                    "recommendation": "non_parametric", 
                    "error": "Nonparametric alternatives are currently disabled. The NonParametricTwoWayANOVA class is available in nonparametricanovas.py but nonparametric fallbacks are disabled.",
                    "parametric_violated": True,
                    "suggested_alternative": "NonParametricTwoWayANOVA class (rank transformation + permutation test)"
                }
                
                # Create standardized results for Excel export
                results = {
                    "test_results": test_results,
                    "transformed_samples": transformed_samples,
                    "samples": filtered_samples,
                    "transformation": test_info.get("transformation"),
                    "normality_tests": test_info["normality_tests"],
                    "variance_test": test_info["variance_test"],
                    "test_type": "non-parametric",
                    "groups": groups,
                    "permutation_test": False,
                    "analysis_log": "Nonparametric Two-Way ANOVA requested but not supported"
                }
                
            elif kwargs.get('test') == 'nonparametric_rm_anova':
                # DISABLED: Nonparametric fallbacks are not yet supported
                # from nonparametricanovas import NonParametricFactory
                
                # Return informational message instead of running nonparametric test  
                test_results = {
                    "test": "Repeated-Measures ANOVA (parametric assumptions violated)",
                    "recommendation": "non_parametric",
                    "error": "Nonparametric alternatives are currently disabled. The NonParametricRMANOVA class is available in nonparametricanovas.py but nonparametric fallbacks are disabled.",
                    "parametric_violated": True,
                    "suggested_alternative": "NonParametricRMANOVA class (rank transformation + permutation test)"
                }
                
                # Create standardized results for Excel export
                results = {
                    "test_results": test_results,
                    "transformed_samples": transformed_samples,
                    "samples": filtered_samples,
                    "transformation": test_info.get("transformation"),
                    "normality_tests": test_info["normality_tests"],
                    "variance_test": test_info["variance_test"],
                    "test_type": "non-parametric",
                    "groups": groups,
                    "analysis_log": "Nonparametric Repeated-Measures ANOVA requested but not supported"
                }
                
                # Handle Excel export if needed
                if not skip_excel:
                    from stats_functions import ResultsExporter
                    ResultsExporter.export_results_to_excel(results, excel_file)
            
            elif kwargs.get('test') == 'nonparametric_mixed_anova':
                # DISABLED: Nonparametric fallbacks are not yet supported
                # from nonparametricanovas import NonParametricMixedANOVA, NonParametricFactory
                
                # Return informational message instead of running nonparametric test
                test_results = {
                    "test": "Mixed-Design ANOVA (parametric assumptions violated)",
                    "recommendation": "non_parametric",
                    "error": "Nonparametric alternatives are currently disabled. The NonParametricMixedANOVA class is available in nonparametricanovas.py but nonparametric fallbacks are disabled.",
                    "parametric_violated": True,
                    "suggested_alternative": "NonParametricMixedANOVA class (rank transformation + permutation test)"
                }
                
                # Create standardized results for Excel export
                results = {
                    "test_results": test_results,
                    "transformed_samples": transformed_samples,
                    "samples": filtered_samples,
                    "transformation": test_info.get("transformation"),
                    "normality_tests": test_info["normality_tests"],
                    "variance_test": test_info["variance_test"],
                    "test_type": "non-parametric",
                    "groups": groups,
                    "analysis_log": "Nonparametric Mixed-Design ANOVA requested but not supported"
                }    

            # Log before transformation
            analysis_log += "\nResults of tests before transformation:\n"
            all_data_normality = test_info["normality_tests"].get("all_data")
            if all_data_normality and all_data_normality.get("p_value") is not None:
                analysis_log += f"Shapiro-Wilk test (normality): p = {all_data_normality['p_value']:.4f} - "
                analysis_log += "Normally distributed\n" if all_data_normality.get('is_normal', False) else "Not normally distributed\n"
            else:
                analysis_log += "Shapiro-Wilk test (normality): Test not performed (insufficient data)\n"

            if test_info["variance_test"].get("p_value") is not None:
                analysis_log += f"Levene test (variance homogeneity): p = {test_info['variance_test']['p_value']:.4f} - "
                analysis_log += "Variances homogeneous\n" if test_info['variance_test'].get('equal_variance', False) else "Variances heterogeneous\n"
            else:
                analysis_log += "Levene test: Not performed (insufficient data)\n"

            # Log transformation
            if test_info.get("transformation"):
                analysis_log += f"\nTransformation: {test_info['transformation'].capitalize()} transformation performed.\n"
                # Log after transformation
                analysis_log += "Results of tests after transformation:\n"
                if test_info["normality_tests"].get("transformed_data", {}).get("p_value") is not None:
                    analysis_log += f"Shapiro-Wilk test (normality): p = {test_info['normality_tests']['transformed_data']['p_value']:.4f} - "
                    analysis_log += "Normally distributed\n" if test_info['normality_tests']['transformed_data'].get('is_normal', False) else "Not normally distributed\n"
                else:
                    analysis_log += "Shapiro-Wilk test (normality): Test not performed (insufficient data)\n"
                if test_info["variance_test"].get("transformed", {}).get("p_value") is not None:
                    analysis_log += f"Levene test (variance homogeneity): p = {test_info['variance_test']['transformed']['p_value']:.4f} - "
                    analysis_log += "Variances homogeneous\n" if test_info['variance_test']['transformed'].get('equal_variance', False) else "Variances heterogeneous\n"
                else:
                    analysis_log += "Levene test: Not performed (insufficient data)\n"
            else:
                analysis_log += "\nTransformation: No transformation performed.\n"

            posthoc_results = None

            if test_results.get('p_value') is not None and test_results['p_value'] < 0.05 and len(groups) > 2:
                # Significant result: perform post-hoc tests
                valid_groups = [g for g in groups if g in transformed_samples and len(transformed_samples[g]) > 0]
                print("DEBUG: valid_groups after filter:", valid_groups)
                print("DEBUG: transformed_samples:", {g: len(transformed_samples[g]) for g in transformed_samples})
                print("DEBUG: original_samples:", {g: len(filtered_samples[g]) for g in filtered_samples})

                test_name = test_results.get('test', '').lower()

                # Check if post-hoc tests have already been performed
                if not test_results.get('pairwise_comparisons'):
                    # Automatic selection for non-parametric tests
                    if 'kruskal' in test_name or 'friedman' in test_name or test_recommendation == 'non_parametric':
                        posthoc_choice = "dunn"
                        analysis_log += "\nAutomatically selected Dunn test as post-hoc for non-parametric test.\n"
                        posthoc_results = StatisticalTester.perform_refactored_posthoc_testing(
                            valid_groups, transformed_samples, test_recommendation,
                            alpha=0.05, posthoc_choice=posthoc_choice
                        )
                    else:
                        # Show dialog for parametric tests
                        posthoc_choice = UIDialogManager.select_posthoc_test_dialog(
                            progress_text=kwargs.get('dialog_progress', None),
                            column_name=kwargs.get('dialog_column', None)
                        )
                        if posthoc_choice and posthoc_choice != "none":
                            control_group = None
                            if posthoc_choice == "dunnett":
                                control_group = UIDialogManager.select_control_group_dialog(valid_groups)
                            posthoc_results = StatisticalTester.perform_refactored_posthoc_testing(
                                valid_groups, transformed_samples, test_recommendation,
                                alpha=0.05, posthoc_choice=posthoc_choice, control_group=control_group
                            )
                    # Process results uniformly - ONLY ONCE here!
                    if posthoc_results:                      
                        if posthoc_choice == "dunnett" and "control_group" in posthoc_results:
                            test_results["control_group"] = posthoc_results["control_group"]
                        if 'pairwise_comparisons' in posthoc_results:
                            import copy
                            test_results['pairwise_comparisons'] = copy.deepcopy(posthoc_results['pairwise_comparisons'])
      
                            print(f"DEBUG: Copied pairwise_comparisons from posthoc_results to test_results")
                            print(f"DEBUG: Same object? {posthoc_results.get('pairwise_comparisons', None) is test_results.get('pairwise_comparisons', None)}")

                        test_results["posthoc_test"] = posthoc_results.get("posthoc_test")

                        # Add debug print to verify
                        print(f"DEBUG: posthoc_results['pairwise_comparisons'] length: {len(posthoc_results.get('pairwise_comparisons', []))}")

                    # HIER DEBUG-AUSGABEN EINFÜGEN
                    print(f"DEBUG: Pairwise comparisons after post-hoc: {len(test_results.get('pairwise_comparisons', []))}")
                    
            # Nach der Post-hoc Verarbeitung, vor test_results.update:
            print(f"DEBUG: Post-hoc results: {posthoc_results.keys() if posthoc_results else None}")
            if posthoc_results and 'error' in posthoc_results and posthoc_results['error']:
                print(f"DEBUG: Post-hoc ERROR: {posthoc_results['error']}")
            print(f"DEBUG: test_results pairwise_comparisons: {len(test_results.get('pairwise_comparisons', []))} items")        


            # Make sure normality and variance test results are explicitly set
            if test_info and "normality_tests" in test_info:
                results["normality_tests"] = test_info["normality_tests"]
            if test_info and "variance_test" in test_info:
                results["variance_test"] = test_info["variance_test"]

            # Make sure test_type/recommendation is set:
            results["recommendation"] = test_recommendation

            # Merge important transformation and test info into results
            results.update(test_results)
            results.update({
                "transformed_samples": transformed_samples,
                "samples": filtered_samples,
                "transformation": test_info.get("transformation"),
                "normality_tests": test_info["normality_tests"],
                "variance_test": test_info["variance_test"],
                "test_type": test_recommendation
            })

            # Nach results.update(test_results):
            print(f"DEBUG: results pairwise_comparisons: {len(results.get('pairwise_comparisons', []))} items")            

            if "boxcox_lambda" in test_info:
                results["boxcox_lambda"] = test_info["boxcox_lambda"]

            analysis_log += f"\nTest performed: {results.get('test', 'Not specified')}\n"
            if 'p_value' in results:
                p_value = results['p_value']
                if isinstance(p_value, (float, int)):
                    analysis_log += f"p-Value: {p_value:.6f}\n"
                    analysis_log += f"Significance: {'Significant (p < 0.05)' if p_value < 0.05 else 'Not significant'}\n"
                else:
                    analysis_log += f"p-Value: {p_value}\n"
                    analysis_log += "Significance: Not determinable\n"

            # For Two-way/Mixed/RM-ANOVA: main effects and interactions in the log
            if "factors" in results:
                for factor in results["factors"]:
                    analysis_log += (
                        f"Main effect {factor['factor']}: "
                        f"F({factor['df1']}, {factor['df2']}) = {factor['F']:.3f}, "
                        f"p = {factor['p_value']:.4f}, "
                        f"Effect size: {factor.get('effect_size', 'N/A')}\n"
                    )
            if "interactions" in results:
                for inter in results["interactions"]:
                    analysis_log += (
                        f"Interaction {inter['factors'][0]} x {inter['factors'][1]}: "
                        f"F({inter['df1']}, {inter['df2']}) = {inter['F']:.3f}, "
                        f"p = {inter['p_value']:.4f}, "
                        f"Effect size: {inter.get('effect_size', 'N/A')}\n"
                    )

            # Post-hoc tests
            if 'pairwise_comparisons' in results and results['pairwise_comparisons']:
                posthoc_test = results.get("posthoc_test", None)
                if posthoc_test and "Tukey HSD" in posthoc_test:
                    posthoc_display = "Tukey HSD Test"
                elif posthoc_test and "Dunnett Test" in posthoc_test:
                    control_group = results.get("control_group", "")
                    posthoc_display = f"Dunnett Test (Control group: {control_group})"
                else:
                    posthoc_display = posthoc_test if posthoc_test else "No post-hoc test performed"

                analysis_log += f"\nPost‑hoc test: {posthoc_display}\n"
                analysis_log += "Pairwise comparisons:\n"
                for comp in results["pairwise_comparisons"]:
                    group1 = str(comp['group1'])
                    group2 = str(comp['group2'])
                    p_val = comp['p_value']
                    significant = comp['significant']
                    p_text = "p < 0.001" if isinstance(p_val, (float, int)) and p_val < 0.001 else safe_format(p_val, "p = {:.4f}")
                    sign_text = "significant" if significant else "not significant"
                    stars = "***" if significant and p_val < 0.001 else "**" if significant and p_val < 0.01 else "*" if significant else ""
                    analysis_log += f"  {group1} vs {group2}: {p_text}, {sign_text} {stars}\n"
            else:
                analysis_log += "\nNo pairwise comparisons were performed or calculated.\n"

            # Define file base based on custom name or groups
            if file_name:
                file_base = file_name
            else:
                file_base = "_".join(map(str, groups))

            excel_file = f"{file_base}_results.xlsx"
            
            results['groups'] = groups
            results['raw_data'] = {g: filtered_samples[g][:] for g in groups}
            if results.get('transformation', 'None') != 'None':
                results['raw_data_transformed'] = {g: transformed_samples[g][:] for g in groups}
            results["variance_homogeneity_test"] = test_info.get("variance_test", {})    

            # Add debug statements before Excel export
            print("DEBUG: Assumption tests before Excel export:")
            print("  Normality tests:", test_info.get("normality_tests", {}))
            print("  Variance tests:", test_info.get("variance_test", {}))
            print("  Test recommendation:", test_recommendation)
                
            # Export to Excel
            if not skip_excel:
                ResultsExporter.export_results_to_excel(results, excel_file, analysis_log)
                analysis_log += f"\nResults were saved to {excel_file}.\n"

            # Create the plot, if not skipped
            if not skip_plots:
                pairwise_comparisons = results.get('pairwise_comparisons', None)
                fig, ax = DataVisualizer.plot_bar(groups, filtered_samples, width=width, height=height, 
                                                  colors=colors, hatches=hatches, compare=compare, 
                                                  test_recommendation=test_recommendation,
                                                  x_label=x_label, y_label=y_label,
                                                  title=title, save_plot=save_plot, error_type=error_type,
                                                  pairwise_results=pairwise_comparisons,
                                                  file_name=file_base)
                analysis_log += f"\nPlots were saved as:\n"
                analysis_log += f"  {file_base}.pdf\n"
                analysis_log += f"  {file_base}.png\n"
                import matplotlib.pyplot as plt
                plt.close(fig)
                results["_file_paths"] = {
                    "excel": os.path.abspath(excel_file),
                    "pdf": os.path.abspath(f"{file_base}.pdf"),
                    "png": os.path.abspath(f"{file_base}.png")
                }
            else:
                results["_file_paths"] = {
                    "excel": os.path.abspath(excel_file)
                }
            # Special visualization for dependent data
            if dependent and not skip_plots:
                try:
                    line_fig, line_ax = DataVisualizer.plot_dependent_samples(
                        groups, filtered_samples, width=width, height=height,
                        colors=colors, title=f"{title} (dependent measurements)" if title else "Dependent measurements",
                        x_label=x_label, y_label=y_label,
                        save_plot=save_plot, file_name=file_base+"_lines",
                        show_individual=show_individual_lines
                    )
                    import matplotlib.pyplot as plt
                    plt.close(line_fig)
                    line_plot_base = file_base+"_lines"
                    results["_file_paths"]["pdf_lines"] = os.path.abspath(f"{line_plot_base}.pdf")
                    results["_file_paths"]["png_lines"] = os.path.abspath(f"{line_plot_base}.png")
                    analysis_log += f"\nAdditional line plot for dependent data created:\n"
                    analysis_log += f"  {line_plot_base}.pdf\n"
                    analysis_log += f"  {line_plot_base}.png\n"
                except Exception as e:
                    analysis_log += f"\nError creating line plot for dependent data: {str(e)}\n"
                    print(f"Error creating line plot: {str(e)}")
            if results.get('transformation', 'None') != 'None':
                results['transformed_data'] = transformed_samples

            # About line 5647, just before "return results"
            params = {
                "file_path": file_path,
                "sheet_name": sheet_name,
                "group_col": group_col,
                "value_cols": value_cols,
                "groups": groups,
                "dependent": dependent,
                "error_type": error_type
            }
            # Build protocol
            def build_analysis_log(results, params):
                log = []
                log.append("ANALYSIS LOG")
                log.append('"This sheet documents the course of the statistical analysis and the decisions made. The log provides a chronological overview of the individual analysis steps, methods used, transformations, test selection, and special notes.\nEach paragraph describes a key step or decision in the analysis process."\n')
                log.append(f"Analysis report\nDate and time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                if 'file_path' in params:
                    log.append(f"File: {params['file_path']}")
                if 'sheet_name' in params:
                    log.append(f"Worksheet: {params['sheet_name']}")
                if 'group_col' in params:
                    log.append(f"Group column: {params['group_col']}")
                if 'value_cols' in params:
                    log.append(f"Value column(s): {', '.join(params['value_cols'])}")
                if 'groups' in params:
                    log.append(f"Groups to analyze: {', '.join(params['groups'])}")
                if 'dependent' in params:
                    log.append(f"Dependent samples: {'Yes' if params['dependent'] else 'No'}")
                if 'error_type' in params:
                    log.append(f"Error bar type: {'SEM (standard error)' if params['error_type']=='se' else 'SD (standard deviation)'}")
                log.append("\n--- ANALYSIS ---\n")
                if results.get('import_status'):
                    log.append("Data imported successfully.")
                if 'group_sizes' in results:
                    log.append("Number of data points per group:")
                    for group, n in results['group_sizes'].items():
                        log.append(f"{group}: {n} data points")
                if 'test_recommendation' in results:
                    log.append(f"Test recommendation: {results['test_recommendation']}")
                if 'normality_p' in results:
                    log.append(f"Shapiro-Wilk test (normality): p = {results['normality_p']:.4f} - {'Normally distributed' if results['normality_p'] > 0.05 else 'Not normally distributed'}")
                if 'levene_p' in results:
                    log.append(f"Levene test (variance homogeneity): p = {results['levene_p']:.4f} - {'Variances homogeneous' if results['levene_p'] > 0.05 else 'Variances heterogeneous'}")
                if 'transformation' in results:
                    log.append(f"Transformation: {results['transformation'] if results['transformation'] else 'No transformation performed.'}")
                if 'test' in results:
                    log.append(f"Test performed: {results['test']}")
                if 'p_value' in results:
                    p_value = results['p_value']
                    if p_value is None:
                        log.append("p-Value: Not available (test may have failed)")
                        if 'error' in results and results['error']:
                            log.append(f"Error: {results['error']}")
                    elif isinstance(p_value, (float, int)):
                        log.append(f"p-Value: {p_value:.6f}")
                        log.append(f"Significance: {'Significant (p < 0.05)' if p_value < 0.05 else 'Not significant'}")
                    else:
                        log.append(f"p-Value: {p_value}")
                        log.append("Significance: Not determinable")
                if "factors" in results:
                    for factor in results["factors"]:
                        log.append(
                            f"Main effect {factor['factor']}: F({factor['df1']}, {factor['df2']}) = {factor['F']:.3f}, "
                            f"p = {factor['p_value']:.4f}, Effect size: {factor.get('effect_size', 'N/A')}"
                        )
                if "interactions" in results:
                    for inter in results["interactions"]:
                        log.append(
                            f"Interaction {inter['factors'][0]} x {inter['factors'][1]}: F({inter['df1']}, {inter['df2']}) = {inter['F']:.3f}, "
                            f"p = {inter['p_value']:.4f}, Effect size: {inter.get('effect_size', 'N/A')}"
                        )
                if 'pairwise_comparisons' in results and results['pairwise_comparisons']:
                    posthoc_test = results.get("posthoc_test", "Post-hoc test")
                    log.append(f"\nPost‑hoc test: {posthoc_test}")
                    log.append("Pairwise comparisons:")
                    for comp in results["pairwise_comparisons"]:
                        group1 = str(comp['group1'])
                        group2 = str(comp['group2'])
                        p_val = comp['p_value']
                        significant = comp['significant']
                        p_text = "p < 0.001" if isinstance(p_val, (float, int)) and p_val < 0.001 else f"p = {p_val:.4f}"
                        sign_text = "significant" if significant else "not significant"
                        stars = "***" if significant and p_val < 0.001 else "**" if significant and p_val < 0.01 else "*" if significant else ""
                        log.append(f"{group1} vs {group2}: {p_text}, {sign_text} {stars}")
                else:
                    log.append("\nNo pairwise comparisons were performed or calculated.")
                return "\n".join(log)
            analysis_log = build_analysis_log(results, params)
            results["analysis_log"] = analysis_log
            UIDialogManager.reset_dialog_cache()
            return results

        except Exception as e:
            error_message = str(e) if str(e) else "Unknown error occurred"
            analysis_log += f"\nERROR: {error_message}\n"
            print(f"Error during analysis: {error_message}")
            import traceback
            traceback.print_exc()
            UIDialogManager.reset_dialog_cache()
            return {"error": error_message, "analysis_log": analysis_log}       

def get_output_path(file_base, ext):
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
    if not os.path.isdir(desktop_path):
        # Fallback: current working directory
        desktop_path = os.getcwd()
    return os.path.join(desktop_path, f"{file_base}.{ext}")

try:
    _QT_APP.exit()
except Exception:
    pass

import os
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from scipy.stats import t
OUTLIER_IMPORTS_AVAILABLE = True
try:
    import seaborn
    import matplotlib.pyplot as plt
    from openpyxl import Workbook
except ImportError:
    OUTLIER_IMPORTS_AVAILABLE = False

class OutlierDetector:
    """
    Detect outliers in grouped data using Grubbs' Test or Dixon's Q-Test.
    Loads an Excel table with columns ['Group', 'Value'], 
    converts all values (German decimal numbers with comma) to float,
    performs Grubbs or Dixon tests iteratively or once for each group separately
    and marks found outliers.
    """

    def __init__(self, df, group_col, value_col):
        """
        Initializes the OutlierDetector with an already loaded DataFrame.
        
        Parameters:
        -----------
        df : pandas.DataFrame
            DataFrame with the data
        group_col : str
            Name of the group column
        value_col : str
            Name of the values column
        """
        print(f"DEBUG: Initializing OutlierDetector with columns: group_col={group_col}, value_col={value_col}")
        print(f"DEBUG: DataFrame shape: {df.shape}, columns: {df.columns.tolist()}")
        
        if not OUTLIER_IMPORTS_AVAILABLE:
            raise ImportError("Required packages for outlier detection are not available. "
                              "Please install: pip install outliers pingouin openpyxl")
        
        # Create copy of DataFrame
        self.df = df.copy()
        self.group_col = group_col
        self.value_col = value_col
        self.active_test = None  # Track which test was run
        self.debug_log = []  # Initialize debug log

        # Ensure columns are present
        if group_col not in self.df.columns or value_col not in self.df.columns:
            raise ValueError(f"Columns '{group_col}' and '{value_col}' must be present in the DataFrame.")

        # Add initialization info to log
        self.debug_log.append(f"*** OUTLIER DETECTION INITIALIZATION ***")
        self.debug_log.append(f"DataFrame shape: {df.shape}")
        self.debug_log.append(f"Group column: {group_col}")
        self.debug_log.append(f"Value column: {value_col}")
        self.debug_log.append(f"Available columns: {df.columns.tolist()}")

        # Convert values column to float (if necessary)
        self._convert_values_to_float()
        
        # Show group statistics after initialization
        self.debug_log.append(f"\n=== GROUP STATISTICS ===")
        self.debug_log.append(f"Number of groups in data: {self.df[group_col].nunique()}")
        
        for group_name, group_data in self.df.groupby(group_col):
            group_stats = self._calculate_group_statistics(group_name, group_data)
            self.debug_log.extend(group_stats)

    def _convert_values_to_float(self):
        """
        Converts the values column to float.
        Handles German decimal separators (comma instead of period).
        """
        try:
            self.debug_log.append(f"\n=== VALUE CONVERSION ===")
            self.debug_log.append(f"Converting column '{self.value_col}' to float")
            
            # Check if the column exists
            if self.value_col not in self.df.columns:
                error_msg = f"Column '{self.value_col}' not found in DataFrame. Available columns: {list(self.df.columns)}"
                self.debug_log.append(f"ERROR: {error_msg}")
                raise ValueError(error_msg)
            
            self.debug_log.append(f"Sample values before conversion: {self.df[self.value_col].head(3).tolist()}")
            self.debug_log.append(f"Data type before conversion: {self.df[self.value_col].dtype}")
            
            # Check if column is already numeric
            if pd.api.types.is_numeric_dtype(self.df[self.value_col]):
                self.debug_log.append("Column is already numeric, no conversion needed")
                return
            
            # Convert German decimal numbers
            self.df[self.value_col] = (
                self.df[self.value_col]
                    .astype(str)
                    .str.replace('.', '', regex=False)   # Remove thousand separators
                    .str.replace(',', '.', regex=False)  # Comma → Period
                    .astype(float)
            )
            
            self.debug_log.append(f"Sample values after conversion: {self.df[self.value_col].head(3).tolist()}")
            self.debug_log.append(f"Data type after conversion: {self.df[self.value_col].dtype}")
            self.debug_log.append(f"Value range after conversion: {self.df[self.value_col].min():.4f} to {self.df[self.value_col].max():.4f}")
        
        except Exception as e:
            error_message = f"Conversion failed: {str(e)}"
            self.debug_log.append(f"ERROR: {error_message}")
            raise ValueError(f"Error converting values column '{self.value_col}' to float: {str(e)}")
        
    def _calculate_group_statistics(self, group_name, group_data):
        """Calculate basic statistics for a group and return as log entries."""
        stats = []
        values = group_data[self.value_col].dropna()
        stats.append(f"\n--- Group '{group_name}' Statistics ---")
        stats.append(f"  Count: {len(values)}")
        
        if len(values) > 0:
            stats.append(f"  Min: {values.min():.4f}")
            stats.append(f"  Max: {values.max():.4f}")
            stats.append(f"  Mean: {values.mean():.4f}")
            stats.append(f"  Median: {values.median():.4f}")
            stats.append(f"  Std Dev: {values.std():.4f}")
        else:
            stats.append("  No valid values in this group")
        
        return stats
    
    @staticmethod
    def _grubbs_iterative(vals: np.ndarray,
                           alpha: float = 0.05,
                           two_sided: bool = True):
        """
        Iteratively remove the most extreme point (if G_obs > G_crit),
        repeat until no further outliers are detected.
        Returns list of relative indices into the original vals array.
        """
        vals = np.asarray(vals, dtype=float)
        # track the original positions
        idxs = np.arange(vals.size)
        out_rel_idxs = []
        
        while vals.size >= 3:
            G_obs, G_crit, rel_idx = OutlierDetector._grubbs_statistic(
                vals, alpha=alpha, two_sided=two_sided
            )
            if G_obs <= G_crit:
                break
            # record and remove
            out_rel_idxs.append(int(idxs[rel_idx]))
            mask = np.ones(vals.size, dtype=bool)
            mask[rel_idx] = False
            vals = vals[mask]
            idxs = idxs[mask]
        
        return out_rel_idxs

    @staticmethod
    def _grubbs_statistic(x: np.ndarray,
                          alpha: float = 0.05,
                          two_sided: bool = True):
        """
        Compute Grubbs' G statistic and critical value for array x.
        Returns (G_obs, G_crit, outlier_rel_index).
        """
        n = x.size
        if n < 3:
            raise ValueError("Grubbs' test requires at least 3 values")
        
        mu = x.mean()
        s  = x.std(ddof=1)
        # maximum absolute deviation
        diffs = np.abs(x - mu)
        rel_idx = int(np.argmax(diffs))
        G_obs = diffs[rel_idx] / s
        
        # two‐sided splits alpha into 2 tails
        tail = 2 if two_sided else 1
        # student-t critical value
        t_crit = t.ppf(1 - alpha/(tail * n), df=n-2)
        # critical G
        G_crit = ((n - 1) / np.sqrt(n)
                  * np.sqrt(t_crit**2 / ( (n - 2) + t_crit**2 )))
        
        return G_obs, G_crit, rel_idx

    def run_grubbs(self, alpha: float = 0.05, iterate: bool = True):
        self.debug_log.append("\n=== GRUBBS OUTLIER DETECTION ===")
        self.debug_log.append(f"alpha={alpha}, iterate={iterate}")

        self.df["Grubbs_Outlier"] = False
        self.active_test = "Grubbs"
        total_out = 0

        for gname, gdf in self.df.groupby(self.group_col):
            self.debug_log.append(f"\n--- Group '{gname}' ({len(gdf)}) ---")

            # original indices and clean values
            idxs = gdf.index.values
            vals = gdf[self.value_col].astype(float).values
            valid = ~np.isnan(vals)
            idxs_valid = idxs[valid]
            vals_valid = vals[valid]

            if len(vals_valid) < 3:
                self.debug_log.append("  n < 3 → skipping Grubbs")
                continue

            # choose iterative vs. single removal
            if iterate:
                rel_outs = OutlierDetector._grubbs_iterative(
                    vals_valid, alpha=alpha, two_sided=True
                )
            else:
                G_obs, G_crit, rel_idx = OutlierDetector._grubbs_statistic(
                    vals_valid, alpha=alpha, two_sided=True
                )
                rel_outs = [rel_idx] if G_obs > G_crit else []

            if not rel_outs:
                self.debug_log.append("  No outliers detected")
            else:
                # map back to global indices
                global_outs = idxs_valid[rel_outs].tolist()
                for gi in global_outs:
                    val = self.df.at[gi, self.value_col]
                    self.debug_log.append(f"  Outlier at idx={gi}, value={val:.4f}")
                self.df.loc[global_outs, "Grubbs_Outlier"] = True
                total_out += len(global_outs)
                self.debug_log.append(f"  ==> {len(global_outs)} outliers in group")

        self.debug_log.append("\n=== GRUBBS SUMMARY ===")
        self.debug_log.append(f"Total outliers: {total_out}")

    def run_mod_z_score(self, threshold=3.5, iterate=False):
        """
        Performs Modified Z-Score outlier detection for each group.
        
        Parameters:
        -----------
        threshold : float
            Threshold for modified Z-scores to be considered outliers (default: 3.5)
        iterate : bool
            Whether to iteratively remove outliers and recompute scores
        """
        self.debug_log.append(f"\n=== MODIFIED Z-SCORE OUTLIER DETECTION EXECUTION ===")
        self.debug_log.append(f"Test parameters: threshold={threshold}, iterate={iterate}")
        self.debug_log.append(f"Test principle: Modified Z-Score uses median absolute deviation (MAD) for robustness against outliers")
        
        # Create the Modified Z-Score column and mark this test as active
        self.df['ModZ_Outlier'] = False
        self.active_test = 'ModZ'
        
        total_outliers = 0
        
        for group_name, group_df in self.df.groupby(self.group_col):
            self.debug_log.append(f"\n--- Processing group '{group_name}' ---")
            self.debug_log.append(f"Group size: {len(group_df)} observations")
            
            indices = group_df.index.tolist()
            values = group_df[self.value_col].values.copy()
            
            # Filter out NaN values
            valid_mask = ~np.isnan(values)
            valid_values = values[valid_mask]
            valid_indices = [idx for mask, idx in zip(valid_mask, indices) if mask]
            
            if len(valid_values) == 0:
                self.debug_log.append(f"Group '{group_name}' has no valid values (all NaN), skipping")
                continue
                
            outlier_indices = []
            
            # Log basic group statistics
            mean_val = np.mean(valid_values)
            median_val = np.median(valid_values)
            self.debug_log.append(f"Group statistics: min={np.min(valid_values):.3f}, max={np.max(valid_values):.3f}, mean={mean_val:.3f}, median={median_val:.3f}")
            
            def detect_single_round(vals, indices_list):
                if len(vals) <= 1:
                    return [], "No outliers found (too few values for MAD calculation)"
                    
                # Calculate median and MAD
                median = np.median(vals)
                mad = np.median(np.abs(vals - median))
                
                # Avoid division by zero
                if mad == 0:
                    return [], "MAD = 0, cannot compute modified Z-scores"
                    
                # Calculate modified Z-scores
                mod_z_scores = 0.6745 * (vals - median) / mad
                
                # Find outliers
                outlier_mask = np.abs(mod_z_scores) > threshold
                outlier_idx = [idx for mask, idx in zip(outlier_mask, indices_list) if mask]
                
                # Log this round
                self.debug_log.append(f"  Modified Z-Score Analysis:")
                self.debug_log.append(f"    Median: {median:.4f}")
                self.debug_log.append(f"    MAD: {mad:.4f}")
                self.debug_log.append(f"    Threshold: {threshold} (absolute value)")
                
                if len(outlier_idx) > 0:
                    self.debug_log.append(f"    {len(outlier_idx)} outliers found in this round")
                    return outlier_idx, None
                else:
                    self.debug_log.append(f"    No outliers found in this round")
                    return [], "No outliers found (all modified Z-scores within threshold)"
            
            # Iterative outlier detection
            vals_copy = valid_values.copy()
            idx_copy = valid_indices.copy() 
            round_count = 0
            
            while True:
                round_count += 1
                self.debug_log.append(f"  Testing round {round_count} with {len(vals_copy)} values")
                
                round_outliers, stop_message = detect_single_round(vals_copy, idx_copy)
                
                if round_outliers:
                    outlier_indices.extend(round_outliers)
                    if iterate:
                        # Remove outliers for next iteration
                        keep_mask = ~np.isin(idx_copy, round_outliers)
                        vals_copy = vals_copy[keep_mask]
                        idx_copy = [idx for keep, idx in zip(keep_mask, idx_copy) if keep]
                    else:
                        self.debug_log.append(f"  No more iterations requested, terminating test")
                        break
                else:
                    self.debug_log.append(f"  {stop_message if stop_message else 'No more outliers found'}, terminating test")
                    break
                    
                if len(vals_copy) <= 1:
                    self.debug_log.append(f"  Too few values left for MAD calculation, terminating test")
                    break
            
            # Mark all found indices in main DataFrame
            if outlier_indices:
                self.df.loc[outlier_indices, 'ModZ_Outlier'] = True
                total_outliers += len(outlier_indices)
                self.debug_log.append(f"  RESULT: {len(outlier_indices)} outliers found in group '{group_name}'")
            else:
                self.debug_log.append(f"  RESULT: No outliers found in group '{group_name}'")
        
        # Final summary
        self.debug_log.append(f"\n=== MODIFIED Z-SCORE TEST SUMMARY ===")
        self.debug_log.append(f"Total outliers detected across all groups: {total_outliers}")
        self.debug_log.append(f"Test completed successfully")


    def save_results(self, output_path):
        """
        Outputs the result to a new Excel file while preserving formulas.
        Only columns for actually performed tests are displayed.
        """
        self.debug_log.append(f"\n=== SAVING RESULTS ===")
        self.debug_log.append(f"Output path: {output_path}")
        self.debug_log.append(f"Active test: {self.active_test}")
        
        temp_file = None
        
        try:
            if self.active_test == 'Grubbs':
                outlier_count = self.df['Grubbs_Outlier'].sum()
                self.debug_log.append(f"Total outliers found with Grubbs method: {outlier_count}")
            elif self.active_test == 'ModZ':
                outlier_count = self.df['ModZ_Outlier'].sum()
                self.debug_log.append(f"Total outliers found with Modified Z-Score method: {outlier_count}")
            
            # Check if file exists and contains formulas we need to preserve
            file_exists = os.path.exists(output_path)
            
            if file_exists:
                # Create a backup of the original file with a timestamp
                from datetime import datetime
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_path = f"{os.path.splitext(output_path)[0]}_{timestamp}_backup.xlsx"
                import shutil
                shutil.copy2(output_path, backup_path)
                self.debug_log.append(f"Created backup of existing file at: {backup_path}")
                
                # Create a new file with a different name for our results
                results_path = f"{os.path.splitext(output_path)[0]}_outliers.xlsx"
                self.debug_log.append(f"Creating new results file at: {results_path}")
            else:
                # Just use the original path if the file doesn't exist
                results_path = output_path
            
            # Create new workbook for our results
            wb = Workbook()

            # 1) Sheet "Raw_Data_with_Outlier_Marking"
            ws_raw = wb.active
            ws_raw.title = "Raw_Data_with_Outlier_Marking"

            # Include only columns for the test that was run
            test_columns = []
            if self.active_test == 'Grubbs' and 'Grubbs_Outlier' in self.df.columns:
                test_columns.append('Grubbs_Outlier')
            elif self.active_test == 'ModZ' and 'ModZ_Outlier' in self.df.columns:
                test_columns.append('ModZ_Outlier')

            # Write headers
            headers = [self.group_col, self.value_col]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws_raw.cell(row=1, column=col_idx, value=str(header))
                cell.font = Font(bold=True)

            # Write data with outlier highlighting
            outlier_count_by_group = {}
            for row_idx, (_, row) in enumerate(self.df.iterrows(), start=2):
                group = row[self.group_col]
                if group not in outlier_count_by_group:
                    outlier_count_by_group[group] = 0
                    
                # Column 1: Group
                group_value = str(group) if group is not None else ""
                ws_raw.cell(row=row_idx, column=1, value=group_value)
                
                # Column 2: Value - mark red if outlier
                value = row[self.value_col]
                if pd.isna(value) or not isinstance(value, (int, float)):
                    value = 0.0
                
                cell_value = ws_raw.cell(row=row_idx, column=2, value=float(value))
                
                # Check for outlier and mark with red highlighting
                is_outlier = any(row.get(col, False) for col in test_columns)
                if is_outlier:
                    outlier_count_by_group[group] += 1
                    # Red font and light red background
                    cell_value.font = Font(color="FF0000", bold=True)
                    cell_value.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    self.debug_log.append(f"Marking outlier: Group={group}, Value={value:.4f}")

            # 2) Sheet "Cleaned"
            ws_clean = wb.create_sheet(title="Cleaned")
            headers_clean = [self.group_col, self.value_col]
            for col_idx, header in enumerate(headers_clean, start=1):
                cell = ws_clean.cell(row=1, column=col_idx, value=str(header))
                cell.font = Font(bold=True)

            # Filter: no outliers from active test
            outlier_mask = pd.Series(False, index=self.df.index)
            for test_col in test_columns:
                outlier_mask |= self.df[test_col]
            
            cleaned_df = self.df[~outlier_mask]
            self.debug_log.append(f"Cleaned data has {len(cleaned_df)} rows (removed {len(self.df) - len(cleaned_df)} outliers)")
            
            for row_idx, (_, row) in enumerate(cleaned_df.iterrows(), start=2):
                group_value = str(row[self.group_col]) if row[self.group_col] is not None else ""
                value = float(row[self.value_col]) if not pd.isna(row[self.value_col]) else 0.0
                
                ws_clean.cell(row=row_idx, column=1, value=group_value)
                ws_clean.cell(row=row_idx, column=2, value=value)

            # 3) Summary sheet with statistics
            ws_summary = wb.create_sheet(title="Summary")
            ws_summary.cell(row=1, column=1, value="Outlier Detection Summary").font = Font(bold=True, size=14)
            
            # Test parameters
            ws_summary.cell(row=3, column=1, value="Test parameters:").font = Font(bold=True)
            ws_summary.cell(row=4, column=1, value="Test type:")
            ws_summary.cell(row=4, column=2, value=str(self.active_test))
            
            # Add test-specific parameters
            if self.active_test == 'Grubbs':
                ws_summary.cell(row=5, column=1, value="Grubbs alpha level:")
                ws_summary.cell(row=5, column=2, value="0.05")  # Default value, adjust if needed
            elif self.active_test == 'ModZ':
                ws_summary.cell(row=5, column=1, value="Z-score threshold:")
                ws_summary.cell(row=5, column=2, value="3.5")  # Default value, adjust if needed
            
            # Group statistics
            row_pos = 7
            ws_summary.cell(row=row_pos, column=1, value="Group statistics:").font = Font(bold=True)
            row_pos += 1
            
            # Headers for statistics table
            stat_headers = ["Group", "Count", "Min", "Max", "Mean", "Std", "Outliers"]
            for col_idx, header in enumerate(stat_headers, start=1):
                cell = ws_summary.cell(row=row_pos, column=col_idx, value=header)
                cell.font = Font(bold=True)
                
            row_pos += 1
            for group_name, group_df in self.df.groupby(self.group_col):
                values = group_df[self.value_col].dropna()
                
                ws_summary.cell(row=row_pos, column=1, value=str(group_name))
                ws_summary.cell(row=row_pos, column=2, value=len(group_df))
                ws_summary.cell(row=row_pos, column=3, value=float(values.min()) if len(values) > 0 else 0.0)
                ws_summary.cell(row=row_pos, column=4, value=float(values.max()) if len(values) > 0 else 0.0)
                ws_summary.cell(row=row_pos, column=5, value=float(values.mean()) if len(values) > 0 else 0.0)
                ws_summary.cell(row=row_pos, column=6, value=float(values.std()) if len(values) > 0 else 0.0)
                ws_summary.cell(row=row_pos, column=7, value=outlier_count_by_group.get(group_name, 0))
                row_pos += 1

            # 4) Enhanced Debug Log sheet with better formatting
            ws_log = wb.create_sheet(title="Debug_Log")
            header_cell = ws_log.cell(row=1, column=1)
            header_cell.value = "Outlier Detection Debug Log"
            header_cell.font = Font(bold=True, size=16, color="0066CC")

            subheader_cell = ws_log.cell(row=2, column=1)
            subheader_cell.value = "This sheet contains detailed information about the outlier detection process."
            subheader_cell.font = Font(italic=True)

            # Set column width for better readability
            ws_log.column_dimensions['A'].width = 150

            # Write debug log entries with enhanced formatting and formula protection
            for idx, log_entry in enumerate(self.debug_log, start=4):
                log_text = str(log_entry)[:32000]
                cell = ws_log.cell(row=idx, column=1)
                cell.value = log_text
                
                # Force inlineStr data type for any text starting with "="
                if log_text.startswith("="):
                    cell.data_type = 'inlineStr'
                
                # Enhanced formatting based on content
                if log_text.startswith("==="):
                    # Main section headers - blue background, white text
                    cell.font = Font(bold=True, size=14, color="FFFFFF")
                    cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
                elif log_text.startswith("---"):
                    # Group headers - light blue background, dark text
                    cell.font = Font(bold=True, size=12, color="000080")
                    cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                elif "ERROR:" in log_text:
                    # Errors - red background, white text
                    cell.font = Font(color="FFFFFF", bold=True, size=11)
                    cell.fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
                elif "OUTLIER DETECTED" in log_text:
                    # Outlier detection - yellow background, red text
                    cell.font = Font(color="CC0000", bold=True, size=11)
                    cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                elif "RESULT:" in log_text:
                    # Results - green background, dark text
                    cell.font = Font(color="006600", bold=True, size=11)
                    cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                elif log_text.startswith("  Testing round") or log_text.startswith("  Running"):
                    # Test round information - light gray background
                    cell.font = Font(size=10, italic=True)
                    cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                elif log_text.startswith("    "):
                    # Detailed information - smaller font, indented
                    cell.font = Font(size=9, color="666666")
                elif log_text.startswith("  "):
                    # General information - smaller font
                    cell.font = Font(size=10)

            # Add visualization
            temp_file = self._add_single_visualization_sheet(wb, self)
            
            # Save the new workbook
            wb.save(results_path)
            self.debug_log.append(f"Results saved to: {results_path}")
            
            # If we created a separate file, inform user about both files
            if file_exists:
                self.debug_log.append(f"Original file with formulas preserved at: {output_path}")
                self.debug_log.append(f"Backup copy created at: {backup_path}")
                print(f"NOTE: To preserve formulas, results were saved to a new file: {results_path}")
                print(f"      Original file with formulas intact: {output_path}")
        
        except Exception as e:
            error_msg = f"Error saving results: {str(e)}"
            self.debug_log.append(f"ERROR: {error_msg}")
            print(f"ERROR: {error_msg}")
            raise
        finally:
            # Clean up temporary files
            if temp_file and os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except:
                    pass  # Ignore cleanup errors

    def get_summary(self):
        """
        Creates a summary of found outliers.
        """
        self.debug_log.append(f"\n=== CREATING SUMMARY ===")
        
        total_rows = len(self.df)
        grubbs_outliers = self.df['Grubbs_Outlier'].sum() if 'Grubbs_Outlier' in self.df.columns else 0
        modz_outliers = self.df['ModZ_Outlier'].sum() if 'ModZ_Outlier' in self.df.columns else 0
        
        self.debug_log.append(f"Total rows: {total_rows}")
        self.debug_log.append(f"Grubbs outliers: {grubbs_outliers}")
        self.debug_log.append(f"Modified Z-Score outliers: {modz_outliers}")
        
        any_outliers = 0
        
        # Check which columns exist before combining
        if 'Grubbs_Outlier' in self.df.columns and 'ModZ_Outlier' in self.df.columns:
            any_outliers = (self.df['Grubbs_Outlier'] | self.df['ModZ_Outlier']).sum()
        elif 'Grubbs_Outlier' in self.df.columns:
            any_outliers = grubbs_outliers
        elif 'ModZ_Outlier' in self.df.columns:
            any_outliers = modz_outliers
        
        self.debug_log.append(f"Combined outliers: {any_outliers}")
        
        # Outliers per group
        group_summary = {}
        for group_name, group_df in self.df.groupby(self.group_col):
            group_total = len(group_df)
            group_grubbs = group_df['Grubbs_Outlier'].sum() if 'Grubbs_Outlier' in group_df.columns else 0
            group_modz = group_df['ModZ_Outlier'].sum() if 'ModZ_Outlier' in group_df.columns else 0
            
            group_any = 0
            if 'Grubbs_Outlier' in group_df.columns and 'ModZ_Outlier' in group_df.columns:
                group_any = (group_df['Grubbs_Outlier'] | group_df['ModZ_Outlier']).sum()
            elif 'Grubbs_Outlier' in group_df.columns:
                group_any = group_grubbs
            elif 'ModZ_Outlier' in group_df.columns:
                group_any = group_modz
            
            self.debug_log.append(f"Group '{group_name}': total={group_total}, outliers={group_any}")
            
            group_summary[group_name] = {
                'total': group_total,
                'grubbs_outliers': group_grubbs,
                'modz_outliers': group_modz,
                'any_outliers': group_any
            }
        
        summary = {
            'total_rows': total_rows,
            'grubbs_outliers': grubbs_outliers,
            'modz_outliers': modz_outliers,
            'any_outliers': any_outliers,
            'groups': group_summary
        }
        
        self.debug_log.append(f"Summary completed: {summary}")
        return summary

    @staticmethod
    def run_multi_dataset_outlier_detection(df, group_col, dataset_columns, alpha=0.05, 
                                            iterate=False, run_grubbs=True, run_modz=True, 
                                            grubbs_alpha=0.05, modz_threshold=3.5,
                                            output_path="multi_dataset_outlier_results.xlsx"):
        """
        Run outlier detection on multiple datasets (columns) and create a combined Excel output.
        """
        print(f"DEBUG: Starting multi-dataset outlier detection")
        print(f"DEBUG: Datasets to analyze: {dataset_columns}")
        print(f"DEBUG: Group column: {group_col}")
        
        all_results = {}
        failed_datasets = {}
        temp_files = []  # Track temporary files for cleanup
        
        # Create main workbook for combined results
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)
        
        try:
            # Analyze each dataset
            for i, dataset_col in enumerate(dataset_columns):
                print(f"DEBUG: Analyzing dataset {i+1}/{len(dataset_columns)}: {dataset_col}")
                
                try:
                    # Create detector for this dataset
                    detector = OutlierDetector(
                        df=df.copy(),
                        group_col=group_col,
                        value_col=dataset_col
                    )
                    
                    # Run requested tests
                    if run_grubbs:
                        print(f"DEBUG: Running Grubbs test for {dataset_col}")
                        detector.run_grubbs(alpha=grubbs_alpha, iterate=iterate)
                    
                    if run_modz:
                        print(f"DEBUG: Running Modified Z-Score test for {dataset_col}")
                        detector.run_mod_z_score(threshold=modz_threshold, iterate=iterate)
                    
                    # Store results
                    all_results[dataset_col] = {
                        'detector': detector,
                        'summary': detector.get_summary()
                    }
                    
                    # Add sheets to combined workbook
                    OutlierDetector._add_dataset_to_workbook(wb, detector, dataset_col, run_grubbs, run_modz)
                    print(f"DEBUG: Successfully analyzed {dataset_col}")
                    
                except Exception as e:
                    error_msg = f"Error analyzing {dataset_col}: {str(e)}"
                    failed_datasets[dataset_col] = error_msg
                    print(f"DEBUG: ERROR: {error_msg}")
                        
            # Create summary sheet
            OutlierDetector._create_multi_summary_sheet(wb, all_results, failed_datasets, dataset_columns)
            
            # Add visualization sheet with swarm plots
            visual_temp_files = OutlierDetector._add_visualization_sheet(wb, all_results, dataset_columns)
            if visual_temp_files:
                temp_files.extend(visual_temp_files)
            
            # Save combined workbook
            wb.save(output_path)
            print(f"DEBUG: Combined results saved to: {output_path}")
            
            # Return summary
            return {
                "type": "multi_dataset_outlier_detection",
                "successful_datasets": list(all_results.keys()),
                "failed_datasets": failed_datasets,
                "total_datasets": len(dataset_columns),
                "output_file": output_path,
                "summary": {
                    "success_count": len(all_results),
                    "failure_count": len(failed_datasets),
                    "success_rate": f"{len(all_results)/len(dataset_columns)*100:.1f}%"
                }
            }
            
        except Exception as e:
            error_msg = f"Critical error in multi-dataset analysis: {str(e)}"
            print(f"DEBUG: CRITICAL ERROR: {error_msg}")
            raise RuntimeError(error_msg)
            
        finally:
            # Clean up temporary files
            for temp_file in temp_files:
                try:
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
                        print(f"DEBUG: Cleaned up temporary file: {temp_file}")
                except Exception as e:
                    print(f"DEBUG: Error cleaning up file {temp_file}: {str(e)}")
    
    @staticmethod
    def _add_dataset_to_workbook(wb, detector, dataset_name, run_grubbs, run_modz):
        """Add analysis sheet for a single dataset with debug log and summary."""
        
        # Create sheet name (Excel sheet names are limited to 31 characters)
        safe_name = dataset_name.replace(' ', '_')[:25]  # Leave room for suffixes
        
        # Create single analysis sheet
        ws = wb.create_sheet(title=f"{safe_name}_Analysis")
        
        # Title
        ws.cell(row=1, column=1, value=f"Outlier Analysis: {dataset_name}").font = Font(bold=True, size=16, color="0066CC")
        
        # Summary section
        summary = detector.get_summary()
        row = 3
        
        ws.cell(row=row, column=1, value="SUMMARY").font = Font(bold=True, size=14)
        row += 1
        
        ws.cell(row=row, column=1, value="Dataset:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=dataset_name)
        row += 1
        
        ws.cell(row=row, column=1, value="Total data points:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=summary['total_rows'])
        row += 1
        
        if run_grubbs:
            ws.cell(row=row, column=1, value="Grubbs outliers:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=summary['grubbs_outliers'])
            row += 1
            
        if run_modz:
            ws.cell(row=row, column=1, value="ModZ outliers:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=summary['modz_outliers'])
            row += 1
        
        ws.cell(row=row, column=1, value="Total outliers:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=summary['any_outliers'])
        row += 2
        
        # Group statistics
        ws.cell(row=row, column=1, value="Group Statistics:").font = Font(bold=True)
        row += 1
        
        group_headers = ["Group", "Total", "Outliers", "Clean"]
        for col_idx, header in enumerate(group_headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(bold=True)
        row += 1
        
        for group_name, group_stats in summary['groups'].items():
            ws.cell(row=row, column=1, value=str(group_name))
            ws.cell(row=row, column=2, value=group_stats['total'])
            ws.cell(row=row, column=3, value=group_stats['any_outliers'])
            ws.cell(row=row, column=4, value=group_stats['total'] - group_stats['any_outliers'])
            row += 1
        
        row += 2
        
        # Debug Log section
        ws.cell(row=row, column=1, value="DEBUG LOG").font = Font(bold=True, size=14, color="0066CC")
        row += 1
        
        ws.cell(row=row, column=1, value="Detailed analysis log:").font = Font(italic=True)
        row += 1
        
        # Set column width for debug log
        ws.column_dimensions['A'].width = 40
        
        # Write debug log entries with enhanced formatting
        for log_entry in detector.debug_log:
            log_text = str(log_entry)[:32000]
            cell = ws.cell(row=row, column=1)
            cell.value = log_text
            
            # Force inlineStr data type for any text starting with "="
            if log_text.startswith("="):
                cell.data_type = 'inlineStr'
            
            # Enhanced formatting based on content
            if log_text.startswith("==="):
                # Main section headers - blue background, white text
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            elif log_text.startswith("---"):
                # Group headers - light blue background, dark text
                cell.font = Font(bold=True, size=10, color="000080")
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            elif "ERROR:" in log_text:
                # Errors - red background, white text
                cell.font = Font(color="FFFFFF", bold=True, size=9)
                cell.fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
            elif "OUTLIER DETECTED" in log_text:
                # Outlier detection - yellow background, red text
                cell.font = Font(color="CC0000", bold=True, size=9)
                cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            elif "RESULT:" in log_text:
                # Results - green background, dark text
                cell.font = Font(color="006600", bold=True, size=9)
                cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            elif log_text.startswith("  Testing round") or log_text.startswith("  Running"):
                # Test round information - light gray background
                cell.font = Font(size=8, italic=True)
                cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            elif log_text.startswith("    "):
                # Detailed information - smaller font, indented
                cell.font = Font(size=8, color="666666")
            elif log_text.startswith("  "):
                # General information - smaller font
                cell.font = Font(size=9)
            
            row += 1

        # Set other column widths
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    @staticmethod
    def _create_multi_summary_sheet(wb, all_results, failed_datasets, dataset_columns):
        """Create a comprehensive summary sheet for all datasets."""
        ws = wb.create_sheet(title="Data_without_outlier", index=0)  # Insert as first sheet
        
        # Get the original DataFrame from the first successful result
        original_df = None
        group_col = None
        for dataset_col, result_data in all_results.items():
            if 'detector' in result_data:
                original_df = result_data['detector'].df
                group_col = result_data['detector'].group_col
                break
        
        if original_df is None:
            # Fallback if no data available
            ws.cell(row=1, column=1, value="No data available").font = Font(bold=True, color="FF0000")
            return
        
        if group_col is None:
            ws.cell(row=1, column=1, value="Group column not found").font = Font(bold=True, color="FF0000")
            return
        
        # Create headers: Sample + all dataset columns
        headers = [group_col] + dataset_columns
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
        
        # Get all unique combinations of group values that exist in the original data
        # We need to preserve the original row structure
        row_idx = 2
        
        # Iterate through the original DataFrame to maintain the exact structure
        for orig_idx, orig_row in original_df.iterrows():
            group_value = orig_row[group_col]
            
            # Column 1: Group/Sample name
            ws.cell(row=row_idx, column=1, value=str(group_value))
            
            # For each dataset column, get the value and check if it's an outlier
            for col_idx, dataset_col in enumerate(dataset_columns, start=2):
                if dataset_col in all_results:
                    detector = all_results[dataset_col]['detector']
                    
                    # Get the value for this specific row
                    if orig_idx < len(detector.df):
                        row_data = detector.df.iloc[orig_idx]
                        value = row_data[dataset_col]
                        
                        # Check if this specific row is marked as an outlier
                        is_grubbs_outlier = row_data.get('Grubbs_Outlier', False) if 'Grubbs_Outlier' in detector.df.columns else False
                        is_modz_outlier = row_data.get('ModZ_Outlier', False) if 'ModZ_Outlier' in detector.df.columns else False
                        is_outlier = is_grubbs_outlier or is_modz_outlier
                        
                        # Add the value to the cell
                        if pd.isna(value):
                            cell = ws.cell(row=row_idx, column=col_idx, value="NaN")
                        else:
                            cell = ws.cell(row=row_idx, column=col_idx, value=float(value))
                            
                            # Mark outliers with red highlighting
                            if is_outlier:
                                cell.font = Font(color="FF0000", bold=True)
                                cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    else:
                        # Row doesn't exist in this detector's data
                        ws.cell(row=row_idx, column=col_idx, value="N/A")
                else:
                    # Dataset failed or not available
                    ws.cell(row=row_idx, column=col_idx, value="N/A")
            
            row_idx += 1
        
        # Set column widths
        for col in range(1, len(headers) + 1):
            if col <= 26:  # A-Z
                ws.column_dimensions[chr(64 + col)].width = 15
            else:  # AA, AB, etc.
                ws.column_dimensions[f"A{chr(64 + col - 26)}"].width = 15
                
    @staticmethod
    def _add_visualization_sheet(wb, all_results, dataset_columns):
        """Add visualization sheet with swarm plots showing outliers."""
        from openpyxl.drawing.image import Image
        import matplotlib.pyplot as plt
        import seaborn as sns
        import tempfile
        import os
        
        # Create a visualization sheet
        ws = wb.create_sheet(title="Visualization")
        
        # Set title
        ws.cell(row=1, column=1, value="Outlier Visualization").font = Font(bold=True, size=16, color="0066CC")
        ws.cell(row=2, column=1, value="Visual representation of data with outliers highlighted").font = Font(italic=True)
        
        # Track row position and temp files
        row = 4
        temp_files = []
        
        # Track if we generated any plots successfully
        plots_added = False
        
        # For each dataset, create a swarm plot with modern styling
        for dataset_col in dataset_columns:
            if dataset_col not in all_results:
                continue
            
            # Get detector for this dataset
            result_data = all_results[dataset_col]
            if 'detector' not in result_data:
                continue
            
            detector = result_data['detector']
            df = detector.df
            group_col = detector.group_col
            value_col = detector.value_col
            
            # Determine which outlier column to use
            outlier_col = None
            if 'Grubbs_Outlier' in df.columns:
                outlier_col = 'Grubbs_Outlier'
                test_name = "Grubbs Test"
            elif 'ModZ_Outlier' in df.columns:
                outlier_col = 'ModZ_Outlier'
                test_name = "Modified Z-Score Test"
            
            if not outlier_col:
                continue
                
            try:
                # Add dataset name as header
                title_text = f"Dataset: {dataset_col}"
                title_cell = ws.cell(row=row, column=1)
                title_cell.value = title_text
                title_cell.font = Font(bold=True, size=14)
                if title_text.startswith("="):
                    title_cell.data_type = 'inlineStr'
                row += 1
                
                # Add test method used
                ws.cell(row=row, column=1, value=f"Method: {test_name}")
                row += 1
                
                # Get outlier count
                outlier_count = df[outlier_col].sum()
                total_count = len(df)
                ws.cell(row=row, column=1, value=f"Found {outlier_count} outliers out of {total_count} data points ({outlier_count/total_count:.1%})")
                row += 2
                
                # ─── Matplotlib/Seaborn ‐ Plot ────────────────────────────────────────────
                # Modern, minimalist style
                sns.set_style("white")  
                plt.rcParams.update({
                    "axes.edgecolor": "#333333",
                    "axes.linewidth": 1.0,
                    "xtick.color": "#333333",
                    "ytick.color": "#333333",
                    "font.size": 12,
                    "axes.titlesize": 14,
                    "axes.labelsize": 12,
                })

                # Larger figure size with less whitespace
                fig, ax = plt.subplots(figsize=(10, 7))
                fig.patch.set_facecolor("white")

                # Boxplot in subtle light gray (no outlier symbols)
                sns.boxplot(
                    x=group_col,
                    y=value_col,
                    data=df,
                    fliersize=0,
                    width=0.5,
                    palette=["#DDDDDD"] * len(df[group_col].unique()),   # Light gray box
                    ax=ax
                )

                # Scatter/Swarm plot: normal points in dark gray, outliers in bright blue
                # Depending on dataset size: stripplot with jitter
                if len(df) < 100:
                    sns.swarmplot(
                        x=group_col,
                        y=value_col,
                        data=df,
                        hue=outlier_col,
                        palette={False: "#555555", True: "#007AFF"},  # Bright blue for outliers
                        size=10,
                        edgecolor="white",
                        linewidth=0.8,
                        ax=ax
                    )
                else:
                    sns.stripplot(
                        x=group_col,
                        y=value_col,
                        data=df,
                        hue=outlier_col,
                        palette={False: "#555555", True: "#007AFF"},
                        size=8,
                        jitter=0.25,
                        alpha=0.8,
                        dodge=False,
                        ax=ax
                    )

                # Remove top and right spines ("despine")
                sns.despine(ax=ax, top=True, right=True, left=False, bottom=False)

                # Axis labels
                ax.set_xlabel(group_col, fontsize=12, color="#333333")
                ax.set_ylabel(value_col, fontsize=12, color="#333333")
                ax.set_title(f"{value_col} by Group (Outliers Highlighted)", fontsize=14, color="#333333", pad=15)

                # Legend only for outliers, no frame, modern positioning
                handles, labels = ax.get_legend_handles_labels()
                # labels come in order [False, True] -> filter accordingly
                new_handles, new_labels = [], []
                for h, lab in zip(handles, labels):
                    if lab == "False":
                        new_handles.append(h)
                        new_labels.append("Normal")
                    elif lab == "True":
                        new_handles.append(h)
                        new_labels.append("Outlier")
                legend = ax.legend(
                    new_handles,
                    new_labels,
                    title="Status",
                    loc="upper right",
                    frameon=False,
                    fontsize=12,
                    title_fontsize=13
                )

                # Subtle grid on y-axis (almost invisible)
                ax.grid(axis="y", color="#f0f0f0", linestyle="-", linewidth=0.7)
                ax.set_axisbelow(True)

                # Tight layout (less margin)
                plt.tight_layout(pad=1)
                
                # Create a temporary file path with high DPI for crisp rendering
                fd, temp_path = tempfile.mkstemp(suffix='.png')
                os.close(fd)
                temp_files.append(temp_path)
                
                # Save with higher DPI for better quality
                plt.savefig(temp_path, dpi=200, bbox_inches="tight", facecolor="white")
                plt.close(fig)
                
                # Add the image to the worksheet
                img = Image(temp_path)
                # Set image size (in pixels)
                img.width = 1000
                img.height = 700
                ws.add_image(img, f'A{row}')
                
                # Mark that we added at least one plot
                plots_added = True
                
                # Move row position down to accommodate the image
                row += 32
                row += 2  # Add some space between plots
                
            except Exception as e:
                error_cell = ws.cell(row=row, column=1)
                error_cell.value = f"Error creating visualization for {dataset_col}: {str(e)}"
                error_cell.data_type = 'inlineStr'  # Ensure error messages aren't interpreted as formulas
                row += 3
        
        # If no plots were added, add a message
        if not plots_added:
            ws.cell(row=row, column=1, value="No visualizations could be generated.")
            row += 1
        
        # Set column width for better display
        ws.column_dimensions['A'].width = 120
        
        # Return list of temp files for later cleanup
        return temp_files
    
    @staticmethod
    def _add_single_visualization_sheet(wb, detector, dataset_name=None):
        """Add visualization sheet with swarm plot showing outliers in modern, minimalist style."""
        from openpyxl.drawing.image import Image
        import matplotlib.pyplot as plt
        import seaborn as sns
        import tempfile
        import os

        # Create a visualization sheet
        ws = wb.create_sheet(title="Visualization")

        # Set title (Excel sheet)
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = "Outlier Visualization"
        title_cell.font = Font(bold=True, size=16, color="0066CC")

        subtitle_cell = ws.cell(row=2, column=1)
        subtitle_cell.value = "Visual representation with highlighted outliers"
        subtitle_cell.font = Font(italic=True)

        # Start row for the image
        row = 4
        temp_path = None

        try:
            df = detector.df
            group_col = detector.group_col
            value_col = detector.value_col

            # Which outlier column?
            if 'Grubbs_Outlier' in df.columns:
                outlier_col = 'Grubbs_Outlier'
                test_name = "Grubbs Test"
            elif 'ModZ_Outlier' in df.columns:
                outlier_col = 'ModZ_Outlier'
                test_name = "Modified Z-Score Test"
            else:
                # If no results
                no_results_cell = ws.cell(row=row, column=1)
                no_results_cell.value = "No outlier results available."
                no_results_cell.data_type = 'inlineStr'
                return None

            # Subheader on the Excel sheet
            title_text = f"Dataset: {dataset_name or value_col}"
            title2_cell = ws.cell(row=row, column=1)
            title2_cell.value = title_text
            title2_cell.font = Font(bold=True, size=14)
            if title_text.startswith("="):
                title2_cell.data_type = 'inlineStr'
            row += 1

            method_cell = ws.cell(row=row, column=1)
            method_cell.value = f"Outlier Method: {test_name}"
            row += 1

            # Outlier counter
            outlier_count = int(df[outlier_col].sum())
            total_count = len(df)
            count_cell = ws.cell(row=row, column=1)
            count_cell.value = f"Found outliers: {outlier_count} of {total_count} ({outlier_count/total_count:.1%})"
            row += 2

            # ─── Matplotlib/Seaborn ‐ Plot ────────────────────────────────────────────
            # Modern, minimalist style
            sns.set_style("white")
            plt.rcParams.update({
                "axes.edgecolor": "#333333",
                "axes.linewidth": 1.0,
                "xtick.color": "#333333",
                "ytick.color": "#333333",
                "font.size": 12,
                "axes.titlesize": 14,
                "axes.labelsize": 12,
            })

            # Larger figure size with less whitespace
            fig, ax = plt.subplots(figsize=(10, 7))
            fig.patch.set_facecolor("white")

            # Boxplot in subtle light gray (no outlier symbols)
            sns.boxplot(
                x=group_col,
                y=value_col,
                data=df,
                fliersize=0,
                width=0.5,
                palette=["#DDDDDD"] * len(df[group_col].unique()),   # Light gray box
                ax=ax
            )

            # Scatter/Swarm plot: normal points in dark gray, outliers in bright blue
            # Depending on dataset size: stripplot with jitter
            if len(df) < 100:
                sns.swarmplot(
                    x=group_col,
                    y=value_col,
                    data=df,
                    hue=outlier_col,
                    palette={False: "#555555", True: "#007AFF"},  # Bright blue for outliers
                    size=10,
                    edgecolor="white",
                    linewidth=0.8,
                    ax=ax
                )
            else:
                sns.stripplot(
                    x=group_col,
                    y=value_col,
                    data=df,
                    hue=outlier_col,
                    palette={False: "#555555", True: "#007AFF"},
                    size=8,
                    jitter=0.25,
                    alpha=0.8,
                    dodge=False,
                    ax=ax
                )

            # Remove top and right spines ("despine")
            sns.despine(ax=ax, top=True, right=True, left=False, bottom=False)

            # Axis labels
            ax.set_xlabel(group_col, fontsize=12, color="#333333")
            ax.set_ylabel(value_col, fontsize=12, color="#333333")
            ax.set_title(f"{value_col} by Group (Outliers Highlighted)", fontsize=14, color="#333333", pad=15)

            # Legend only for outliers, no frame, modern positioning
            handles, labels = ax.get_legend_handles_labels()
            # labels come in order [False, True] -> filter accordingly
            new_handles, new_labels = [], []
            for h, lab in zip(handles, labels):
                if lab == "False":
                    new_handles.append(h)
                    new_labels.append("Normal")
                elif lab == "True":
                    new_handles.append(h)
                    new_labels.append("Outlier")
            legend = ax.legend(
                new_handles,
                new_labels,
                title="Status",
                loc="upper right",
                frameon=False,
                fontsize=12,
                title_fontsize=13
            )

            # Subtle grid on y-axis (almost invisible)
            ax.grid(axis="y", color="#f0f0f0", linestyle="-", linewidth=0.7)
            ax.set_axisbelow(True)

            # Tight layout (less margin)
            plt.tight_layout(pad=1)

            # Save as PNG temporarily
            fd, temp_path = tempfile.mkstemp(suffix=".png")
            os.close(fd)
            plt.savefig(temp_path, dpi=200, bbox_inches="tight", facecolor="white")
            plt.close(fig)

            # Insert image into Excel sheet
            img = Image(temp_path)
            img.width = 1000  # in pixels
            img.height = 700
            ws.add_image(img, f"A{row}")

            # ─── Group Statistics Table ───────────────────────────────────────────
            row += 32  # Leave space for the image
            ws.cell(row=row, column=1, value="Group Statistics:").font = Font(bold=True)
            row += 1

            # Table headers
            stats_headers = ["Group", "Count", "Mean", "StdDev", "Median", "Min", "Max", "Outliers"]
            for col_idx, header in enumerate(stats_headers, start=1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = Font(bold=True)
            row += 1

            # Values per group
            for group_name, group_df in df.groupby(group_col):
                vals = group_df[value_col].dropna()
                out_count = int(group_df[outlier_col].sum())

                # Group
                grp_cell = ws.cell(row=row, column=1)
                grp_val = str(group_name)
                grp_cell.value = grp_val
                if grp_val.startswith("="):
                    grp_cell.data_type = 'inlineStr'

                # Numeric cells
                ws.cell(row=row, column=2, value=len(vals))
                ws.cell(row=row, column=3, value=float(vals.mean()) if len(vals) > 0 else 0.0)
                ws.cell(row=row, column=4, value=float(vals.std()) if len(vals) > 0 else 0.0)
                ws.cell(row=row, column=5, value=float(vals.median()) if len(vals) > 0 else 0.0)
                ws.cell(row=row, column=6, value=float(vals.min()) if len(vals) > 0 else 0.0)
                ws.cell(row=row, column=7, value=float(vals.max()) if len(vals) > 0 else 0.0)

                out_cell = ws.cell(row=row, column=8, value=out_count)
                # Highlight if there are outliers
                if out_count > 0:
                    out_cell.font = Font(color="FF0000", bold=True)
                    for c in range(1, 9):
                        ws.cell(row=row, column=c).fill = PatternFill(
                            start_color="FFEEEE",
                            end_color="FFEEEE",
                            fill_type="solid"
                        )
                row += 1

        except Exception as e:
            # If an error occurs during plotting
            err_cell = ws.cell(row=row, column=1)
            err_cell.value = f"Error in visualization: {str(e)}"
            err_cell.data_type = 'inlineStr'

        # Adjust column widths
        for col in range(1, 9):
            if col == 1:
                ws.column_dimensions[chr(64 + col)].width = 30
            else:
                ws.column_dimensions[chr(64 + col)].width = 15

        return temp_path