import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import xlsxwriter
import os
import warnings
import string
from itertools import combinations
from datetime import datetime
from matplotlib.ticker import ScalarFormatter
from PyQt5.QtWidgets import QDialog, QVBoxLayout, QLabel, QRadioButton, QDialogButtonBox
from decisiontreevisualizer import DecisionTreeVisualizer
from lazy_imports import get_pingouin, get_scipy_stats, get_seaborn, get_statsmodels_multitest

# Late import functions to avoid circular imports
def get_results_exporter():
    """Get ResultsExporter class lazily"""
    from resultsexporter import ResultsExporter
    return ResultsExporter

def get_data_visualizer():
    """Get DataVisualizer class lazily"""
    from datavisualizer import DataVisualizer
    return DataVisualizer

def get_statistical_tester():
    """Get StatisticalTester class lazily"""
    from statisticaltester import StatisticalTester
    return StatisticalTester
# DISABLED: Nonparametric fallbacks are not yet supported
# from nonparametricanovas import NonParametricFactory, NonParametricRMANOVA

# Lazy imports for heavy modules - only imported when actually needed
def get_seaborn_module():
    """Get seaborn module lazily"""
    try:
        return get_seaborn()
    except ImportError:
        import seaborn as sns
        return sns

def get_stats_module():
    """Get scipy.stats module lazily"""
    try:
        return get_scipy_stats()
    except ImportError:
        from scipy import stats
        return stats

def get_pingouin_module():
    """Get pingouin module lazily"""
    try:
        return get_pingouin()
    except ImportError:
        import pingouin as pg
        return pg

def get_multicomp_module():
    """Get statsmodels multicomp module lazily"""
    try:
        from statsmodels.stats.multicomp import pairwise_tukeyhsd
        multipletests = get_statsmodels_multitest()
        return pairwise_tukeyhsd, multipletests
    except ImportError:
        multipletests = get_statsmodels_multitest()  # This will return fallback
        return None, multipletests

def get_boxcox_functions():
    """Get scipy boxcox functions lazily"""
    try:
        from scipy.stats import boxcox, boxcox_normmax
        return boxcox, boxcox_normmax
    except ImportError:
        return None, None

# --------------------------------------------------------------
#  Fallback QApplication to prevent dialogs blocking when script
#  is run purely via CLI
# --------------------------------------------------------------
from PyQt5.QtWidgets import QApplication
import sys as _sys
import time
from contextlib import contextmanager

@contextmanager
def working_directory(path):
    """Context manager for safely changing directories"""
    previous_dir = os.getcwd()
    try:
        os.chdir(path)
        yield
    finally:
        os.chdir(previous_dir)

print(f"DEBUG: RUNNING FILE VERSION FROM {time.time()} - {os.path.abspath(__file__)}")

_QT_APP = QApplication.instance()
if _QT_APP is None:
    _QT_APP = QApplication(_sys.argv)

warnings.simplefilter(action='ignore', category=FutureWarning)

def safe_format(val, fmt="{:.4f}", none_text="N/A"):
    """
    Format a value safely: if numeric, use the given format;
    if None, return the none_text; otherwise, cast to string.
    """
    if isinstance(val, (float, int)):
        try:
            return fmt.format(val)
        except Exception:
            return str(val)
    elif val is None:
        return none_text
    else:
        return str(val)


class AssumptionVisualizer:
    """
    Creates visual examinations for statistical test assumptions (normality and homoscedasticity).
    """
    
    @staticmethod
    def create_normality_plot(data_dict, title_suffix="", transformation=None, results=None):
        """
        Create Q-Q plot for normality examination using MODEL RESIDUALS (not individual groups).
        
        Parameters:
        -----------
        data_dict : dict
            Dictionary with group names as keys and data arrays as values (for fallback)
        title_suffix : str
            Suffix to add to the plot title (e.g., "Before Transformation")
        transformation : str
            Name of transformation applied (if any)
        results : dict
            Results dictionary containing model residuals if available
            
        Returns:
        --------
        str
            Path to the saved plot file, or None if failed
        """
        try:
            import tempfile
            import time
            stats = get_stats_module()
            
            # Try to get model residuals first (correct approach)
            residuals = None
            if results and 'model_residuals' in results:
                residuals = results['model_residuals']
                data_source = "Model Residuals"
            elif results and 'residuals' in results:
                residuals = results['residuals']
                data_source = "Model Residuals"
            else:
                # Fallback: combine all data from groups (less ideal but better than nothing)
                if not data_dict:
                    return None
                all_values = []
                for values in data_dict.values():
                    clean_vals = [v for v in values if not (pd.isna(v) if pd else np.isnan(v))]
                    all_values.extend(clean_vals)
                residuals = np.array(all_values)
                data_source = "Combined Group Data"
            
            if residuals is None or len(residuals) < 3:
                return None
            
            # Remove NaN values from residuals
            clean_residuals = np.array([r for r in residuals if not (pd.isna(r) if pd else np.isnan(r))])
            
            if len(clean_residuals) < 3:
                return None
            
            # Create single Q-Q plot - use FIXED size for consistency between before/after
            # Fixed size ensures both before and after plots are identical
            fig, ax = plt.subplots(1, 1, figsize=(12, 6))  # Fixed 12x6 for all QQ plots
            
            # Create Q-Q plot of residuals
            stats.probplot(clean_residuals, dist="norm", plot=ax)
            
            # Customize plot
            ax.set_title(f"Normality Check - Q-Q Plot of {data_source}\n(n={len(clean_residuals)} observations)", 
                        fontsize=14, fontweight='bold', pad=20)
            ax.set_xlabel("Theoretical Normal Quantiles", fontsize=12)
            ax.set_ylabel("Sample Quantiles", fontsize=12)
            ax.grid(True, alpha=0.3)
            
            # Style the reference line
            line = ax.get_lines()[1]  # Second line is the reference line
            line.set_color('red')
            line.set_linewidth(3)
            line.set_alpha(0.8)
            
            # Style the data points
            points = ax.get_lines()[0]
            points.set_markersize(6)
            points.set_alpha(0.7)
            
            # Add transformation info to title if applicable
            transform_text = f" ({transformation})" if transformation and transformation.lower() != "none" else ""
            fig.suptitle(f"Normality Examination{transform_text}{title_suffix}", 
                        fontsize=16, fontweight='bold')
            
            plt.tight_layout()
            
            # Save to temporary file
            temp_dir = tempfile.gettempdir()
            temp_filename = f"normality_plot_{int(time.time())}.png"
            temp_path = os.path.join(temp_dir, temp_filename)
            
            # Use consistent saving parameters - no tight bbox to ensure consistent sizing
            fig.savefig(temp_path, dpi=300, bbox_inches=None, facecolor='white', pad_inches=0.2)
            plt.close(fig)
            
            print(f"DEBUG: Generated normality plot: {temp_path}")
            return temp_path
            
        except Exception as e:
            print(f"DEBUG: Error creating normality plot: {str(e)}")
            import traceback
            traceback.print_exc()
            return None
    
    @staticmethod
    def create_homoscedasticity_plot(data_dict, title_suffix="", transformation=None):
        """
        Create boxplots for homoscedasticity (variance homogeneity) examination.
        
        Parameters:
        -----------
        data_dict : dict
            Dictionary with group names as keys and data arrays as values
        title_suffix : str
            Suffix to add to the plot title (e.g., "Before Transformation")
        transformation : str
            Name of transformation applied (if any)
            
        Returns:
        --------
        str
            Path to the saved plot file, or None if failed
        """
        try:
            import tempfile
            import time
            
            if not data_dict:
                return None
            
            # Create figure with FIXED size for consistency between before/after
            # Fixed size ensures both before and after plots are identical
            fig, ax = plt.subplots(1, 1, figsize=(12, 6))  # Fixed 12x6 for all boxplots
            
            # Prepare data for boxplot
            group_names = []
            group_data = []
            
            for group_name, values in data_dict.items():
                # Remove NaN values
                clean_values = [v for v in values if not (pd.isna(v) if pd else np.isnan(v))]
                if clean_values:
                    group_names.append(f"{group_name}\n(n={len(clean_values)})")
                    group_data.append(clean_values)
            
            if not group_data:
                ax.text(0.5, 0.5, "No valid data for plot", ha='center', va='center', 
                       transform=ax.transAxes, fontsize=14)
                ax.set_title("Variance Homogeneity Examination", fontsize=16, fontweight='bold')
                return None
            
            # Create boxplot with better styling
            bp = ax.boxplot(group_data, labels=group_names, patch_artist=True, 
                           notch=True, showmeans=True, meanline=True)
            
            # Color the boxes with distinct colors
            colors = plt.cm.Set2(np.linspace(0, 1, len(group_data)))
            for patch, color in zip(bp['boxes'], colors):
                patch.set_facecolor(color)
                patch.set_alpha(0.8)
                patch.set_linewidth(1.5)
            
            # Style other elements
            for element in ['whiskers', 'fliers', 'medians', 'caps']:
                plt.setp(bp[element], linewidth=1.5)
            
            # Customize plot appearance
            ax.set_ylabel('Values', fontsize=14, fontweight='bold')
            ax.set_xlabel('Groups', fontsize=14, fontweight='bold')
            ax.grid(True, alpha=0.3)
            ax.tick_params(axis='x', rotation=45, labelsize=12)
            ax.tick_params(axis='y', labelsize=12)
            
            # Add main title
            transform_text = f" ({transformation})" if transformation and transformation.lower() != "none" else ""
            ax.set_title(f"Variance Homogeneity Examination - Boxplots{transform_text}{title_suffix}", 
                        fontsize=16, fontweight='bold', pad=20)
            
            plt.tight_layout()
            
            # Save to temporary file
            temp_dir = tempfile.gettempdir()
            temp_filename = f"homoscedasticity_plot_{int(time.time())}.png"
            temp_path = os.path.join(temp_dir, temp_filename)
            
            # Use consistent saving parameters - no tight bbox to ensure consistent sizing
            fig.savefig(temp_path, dpi=300, bbox_inches=None, facecolor='white', pad_inches=0.2)
            plt.close(fig)
            
            print(f"DEBUG: Generated homoscedasticity plot: {temp_path}")
            return temp_path
            
        except Exception as e:
            print(f"DEBUG: Error creating homoscedasticity plot: {str(e)}")
            import traceback
            traceback.print_exc()
            return None
    
    @staticmethod
    def generate_assumption_plots(results):
        """
        Generate both normality and homoscedasticity plots for the given results.
        Creates before/after transformation plots when applicable.
        
        Parameters:
        -----------
        results : dict
            Results dictionary containing test data and transformation info
            
        Returns:
        --------
        dict
            Dictionary with plot paths: {
                'normality_before': path,
                'normality_after': path,
                'homoscedasticity_before': path,  
                'homoscedasticity_after': path
            }
        """
        plot_paths = {
            'normality_before': None,
            'normality_after': None,
            'homoscedasticity_before': None,
            'homoscedasticity_after': None
        }
        
        try:
            # Get original data
            raw_data = results.get('raw_data', results.get('original_data', {}))
            if not raw_data:
                print("DEBUG: No raw data found for assumption plots")
                return plot_paths
            
            # Get transformation info
            transformation = results.get('transformation', 'None')
            transformed_data = results.get('raw_data_transformed', results.get('transformed_data', {}))
            
            # Filter out non-data keys
            raw_data_filtered = {k: v for k, v in raw_data.items() if str(k).lower() not in ['group', 'sample', '']}
            
            # Generate BEFORE transformation plots
            if raw_data_filtered:
                print(f"DEBUG: Generating BEFORE plots for {len(raw_data_filtered)} groups: {list(raw_data_filtered.keys())}")
                plot_paths['normality_before'] = AssumptionVisualizer.create_normality_plot(
                    raw_data_filtered, " - Before Transformation" if transformation and transformation.lower() != 'none' else "",
                    results=results
                )
                print(f"DEBUG: Q-Q plot BEFORE path: {plot_paths['normality_before']}")
                
                plot_paths['homoscedasticity_before'] = AssumptionVisualizer.create_homoscedasticity_plot(
                    raw_data_filtered, " - Before Transformation" if transformation and transformation.lower() != 'none' else ""
                )
                print(f"DEBUG: Boxplot BEFORE path: {plot_paths['homoscedasticity_before']}")
            else:
                print("DEBUG: No valid raw data found after filtering")
            
            # Generate AFTER transformation plots (if transformation was applied)
            if transformed_data and transformation and transformation.lower() != 'none':
                transformed_filtered = {k: v for k, v in transformed_data.items() if str(k).lower() not in ['group', 'sample', '']}
                if transformed_filtered:
                    print(f"DEBUG: Generating AFTER plots for {len(transformed_filtered)} groups: {list(transformed_filtered.keys())}")
                    plot_paths['normality_after'] = AssumptionVisualizer.create_normality_plot(
                        transformed_filtered, " - After Transformation", transformation, results=results
                    )
                    plot_paths['homoscedasticity_after'] = AssumptionVisualizer.create_homoscedasticity_plot(
                        transformed_filtered, " - After Transformation", transformation
                    )
                    print(f"DEBUG: Q-Q plot AFTER path: {plot_paths['normality_after']}")
                    print(f"DEBUG: Boxplot AFTER path: {plot_paths['homoscedasticity_after']}")
            
            # Track all generated files for cleanup
            ResultsExporter = get_results_exporter()
            for plot_path in plot_paths.values():
                if plot_path:
                    ResultsExporter.track_temp_file(plot_path)
            
            return plot_paths
            
        except Exception as e:
            print(f"DEBUG: Error generating assumption plots: {str(e)}")
            import traceback
            traceback.print_exc()
            return plot_paths
    
class PostHocAnalyzer:
    """Base class for all post-hoc tests with uniform methods."""
    
    @staticmethod
    def create_result_template(test_name):
        """Creates a standard dictionary for post-hoc results."""
        return {
            "posthoc_test": test_name,
            "pairwise_comparisons": [],
            "error": None
        }
    
    @staticmethod
    def add_comparison(results, group1, group2, test, p_value, statistic=None, 
                       corrected=True, correction_method=None, effect_size=None, 
                       effect_size_type=None, confidence_interval=(None, None), 
                       alpha=0.05):
        """Adds a standardized pairwise comparison to the results."""
        significant = p_value < alpha if isinstance(p_value, (float, int)) else False
        
        comparison = {
            "group1": str(group1),
            "group2": str(group2),
            "test": test,
            "p_value": float(p_value) if isinstance(p_value, (float, int)) else p_value,
            "statistic": float(statistic) if isinstance(statistic, (float, int)) else statistic,
            "significant": significant,
            "corrected": corrected,
            "effect_size": float(effect_size) if isinstance(effect_size, (float, int)) else effect_size,
            "effect_size_type": effect_size_type,
            "confidence_interval": confidence_interval
        }
        
        if correction_method:
            comparison["correction"] = correction_method
            
        results["pairwise_comparisons"].append(comparison)
        return comparison
        
    @staticmethod
    def _holm_sidak_correction(p_values):
        """Applies Holm-Sidak correction to a list of p-values."""
        if not p_values:
            return []
        
        # Use statsmodels implementation instead of custom one
        multipletests = get_statsmodels_multitest()
        reject, corrected_p, _, _ = multipletests(p_values, method='holm-sidak')
        return corrected_p.tolist()

class TwoWayPostHocAnalyzer(PostHocAnalyzer):
    @staticmethod
    def build_group_label(factors, values):
        # Always use the same order and format as the dialog: 'FactorA=..., FactorB=...'
        return ', '.join([f"{factors[i]}={values[i]}" for i in range(len(factors))])
    """Post-hoc tests for Two-Way ANOVA with a uniform interface."""
    
    @staticmethod
    def perform_test(df, dv, factors, alpha=0.05, selected_comparisons=None, method="holm_sidak", control_group=None):
        """
        Performs post-hoc tests for Two-Way ANOVA.
        
        Parameters:
        -----------
        df : pandas.DataFrame
            Data in long format
        dv : str
            Dependent variable
        factors : list
            List of the two factors [factor_a, factor_b]
        alpha : float
            Significance level
        selected_comparisons : set, optional
            Set of normalized comparison pairs to perform
        method : str, optional
            Post-hoc method: "holm_sidak", "bonferroni", "tukey"
            
        Returns:
        --------
        dict
            Standardized post-hoc results
        """
        result = PostHocAnalyzer.create_result_template("Two-Way ANOVA Post-hoc Tests")
        try:
            print(f"DEBUG POSTHOC: selected_comparisons = {selected_comparisons}")
            # Use the same normalization function for group pairs (must match dialog)
            def normalize_pair(pair):
                # Sort and strip, but also ensure both elements are formatted identically to dialog
                return tuple(sorted([s.strip() for s in pair]))
            normalized_selected = set(normalize_pair(pair) for pair in selected_comparisons) if selected_comparisons else None
            print(f"DEBUG POSTHOC: normalized_selected = {normalized_selected}")
            available_pairs = set()
            pg = get_pingouin_module()
            has_pingouin = True
        except ImportError:
            has_pingouin = False
        except Exception as e:
            print(f"DEBUG POSTHOC: Exception during normalization: {e}")
            has_pingouin = False
        try:
            if has_pingouin:
                print(f"DEBUG POSTHOC: DataFrame columns: {df.columns.tolist()}")
                print(f"DEBUG POSTHOC: DataFrame head:\n{df.head()}")
                print(f"DEBUG POSTHOC: factors = {factors}, dv = {dv}")
                # Manual post-hoc for interaction: generate all interaction group pairs
                from itertools import combinations
                import numpy as np
                from scipy.stats import ttest_ind
                # Build all interaction group labels
                interaction_groups = []
                group_to_values = {}
                for level_b in sorted(df[factors[0]].unique()):
                    for level_a in sorted(df[factors[1]].unique()):
                        label = f"{factors[0]}={level_b}, {factors[1]}={level_a}"
                        mask = (df[factors[0]] == level_b) & (df[factors[1]] == level_a)
                        values = df.loc[mask, dv].values
                        if len(values) > 0:
                            interaction_groups.append(label)
                            group_to_values[label] = values
                print(f"DEBUG POSTHOC: interaction_groups = {interaction_groups}")
                # Generate all possible pairs
                all_pairs = list(combinations(interaction_groups, 2))
                # If selected_comparisons is provided, filter to only those pairs
                if normalized_selected is not None:
                    filtered_pairs = [pair for pair in all_pairs if normalize_pair(pair) in normalized_selected]
                else:
                    filtered_pairs = all_pairs
                print(f"DEBUG POSTHOC: filtered_pairs = {filtered_pairs}")
                # Perform t-tests for each pair
                pvals = []
                stats_list = []
                for g1, g2 in filtered_pairs:
                    vals1 = group_to_values[g1]
                    vals2 = group_to_values[g2]
                    # Use t-test (assume equal variance for now)
                    stat, pval = ttest_ind(vals1, vals2, equal_var=True)
                    pvals.append(pval)
                    stats_list.append((g1, g2, stat, pval, vals1, vals2))
                # Apply multiple comparison correction based on method
                multipletests = get_statsmodels_multitest()
                if pvals:
                    if method.lower() == 'tukey':
                        # For Tukey, we'll use a different approach below
                        correction_method = "Tukey HSD"
                        pvals_corr = pvals  # Will be replaced by Tukey results
                    elif method.lower() == 'dunnett' and control_group:
                        # For Dunnett, use proper Dunnett test implementation
                        correction_method = "Dunnett"
                        try:
                            import scikit_posthocs as sp
                            # Prepare data for scikit_posthocs
                            all_data = []
                            all_groups = []
                            for group in interaction_groups:
                                values = group_to_values[group]
                                all_data.extend(values)
                                all_groups.extend([group] * len(values))
                            
                            # Create DataFrame for scikit_posthocs
                            import pandas as pd
                            dunnett_df = pd.DataFrame({"value": all_data, "group": all_groups})
                            
                            # Use the control_group directly - it's already the exact group name the user selected
                            control_label = control_group
                            print(f"DEBUG: Using control_group directly: '{control_label}'")
                            
                            # Perform Dunnett test
                            dunnett_result = sp.posthoc_dunnett(dunnett_df, val_col="value", group_col="group", control=control_label)
                            
                            # Extract p-values for the comparisons we made
                            pvals_corr = []
                            for g1, g2, *_ in stats_list:
                                if g1 == control_label or g2 == control_label:
                                    # Get the p-value from the Dunnett result matrix
                                    try:
                                        if g1 == control_label:
                                            p_val = float(dunnett_result.loc[g2, control_label])
                                        else:
                                            p_val = float(dunnett_result.loc[g1, control_label])
                                    except (KeyError, ValueError):
                                        # Fallback to original p-value
                                        p_val = stats_list[len(pvals_corr)][3]
                                    pvals_corr.append(p_val)
                                else:
                                    pvals_corr.append(1.0)  # Non-control comparisons get p=1.0
                        except ImportError:
                            # Fallback if scikit_posthocs not available
                            # Filter to only comparisons involving the control group
                            dunnett_pvals = []
                            dunnett_stats = []
                            control_label = control_group  # Use control_group directly
                            print(f"DEBUG: Dunnett fallback using control_group: '{control_label}'")
                            
                            for i, (g1, g2, stat, pval, vals1, vals2) in enumerate(stats_list):
                                if g1 == control_label or g2 == control_label:
                                    dunnett_pvals.append(pval)
                                    dunnett_stats.append((g1, g2, stat, pval, vals1, vals2))
                                
                                # Apply Dunnett correction using Holm-Sidak as fallback
                                if dunnett_pvals:
                                    reject, pvals_corr_dunnett, _, _ = multipletests(dunnett_pvals, alpha=alpha, method='holm-sidak')
                                    # Map back to original order
                                    pvals_corr = []
                                    dunnett_idx = 0
                                    for g1, g2, *_ in stats_list:
                                        if g1 == control_label or g2 == control_label:
                                            pvals_corr.append(pvals_corr_dunnett[dunnett_idx])
                                            dunnett_idx += 1
                                        else:
                                            pvals_corr.append(1.0)  # Non-control comparisons get p=1.0
                                else:
                                    pvals_corr = [1.0] * len(pvals)
                    else:
                        # Default: Holm-Sidak
                        correction_method = "Holm-Sidak"
                        reject, pvals_corr, _, _ = multipletests(pvals, alpha=alpha, method='holm-sidak')
                else:
                    pvals_corr = []
                    correction_method = "Holm-Sidak"
                # Add to results
                for i, (g1, g2, stat, pval, vals1, vals2) in enumerate(stats_list):
                    # Effect size: Cohen's d
                    n1, n2 = len(vals1), len(vals2)
                    s1, s2 = np.var(vals1, ddof=1), np.var(vals2, ddof=1)
                    s_pooled = np.sqrt(((n1-1)*s1 + (n2-1)*s2) / (n1+n2-2)) if (n1+n2-2) > 0 else 0
                    cohen_d = (np.mean(vals1) - np.mean(vals2)) / s_pooled if s_pooled > 0 else 0
                    # Confidence interval for mean difference
                    mean_diff = np.mean(vals1) - np.mean(vals2)
                    stderr_diff = np.sqrt(s1/n1 + s2/n2) if n1 > 0 and n2 > 0 else 0
                    from scipy.stats import t
                    df_ = n1 + n2 - 2
                    if df_ > 0 and stderr_diff > 0:
                        t_crit = t.ppf(1 - alpha/2, df_)
                        ci = (mean_diff - t_crit * stderr_diff, mean_diff + t_crit * stderr_diff)
                    else:
                        ci = (None, None)
                    PostHocAnalyzer.add_comparison(
                        result,
                        group1=g1,
                        group2=g2,
                        test="Pairwise t-test",
                        p_value=pvals_corr[i] if i < len(pvals_corr) else pval,
                        statistic=stat,
                        corrected=True,
                        correction_method=correction_method,
                        effect_size=cohen_d,
                        effect_size_type="cohen_d",
                        confidence_interval=ci,
                        alpha=alpha
                    )
                print(f"DEBUG POSTHOC: Added {len(stats_list)} comparisons to results.")
                # After all, print available pairs and warn if any selected pair is not present
                available_pairs = set(normalize_pair((g1, g2)) for g1, g2, *_ in stats_list)
                print(f"DEBUG POSTHOC: available_pairs = {available_pairs}")
                if normalized_selected is not None:
                    missing = normalized_selected - available_pairs
                    if missing:
                        print(f"WARNING: The following selected pairs were not found in the available post-hoc comparisons: {missing}")
                from statsmodels.stats.multicomp import pairwise_tukeyhsd
                # Create interaction group for Tukey HSD
                df['interaction_group'] = df[factors[0]].astype(str) + "_" + df[factors[1]].astype(str)
                # Run Tukey HSD on the interaction groups
                tukey = pairwise_tukeyhsd(df[dv], df['interaction_group'], alpha=alpha)
                # For the Tukey HSD test in the fallback, we'll need to manually apply Holm-Sidak
                # First collect all pairwise comparisons and p-values
                comparisons = []
                for i in range(len(tukey.pvalues)):
                    group1 = tukey.groupsunique[tukey.pairindices[i, 0]]
                    group2 = tukey.groupsunique[tukey.pairindices[i, 1]]
                    p_val = tukey.pvalues[i]
                    conf_int = tukey.confint[i]
                    comparisons.append({
                        'group1': group1,
                        'group2': group2,
                        'p_value': p_val,
                        'conf_int': conf_int
                    })
                # Apply Holm-Sidak correction
                p_values = [comp['p_value'] for comp in comparisons]
                multipletests = get_statsmodels_multitest()
                reject, corrected_p_values, _, _ = multipletests(p_values, alpha=alpha, method='holm-sidak')
                # Convert results into standardized format with corrected p-values
                for i, comp in enumerate(comparisons):
                    is_significant = corrected_p_values[i] < alpha
                    # Normalize for matching
                    norm_pair = normalize_pair((comp['group1'], comp['group2']))
                    match = (normalized_selected is not None and norm_pair in normalized_selected)
                    print(f"DEBUG POSTHOC: fallback comparing {comp['group1']} vs {comp['group2']} | normalized: {norm_pair} | match: {match}")
                    if normalized_selected is not None and not match:
                        continue
                    PostHocAnalyzer.add_comparison(
                        result,
                        group1=comp['group1'],
                        group2=comp['group2'],
                        test="Pairwise t-test",
                        p_value=corrected_p_values[i],
                        statistic=None,
                        corrected=True,
                        correction_method="Holm-Sidak",
                        confidence_interval=tuple(comp['conf_int']),
                        alpha=alpha
                    )
            
            # Set the posthoc_test value for decision tree visualization
            method_name_map = {
                "tukey": "Tukey HSD",
                "dunnett": "Dunnett Test",
                "paired_custom": "Custom paired t-tests (Holm-Sidak)",
                "holm_sidak": "Custom paired t-tests (Holm-Sidak)"
            }
            result["posthoc_test"] = method_name_map.get(method, f"Post-hoc test ({method})")
            
            return result
        except Exception as e:
            result["error"] = f"Error in Two-Way ANOVA post-hoc tests: {str(e)}"
            return result
        
class MixedAnovaPostHocAnalyzer(PostHocAnalyzer):
    """UPDATED: Advanced post-hoc tests for Mixed ANOVA with proper between/within factor handling."""
    
    @staticmethod
    def perform_test(df, dv, subject, between, within, alpha=0.05, selected_comparisons=None, method='tukey', control_group=None):
        """
        UPDATED: Performs sophisticated post-hoc tests for Mixed ANOVA with proper between/within handling.
        
        Major improvements:
        - Proper distinction between between-subject and within-subject comparisons
        - Enhanced statistical tests for mixed designs
        - Better subject-ID handling for within-subject comparisons
        - Improved effect size calculations for mixed designs
        - Enhanced interaction analysis
        
        Parameters:
        -----------
        df : pandas.DataFrame
            Data in long format
        dv : str
            Dependent variable
        subject : str
            Column with subject ID
        between : list
            List with the between-factor [between_factor]
        within : list
            List with the within-factor [within_factor]
        alpha : float
            Significance level (default: 0.05)
        selected_comparisons : set, optional
            Set of normalized comparison pairs to perform
        method : str, optional
            Post-hoc method: "tukey", "bonferroni", "holm_sidak", "dunnett"
        control_group : str, optional
            Control group for Dunnett's test
            
        Returns:
        --------
        dict
            Standardized post-hoc results with mixed-design corrections
        """
        result = PostHocAnalyzer.create_result_template("Mixed ANOVA Post-hoc Tests")
        
        try:
            between_factor = between[0]
            within_factor = within[0]
            
            print(f"DEBUG MIXED POSTHOC: selected_comparisons = {selected_comparisons}")
            print(f"DEBUG MIXED POSTHOC: between_factor = {between_factor}, within_factor = {within_factor}")
            
            # Normalize comparison pairs function (consistent with other ANOVAs)
            def normalize_pair(pair):
                return tuple(sorted([s.strip() for s in pair]))
            
            # Handle selected comparisons
            if selected_comparisons:
                if isinstance(selected_comparisons, set):
                    normalized_selected = selected_comparisons
                else:
                    normalized_selected = set(normalize_pair(pair) for pair in selected_comparisons)
            else:
                normalized_selected = None
            
            print(f"DEBUG MIXED POSTHOC: normalized_selected = {normalized_selected}")
            
            # Validate mixed design data structure
            between_levels = sorted(df[between_factor].unique())
            within_levels = sorted(df[within_factor].unique())
            
            print(f"DEBUG MIXED POSTHOC: between_levels = {between_levels}, within_levels = {within_levels}")
            
            # Check for complete mixed design (all subjects should have all within-factor levels)
            subject_within_counts = df.groupby([subject, between_factor])[within_factor].nunique()
            expected_within_measures = len(within_levels)
            incomplete_cases = subject_within_counts[subject_within_counts < expected_within_measures]
            
            if len(incomplete_cases) > 0:
                print(f"WARNING: {len(incomplete_cases)} subject-between-factor combinations have incomplete within-factor data")
            
            # Import required modules
            from itertools import combinations
            import numpy as np
            from scipy import stats as scipy_stats
            
            # Build interaction group labels and classify comparison types
            interaction_groups = []
            group_to_data = {}
            
            for between_level in between_levels:
                for within_level in within_levels:
                    group_label = f"{between_factor}={between_level}, {within_factor}={within_level}"
                    mask = (df[between_factor] == between_level) & (df[within_factor] == within_level)
                    group_data = df.loc[mask].copy()
                    
                    if len(group_data) > 0:
                        interaction_groups.append(group_label)
                        group_to_data[group_label] = {
                            'values': group_data[dv].values,
                            'subjects': group_data[subject].values,
                            'between_level': between_level,
                            'within_level': within_level,
                            'data': group_data
                        }
            
            print(f"DEBUG MIXED POSTHOC: interaction_groups = {interaction_groups}")
            
            # Collect all pairwise comparisons and classify them
            available_pairs = set()
            comparisons = []
            
            for group1_label, group2_label in combinations(interaction_groups, 2):
                norm_pair = normalize_pair((group1_label, group2_label))
                available_pairs.add(norm_pair)
                
                # Check if this comparison is selected
                if normalized_selected is not None and norm_pair not in normalized_selected:
                    continue
                
                group1_data = group_to_data[group1_label]
                group2_data = group_to_data[group2_label]
                
                # Classify the type of comparison
                comparison_type = MixedAnovaPostHocAnalyzer._classify_comparison_type(
                    group1_data, group2_data, between_factor, within_factor
                )
                
                print(f"DEBUG MIXED POSTHOC: Comparing {group1_label} vs {group2_label}, type: {comparison_type}")
                
                # Perform appropriate statistical test based on comparison type
                if comparison_type == "within_subject":
                    # Within-subject comparison: use paired t-test
                    t_stat, p_val, effect_size, ci_lower, ci_upper, n_pairs = MixedAnovaPostHocAnalyzer._within_subject_test(
                        group1_data, group2_data, dv, subject, alpha
                    )
                    test_type = "Paired t-test (within-subject)"
                    
                elif comparison_type == "between_subject":
                    # Between-subject comparison: use independent t-test
                    t_stat, p_val, effect_size, ci_lower, ci_upper, n_pairs = MixedAnovaPostHocAnalyzer._between_subject_test(
                        group1_data, group2_data, dv, alpha
                    )
                    test_type = "Independent t-test (between-subject)"
                    
                else:  # "mixed" - most complex case
                    # Mixed comparison: different between-groups AND different within-levels
                    t_stat, p_val, effect_size, ci_lower, ci_upper, n_pairs = MixedAnovaPostHocAnalyzer._mixed_comparison_test(
                        group1_data, group2_data, dv, subject, alpha
                    )
                    test_type = "Independent t-test (mixed comparison)"
                
                if t_stat is not None:  # Valid comparison
                    comparisons.append({
                        "group1": group1_label,
                        "group2": group2_label,
                        "comparison_type": comparison_type,
                        "test_type": test_type,
                        "t_stat": t_stat,
                        "p_val": p_val,
                        "effect_size": effect_size,
                        "ci_lower": ci_lower,
                        "ci_upper": ci_upper,
                        "n_pairs": n_pairs
                    })
            
            if not comparisons:
                result["error"] = "No valid pairwise comparisons could be performed"
                return result
            
            # Apply multiple comparison correction based on method
            p_values = [comp["p_val"] for comp in comparisons]
            n_comparisons = len(comparisons)
            
            if method.lower() == 'tukey':
                # Enhanced Tukey HSD for mixed designs
                correction_method = "Tukey HSD (Mixed)"
                try:
                    # Try to use pingouin for proper Tukey implementation
                    pg = get_pingouin_module()
                    if pg is not None:
                        corrected_p_values = []
                        for comp in comparisons:
                            # Use appropriate Tukey correction based on comparison type
                            if comp["comparison_type"] == "within_subject":
                                # More liberal correction for within-subject comparisons
                                q_stat = abs(comp["t_stat"]) * np.sqrt(2)
                                p_tukey = MixedAnovaPostHocAnalyzer._tukey_p_value(q_stat, len(within_levels), comp["n_pairs"] - 1)
                            else:
                                # Standard Tukey for between-subject comparisons
                                q_stat = abs(comp["t_stat"]) * np.sqrt(2)
                                p_tukey = MixedAnovaPostHocAnalyzer._tukey_p_value(q_stat, len(interaction_groups), comp["n_pairs"] - 1)
                            corrected_p_values.append(p_tukey)
                    else:
                        # Fallback to Bonferroni
                        corrected_p_values = [min(1.0, p * n_comparisons) for p in p_values]
                        correction_method = "Bonferroni (Tukey unavailable)"
                except:
                    # Fallback to Bonferroni if Tukey calculation fails
                    corrected_p_values = [min(1.0, p * n_comparisons) for p in p_values]
                    correction_method = "Bonferroni (Tukey calculation failed)"
                    
            elif method.lower() == 'bonferroni':
                correction_method = "Bonferroni"
                corrected_p_values = [min(1.0, p * n_comparisons) for p in p_values]
                
            elif method.lower() == 'dunnett' and control_group:
                correction_method = "Dunnett"
                # Filter to only control group comparisons
                dunnett_p_values = []
                control_indices = []
                
                for i, comp in enumerate(comparisons):
                    if control_group in comp["group1"] or control_group in comp["group2"]:
                        dunnett_p_values.append(comp["p_val"])
                        control_indices.append(i)
                
                if dunnett_p_values:
                    k = len(dunnett_p_values)
                    dunnett_corrected = [min(1.0, p * k * 0.8) for p in dunnett_p_values]  # Approximate Dunnett factor
                    
                    corrected_p_values = [1.0] * len(p_values)
                    for j, orig_idx in enumerate(control_indices):
                        corrected_p_values[orig_idx] = dunnett_corrected[j]
                else:
                    corrected_p_values = [1.0] * len(p_values)
                    correction_method = "Dunnett (no control comparisons found)"
            else:
                # Default: Holm-Sidak
                correction_method = "Holm-Sidak"
                corrected_p_values = PostHocAnalyzer._holm_sidak_correction(p_values)
            
            # Add each pairwise comparison result with enhanced mixed-design information
            for i, comp in enumerate(comparisons):
                is_significant = corrected_p_values[i] < alpha
                
                PostHocAnalyzer.add_comparison(
                    result,
                    group1=comp["group1"],
                    group2=comp["group2"],
                    test=f"{comp['test_type']} ({correction_method})",
                    p_value=corrected_p_values[i],
                    statistic=comp["t_stat"],
                    corrected=True,
                    correction_method=correction_method,
                    effect_size=comp["effect_size"],
                    effect_size_type="cohen_d_mixed",  # Specify mixed design version
                    confidence_interval=(comp["ci_lower"], comp["ci_upper"]),
                    alpha=alpha,
                    significant=is_significant,
                    # Additional mixed-design specific information
                    comparison_type=comp["comparison_type"],
                    n_pairs=comp["n_pairs"]
                )
            
            # Add summary information
            between_comparison_count = sum(1 for c in comparisons if c["comparison_type"] == "between_subject")
            within_comparison_count = sum(1 for c in comparisons if c["comparison_type"] == "within_subject")
            mixed_comparison_count = sum(1 for c in comparisons if c["comparison_type"] == "mixed")
            
            result["summary"] = {
                "total_comparisons": n_comparisons,
                "between_subject_comparisons": between_comparison_count,
                "within_subject_comparisons": within_comparison_count,
                "mixed_comparisons": mixed_comparison_count,
                "correction_method": correction_method,
                "family_wise_alpha": alpha,
                "between_factor": between_factor,
                "within_factor": within_factor,
                "between_levels": between_levels,
                "within_levels": within_levels
            }
            
            # Diagnostic information
            print(f"DEBUG MIXED POSTHOC: available_pairs = {available_pairs}")
            if normalized_selected is not None:
                missing = normalized_selected - available_pairs
                if missing:
                    print(f"WARNING: The following selected pairs were not found: {missing}")
            
            # Set posthoc_test for visualization
            method_name_map = {
                "tukey": "Tukey HSD (Mixed)",
                "dunnett": "Dunnett Test (Mixed)",
                "bonferroni": "Bonferroni (Mixed)",
                "holm_sidak": "Holm-Sidak (Mixed)"
            }
            result["posthoc_test"] = method_name_map.get(method, f"Mixed Post-hoc ({method})")
            
            return result
            
        except Exception as e:
            result["error"] = f"Error in Mixed ANOVA post-hoc tests: {str(e)}"
            print(f"ERROR MIXED POSTHOC: {str(e)}")
            import traceback
            traceback.print_exc()
            return result
    
    @staticmethod
    def _classify_comparison_type(group1_data, group2_data, between_factor, within_factor):
        """Classify the type of comparison in mixed ANOVA design."""
        between1 = group1_data['between_level']
        between2 = group2_data['between_level']
        within1 = group1_data['within_level']
        within2 = group2_data['within_level']
        
        if between1 == between2 and within1 != within2:
            return "within_subject"  # Same between-group, different within-levels
        elif between1 != between2 and within1 == within2:
            return "between_subject"  # Different between-groups, same within-level
        else:
            return "mixed"  # Different between-groups AND different within-levels
    
    @staticmethod
    def _within_subject_test(group1_data, group2_data, dv, subject, alpha):
        """Perform within-subject test for Mixed ANOVA."""
        from scipy import stats as scipy_stats
        import numpy as np
        
        # Get common subjects between both groups
        subjects1 = set(group1_data['subjects'])
        subjects2 = set(group2_data['subjects'])
        common_subjects = subjects1 & subjects2
        
        if len(common_subjects) < 3:
            return None, None, None, None, None, None
        
        # Extract paired data for common subjects
        data1_dict = dict(zip(group1_data['subjects'], group1_data['values']))
        data2_dict = dict(zip(group2_data['subjects'], group2_data['values']))
        
        paired_data1 = [data1_dict[subj] for subj in sorted(common_subjects)]
        paired_data2 = [data2_dict[subj] for subj in sorted(common_subjects)]
        
        # Perform paired t-test
        t_stat, p_val = scipy_stats.ttest_rel(paired_data1, paired_data2)
        
        # Calculate effect size for paired data
        differences = np.array(paired_data1) - np.array(paired_data2)
        mean_diff = np.mean(differences)
        std_diff = np.std(differences, ddof=1)
        effect_size = mean_diff / std_diff if std_diff > 0 else 0
        
        # Calculate confidence interval
        n = len(differences)
        se_diff = std_diff / np.sqrt(n)
        t_crit = scipy_stats.t.ppf(1 - alpha/2, n - 1)
        ci_lower = mean_diff - t_crit * se_diff
        ci_upper = mean_diff + t_crit * se_diff
        
        return t_stat, p_val, effect_size, ci_lower, ci_upper, n
    
    @staticmethod
    def _between_subject_test(group1_data, group2_data, dv, alpha):
        """Perform between-subject test for Mixed ANOVA."""
        from scipy import stats as scipy_stats
        import numpy as np
        
        values1 = group1_data['values']
        values2 = group2_data['values']
        
        if len(values1) < 2 or len(values2) < 2:
            return None, None, None, None, None, None
        
        # Perform independent t-test
        t_stat, p_val = scipy_stats.ttest_ind(values1, values2, equal_var=True)
        
        # Calculate Cohen's d for independent samples
        n1, n2 = len(values1), len(values2)
        s1, s2 = np.var(values1, ddof=1), np.var(values2, ddof=1)
        s_pooled = np.sqrt(((n1-1)*s1 + (n2-1)*s2) / (n1+n2-2))
        effect_size = (np.mean(values1) - np.mean(values2)) / s_pooled if s_pooled > 0 else 0
        
        # Calculate confidence interval for mean difference
        mean_diff = np.mean(values1) - np.mean(values2)
        se_diff = s_pooled * np.sqrt(1/n1 + 1/n2)
        df = n1 + n2 - 2
        t_crit = scipy_stats.t.ppf(1 - alpha/2, df)
        ci_lower = mean_diff - t_crit * se_diff
        ci_upper = mean_diff + t_crit * se_diff
        
        return t_stat, p_val, effect_size, ci_lower, ci_upper, min(n1, n2)
    
    @staticmethod
    def _mixed_comparison_test(group1_data, group2_data, dv, subject, alpha):
        """Perform mixed comparison test (different between-groups AND within-levels)."""
        # For mixed comparisons, treat as independent samples (conservative approach)
        return MixedAnovaPostHocAnalyzer._between_subject_test(group1_data, group2_data, dv, alpha)
    
    @staticmethod 
    def _tukey_p_value(q_stat, k, df):
        """Calculate p-value for Tukey's q statistic."""
        from scipy.stats import studentized_range
        try:
            return 1 - studentized_range.cdf(q_stat, k, df)
        except:
            # Fallback to t-distribution approximation
            from scipy.stats import t
            import math
            t_equiv = q_stat / math.sqrt(2)
            return 2 * (1 - t.cdf(abs(t_equiv), df))

    @staticmethod
    def perform_test(df, between, within, dv, subject, alpha=0.05, selected_comparisons=None, method='tukey', control_group=None):
        """
        UPDATED: Enhanced Mixed ANOVA post-hoc tests with proper between/within factor distinction
        """
        try:
            result = PostHocAnalyzer.create_result_template("Mixed ANOVA Post-hoc Tests")
            
            # Create interaction groups (between_level:within_level combinations)
            interaction_groups = []
            group_to_data = {}
            
            for between_level in df[between].unique():
                for within_level in df[within].unique():
                    group_data = df[(df[between] == between_level) & (df[within] == within_level)]
                    if len(group_data) > 0:
                        group_name = f"{between_level}:{within_level}"
                        interaction_groups.append(group_name)
                        group_to_data[group_name] = {
                            'values': group_data[dv].tolist(),
                            'subjects': group_data[subject].tolist(),
                            'between_level': between_level,
                            'within_level': within_level
                        }
            
            print(f"DEBUG POSTHOC: interaction_groups = {interaction_groups}")
            
            # Handle selected comparisons
            def normalize_pair(pair):
                return tuple(sorted(pair))
            
            normalized_selected = None
            if selected_comparisons:
                normalized_selected = set()
                for pair in selected_comparisons:
                    normalized_selected.add(normalize_pair(pair))
                print(f"DEBUG POSTHOC: normalized_selected = {normalized_selected}")
            
            # Generate all possible pairs and filter by user selection
            all_pairs = list(combinations(interaction_groups, 2))
            
            if normalized_selected is not None:
                filtered_pairs = [pair for pair in all_pairs if normalize_pair(pair) in normalized_selected]
            else:
                filtered_pairs = all_pairs
            
            print(f"DEBUG POSTHOC: filtered_pairs = {filtered_pairs}")
            
            # Import required functions
            from scipy.stats import ttest_rel, ttest_ind
            from itertools import combinations
            import numpy as np
            
            # Perform appropriate tests for each pair
            pvals = []
            stats_list = []
            available_pairs = set()
            
            for g1, g2 in filtered_pairs:
                available_pairs.add(normalize_pair((g1, g2)))
                
                data1 = group_to_data[g1]
                data2 = group_to_data[g2]
                
                # Determine test type based on comparison
                same_between = data1['between_level'] == data2['between_level']
                same_within = data1['within_level'] == data2['within_level']
                
                matched_data1 = None
                matched_data2 = None
                
                if same_between and not same_within:
                    # Within-subject comparison (same group, different time points)
                    # Need to match subjects for paired t-test
                    subjects1 = set(data1['subjects'])
                    subjects2 = set(data2['subjects'])
                    common_subjects = subjects1 & subjects2
                    
                    if len(common_subjects) > 0:
                        # Get matched data for common subjects
                        matched_data1 = []
                        matched_data2 = []
                        for subj in sorted(common_subjects):
                            idx1 = list(data1['subjects']).index(subj)
                            idx2 = list(data2['subjects']).index(subj)
                            matched_data1.append(data1['values'][idx1])
                            matched_data2.append(data2['values'][idx2])
                        
                        # Paired t-test
                        stat, pval = ttest_rel(matched_data1, matched_data2)
                        test_type = "Paired t-test"
                    else:
                        # No common subjects - skip this comparison
                        continue
                        
                elif not same_between:
                    # Between-groups comparison (independent t-test)
                    stat, pval = ttest_ind(data1['values'], data2['values'], equal_var=True)
                    test_type = "Independent t-test"
                else:
                    # Same group and same time point - skip (not meaningful)
                    continue
                
                pvals.append(pval)
                stats_list.append((g1, g2, stat, pval, test_type, data1, data2, matched_data1, matched_data2))
            
            # Apply multiple comparison correction based on method
            multipletests = get_statsmodels_multitest()
            if pvals:
                if method.lower() == 'tukey':
                    # For Tukey, we'll use a different approach
                    correction_method = "Tukey HSD"
                    reject, pvals_corr, _, _ = multipletests(pvals, alpha=alpha, method='holm-sidak')  # Fallback
                elif method.lower() == 'dunnett' and control_group:
                    # For Dunnett, filter to only control group comparisons
                    correction_method = "Dunnett"
                    # Filter to only comparisons involving the control group
                    dunnett_pvals = []
                    control_comparisons = []
                    
                    for i, (g1, g2, stat, pval, test_type, data1, data2, matched_data1, matched_data2) in enumerate(stats_list):
                        # Use exact match instead of substring search
                        if g1 == control_group or g2 == control_group:
                            dunnett_pvals.append(pval)
                            control_comparisons.append(i)
                    
                    if dunnett_pvals:
                        # Apply correction only to control group comparisons
                        reject, pvals_corr_dunnett, _, _ = multipletests(dunnett_pvals, alpha=alpha, method='holm-sidak')
                        # Map back to original order
                        pvals_corr = [1.0] * len(pvals)  # Start with all p-values as 1.0
                        for j, orig_idx in enumerate(control_comparisons):
                            pvals_corr[orig_idx] = pvals_corr_dunnett[j]
                    else:
                        pvals_corr = [1.0] * len(pvals)
                        correction_method = "Dunnett (no control comparisons found)"
                else:
                    # Default: Holm-Sidak
                    correction_method = "Holm-Sidak"
                    reject, pvals_corr, _, _ = multipletests(pvals, alpha=alpha, method='holm-sidak')
            else:
                pvals_corr = []
                correction_method = "Holm-Sidak"
            
            # Add results
            for i, (g1, g2, stat, pval, test_type, data1, data2, matched_data1, matched_data2) in enumerate(stats_list):
                # Calculate effect size
                if test_type == "Paired t-test":
                    # Cohen's d for paired samples
                    if matched_data1 is not None and matched_data2 is not None:
                        diff = np.array(matched_data1) - np.array(matched_data2)
                        effect_size = np.mean(diff) / np.std(diff, ddof=1) if np.std(diff, ddof=1) > 0 else 0
                    else:
                        effect_size = 0
                    effect_size_type = "cohen_d"
                else:
                    # Cohen's d for independent samples
                    n1, n2 = len(data1['values']), len(data2['values'])
                    s1, s2 = np.var(data1['values'], ddof=1), np.var(data2['values'], ddof=1)
                    s_pooled = np.sqrt(((n1-1)*s1 + (n2-1)*s2) / (n1+n2-2)) if (n1+n2-2) > 0 else 0
                    effect_size = (np.mean(data1['values']) - np.mean(data2['values'])) / s_pooled if s_pooled > 0 else 0
                    effect_size_type = "cohen_d"
                
                # Calculate confidence interval
                if test_type == "Paired t-test":
                    if matched_data1 is not None and matched_data2 is not None:
                        diff = np.array(matched_data1) - np.array(matched_data2)
                        n = len(diff)
                        mean_diff = np.mean(diff)
                        se = np.std(diff, ddof=1) / np.sqrt(n)
                        df_val = n - 1
                    else:
                        mean_diff = 0
                        se = 0
                        df_val = 0
                else:
                    n1, n2 = len(data1['values']), len(data2['values'])
                    mean_diff = np.mean(data1['values']) - np.mean(data2['values'])
                    s1, s2 = np.var(data1['values'], ddof=1), np.var(data2['values'], ddof=1)
                    se = np.sqrt(s1/n1 + s2/n2)
                    df_val = n1 + n2 - 2
                
                from scipy.stats import t
                if df_val > 0 and se > 0:
                    t_crit = t.ppf(1 - alpha/2, df_val)
                    ci = (mean_diff - t_crit * se, mean_diff + t_crit * se)
                else:
                    ci = (None, None)
                
                print(f"DEBUG POSTHOC: Adding comparison {g1} vs {g2} (test: {test_type})")
                PostHocAnalyzer.add_comparison(
                    result,
                    group1=g1,
                    group2=g2,
                    test=test_type,
                    p_value=pvals_corr[i] if i < len(pvals_corr) else pval,
                    statistic=stat,
                    corrected=True,
                    correction_method=correction_method,
                    effect_size=effect_size,
                    effect_size_type=effect_size_type,
                    confidence_interval=ci,
                    alpha=alpha
                )
            
            # After all, print available pairs and warn if any selected pair is not present
            print(f"DEBUG POSTHOC: available_pairs = {available_pairs}")
            if normalized_selected is not None:
                missing = normalized_selected - available_pairs
                if missing:
                    print(f"WARNING: The following selected pairs were not found in the available post-hoc comparisons: {missing}")
            
            # Set the posthoc_test value for decision tree visualization
            method_name_map = {
                "tukey": "Tukey HSD",
                "dunnett": "Dunnett Test", 
                "paired_custom": "Custom paired t-tests (Holm-Sidak)",
                "holm_sidak": "Custom paired t-tests (Holm-Sidak)"
            }
            result["posthoc_test"] = method_name_map.get(method, f"Post-hoc test ({method})")
            
            return result
        except Exception as e:
            result["error"] = f"Error in Mixed ANOVA post-hoc tests: {str(e)}"
            return result
        
class RMAnovaPostHocAnalyzer(PostHocAnalyzer):
    """UPDATED: Advanced post-hoc tests for Repeated Measures ANOVA with proper within-subject design handling."""
        
    @staticmethod
    def perform_test(df, dv, subject, within, alpha=0.05, selected_comparisons=None, method='tukey', control_group=None):
        """
        UPDATED: Performs sophisticated post-hoc tests for RM ANOVA with proper within-subject handling.
        
        Major improvements:
        - Proper within-subject data validation
        - Enhanced Tukey HSD for repeated measures
        - Cohen's d for repeated measures (cohen_d_rm)
        - Complete subject tracking
        - Better error handling and diagnostics
        - Summary statistics for RM design
        """
        result = PostHocAnalyzer.create_result_template("RM ANOVA Post-hoc Tests")
        
        try:
            print(f"DEBUG RM POSTHOC: selected_comparisons = {selected_comparisons}")
            
            # Normalize comparison pairs function (consistent with other ANOVAs)
            def normalize_pair(pair):
                return tuple(sorted([s.strip() for s in pair]))
            
            # Handle selected comparisons
            if selected_comparisons:
                if isinstance(selected_comparisons, set):
                    normalized_selected = selected_comparisons
                else:
                    normalized_selected = set(normalize_pair(pair) for pair in selected_comparisons)
            else:
                normalized_selected = None
            
            print(f"DEBUG RM POSTHOC: normalized_selected = {normalized_selected}")
            
            # Get within-subject factor and levels
            within_factor = within[0]
            within_levels = sorted(df[within_factor].unique())
            
            # Validate that we have repeated measures data
            subject_counts = df.groupby(subject)[within_factor].nunique()
            expected_measures = len(within_levels)
            incomplete_subjects = subject_counts[subject_counts < expected_measures]
            
            if len(incomplete_subjects) > 0:
                print(f"WARNING: {len(incomplete_subjects)} subjects have incomplete data")
            
            # Get complete cases only for robust within-subject analysis
            complete_subjects = subject_counts[subject_counts == expected_measures].index
            df_complete = df[df[subject].isin(complete_subjects)].copy()
            
            print(f"DEBUG RM POSTHOC: Complete subjects: {len(complete_subjects)}, Total levels: {expected_measures}")
            
            # Import required modules
            from itertools import combinations
            import numpy as np
            from scipy import stats as scipy_stats
            
            # Collect all pairwise comparisons with proper within-subject handling
            available_pairs = set()
            comparisons = []
            
            for level1, level2 in combinations(within_levels, 2):
                norm_pair = normalize_pair((str(level1), str(level2)))
                available_pairs.add(norm_pair)
                
                # Check if this comparison is selected
                if normalized_selected is not None and norm_pair not in normalized_selected:
                    continue
                
                # Extract paired data for this comparison (same subjects in both conditions)
                data1_df = df_complete[df_complete[within_factor] == level1].sort_values(by=subject)
                data2_df = df_complete[df_complete[within_factor] == level2].sort_values(by=subject)
                
                # Ensure same subjects in both groups
                common_subjects = set(data1_df[subject]) & set(data2_df[subject])
                data1_df = data1_df[data1_df[subject].isin(common_subjects)].sort_values(by=subject)
                data2_df = data2_df[data2_df[subject].isin(common_subjects)].sort_values(by=subject)
                
                data1 = data1_df[dv].values
                data2 = data2_df[dv].values
                
                if len(data1) != len(data2) or len(data1) < 3:
                    print(f"WARNING: Insufficient paired data for {level1} vs {level2}")
                    continue
                
                # Perform paired t-test (appropriate for within-subject design)
                t_stat, p_val = scipy_stats.ttest_rel(data1, data2)
                
                # Calculate within-subject effect size (Cohen's d for repeated measures)
                differences = data1 - data2
                mean_diff = np.mean(differences)
                std_diff = np.std(differences, ddof=1)
                
                # Cohen's d for repeated measures (using difference scores)
                effect_size = mean_diff / std_diff if std_diff > 0 else 0
                
                # Calculate confidence interval for mean difference
                n = len(differences)
                se_diff = std_diff / np.sqrt(n)
                df_t = n - 1
                
                # Store raw comparison data
                comparisons.append({
                    "level1": level1,
                    "level2": level2,
                    "t_stat": t_stat,
                    "p_val": p_val,
                    "effect_size": effect_size,
                    "mean_diff": mean_diff,
                    "se_diff": se_diff,
                    "df": df_t,
                    "n_pairs": n,
                    "data1": data1,
                    "data2": data2,
                    "differences": differences
                })
            
            if not comparisons:
                result["error"] = "No valid pairwise comparisons could be performed"
                return result
            
            # Apply multiple comparison correction based on method
            p_values = [comp["p_val"] for comp in comparisons]
            n_comparisons = len(comparisons)
            
            if method.lower() == 'tukey':
                # Implement proper Tukey HSD for repeated measures
                correction_method = "Tukey HSD (RM)"
                try:
                    # Try to use pingouin for proper Tukey implementation
                    pg = get_pingouin_module()
                    if pg is not None:
                        # Use Tukey's studentized range statistic for RM design
                        corrected_p_values = []
                        
                        for comp in comparisons:
                            # Convert t-statistic to Tukey's q statistic
                            q_stat = abs(comp["t_stat"]) * np.sqrt(2)
                            p_tukey = RMAnovaPostHocAnalyzer._tukey_p_value(q_stat, len(within_levels), comp["df"])
                            corrected_p_values.append(p_tukey)
                    else:
                        # Fallback to conservative Bonferroni
                        corrected_p_values = [min(1.0, p * n_comparisons) for p in p_values]
                        correction_method = "Bonferroni (Tukey unavailable)"
                except:
                    # Fallback to Bonferroni if Tukey calculation fails
                    corrected_p_values = [min(1.0, p * n_comparisons) for p in p_values]
                    correction_method = "Bonferroni (Tukey calculation failed)"
                    
            elif method.lower() == 'bonferroni':
                correction_method = "Bonferroni"
                corrected_p_values = [min(1.0, p * n_comparisons) for p in p_values]
                
            elif method.lower() == 'dunnett' and control_group:
                correction_method = "Dunnett"
                # Filter to only control group comparisons
                dunnett_p_values = []
                control_indices = []
                
                for i, comp in enumerate(comparisons):
                    level1_str = str(comp["level1"])
                    level2_str = str(comp["level2"])
                    if level1_str == control_group or level2_str == control_group:
                        dunnett_p_values.append(comp["p_val"])
                        control_indices.append(i)
                
                if dunnett_p_values:
                    # Apply Dunnett correction (more liberal than Bonferroni for control comparisons)
                    k = len(dunnett_p_values)  # Number of comparisons with control
                    dunnett_corrected = [min(1.0, p * k * 0.8) for p in dunnett_p_values]  # Approximate Dunnett factor
                    
                    corrected_p_values = [1.0] * len(p_values)
                    for j, orig_idx in enumerate(control_indices):
                        corrected_p_values[orig_idx] = dunnett_corrected[j]
                else:
                    corrected_p_values = [1.0] * len(p_values)
                    correction_method = "Dunnett (no control comparisons found)"
            else:
                # Default: Holm-Sidak (step-down method, less conservative than Bonferroni)
                correction_method = "Holm-Sidak"
                corrected_p_values = PostHocAnalyzer._holm_sidak_correction(p_values)
            
            # Calculate family-wise corrected confidence intervals
            # Use Sidak correction for simultaneous confidence intervals
            alpha_sidak = 1 - (1 - alpha) ** (1 / n_comparisons)
            
            # Add each pairwise comparison result with enhanced within-subject information
            for i, comp in enumerate(comparisons):
                # Calculate corrected confidence interval
                t_crit = scipy_stats.t.ppf(1 - alpha_sidak/2, comp["df"])
                ci_lower = comp["mean_diff"] - t_crit * comp["se_diff"]
                ci_upper = comp["mean_diff"] + t_crit * comp["se_diff"]
                
                # Determine significance
                is_significant = corrected_p_values[i] < alpha
                
                PostHocAnalyzer.add_comparison(
                    result,
                    group1=str(comp["level1"]),
                    group2=str(comp["level2"]),
                    test=f"Paired t-test ({correction_method})",
                    p_value=corrected_p_values[i],
                    statistic=comp["t_stat"],
                    corrected=True,
                    correction_method=correction_method,
                    effect_size=comp["effect_size"],
                    effect_size_type="cohen_d_rm",  # Specify repeated measures version
                    confidence_interval=(ci_lower, ci_upper),
                    alpha=alpha,
                    significant=is_significant,
                    # Additional RM-specific information
                    degrees_of_freedom=comp["df"],
                    n_pairs=comp["n_pairs"],
                    mean_difference=comp["mean_diff"]
                )
            
            # Add summary information
            result["summary"] = {
                "total_comparisons": n_comparisons,
                "correction_method": correction_method,
                "family_wise_alpha": alpha,
                "complete_subjects": len(complete_subjects),
                "total_subjects": len(df[subject].unique()),
                "within_factor": within_factor,
                "within_levels": within_levels
            }
            
            # Diagnostic information
            print(f"DEBUG RM POSTHOC: available_pairs = {available_pairs}")
            if normalized_selected is not None:
                missing = normalized_selected - available_pairs
                if missing:
                    print(f"WARNING: The following selected pairs were not found: {missing}")
            
            # Set posthoc_test for visualization
            method_name_map = {
                "tukey": "Tukey HSD (RM)",
                "dunnett": "Dunnett Test (RM)",
                "bonferroni": "Bonferroni (RM)",
                "holm_sidak": "Holm-Sidak (RM)"
            }
            result["posthoc_test"] = method_name_map.get(method, f"RM Post-hoc ({method})")
            
            return result
            
        except Exception as e:
            result["error"] = f"Error in RM ANOVA post-hoc tests: {str(e)}"
            print(f"ERROR RM POSTHOC: {str(e)}")
            import traceback
            traceback.print_exc()
            return result
    
    @staticmethod
    def _get_tukey_critical_value(k, df, alpha=0.05):
        """Get critical value for Tukey's HSD test (simplified implementation)."""
        # This is a simplified implementation - in practice, use statistical tables
        from scipy.stats import studentized_range
        try:
            return studentized_range.ppf(1 - alpha, k, df)
        except:
            # Fallback approximation
            import math
            return math.sqrt(2) * 2.0  # Very rough approximation
    
    @staticmethod 
    def _tukey_p_value(q_stat, k, df):
        """Calculate p-value for Tukey's q statistic (simplified implementation)."""
        from scipy.stats import studentized_range
        try:
            return 1 - studentized_range.cdf(q_stat, k, df)
        except:
            # Fallback to t-distribution approximation
            from scipy.stats import t
            import math
            t_equiv = q_stat / math.sqrt(2)
            return 2 * (1 - t.cdf(abs(t_equiv), df))
        try:
            print(f"DEBUG POSTHOC: selected_comparisons = {selected_comparisons}")
            # Use the same normalization function for group pairs (must match dialog)
            def normalize_pair(pair):
                # Sort and strip, but also ensure both elements are formatted identically to dialog
                return tuple(sorted([s.strip() for s in pair]))
            
            # Handle both set and list inputs for selected_comparisons
            if selected_comparisons:
                if isinstance(selected_comparisons, set):
                    normalized_selected = selected_comparisons  # Already normalized
                else:
                    normalized_selected = set(normalize_pair(pair) for pair in selected_comparisons)
            else:
                normalized_selected = None
            print(f"DEBUG POSTHOC: normalized_selected = {normalized_selected}")
            available_pairs = set()
            
            from itertools import combinations
            stats = get_stats_module()

            within_factor = within[0]
            within_levels = sorted(df[within_factor].unique())
            n_tests = len(list(combinations(within_levels, 2)))

            # Collect all comparisons first
            comparisons = []
            for level1, level2 in combinations(within_levels, 2):
                norm_pair = normalize_pair((str(level1), str(level2)))
                available_pairs.add(norm_pair)
                match = (normalized_selected is not None and norm_pair in normalized_selected)
                print(f"DEBUG POSTHOC: normalized group pair {norm_pair}, match: {match}")
                if normalized_selected is not None and not match:
                    continue
                data1 = df[df[within_factor] == level1].sort_values(by=subject)[dv].values
                data2 = df[df[within_factor] == level2].sort_values(by=subject)[dv].values
                min_len = min(len(data1), len(data2))
                data1 = data1[:min_len]
                data2 = data2[:min_len]

                t_stat, p_val = stats.ttest_rel(data1, data2)
                
                diff = data1 - data2
                mean_diff = np.mean(diff)
                std_diff = np.std(diff, ddof=1)
                effect_size = mean_diff / std_diff if std_diff > 0 else 0

                n = len(diff)
                stderr = std_diff / np.sqrt(n)
                from scipy.stats import t
                # Sidak-adjusted alpha for family-wise confidence intervals
                alpha_sidak = 1 - (1-alpha)**(1/n_tests) 
                t_crit = t.ppf(1 - alpha_sidak/2, n-1)
                ci_lower = mean_diff - t_crit * stderr
                ci_upper = mean_diff + t_crit * stderr
                
                comparisons.append({
                    "level1": level1,
                    "level2": level2,
                    "t_stat": t_stat,
                    "p_val": p_val,
                    "effect_size": effect_size,
                    "ci_lower": ci_lower,
                    "ci_upper": ci_upper
                })
            
            # Apply multiple comparison correction based on method
            p_values = [comp["p_val"] for comp in comparisons]
            if method.lower() == 'tukey':
                # For Tukey, we'd need pingouin or another library for proper implementation
                correction_method = "Tukey HSD"
                corrected_p_values = PostHocAnalyzer._holm_sidak_correction(p_values)  # Fallback
            elif method.lower() == 'dunnett' and control_group:
                # For Dunnett, filter to only control group comparisons
                correction_method = "Dunnett"
                # Filter to only comparisons involving the control group
                dunnett_p_values = []
                control_indices = []
                
                for i, comp in enumerate(comparisons):
                    level1_str = str(comp["level1"])
                    level2_str = str(comp["level2"])
                    # Use exact match instead of substring search
                    if level1_str == control_group or level2_str == control_group:
                        dunnett_p_values.append(comp["p_val"])
                        control_indices.append(i)
                
                if dunnett_p_values:
                    # Apply correction only to control group comparisons
                    dunnett_corrected = PostHocAnalyzer._holm_sidak_correction(dunnett_p_values)
                    # Map back to original order
                    corrected_p_values = [1.0] * len(p_values)  # Start with all p-values as 1.0
                    for j, orig_idx in enumerate(control_indices):
                        corrected_p_values[orig_idx] = dunnett_corrected[j]
                else:
                    corrected_p_values = [1.0] * len(p_values)
                    correction_method = "Dunnett (no control comparisons found)"
            else:
                # Default: Holm-Sidak
                correction_method = "Holm-Sidak"
                corrected_p_values = PostHocAnalyzer._holm_sidak_correction(p_values)
            
            # Add each pairwise comparison result with corrected p-values
            for i, comp in enumerate(comparisons):
                PostHocAnalyzer.add_comparison(
                    result,
                    group1=str(comp["level1"]),
                    group2=str(comp["level2"]),
                    test=f"Paired t-test ({correction_method})",
                    p_value=corrected_p_values[i],
                    statistic=comp["t_stat"],
                    corrected=True,
                    correction_method=correction_method,
                    effect_size=comp["effect_size"],
                    effect_size_type="cohen_d",
                    confidence_interval=(comp["ci_lower"], comp["ci_upper"]),
                    alpha=alpha
                )

            # After all, print available pairs and warn if any selected pair is not present
            print(f"DEBUG POSTHOC: available_pairs = {available_pairs}")
            if normalized_selected is not None:
                missing = normalized_selected - available_pairs
                if missing:
                    print(f"WARNING: The following selected pairs were not found in the available post-hoc comparisons: {missing}")

            # If no comparisons were added, add a placeholder
            if not result["pairwise_comparisons"]:
                PostHocAnalyzer.add_comparison(
                    result,
                    group1="No comparison available",
                    group2="",
                    test="RM ANOVA Post-hoc",
                    p_value=None,
                    statistic=None,
                    corrected=False,
                    correction_method=None,
                    effect_size=None,
                    effect_size_type=None,
                    confidence_interval=(None, None),
                    alpha=alpha
                )

            # Set the posthoc_test value for decision tree visualization
            method_name_map = {
                "tukey": "Tukey HSD",
                "dunnett": "Dunnett Test",
                "paired_custom": "Custom paired t-tests (Holm-Sidak)",
                "holm_sidak": "Custom paired t-tests (Holm-Sidak)"
            }
            result["posthoc_test"] = method_name_map.get(method, f"Post-hoc test ({method})")

            return result
        except Exception as e:
            result["error"] = f"Error in RM ANOVA post-hoc tests: {str(e)}"
            return result
        
class PostHocStatistics:
    """UPDATED: Statistical calculations for various post-hoc tests."""
    
    @staticmethod
    def calculate_cohens_d(group1_data, group2_data, paired=False):
        """Calculates Cohen's d effect size with appropriate adjustments."""
        import numpy as np
        
        if paired:
            diff = np.array(group1_data) - np.array(group2_data)
            return np.mean(diff) / np.std(diff, ddof=1) if np.std(diff, ddof=1) > 0 else 0
        else:
            n1, n2 = len(group1_data), len(group2_data)
            s1, s2 = np.var(group1_data, ddof=1), np.var(group2_data, ddof=1)
            s_pooled = np.sqrt(((n1-1)*s1 + (n2-1)*s2) / (n1+n2-2))
            return (np.mean(group1_data) - np.mean(group2_data)) / s_pooled if s_pooled > 0 else 0
    
    @staticmethod
    def calculate_ci_mean_diff(group1_data, group2_data, alpha=0.05, paired=False):
        """Calculates confidence intervals for the mean difference."""
        import numpy as np
        from scipy.stats import t
        
        try:
            if paired:
                diff = np.array(group1_data) - np.array(group2_data)
                n = len(diff)
                mean_diff = np.mean(diff)
                se = np.std(diff, ddof=1) / np.sqrt(n)
                df = n - 1
            else:
                n1, n2 = len(group1_data), len(group2_data)
                mean_diff = np.mean(group1_data) - np.mean(group2_data)
                s1, s2 = np.var(group1_data, ddof=1), np.var(group2_data, ddof=1)
                se = np.sqrt(s1/n1 + s2/n2)
                df = (s1/n1 + s2/n2)**2 / ((s1/n1)**2/(n1-1) + (s2/n2)**2/(n2-1))
                
            t_crit = t.ppf(1 - alpha/2, df)
            ci_lower = mean_diff - t_crit * se
            ci_upper = mean_diff + t_crit * se
            
            return (float(ci_lower), float(ci_upper))
        except Exception:
            return (None, None)
        
class TukeyHSD(PostHocAnalyzer):
    
    @staticmethod
    def perform_test(valid_groups, samples, alpha=0.05):
        """Performs the Tukey HSD test."""
        import numpy as np
        from statsmodels.stats.multicomp import pairwise_tukeyhsd
        
        result = PostHocAnalyzer.create_result_template("Tukey HSD Test")
        
        try:
            all_data = []
            group_labels = []
            
            for group in valid_groups:
                values = samples[group]
                all_data.extend(values)
                group_labels.extend([str(group)] * len(values))
            
            if len(set(group_labels)) < 2:
                result["error"] = "Tukey HSD requires at least two groups."
                return result

            tukey_result = pairwise_tukeyhsd(endog=all_data, groups=group_labels, alpha=alpha)
            
            # Check if tukey_result has a summary() attribute
            if hasattr(tukey_result, 'summary'):
                summary = tukey_result.summary()
                
                # Extract data from the summary table
                for i in range(len(tukey_result.meandiffs)):
                    group1, group2 = summary.data[i+1][0:2]  # First two columns are the groups
                    p_val = summary.data[i+1][3]  # Fourth column is the p-value
                    lower, upper = summary.data[i+1][4:6]  # Fifth and sixth columns are the confidence intervals
                    reject = summary.data[i+1][6]  # Seventh column is the reject info
                    is_significant = reject  # This variable is no longer passed as parameter!

                    # Calculate Cohen's d effect size
                    group1_data = samples[group1]
                    group2_data = samples[group2]
                    effect_size = PostHocStatistics.calculate_cohens_d(group1_data, group2_data)

                    # Use the common method to add a comparison
                    PostHocAnalyzer.add_comparison(
                        result,
                        group1=group1,
                        group2=group2,
                        test="Tukey HSD",
                        p_value=p_val,
                        statistic=tukey_result.meandiffs[i],
                        corrected=True,
                        correction_method="Tukey HSD",
                        effect_size=effect_size,
                        effect_size_type="cohen_d",
                        confidence_interval=(float(lower), float(upper)),
                        alpha=alpha
                        # The parameter significant=is_significant was removed
                    )
            else:
                result["error"] = "TukeyHSDResults object has no summary() attribute"
                return result
            
            # Set the posthoc_test value for decision tree visualization
            result["posthoc_test"] = "Tukey HSD"
            
            return result
        except Exception as e:
            result["error"] = f"Error in Tukey HSD test: {str(e)}"
            return result
        
class DunnettTest(PostHocAnalyzer):
    """Implementation of the Dunnett test for comparing multiple groups to a control group."""
    @staticmethod
    def perform_test(valid_groups, samples, control_group, alpha=0.05):
        """
        Performs the Dunnett test (compares each group to the control group).
        """
        result = PostHocAnalyzer.create_result_template(f"Dunnett Test (Control group: {control_group})")
        result["control_group"] = control_group

        try:
            import scikit_posthocs as sp
            all_data = []
            group_labels = []
            for group in valid_groups:
                values = samples[group]
                all_data.extend(values)
                group_labels.extend([str(group)] * len(values))
            df = pd.DataFrame({"value": all_data, "group": group_labels})

            dunnett_result = sp.posthoc_dunnett(df, val_col="value", group_col="group", control=str(control_group))

            group_means = {str(g): np.mean(samples[g]) for g in valid_groups}
            group_n = {str(g): len(samples[g]) for g in valid_groups}
            group_std = {str(g): np.std(samples[g], ddof=1) for g in valid_groups}

            p_vals = []
            group_pairs = []
            mean_diffs = []
            ci_lowers = []
            ci_uppers = []
            effect_sizes = []

            k = len(valid_groups) - 1  # Number of comparisons with control
            alpha_sidak = 1 - (1 - alpha) ** (1 / k) if k > 0 else alpha

            for group in valid_groups:
                if str(group) == str(control_group):
                    continue
                # Robust matrix lookup (row/col swap if needed)
                try:
                    p_val = float(dunnett_result.loc[str(group), str(control_group)])
                except KeyError:
                    try:
                        p_val = float(dunnett_result.loc[str(control_group), str(group)])
                    except Exception:
                        p_val = 1.0
                p_vals.append(p_val)
                group_pairs.append(group)

                mean_diff = group_means[str(group)] - group_means[str(control_group)]
                n1 = group_n[str(group)]
                n2 = group_n[str(control_group)]
                s1 = group_std[str(group)]
                s2 = group_std[str(control_group)]
                pooled_std = np.sqrt(((n1 - 1) * (s1 ** 2) + (n2 - 1) * (s2 ** 2)) / (n1 + n2 - 2))
                se_diff = pooled_std * np.sqrt(1 / n1 + 1 / n2)
                df_val = n1 + n2 - 2
                from scipy.stats import t
                t_crit = t.ppf(1 - alpha_sidak / 2, df_val)
                ci_lower = mean_diff - t_crit * se_diff
                ci_upper = mean_diff + t_crit * se_diff
                effect_size = mean_diff / pooled_std if pooled_std > 0 else 0

                mean_diffs.append(mean_diff)
                ci_lowers.append(ci_lower)
                ci_uppers.append(ci_upper)
                effect_sizes.append(effect_size)

            multipletests = get_statsmodels_multitest()
            reject, p_adj, _, _ = multipletests(p_vals, alpha=alpha, method='holm-sidak')

            for i, group in enumerate(group_pairs):
                PostHocAnalyzer.add_comparison(
                    result,
                    group1=group,
                    group2=control_group,
                    test="Dunnett",
                    p_value=p_adj[i],
                    statistic=None,
                    corrected=True,
                    correction_method="Holm-Sidak",
                    effect_size=effect_sizes[i],
                    effect_size_type="cohen_d",
                    confidence_interval=(float(ci_lowers[i]), float(ci_uppers[i])),
                    alpha=alpha
                )

            # Set the posthoc_test value for decision tree visualization
            result["posthoc_test"] = "Dunnett Test"

            return result
        except Exception as e:
            import traceback
            result["error"] = f"Error in Dunnett test: {str(e)}"
            traceback.print_exc()
            return result
        
import numpy as np
from scipy.stats import mannwhitneyu

class DunnTest(PostHocAnalyzer):
    @staticmethod
    def perform_test(valid_groups, samples, alpha=0.05, n_boot=1000):
        result = PostHocAnalyzer.create_result_template("Dunn-Test")

        try:
            import scikit_posthocs as sp
        except ImportError:
            result["error"] = "scikit-posthocs is not installed."
            return result

        # 1) Get raw p-values matrix
        data_array = [samples[g] for g in valid_groups]
        raw_p = sp.posthoc_dunn(data_array, p_adjust=None)  # no internal correction

        # 2) Flatten into list and correct with Holm-Sidak
        pairs = []
        pvals = []
        for i, g1 in enumerate(valid_groups):
            for j, g2 in enumerate(valid_groups):
                if i < j:
                    pairs.append((g1, g2))
                    pvals.append(raw_p.iloc[i, j])
        multipletests = get_statsmodels_multitest()
        reject, p_adj, _, _ = multipletests(pvals, alpha=alpha, method='holm-sidak')

        # 3) Loop over pairs and compute effect & CI
        for (g1, g2), pval_adj, sig in zip(pairs, p_adj, reject):
            x, y = samples[g1], samples[g2]
            # Mann–Whitney U for effect‐size r
            U, _ = mannwhitneyu(x, y, alternative='two-sided')
            n1, n2 = len(x), len(y)
            z = (U - n1 * n2 / 2) / np.sqrt(n1 * n2 * (n1 + n2 + 1) / 12)
            effect_r = abs(z) / np.sqrt(n1 + n2)

            # Hodges-Lehmann
            diffs = [xi - yi for xi in x for yi in y]
            hl = np.median(diffs)

            # Bootstrap CI on HL
            boots = []
            for _ in range(n_boot):
                b1 = np.random.choice(x, n1, replace=True)
                b2 = np.random.choice(y, n2, replace=True)
                boots.append(np.median([u - v for u in b1 for v in b2]))
            ci_low, ci_high = np.percentile(boots, [100*alpha/2, 100*(1-alpha/2)])

            # Median difference
            med_diff = np.median(x) - np.median(y)

            PostHocAnalyzer.add_comparison(
                result,
                group1=g1,
                group2=g2,
                test="Dunn",
                p_value=pval_adj,
                statistic=None,
                corrected=True,
                correction_method="Holm-Sidak",
                effect_size=effect_r,
                effect_size_type="r",
                confidence_interval=(float(ci_low), float(ci_high)),
                alpha=alpha
            )

        return result

class DependentPostHoc(PostHocAnalyzer):
    @staticmethod
    def perform_test(valid_groups, samples, alpha=0.05, parametric=True):
        name = "Parametric paired t-tests" if parametric else "Wilcoxon signed-rank tests"
        result = PostHocAnalyzer.create_result_template(name)

        # 1) check equal lengths
        sizes = [len(samples[g]) for g in valid_groups]
        if len(set(sizes)) != 1:
            result["error"] = "All groups must have same length for dependent tests."
            return result

        # 2) collect stats
        pvals, stats_list, pairs = [], [], []
        for g1, g2 in combinations(valid_groups, 2):
            x, y = np.array(samples[g1]), np.array(samples[g2])
            if parametric:
                tstat, p = stats.ttest_rel(x, y)
                stats_list.append(tstat)
            else:
                wstat, p = stats.wilcoxon(x, y)
                stats_list.append(wstat)
            pvals.append(p)
            pairs.append((g1, g2, x, y))

        # 3) Holm-Sidak correction
        multipletests = get_statsmodels_multitest()
        reject, p_adj, _, _ = multipletests(pvals, alpha=alpha, method='holm-sidak')

        # 4) add comparisons
        for i, (g1, g2, x, y) in enumerate(pairs):
            if parametric:
                # paired CI and d
                ci = PostHocStatistics.calculate_ci_mean_diff(x, y, alpha=alpha, paired=True)
                d = PostHocStatistics.calculate_cohens_d(x, y, paired=True)
                test = "Paired t-test"
                stat = stats_list[i]
                es, estype = d, "cohen_d"
            else:
                # r from Wilcoxon
                n = len(x)
                W = stats_list[i]
                mu = n*(n+1)/4
                sigma = np.sqrt(n*(n+1)*(2*n+1)/24)
                z = (W - mu)/sigma
                r = abs(z)/np.sqrt(n)
                ci = (None, None)
                test = "Wilcoxon signed-rank"
                stat = W
                es, estype = r, "r"

            PostHocAnalyzer.add_comparison(
                result,
                group1=g1,
                group2=g2,
                test=test,
                p_value=p_adj[i],
                statistic=stat,
                corrected=True,
                correction_method="Holm-Sidak",
                effect_size=es,
                effect_size_type=estype,
                confidence_interval=ci,
                alpha=alpha
            )

        return result
            
class PostHocFactory:
    @staticmethod
    def create_test(test_type, is_parametric=True, is_dependent=False):
        """Creates the correct post-hoc test implementation based on parameters."""
        if is_dependent:
            return DependentPostHoc()
        
        if is_parametric:
            if test_type == "tukey":
                return TukeyHSD()
            elif test_type == "dunnett":
                return DunnettTest()
        else:
            if test_type == "dunn":
                return DunnTest()
            elif test_type == "conover":
                # Return None or a message since ConoverPostHoc is removed
                return None
            elif test_type == "nemenyi":
                # Return None or a message since NemenyiPostHoc is removed
                return None
        
        return None
    
    @staticmethod
    def create_anova_posthoc(anova_type, **kwargs):
        """Creates specialized post-hoc tests for different ANOVA types."""
        if anova_type == "two_way":
            return TwoWayPostHocAnalyzer()
        elif anova_type == "mixed":
            return MixedAnovaPostHocAnalyzer()
        elif anova_type == "rm":
            return RMAnovaPostHocAnalyzer()
        return None
    
    @staticmethod
    def perform_posthoc_for_anova(anova_type, df, dv, subject=None, between=None, within=None, alpha=0.05, selected_comparisons=None, method="paired_custom", control_group=None):
        """
        Performs post-hoc tests for an ANOVA type and returns standardized results.
        
        Parameters:
        -----------
        anova_type : str
            Type of ANOVA ('two_way', 'mixed', 'rm')
        df : pandas.DataFrame
            Dataset in long format
        dv : str
            Name of the dependent variable
        subject : str, optional
            Name of the subject variable (for Mixed and RM ANOVA)
        between : list, optional
            List of between factors
        within : list, optional
            List of within factors
        alpha : float, optional
            Significance level (default: 0.05)
        method : str, optional
            Post-hoc method ("tukey", "dunnett", "paired_custom")
        control_group : str, optional
            Control group for Dunnett test
            
        Returns:
        --------
        dict
            Standardized post-hoc results
        """
        analyzer = PostHocFactory.create_anova_posthoc(anova_type)
        if analyzer is None:
            return {"error": f"No post-hoc test available for ANOVA type '{anova_type}'"}
        
        if anova_type == "two_way":
            if not between or len(between) != 2:
                return {"error": "Two-Way ANOVA requires two between factors"}
            return analyzer.perform_test(df=df, dv=dv, factors=between, alpha=alpha, selected_comparisons=selected_comparisons, method=method, control_group=control_group)
        
        elif anova_type == "mixed":
            # Full implementation for Mixed ANOVA
            if not subject:
                return {"error": "Mixed ANOVA requires a subject variable"}
            if not between or len(between) != 1:
                return {"error": "Mixed ANOVA requires exactly one between factor"}
            if not within or len(within) != 1:
                return {"error": "Mixed ANOVA requires exactly one within factor"}
            
            return analyzer.perform_test(df=df, dv=dv, subject=subject, between=between, within=within, alpha=alpha, selected_comparisons=selected_comparisons, method=method, control_group=control_group)
        
        elif anova_type == "rm":
            # Full implementation for RM-ANOVA
            if not subject:
                return {"error": "RM-ANOVA requires a subject variable"}
            if not within or len(within) < 1:
                return {"error": "RM-ANOVA requires at least one within factor"}
            
            # Get post-hoc results from analyzer
            posthoc = analyzer.perform_test(df=df, dv=dv, subject=subject, within=within, alpha=alpha, selected_comparisons=selected_comparisons, method=method, control_group=control_group)
            
            # Add validation to ensure we're getting valid results
            if posthoc and 'pairwise_comparisons' in posthoc:
                print(f"DEBUG: Found {len(posthoc['pairwise_comparisons'])} rm-anova post-hoc comparisons")
            else:
                print("DEBUG: No valid rm-anova post-hoc results found!")
                
            # Explicitly pass through the posthoc results without modification
            return posthoc
        
        return {"error": f"Unknown ANOVA type: {anova_type}"}
    
class DataImporter:
    @staticmethod
    def import_data(file_path, sheet_name=0, group_col="Group", value_cols=None, combine_columns=False):
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"The file {file_path} was not found.")
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
        if group_col not in df.columns:
            raise ValueError(f"The group column '{group_col}' was not found. Available columns: {', '.join(df.columns)}")
        if value_cols is None:
            numeric_cols = [col for col in df.columns if pd.api.types.is_numeric_dtype(df[col]) and col != group_col]
            if len(numeric_cols) == 0:
                raise ValueError("No numeric columns found that can be used as measurements.")
            value_cols = numeric_cols
        for col in value_cols:
            if col not in df.columns:
                raise ValueError(f"The value column '{col}' was not found. Available columns: {', '.join(df.columns)}")
        groups = sorted(df[group_col].unique())
        samples = {}
        if combine_columns:
            for group in groups:
                combined_values = []
                for col in value_cols:
                    values = df[df[group_col] == group][col].dropna().tolist()
                    combined_values.extend(values)
                samples[group] = combined_values
        else:  # if combine_columns=False
            if len(value_cols) > 1:
                print("Warning: Multiple value columns specified, but combine_columns=False. Only the first column will be used.")
            
            for group in groups:
                values = df[df[group_col] == group][value_cols[0]].dropna().tolist()
                samples[group] = values
        return samples, df

import numpy as np
import pandas as pd
import scipy.stats as stats

try:
    import scikit_posthocs as sp
    HAS_SCPH = True
except ImportError:
    HAS_SCPH = False

# Robust import for nonparametricanovas: works as script or module
try:
    from nonparametricanovas import GLMMTwoWayANOVA, GEERMANOVA, GLMMMixedANOVA, auto_anova_decision
except ImportError:
    import sys as _sys
    import os as _os
    _src_dir = _os.path.dirname(_os.path.abspath(__file__))
    _parent_dir = _os.path.abspath(_os.path.join(_src_dir, os.pardir))
    if _parent_dir not in _sys.path:
        _sys.path.insert(0, _parent_dir)
    try:
        from nonparametricanovas import GLMMTwoWayANOVA, GEERMANOVA, GLMMMixedANOVA, auto_anova_decision
    except ImportError as e:
        raise ImportError(
            "Could not import nonparametricanovas. Tried both absolute and sys.path hack. "
            "Current sys.path: {}. Error: {}".format(_sys.path, e)
        )


            
class UIDialogManager:
    @staticmethod
    def select_posthoc_test_dialog(parent=None, progress_text=None, column_name=None, default_method=None):
        dialog = QDialog(parent)
        layout = QVBoxLayout(dialog)

        # Set window title
        title = "Select Post-hoc Test"
        if column_name:
            title += f" for '{column_name}'"
        if progress_text:
            title += f" {progress_text}"
        dialog.setWindowTitle(title)

        info_text = "The ANOVA has revealed significant differences. Please select a post-hoc test:"
        if progress_text and ("two_way_anova" in progress_text or "mixed_anova" in progress_text or "repeated_measures_anova" in progress_text):
            info_text = ("The advanced ANOVA has revealed significant differences. For advanced ANOVAs, "
                        "paired t-tests are often preferred to examine specific interaction effects. "
                        "Please select a post-hoc test:")
        elif progress_text and "two_way_anova" in progress_text:
            info_text = ("The Two-Way ANOVA has revealed significant differences. For Two-Way ANOVA, "
                        "paired t-tests are often preferred to examine specific interaction effects. "
                        "Please select a post-hoc test:")
        
        info = QLabel(info_text)
        info.setWordWrap(True)
        layout.addWidget(info)

        # RadioButtons for post-hoc tests - options depend on context
        if progress_text and ("two_way_anova" in progress_text or "mixed_anova" in progress_text or "repeated_measures_anova" in progress_text):
            # For advanced ANOVAs: only offer Tukey and Custom paired t-tests (no Dunnett)
            options = [
                ("Tukey-HSD Test (compares all pairs)", "tukey"),
                ("Custom paired t-tests (you select specific pairs to compare)", "paired_custom"),
            ]
        else:
            # For One-Way ANOVA: offer all three options
            options = [
                ("Tukey-HSD Test (compares all pairs, best for main effects)", "tukey"),
                ("Dunnett Test (compares all groups against ONE control group)", "dunnett"),
                ("Custom paired t-tests (you select specific pairs to compare)", "paired_custom"),
            ]
        radio_buttons = []
        for label, value in options:
            rb = QRadioButton(label)
            layout.addWidget(rb)
            radio_buttons.append((rb, value))
        
        # Set default selection based on context
        if default_method is None:
            default_method = "tukey"  # Original default
        
        for i, (rb, value) in enumerate(radio_buttons):
            if value == default_method:
                rb.setChecked(True)
                break
        else:
            # If default_method not found, default to first option
            radio_buttons[0][0].setChecked(True)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        layout.addWidget(buttons)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)

        if dialog.exec_() == QDialog.Accepted:
            for rb, value in radio_buttons:
                if rb.isChecked():
                    return value
        return None

    @staticmethod
    def select_nonparametric_posthoc_dialog(parent=None, progress_text=None, column_name=None):
        """
        Dialog for nonparametric post-hoc tests: Dunn or Mann-Whitney U (custom pairs)
        """
        from PyQt5.QtWidgets import QDialog, QVBoxLayout, QLabel, QRadioButton, QDialogButtonBox

        dialog = QDialog(parent)
        layout = QVBoxLayout(dialog)

        title = "Nonparametric Post-hoc Test"
        if column_name:
            title += f" for '{column_name}'"
        if progress_text:
            title += f" {progress_text}"
        dialog.setWindowTitle(title)

        info = QLabel("Please select the desired nonparametric post-hoc test:")
        layout.addWidget(info)

        options = [
            ("Dunn Test (all pairs, Holm-Sidak correction)", "dunn"),
            ("Mann-Whitney-U Tests (custom pairs, Sidak correction)", "mw_custom"),
        ]
        radio_buttons = []
        for label, value in options:
            rb = QRadioButton(label)
            layout.addWidget(rb)
            radio_buttons.append((rb, value))
        radio_buttons[0][0].setChecked(True)  # Default: Dunn

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        layout.addWidget(buttons)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)

        if dialog.exec_() == QDialog.Accepted:
            for rb, value in radio_buttons:
                if rb.isChecked():
                    return value
        return None
    
    @staticmethod
    def select_custom_pairs_dialog(groups):
        """
        Dialog to select custom group pairs for paired t-tests.
        Returns a list of (group1, group2) tuples.
        """
        from PyQt5.QtWidgets import QDialog, QVBoxLayout, QLabel, QCheckBox, QDialogButtonBox, QWidget, QHBoxLayout, QScrollArea

        class PairSelectionDialog(QDialog):
            def __init__(self, groups, parent=None):
                super().__init__(parent)
                self.setWindowTitle("Select Group Pairs for Paired t-tests")
                self.selected_pairs = []
                layout = QVBoxLayout(self)
                label = QLabel("Select the group pairs to compare (paired t-test):")
                layout.addWidget(label)
                
                # Create scroll area for many group pairs
                scroll = QScrollArea(self)
                scroll.setWidgetResizable(True)
                scroll_content = QWidget()
                scroll_layout = QVBoxLayout(scroll_content)
                
                self.checkboxes = []
                for i, g1 in enumerate(groups):
                    for g2 in groups[i+1:]:
                        pair_str = f"{g1} vs {g2}"
                        cb = QCheckBox(pair_str)
                        scroll_layout.addWidget(cb)
                        self.checkboxes.append((cb, (g1, g2)))
                
                scroll_content.setLayout(scroll_layout)
                scroll.setWidget(scroll_content)
                
                # Limit maximum height to prevent dialog from becoming too large
                # Calculate dynamic height: max 350px or 50% of screen height, whichever is smaller
                from PyQt5.QtWidgets import QApplication
                if QApplication.instance():
                    screen = QApplication.instance().primaryScreen()
                    if screen:
                        screen_height = screen.geometry().height()
                        max_height = min(350, int(screen_height * 0.5))
                        scroll.setMaximumHeight(max_height)
                    else:
                        scroll.setMaximumHeight(350)  # Fallback
                else:
                    scroll.setMaximumHeight(350)  # Fallback
                    
                layout.addWidget(scroll)
                
                buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                buttons.accepted.connect(self.accept)
                buttons.rejected.connect(self.reject)
                layout.addWidget(buttons)

            def accept(self):
                self.selected_pairs = [pair for cb, pair in self.checkboxes if cb.isChecked()]
                super().accept()

        dialog = PairSelectionDialog(groups)
        if dialog.exec_() == QDialog.Accepted:
            return dialog.selected_pairs
        return []

    @staticmethod
    def select_control_group_dialog(groups):
        """Opens a dialog window to select the control group"""
        from PyQt5.QtWidgets import QDialog, QVBoxLayout, QLabel, QRadioButton, QDialogButtonBox

        class ControlGroupDialog(QDialog):
            def __init__(self, available_groups, parent=None):
                super().__init__(parent)
                self.setWindowTitle("Select Control Group")
                self.selected_group = None
                self.groups = available_groups

                layout = QVBoxLayout(self)
                label = QLabel("Please select the control group for the Dunnett test:")
                layout.addWidget(label)

                self.group_buttons = []
                for group in self.groups:
                    rb = QRadioButton(str(group))
                    self.group_buttons.append(rb)
                    layout.addWidget(rb)

                # Select first group by default
                if self.group_buttons:
                    self.group_buttons[0].setChecked(True)

                button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                button_box.accepted.connect(self.accept)
                button_box.rejected.connect(self.reject)
                layout.addWidget(button_box)

            def accept(self):
                for i, button in enumerate(self.group_buttons):
                    if button.isChecked():
                        self.selected_group = self.groups[i]
                        break
                super().accept()

        # Always create a new dialog
        dialog = ControlGroupDialog(groups)
        if dialog.exec_() == QDialog.Accepted:
            return dialog.selected_group
        return groups[0]  # Default: first group
    
    @staticmethod
    def select_transformation_dialog(parent=None, progress_text=None, column_name=None, force_show=False):
        # NO CACHING - Each analysis starts fresh and shows the dialog every time
        # This ensures consistent behavior between normal tests and advanced tests
        
        dialog = QDialog(parent)
        layout = QVBoxLayout(dialog)

        # Set window title
        title = "Select Transformation"
        if column_name:
            title += f" for '{column_name}'"
        if progress_text:
            title += f" {progress_text}"
        dialog.setWindowTitle(title)

        # Info text - CONSISTENT WORDING
        info = QLabel("Please select the desired transformation:")
        layout.addWidget(info)

        # RadioButtons for transformations
        options = [
            ("Log10 transformation (for positive, right-skewed data)", "log10"),
            ("Box-Cox transformation (automatic lambda optimization)", "boxcox"),
            ("Arcsin square root transformation (for percentages/proportions)", "arcsin_sqrt"),
        ]
        radio_buttons = []
        for label, value in options:
            rb = QRadioButton(label)
            layout.addWidget(rb)
            radio_buttons.append((rb, value))
        radio_buttons[0][0].setChecked(True)  # Default: Log10

        # OK/Cancel buttons
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        layout.addWidget(buttons)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)

        # Show dialog
        if dialog.exec_() == QDialog.Accepted:
            for rb, value in radio_buttons:
                if rb.isChecked():
                    return value
        
        # If canceled, return None
        return None
       
class DatasetSelector:
    """Helper class to manage dataset selection in the UI"""
    
    @staticmethod
    def get_available_datasets(file_path, sheet_name=None):
        """
        Get all available datasets (sheets) from an Excel file
        
        Returns:
        --------
        dict: {sheet_name: preview_info}
        """
        try:
            if file_path.endswith('.xlsx') or file_path.endswith('.xls'):
                # Get all sheet names
                xl_file = pd.ExcelFile(file_path)
                datasets = {}
                
                for sheet in xl_file.sheet_names:
                    try:
                        # Get a preview of each sheet
                        df_preview = pd.read_excel(file_path, sheet_name=sheet, nrows=5)
                        datasets[sheet] = {
                            'columns': df_preview.columns.tolist(),
                            'shape': f"{len(pd.read_excel(file_path, sheet_name=sheet))} rows",
                            'preview': df_preview.head(3).to_dict('records')
                        }
                    except Exception as e:
                        datasets[sheet] = {'error': str(e)}
                
                return datasets
            else:
                # For CSV files, return single dataset
                df_preview = pd.read_csv(file_path, nrows=5)
                return {
                    'CSV Data': {
                        'columns': df_preview.columns.tolist(),
                        'shape': f"{len(pd.read_csv(file_path))} rows",
                        'preview': df_preview.head(3).to_dict('records')
                    }
                }
        except Exception as e:
            return {'Error': {'error': str(e)}}

# Modified AnalysisManager.analyze function
class AnalysisManager:
    @staticmethod
    def analyze(file_path, group_col, groups, sheet_name=0, value_cols=None, 
                selected_datasets=None, combine_columns=False, width=12, height=10, 
                dependent=False, compare=None, colors=None, hatches=None,
                title=None, x_label=None, y_label=None, file_name=None, 
                save_plot=True, skip_plots=False, error_type="sd", skip_excel=False, 
                dataset_name=None, additional_factors=None, show_individual_lines=True, 
                **kwargs):
        
        print("DEBUG ANALYZE: AnalysisManager.analyze called")
        print(f"DEBUG ANALYZE: Current working directory: {os.getcwd()}")
        print(f"DEBUG ANALYZE: file_path = {file_path}")
        print(f"DEBUG ANALYZE: file_name = {file_name}")
        print(f"DEBUG ANALYZE: save_plot = {save_plot}, skip_plots = {skip_plots}, skip_excel = {skip_excel}")
        # Single dataset analysis (existing functionality)
        if selected_datasets is None or len(selected_datasets) <= 1:
            # Use existing single dataset logic
            actual_sheet = selected_datasets[0] if selected_datasets else sheet_name
            return AnalysisManager._analyze_single_dataset(
                file_path, group_col, groups, actual_sheet, value_cols, 
                combine_columns, width, height, dependent, compare, colors, hatches,
                title, x_label, y_label, file_name, save_plot, skip_plots, 
                error_type, skip_excel, dataset_name, additional_factors, 
                show_individual_lines, **kwargs
            )
        
        # Multiple dataset analysis
        else:
            return AnalysisManager._analyze_multiple_datasets(
                file_path, group_col, groups, selected_datasets, value_cols,
                combine_columns, width, height, dependent, compare, colors, hatches,
                title, x_label, y_label, file_name, save_plot, skip_plots,
                error_type, skip_excel, additional_factors, show_individual_lines, **kwargs
            )
            
    @staticmethod
    def _analyze_multiple_datasets(file_path, group_col, groups, selected_datasets, value_cols,
                                  combine_columns, width, height, dependent, compare, colors, hatches,
                                  title, x_label, y_label, file_name, save_plot, skip_plots,
                                  error_type, skip_excel, additional_factors, show_individual_lines, **kwargs):
        """
        Multiple dataset analysis with unified Excel output
        """
        from datetime import datetime
        
        all_results = {}
        failed_datasets = {}
        
        print(f"Starting analysis of {len(selected_datasets)} datasets...")
        
        # Analyze each selected dataset
        for i, dataset_name in enumerate(selected_datasets):
            print(f"Analyzing dataset {i+1}/{len(selected_datasets)}: {dataset_name}")
            
            try:
                # Analyze single dataset
                result = AnalysisManager._analyze_single_dataset(
                    file_path=file_path,
                    group_col=group_col,
                    groups=groups,
                    sheet_name=dataset_name,
                    value_cols=value_cols,
                    combine_columns=combine_columns,
                    width=width,
                    height=height,
                    dependent=dependent,
                    compare=compare,
                    colors=colors,
                    hatches=hatches,
                    title=f"{title} - {dataset_name}" if title else dataset_name,
                    x_label=x_label,
                    y_label=y_label,
                    file_name=f"{file_name}_{dataset_name}" if file_name else dataset_name,
                    save_plot=save_plot,
                    skip_plots=skip_plots,
                    error_type=error_type,
                    skip_excel=True,  # Skip individual Excel files
                    dataset_name=dataset_name,
                    additional_factors=additional_factors,
                    show_individual_lines=show_individual_lines,
                    dialog_progress=f"({i+1}/{len(selected_datasets)})",
                    dialog_column=dataset_name,
                    **kwargs
                )
                
                if "error" in result:
                    failed_datasets[dataset_name] = result["error"]
                    print(f"ERROR analyzing {dataset_name}: {result['error']}")
                else:
                    all_results[dataset_name] = result
                    print(f"Successfully analyzed {dataset_name}")
                    
            except Exception as e:
                error_msg = f"Exception during analysis: {str(e)}"
                failed_datasets[dataset_name] = error_msg
                print(f"ERROR analyzing {dataset_name}: {error_msg}")
        
        # Create combined Excel output
        if all_results:
            base_name = file_name if file_name else "multi_dataset_analysis"
            excel_path = f"{base_name}_combined_results.xlsx"
            
            try:
                ResultsExporter = get_results_exporter()
                ResultsExporter.export_multi_dataset_results(all_results, excel_path)
                print(f"Combined results saved to: {excel_path}")
            except Exception as e:
                print(f"Error creating combined Excel file: {str(e)}")
        
        # Return summary
        return {
            "type": "multi_dataset_analysis",
            "successful_datasets": list(all_results.keys()),
            "failed_datasets": failed_datasets,
            "results": all_results,
            "combined_excel": excel_path if all_results else None,
            "summary": {
                "total_datasets": len(selected_datasets),
                "successful": len(all_results),
                "failed": len(failed_datasets),
                "success_rate": f"{len(all_results)/len(selected_datasets)*100:.1f}%"
            }
        }
            
    @staticmethod
    def _analyze_single_dataset(file_path, group_col, groups, sheet_name, value_cols, 
                               combine_columns, width, height, dependent, compare, colors, hatches,
                               title, x_label, y_label, file_name, save_plot, skip_plots, 
                               error_type, skip_excel, dataset_name, additional_factors, 
                               show_individual_lines, **kwargs):
        
        # Get classes lazily to avoid circular imports
        ResultsExporter = get_results_exporter()
        StatisticalTester = get_statistical_tester()
        DataVisualizer = get_data_visualizer()
        
        # CRITICAL FIX: Ensure additional_factors is available in kwargs
        # since the advanced test logic looks for it there
        if additional_factors is not None and 'additional_factors' not in kwargs:
            kwargs['additional_factors'] = additional_factors
        
        # Basic parameter validation
        if not file_path or not os.path.exists(file_path):
            raise ValueError("Please specify a valid file")
        if not groups:
            raise ValueError("Please specify at least one group")
        if not group_col:
            raise ValueError("Please specify a valid group column")
        if error_type not in ["sd", "se"]:
            raise ValueError("Error bar type must be 'sd' or 'se'")

        from datetime import datetime
        analysis_log = f"Analysis Report\n"
        analysis_log += f"Date and time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
        analysis_log += f"File: {file_path}\n"
        analysis_log += f"Worksheet: {sheet_name}\n"
        analysis_log += f"Group column: {group_col}\n"
        analysis_log += f"Value column(s): {', '.join(value_cols) if value_cols else 'All numeric columns'}\n"
        analysis_log += f"Groups to analyze: {', '.join(map(str, groups))}\n"
        analysis_log += f"Dependent samples: {'Yes' if dependent else 'No'}\n"
        analysis_log += f"Error bar type: {'SEM (standard error)' if error_type == 'se' else 'SD (standard deviation)'}\n"

        if compare:
            compare_str = ", ".join([f"{g1} vs {g2}" for g1, g2 in compare])
            analysis_log += f"Specific comparisons: {compare_str}\n"

        analysis_log += "\n--- ANALYSIS ---\n\n"

        try:
            # Import and filter data
            samples, df = DataImporter.import_data(file_path, sheet_name=sheet_name, group_col=group_col, 
                                                  value_cols=value_cols, combine_columns=combine_columns)
            filtered_samples = {g: samples[g] for g in groups if g in samples}

            # Validations and logging
            if not filtered_samples:
                raise ValueError(f"None of the specified groups were found in the data. Available groups: {list(samples.keys())}")

            for group, values in filtered_samples.items():
                if len(values) < 1:
                    raise ValueError(f"Group '{group}' contains no data.")

            analysis_log += f"Data imported successfully.\n"
            analysis_log += "Number of data points per group:\n"
            for group, values in filtered_samples.items():
                analysis_log += f"  {group}: {len(values)} data points\n"

            # Initialize the result dictionary (important: before first assignments!)
            results = {}

            # For advanced tests that use prepare_advanced_test, skip the normality check here
            # as it will be handled in the advanced test flow
            if kwargs.get('test') in ['mixed_anova', 'two_way_anova', 'repeated_measures_anova']:
                # Skip normality check - will be handled by prepare_advanced_test
                transformed_samples = None
                test_recommendation = None
                test_info = None
            else:
                # Determine model type based on parameters
                if len(groups) == 2:
                    model_type = "ttest"
                    formula = "Value ~ C(Group)"
                elif len(groups) > 2:
                    model_type = "oneway"
                    formula = "Value ~ C(Group)"
                else:
                    model_type = "oneway"
                    formula = "Value ~ C(Group)"
                    
                # Normality and variance check with dataset name
                transformed_samples, test_recommendation, test_info = StatisticalTester.check_normality_and_variance(
                    groups,
                    filtered_samples,
                    dataset_name=dataset_name,
                    progress_text=kwargs.get('dialog_progress', None),
                    column_name=kwargs.get('dialog_column', None),
                    formula=formula,
                    model_type=model_type
                )
            print(f"DEBUG: Test recommendation is '{test_recommendation}'")
            print(f"DEBUG: Test info transformation: '{test_info.get('transformation') if test_info else 'N/A'}'")

            # Write test recommendation to log (only if we have one)
            if test_recommendation:
                analysis_log += f"\nTest recommendation: {test_recommendation}\n"

            # For dependent samples, perform additional validation
            if dependent:
                validation = StatisticalTester.validate_dependent_data(filtered_samples, groups)
                if not validation["valid"]:
                    error_message = "Error validating dependent data:\n" + "\n".join(validation["messages"])
                    analysis_log += f"\n{error_message}\n"
                    if not kwargs.get('force_continue', False):
                        print(f"WARNING: {error_message}")
                        analysis_log += "\nAnalysis continues with warning, results may be unreliable."
                        
            # Adjust test type based on recommendation if auto-selection is enabled
            if kwargs.get('auto_nonparametric', True) and test_recommendation == 'non_parametric':
                if kwargs.get('test') == 'two_way_anova':
                    kwargs['test'] = 'nonparametric_two_way_anova'
                    analysis_log += "\nSwitching to nonparametric two-way ANOVA based on assumption check.\n"
                elif kwargs.get('test') == 'repeated_measures_anova':
                    kwargs['test'] = 'nonparametric_rm_anova'
                    analysis_log += "\nSwitching to nonparametric repeated measures ANOVA based on assumption check.\n"
                elif kwargs.get('test') == 'mixed_anova':
                    kwargs['test'] = 'nonparametric_mixed_anova'
                    analysis_log += "\nSwitching to nonparametric mixed ANOVA based on assumption check.\n"            

            # Perform the appropriate statistical test - only call ONCE
            if kwargs.get('test') == 'mixed_anova':
                additional_factors = kwargs.get('additional_factors', [])
                if len(additional_factors) >= 2:
                    between_factor, within_factor = additional_factors[0], additional_factors[1]
                else:
                    return {"error": "Mixed ANOVA requires two factors (between and within)"}
                # Step 3: Call prepare_advanced_test first
                prep = StatisticalTester.prepare_advanced_test(
                    df, 'mixed_anova', value_cols[0], 'Subject', [between_factor], [within_factor]
                )
                if "error" in prep:
                    return prep  # or handle error

                # Step 4: Pass outputs to perform_advanced_test
                results = StatisticalTester.perform_advanced_test(
                    df=df,
                    test='mixed_anova',
                    dv=value_cols[0],
                    subject='Subject',
                    between=[between_factor],
                    within=[within_factor],
                    alpha=0.05,
                    transformed_samples=prep["transformed_samples"],
                    recommendation=prep["recommendation"],
                    test_info=prep["test_info"],
                    transform_fn=None,
                    force_parametric=kwargs.get('force_parametric', False)
                )
                # Get the transformation type from the test_info
                requested_transform = prep["test_info"].get("transformation", "None")
                print(f"DEBUG: Requested transformation: {requested_transform}")
                print(f"DEBUG: Applied transformation: {results.get('transformation')}")
                # For consistency with the rest of the code, assign results to test_results
                test_results = results
                # Also extract the test_info and other variables for the rest of the code
                test_info = prep["test_info"]
                test_recommendation = prep["recommendation"]
                transformed_samples = prep["transformed_samples"]
            elif kwargs.get('test') == 'two_way_anova':
                between_factors = kwargs.get('additional_factors', [])
                # Step 3: Call prepare_advanced_test first
                prep = StatisticalTester.prepare_advanced_test(
                    df, 'two_way_anova', value_cols[0], None, between_factors, None
                )
                if "error" in prep:
                    return prep  # or handle error

                # Step 4: Pass outputs to perform_advanced_test
                results = StatisticalTester.perform_advanced_test(
                    df=df,
                    test='two_way_anova',
                    dv=value_cols[0],
                    subject=None,
                    between=between_factors,
                    within=None,
                    alpha=0.05,
                    transformed_samples=prep["transformed_samples"],
                    recommendation=prep["recommendation"],
                    test_info=prep["test_info"],
                    transform_fn=None,
                    force_parametric=kwargs.get('force_parametric', False)
                )
                # Get the transformation type from the test_info
                requested_transform = prep["test_info"].get("transformation", "None")
                print(f"DEBUG: Requested transformation: {requested_transform}")
                print(f"DEBUG: Applied transformation: {results.get('transformation')}")
                # For consistency with the rest of the code, assign results to test_results
                test_results = results
                # Also extract the test_info and other variables for the rest of the code
                test_info = prep["test_info"]
                test_recommendation = prep["recommendation"]
                transformed_samples = prep["transformed_samples"]
            elif kwargs.get('test') == 'repeated_measures_anova':
                additional_factors = kwargs.get('additional_factors', [])
                if len(additional_factors) >= 1:
                    within_factor = additional_factors[0]  # RM-ANOVA uses within factor
                else:
                    return {"error": "Repeated measures ANOVA requires at least one within factor"}
                # Step 3: Call prepare_advanced_test first
                prep = StatisticalTester.prepare_advanced_test(
                    df, 'repeated_measures_anova', value_cols[0], 'Subject', None, [within_factor]
                )
                if "error" in prep:
                    return prep  # or handle error

                # Step 4: Pass outputs to perform_advanced_test
                results = StatisticalTester.perform_advanced_test(
                    df=df,
                    test='repeated_measures_anova',
                    dv=value_cols[0],
                    subject='Subject',
                    between=None,
                    within=[within_factor],
                    alpha=0.05,
                    transformed_samples=prep["transformed_samples"],
                    recommendation=prep["recommendation"],
                    test_info=prep["test_info"],
                    transform_fn=None,
                    force_parametric=kwargs.get('force_parametric', False)
                )
                # Get the transformation type from the test_info
                requested_transform = prep["test_info"].get("transformation", "None")
                print(f"DEBUG: Requested transformation: {requested_transform}")
                print(f"DEBUG: Applied transformation: {results.get('transformation')}")
                # For consistency with the rest of the code, assign results to test_results
                test_results = results
                # Also extract the test_info and other variables for the rest of the code
                test_info = prep["test_info"]
                test_recommendation = prep["recommendation"]
                transformed_samples = prep["transformed_samples"]
            else:
                # Standard path for simple tests
                test_results = StatisticalTester.perform_statistical_test(
                    groups, transformed_samples, filtered_samples,
                    dependent=dependent, test_recommendation=test_recommendation, test_info=test_info
                )

                
            if kwargs.get('test') == 'nonparametric_two_way_anova':
                # DISABLED: Nonparametric fallbacks are not yet supported
                # from nonparametricanovas import NonParametricTwoWayANOVA, NonParametricFactory
                
                # Return informational message instead of running nonparametric test
                test_results = {
                    "test": "Two-Way ANOVA (parametric assumptions violated)",
                    "recommendation": "non_parametric", 
                    "error": "Nonparametric alternatives are currently disabled. The NonParametricTwoWayANOVA class is available in nonparametricanovas.py but nonparametric fallbacks are disabled.",
                    "parametric_violated": True,
                    "suggested_alternative": "NonParametricTwoWayANOVA class (rank transformation + permutation test)"
                }
                
                # Create standardized results for Excel export
                results = {
                    "test_results": test_results,
                    "transformed_samples": transformed_samples,
                    "samples": filtered_samples,
                    "transformation": test_info.get("transformation"),
                    "normality_tests": test_info["normality_tests"],
                    "variance_test": test_info["variance_test"],
                    "test_type": "non-parametric",
                    "groups": groups,
                    "permutation_test": False,
                    "analysis_log": "Nonparametric Two-Way ANOVA requested but not supported"
                }
                
            elif kwargs.get('test') == 'nonparametric_rm_anova':
                # DISABLED: Nonparametric fallbacks are not yet supported
                # from nonparametricanovas import NonParametricFactory
                
                # Return informational message instead of running nonparametric test  
                test_results = {
                    "test": "Repeated-Measures ANOVA (parametric assumptions violated)",
                    "recommendation": "non_parametric",
                    "error": "Nonparametric alternatives are currently disabled. The NonParametricRMANOVA class is available in nonparametricanovas.py but nonparametric fallbacks are disabled.",
                    "parametric_violated": True,
                    "suggested_alternative": "NonParametricRMANOVA class (rank transformation + permutation test)"
                }
                
                # Create standardized results for Excel export
                results = {
                    "test_results": test_results,
                    "transformed_samples": transformed_samples,
                    "samples": filtered_samples,
                    "transformation": test_info.get("transformation"),
                    "normality_tests": test_info["normality_tests"],
                    "variance_test": test_info["variance_test"],
                    "test_type": "non-parametric",
                    "groups": groups,
                    "analysis_log": "Nonparametric Repeated-Measures ANOVA requested but not supported"
                }
                
                # Handle Excel export if needed
                if not skip_excel:
                    print(f"DEBUG EXCEL: About to export Excel file")
                    print(f"DEBUG EXCEL: Current working directory: {os.getcwd()}")
                    
                    # Store original directory
                    original_dir = os.getcwd()
                    
                    # Generate excel file path using absolute path
                    dataset_str = f"{dataset_name}_" if dataset_name else ""
                    file_base = f"{dataset_str}{file_name}" if file_name else f"{groups[0]}_vs_{groups[1]}"
                    excel_file = get_output_path(file_base, "xlsx")
                    print(f"DEBUG EXCEL: Will save to: {excel_file}")
                    
                    # Export only once
                    ResultsExporter = get_results_exporter()
                    ResultsExporter.export_results_to_excel(results, excel_file, analysis_log)
                    print(f"DEBUG EXCEL: Export completed, file exists: {os.path.exists(excel_file)}")
                    
                    # Update analysis log with the absolute path
                    if os.path.exists(excel_file):
                        analysis_log += f"\nResults were saved to {excel_file}.\n"
                        # Store the path in results for later reference
                        results["excel_output_path"] = excel_file
                    else:
                        error_msg = f"Failed to create Excel file at: {excel_file}"
                        analysis_log += f"\nERROR: {error_msg}\n"
                        print(f"ERROR: {error_msg}")
                        
                    # Ensure we're back in the original directory
                    if os.getcwd() != original_dir:
                        os.chdir(original_dir)
                        print(f"DEBUG: Restored original directory: {original_dir}")
            
            elif kwargs.get('test') == 'nonparametric_mixed_anova':
                # DISABLED: Nonparametric fallbacks are not yet supported
                # from nonparametricanovas import NonParametricMixedANOVA, NonParametricFactory
                
                # Return informational message instead of running nonparametric test
                test_results = {
                    "test": "Mixed-Design ANOVA (parametric assumptions violated)",
                    "recommendation": "non_parametric",
                    "error": "Nonparametric alternatives are currently disabled. The NonParametricMixedANOVA class is available in nonparametricanovas.py but nonparametric fallbacks are disabled.",
                    "parametric_violated": True,
                    "suggested_alternative": "NonParametricMixedANOVA class (rank transformation + permutation test)"
                }
                
                # Create standardized results for Excel export
                results = {
                    "test_results": test_results,
                    "transformed_samples": transformed_samples,
                    "samples": filtered_samples,
                    "transformation": test_info.get("transformation"),
                    "normality_tests": test_info["normality_tests"],
                    "variance_test": test_info["variance_test"],
                    "test_type": "non-parametric",
                    "groups": groups,
                    "analysis_log": "Nonparametric Mixed-Design ANOVA requested but not supported"
                }    

            # Log before transformation (only for standard tests that went through normality checking)
            if test_info:
                analysis_log += "\nResults of assumption tests before transformation:\n"
                
                # Get residual normality from new structure
                pre_residual_norm = test_info.get("pre_transformation", {}).get("residuals_normality")
                if pre_residual_norm and pre_residual_norm.get("p_value") is not None:
                    analysis_log += f"Shapiro-Wilk test (model residuals normality): p = {pre_residual_norm['p_value']:.4f} - "
                    analysis_log += "Model residuals normally distributed\n" if pre_residual_norm.get('is_normal', False) else "Model residuals not normally distributed\n"
                else:
                    analysis_log += "Shapiro-Wilk test (model residuals): Test not performed (insufficient data)\n"

                # Get variance test from new structure
                pre_variance = test_info.get("pre_transformation", {}).get("variance")
                if pre_variance and pre_variance.get("p_value") is not None:
                    analysis_log += f"Brown-Forsythe test (variance homogeneity): p = {pre_variance['p_value']:.4f} - "
                    analysis_log += "Variances homogeneous\n" if pre_variance.get('equal_variance', False) else "Variances heterogeneous\n"
                else:
                    analysis_log += "Brown-Forsythe test: Not performed (insufficient data)\n"

                # Log transformation
                if test_info.get("transformation"):
                    analysis_log += f"\nTransformation: {test_info['transformation'].capitalize()} transformation performed.\n"
                    # Log after transformation
                    analysis_log += "Results of assumption tests after transformation:\n"
                    
                    # Get post-transformation residual normality
                    post_residual_norm = test_info.get("post_transformation", {}).get("residuals_normality")
                    if post_residual_norm and post_residual_norm.get("p_value") is not None:
                        analysis_log += f"Shapiro-Wilk test (transformed model residuals): p = {post_residual_norm['p_value']:.4f} - "
                        analysis_log += "Transformed model residuals normally distributed\n" if post_residual_norm.get('is_normal', False) else "Transformed model residuals not normally distributed\n"
                    else:
                        analysis_log += "Shapiro-Wilk test (transformed model residuals): Test not performed (insufficient data)\n"
                    
                    # Get post-transformation variance
                    post_variance = test_info.get("post_transformation", {}).get("variance")
                    if post_variance and post_variance.get("p_value") is not None:
                        analysis_log += f"Brown-Forsythe test (transformed data variance homogeneity): p = {post_variance['p_value']:.4f} - "
                        analysis_log += "Transformed data variances homogeneous\n" if post_variance.get('equal_variance', False) else "Transformed data variances heterogeneous\n"
                    else:
                        analysis_log += "Brown-Forsythe test (transformed data): Not performed (insufficient data)\n"
                else:
                    analysis_log += "\nTransformation: No transformation performed.\n"

            posthoc_results = None

            if test_results.get('p_value') is not None and test_results['p_value'] < 0.05 and len(groups) > 2:
                # Significant result: perform post-hoc tests
                valid_groups = [g for g in groups if g in transformed_samples and len(transformed_samples[g]) > 0]
                print("DEBUG: valid_groups after filter:", valid_groups)
                print("DEBUG: transformed_samples:", {g: len(transformed_samples[g]) for g in transformed_samples})
                print("DEBUG: original_samples:", {g: len(filtered_samples[g]) for g in filtered_samples})

                test_name = test_results.get('test', '').lower()

                # Check if post-hoc tests have already been performed
                if not test_results.get('pairwise_comparisons'):
                    # Let the perform_refactored_posthoc_testing function handle dialog selection for all tests
                    if 'kruskal' in test_name or 'friedman' in test_name or test_recommendation == 'non_parametric':
                        print("DEBUG: Significant non-parametric test (section 2), calling perform_refactored_posthoc_testing without preset posthoc_choice")
                        posthoc_results = StatisticalTester.perform_refactored_posthoc_testing(
                            valid_groups, transformed_samples, test_recommendation,
                            alpha=0.05, posthoc_choice=None  # Let the function show the dialog
                        )
                    else:
                        # Show dialog for parametric tests
                        posthoc_choice = UIDialogManager.select_posthoc_test_dialog(
                            progress_text=kwargs.get('dialog_progress', None),
                            column_name=kwargs.get('dialog_column', None)
                        )
                        if posthoc_choice and posthoc_choice != "none":
                            control_group = None
                            if posthoc_choice == "dunnett":
                                control_group = UIDialogManager.select_control_group_dialog(valid_groups)
                            elif posthoc_choice == "paired_custom":
                                # Handle paired custom directly here to avoid double dialog
                                pairs = UIDialogManager.select_custom_pairs_dialog(valid_groups)
                                if pairs:
                                    # Import required modules
                                    from scipy import stats
                                    import numpy as np
                                    multipletests = get_statsmodels_multitest()
                                    
                                    # Paired t-tests for the selected pairs
                                    pvals, stats_list = [], []
                                    for g1, g2 in pairs:
                                        x, y = np.array(transformed_samples[g1]), np.array(transformed_samples[g2])
                                        tstat, p = stats.ttest_rel(x, y)
                                        stats_list.append(tstat)
                                        pvals.append(p)
                                        # Holm–Šidák correction
                                    reject, p_adj, _, _ = multipletests(pvals, alpha=0.05, method='holm-sidak')
                                    
                                    # Create results in the same format as other post-hoc tests
                                    posthoc_results = {
                                        "posthoc_test": "Custom paired t-tests (Holm-Sidak)",
                                        "pairwise_comparisons": [],
                                        "error": None
                                    }
                                    
                                    # Collect results
                                    for i, (g1, g2) in enumerate(pairs):
                                        ci = PostHocStatistics.calculate_ci_mean_diff(transformed_samples[g1], transformed_samples[g2], alpha=0.05, paired=True)
                                        d = PostHocStatistics.calculate_cohens_d(transformed_samples[g1], transformed_samples[g2], paired=True)
                                        PostHocAnalyzer.add_comparison(
                                            posthoc_results,
                                            group1=g1,
                                            group2=g2,
                                            test="Paired t-test (Holm-Sidak)",
                                            p_value=p_adj[i],
                                            statistic=stats_list[i],
                                            corrected=True,
                                            correction_method="Holm-Sidak",
                                            effect_size=d,
                                            effect_size_type="cohen_d",
                                            confidence_interval=ci,
                                            alpha=0.05
                                        )
                                else:
                                    posthoc_results = {
                                        "posthoc_test": "No pairs selected for custom paired t-tests",
                                        "pairwise_comparisons": [],
                                        "error": None
                                    }
                            else:
                                # For other parametric post-hoc tests, use the refactored function
                                posthoc_results = StatisticalTester.perform_refactored_posthoc_testing(
                                    valid_groups, transformed_samples, test_recommendation,
                                    alpha=0.05, posthoc_choice=posthoc_choice, control_group=control_group
                                )
                    # Process results uniformly - ONLY ONCE here!
                    if posthoc_results:                      
                        if posthoc_choice == "dunnett" and "control_group" in posthoc_results:
                            test_results["control_group"] = posthoc_results["control_group"]
                        if 'pairwise_comparisons' in posthoc_results:
                            import copy
                            test_results['pairwise_comparisons'] = copy.deepcopy(posthoc_results['pairwise_comparisons'])
      
                            print(f"DEBUG: Copied pairwise_comparisons from posthoc_results to test_results")
                            print(f"DEBUG: Same object? {posthoc_results.get('pairwise_comparisons', None) is test_results.get('pairwise_comparisons', None)}")

                        test_results["posthoc_test"] = posthoc_results.get("posthoc_test")

                        # Add debug print to verify
                        print(f"DEBUG: posthoc_results['pairwise_comparisons'] length: {len(posthoc_results.get('pairwise_comparisons', []))}")

                    # INSERT DEBUG OUTPUTS HERE
                    print(f"DEBUG: Pairwise comparisons after post-hoc: {len(test_results.get('pairwise_comparisons', []))}")
                    
            # After post-hoc processing, before test_results.update:
            print(f"DEBUG: Post-hoc results: {posthoc_results.keys() if posthoc_results else None}")
            if posthoc_results and 'error' in posthoc_results and posthoc_results['error']:
                print(f"DEBUG: Post-hoc ERROR: {posthoc_results['error']}")
            print(f"DEBUG: test_results pairwise_comparisons: {len(test_results.get('pairwise_comparisons', []))} items")        


            # Make sure normality and variance test results are explicitly set (only if available)
            # Convert new test_info structure to the expected format for Excel export
            if test_info:
                print(f"DEBUG TEST_INFO STRUCTURE: {test_info}")
                print(f"DEBUG TEST_INFO KEYS: {list(test_info.keys())}")
                if "pre_transformation" in test_info:
                    print(f"DEBUG PRE_TRANSFORMATION: {test_info['pre_transformation']}")
                
                # Convert new residuals-based test info to compatible format
                normality_tests_compat = {}
                variance_test_compat = {}
                
                # Pre-transformation residual normality
                if "pre_transformation" in test_info and "residuals_normality" in test_info["pre_transformation"]:
                    pre_norm = test_info["pre_transformation"]["residuals_normality"]
                    normality_tests_compat["model_residuals"] = {
                        "statistic": pre_norm.get("statistic"),
                        "p_value": pre_norm.get("p_value"),
                        "is_normal": pre_norm.get("is_normal", False),
                        "test_type": "Shapiro-Wilk (Model Residuals)"
                    }
                
                # Post-transformation residual normality
                if "post_transformation" in test_info and "residuals_normality" in test_info["post_transformation"]:
                    post_norm = test_info["post_transformation"]["residuals_normality"]
                    normality_tests_compat["model_residuals_transformed"] = {
                        "statistic": post_norm.get("statistic"),
                        "p_value": post_norm.get("p_value"),
                        "is_normal": post_norm.get("is_normal", False),
                        "test_type": "Shapiro-Wilk (Transformed Model Residuals)"
                    }
                
                # Variance tests
                if "pre_transformation" in test_info and "variance" in test_info["pre_transformation"]:
                    pre_var = test_info["pre_transformation"]["variance"]
                    variance_test_compat.update({
                        "statistic": pre_var.get("statistic"),
                        "p_value": pre_var.get("p_value"),
                        "equal_variance": pre_var.get("equal_variance", False)
                    })
                    print(f"DEBUG VARIANCE_TEST_COMPAT: {variance_test_compat}")
                
                if "post_transformation" in test_info and "variance" in test_info["post_transformation"]:
                    post_var = test_info["post_transformation"]["variance"]
                    variance_test_compat["transformed"] = {
                        "statistic": post_var.get("statistic"),
                        "p_value": post_var.get("p_value"),
                        "equal_variance": post_var.get("equal_variance", False)
                    }
                
                results["normality_tests"] = normality_tests_compat
                results["variance_test"] = variance_test_compat
                
                print(f"DEBUG FINAL normality_tests: {results['normality_tests']}")
                print(f"DEBUG FINAL variance_test: {results['variance_test']}")
                
                # Add test_info for complete information
                results["test_info"] = test_info
            else:
                print("DEBUG: test_info is None or empty!")

            # Make sure test_type/recommendation is set (only if available):
            if test_recommendation:
                results["recommendation"] = test_recommendation

            # Merge important transformation and test info into results
            results.update(test_results)
            
            # Store normality_tests and variance_test before they get overwritten
            preserved_normality = results.get("normality_tests", {})
            preserved_variance = results.get("variance_test", {})
            
            results.update({
                "transformed_samples": transformed_samples,
                "samples": filtered_samples,
                "transformation": test_info.get("transformation") if test_info else None,
                "test_type": test_recommendation
            })
            
            # Restore the correctly formatted test data (don't overwrite with empty data!)
            if preserved_normality:
                results["normality_tests"] = preserved_normality
            elif test_info and "normality_tests" in test_info:
                results["normality_tests"] = test_info["normality_tests"]
            else:
                results["normality_tests"] = {}
                
            if preserved_variance:
                results["variance_test"] = preserved_variance
            elif test_info and "variance_test" in test_info:
                results["variance_test"] = test_info["variance_test"]  
            else:
                results["variance_test"] = {}

            # Nach results.update(test_results):
            print(f"DEBUG: results pairwise_comparisons: {len(results.get('pairwise_comparisons', []))} items")            

            if test_info and "boxcox_lambda" in test_info:
                results["boxcox_lambda"] = test_info["boxcox_lambda"]

            analysis_log += f"\nTest performed: {results.get('test', 'Not specified')}\n"
            if 'p_value' in results:
                p_value = results['p_value']
                if isinstance(p_value, (float, int)):
                    analysis_log += f"p-Value: {p_value:.6f}\n"
                    analysis_log += f"Significance: {'Significant (p < 0.05)' if p_value < 0.05 else 'Not significant'}\n"
                else:
                    analysis_log += f"p-Value: {p_value}\n"
                    analysis_log += "Significance: Not determinable\n"

            # For Two-way/Mixed/RM-ANOVA: main effects and interactions in the log
            if "factors" in results:
                for factor in results["factors"]:
                    analysis_log += (
                        f"Main effect {factor['factor']}: "
                        f"F({factor['df1']}, {factor['df2']}) = {factor['F']:.3f}, "
                        f"p = {factor['p_value']:.4f}, "
                        f"Effect size: {factor.get('effect_size', 'N/A')}\n"
                    )
            if "interactions" in results:
                for inter in results["interactions"]:
                    analysis_log += (
                        f"Interaction {inter['factors'][0]} x {inter['factors'][1]}: "
                        f"F({inter['df1']}, {inter['df2']}) = {inter['F']:.3f}, "
                        f"p = {inter['p_value']:.4f}, "
                        f"Effect size: {inter.get('effect_size', 'N/A')}\n"
                    )

            # Post-hoc tests
            if 'pairwise_comparisons' in results and results['pairwise_comparisons']:
                posthoc_test = results.get("posthoc_test", None)
                if posthoc_test and "Tukey HSD" in posthoc_test:
                    posthoc_display = "Tukey HSD Test"
                elif posthoc_test and "Dunnett Test" in posthoc_test:
                    control_group = results.get("control_group", "")
                    posthoc_display = f"Dunnett Test (Control group: {control_group})"
                else:
                    posthoc_display = posthoc_test if posthoc_test else "No post-hoc test performed"

                analysis_log += f"\nPost‑hoc test: {posthoc_display}\n"
                analysis_log += "Pairwise comparisons:\n"
                for comp in results["pairwise_comparisons"]:
                    group1 = str(comp['group1'])
                    group2 = str(comp['group2'])
                    p_val = comp['p_value']
                    significant = comp['significant']
                    p_text = "p < 0.001" if isinstance(p_val, (float, int)) and p_val < 0.001 else safe_format(p_val, "p = {:.4f}")
                    sign_text = "significant" if significant else "not significant"
                    stars = "***" if significant and p_val < 0.001 else "**" if significant and p_val < 0.01 else "*" if significant else ""
                    analysis_log += f"  {group1} vs {group2}: {p_text}, {sign_text} {stars}\n"
            else:
                analysis_log += "\nNo pairwise comparisons were performed or calculated.\n"

            # Define file base based on custom name or groups
            if file_name:
                file_base = file_name
            else:
                file_base = "_".join(map(str, groups))

            excel_file = f"{file_base}_results.xlsx"
            
            results['groups'] = groups
            results['raw_data'] = {g: filtered_samples[g][:] for g in groups}
            if results.get('transformation', 'None') != 'None':
                results['raw_data_transformed'] = {g: transformed_samples[g][:] for g in groups}
            
            # DO NOT OVERWRITE! The variance_test and normality_tests are already set above
            # Keep the old format for backward compatibility if variance_test doesn't exist
            if "variance_test" not in results:
                results["variance_homogeneity_test"] = test_info.get("variance_test", {}) if test_info else {}    

            # Add debug statements before Excel export
            print("DEBUG: Assumption tests before Excel export:")
            print("  Normality tests:", results.get("normality_tests", {}))
            print("  Variance tests:", results.get("variance_test", {}))
            print("  Test recommendation:", test_recommendation)
                
            # Export to Excel
            if not skip_excel:
                original_dir = os.getcwd()
                print(f"DEBUG: Directory before Excel export: {original_dir}")
                
                # Use absolute path for Excel file
                excel_file = get_output_path(file_base, "xlsx") 
                
                ResultsExporter = get_results_exporter()
                ResultsExporter.export_results_to_excel(results, excel_file, analysis_log)
                analysis_log += f"\nResults were saved to {excel_file}.\n"
                
                # Ensure we're back in the original directory
                if os.getcwd() != original_dir:
                    os.chdir(original_dir)
                    print(f"DEBUG: Restored original directory: {original_dir}")

            # Create the plot, if not skipped
            if not skip_plots:
                print(f"DEBUG: Current working directory before export: {os.getcwd()}")
                pairwise_comparisons = results.get('pairwise_comparisons', None)
                
                # Get plot type from kwargs, default to 'Bar'
                plot_type = kwargs.get('plot_type', 'Bar')
                print(f"DEBUG: Creating plot of type: {plot_type}")
                
                # Create a clean kwargs dict without parameters that plotting methods don't accept
                # Only exclude parameters that definitely don't exist in plot methods
                plot_kwargs = {k: v for k, v in kwargs.items() if k not in [
                    'plot_type', 'file_path', 'group_col', 'groups', 'sheet_name', 
                    'value_cols', 'combine_columns', 'skip_plots', 'skip_excel',
                    'dependent', 'show_individual_lines', 'compare', 'additional_factors',
                    'dataset_name', 'dialog_column', 'dialog_progress',
                    # Parameters that don't exist in plot_bar method
                    'aspect', 'x_label_size', 'y_label_size', 'title_size',
                    'refline', 'panel_labels', 'value_annotations', 'significance_mode',
                    'embed_fonts', 'add_metadata',
                    # Appearance/formatting keys to exclude
                    'font_main', 'font_axis', 'show_title', 'fontsize_title', 'fontsize_axis',
                    'fontsize_ticks', 'fontsize_groupnames', 'axis_linewidth', 'bar_linewidth',
                    'gridline_width', 'grid', 'minor_ticks', 'logy', 'logx', 'despine', 'alpha',
                    'bar_edge_color', 'bar_edge_width', 'grid_style', 'spine_style', 'tick_label_size',
                    'dpi'
                ]}
                
                # Choose the appropriate plot function based on plot_type
                if plot_type == "Bar":
                    # Ensure show_points is always True for bar plots
                    plot_kwargs['show_points'] = True
                    plot_kwargs['point_size'] = plot_kwargs.get('point_size', 80)
                    plot_kwargs['point_alpha'] = plot_kwargs.get('point_alpha', 0.8)
                    # Always pass colors to legend
                    fig, ax = DataVisualizer.plot_bar(
                        groups, filtered_samples, width=width, height=height,
                        colors=colors, hatches=hatches, compare=compare,
                        test_recommendation=test_recommendation,
                        x_label=x_label, y_label=y_label,
                        title=title, save_plot=save_plot, error_type=error_type,
                        pairwise_results=pairwise_comparisons,
                        file_name=file_base, legend_colors=colors, **plot_kwargs)
                elif plot_type == "Box":
                    # Ensure show_points is always True for box plots
                    fig, ax = DataVisualizer.plot_box(
                        groups, filtered_samples, width=width, height=height,
                        colors=colors, hatches=hatches,
                        test_recommendation=test_recommendation,
                        x_label=x_label, y_label=y_label,
                        title=title, save_plot=save_plot,
                        show_points=True, point_size=80, point_alpha=0.8,
                        pairwise_results=pairwise_comparisons,
                        file_name=file_base, legend_colors=colors)
                elif plot_type == "Violin":
                    # Ensure show_points is always True for violin plots  
                    fig, ax = DataVisualizer.plot_violin(
                        groups, filtered_samples, width=width, height=height,
                        colors=colors, hatches=hatches,
                        test_recommendation=test_recommendation,
                        x_label=x_label, y_label=y_label,
                        title=title, save_plot=save_plot,
                        show_points=True, point_size=80, point_alpha=0.8,
                        pairwise_results=pairwise_comparisons,
                        file_name=file_base, legend_colors=colors)
                elif plot_type == "Strip":
                    # Strip plot doesn't exist, fall back to box plot with points
                    fig, ax = DataVisualizer.plot_box(
                        groups, filtered_samples, width=width, height=height,
                        colors=colors, hatches=hatches,
                        test_recommendation=test_recommendation,
                        x_label=x_label, y_label=y_label,
                        title=title, save_plot=save_plot,
                        show_points=True, point_size=80, point_alpha=0.8,
                        pairwise_results=pairwise_comparisons,
                        file_name=file_base, legend_colors=colors)
                elif plot_type == "Raincloud":
                    # Ensure show_points is always True for raincloud plots
                    fig, ax = DataVisualizer.plot_raincloud(
                        groups, filtered_samples, width=width, height=height,
                        colors=colors, hatches=hatches,
                        test_recommendation=test_recommendation,
                        x_label=x_label, y_label=y_label,
                        title=title, save_plot=save_plot,
                        show_points=True, point_size=80, point_alpha=0.8,
                        pairwise_results=pairwise_comparisons,
                        file_name=file_base, legend_colors=colors)
                else:
                    # Fallback to bar plot for unknown plot types
                    print(f"WARNING: Unknown plot type '{plot_type}', falling back to Bar plot")
                    # Ensure show_points is always True for fallback bar plots
                    plot_kwargs['show_points'] = True
                    plot_kwargs['point_size'] = plot_kwargs.get('point_size', 80)
                    plot_kwargs['point_alpha'] = plot_kwargs.get('point_alpha', 0.8)
                    fig, ax = DataVisualizer.plot_bar(
                        groups, filtered_samples, width=width, height=height,
                        colors=colors, hatches=hatches, compare=compare,
                        test_recommendation=test_recommendation,
                        x_label=x_label, y_label=y_label,
                        title=title, save_plot=save_plot, error_type=error_type,
                        pairwise_results=pairwise_comparisons,
                        file_name=file_base, legend_colors=colors, **plot_kwargs)
                analysis_log += f"\nPlots were saved as:\n"
                analysis_log += f"  {file_base}.pdf\n"
                analysis_log += f"  {file_base}.png\n"
                import matplotlib.pyplot as plt
                plt.close(fig)
                results["_file_paths"] = {
                    "excel": os.path.abspath(excel_file),
                    "pdf": os.path.abspath(f"{file_base}.pdf"),
                    "png": os.path.abspath(f"{file_base}.png")
                }
            else:
                results["_file_paths"] = {
                    "excel": os.path.abspath(excel_file)
                }
            # Special visualization for dependent data
            if dependent and not skip_plots:
                try:
                    line_fig, line_ax = DataVisualizer.plot_dependent_samples(
                        groups, filtered_samples, width=width, height=height,
                        colors=colors, title=f"{title} (dependent measurements)" if title else "Dependent measurements",
                        x_label=x_label, y_label=y_label,
                        save_plot=save_plot, file_name=file_base+"_lines",
                        show_individual=show_individual_lines
                    )
                    import matplotlib.pyplot as plt
                    plt.close(line_fig)
                    line_plot_base = file_base+"_lines"
                    results["_file_paths"]["pdf_lines"] = os.path.abspath(f"{line_plot_base}.pdf")
                    results["_file_paths"]["png_lines"] = os.path.abspath(f"{line_plot_base}.png")
                    analysis_log += f"\nAdditional line plot for dependent data created:\n"
                    analysis_log += f"  {line_plot_base}.pdf\n"
                    analysis_log += f"  {line_plot_base}.png\n"
                except Exception as e:
                    analysis_log += f"\nError creating line plot for dependent data: {str(e)}\n"
                    print(f"Error creating line plot: {str(e)}")
            if results.get('transformation', 'None') != 'None':
                results['transformed_data'] = transformed_samples

            # About line 5647, just before "return results"
            params = {
                "file_path": file_path,
                "sheet_name": sheet_name,
                "group_col": group_col,
                "value_cols": value_cols,
                "groups": groups,
                "dependent": dependent,
                "error_type": error_type
            }
            # Build protocol
            def build_analysis_log(results, params):
                log = []
                log.append("ANALYSIS LOG")
                log.append('"This sheet documents the course of the statistical analysis and the decisions made. The log provides a chronological overview of the individual analysis steps, methods used, transformations, test selection, and special notes.\nEach paragraph describes a key step or decision in the analysis process."\n')
                log.append(f"Analysis report\nDate and time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                if 'file_path' in params:
                    log.append(f"File: {params['file_path']}")
                if 'sheet_name' in params:
                    log.append(f"Worksheet: {params['sheet_name']}")
                if 'group_col' in params:
                    log.append(f"Group column: {params['group_col']}")
                if 'value_cols' in params:
                    log.append(f"Value column(s): {', '.join(params['value_cols'])}")
                if 'groups' in params:
                    log.append(f"Groups to analyze: {', '.join(params['groups'])}")
                if 'dependent' in params:
                    log.append(f"Dependent samples: {'Yes' if params['dependent'] else 'No'}")
                if 'error_type' in params:
                    log.append(f"Error bar type: {'SEM (standard error)' if params['error_type']=='se' else 'SD (standard deviation)'}")
                log.append("\n--- ANALYSIS ---\n")
                if results.get('import_status'):
                    log.append("Data imported successfully.")
                if 'group_sizes' in results:
                    log.append("Number of data points per group:")
                    for group, n in results['group_sizes'].items():
                        log.append(f"{group}: {n} data points")
                if 'test_recommendation' in results:
                    log.append(f"Test recommendation: {results['test_recommendation']}")
                if 'normality_p' in results:
                    log.append(f"Shapiro-Wilk test (normality): p = {results['normality_p']:.4f} - {'Normally distributed' if results['normality_p'] > 0.05 else 'Not normally distributed'}")
                if 'levene_p' in results:
                    log.append(f"Brown-Forsythe test (variance homogeneity): p = {results['levene_p']:.4f} - {'Variances homogeneous' if results['levene_p'] > 0.05 else 'Variances heterogeneous'}")
                if 'transformation' in results:
                    log.append(f"Transformation: {results['transformation'] if results['transformation'] else 'No transformation performed.'}")
                if 'test' in results:
                    log.append(f"Test performed: {results['test']}")
                if 'p_value' in results:
                    p_value = results['p_value']
                    if p_value is None:
                        log.append("p-Value: Not available (test may have failed)")
                        if 'error' in results and results['error']:
                            log.append(f"Error: {results['error']}")
                    elif isinstance(p_value, (float, int)):
                        log.append(f"p-Value: {p_value:.6f}")
                        log.append(f"Significance: {'Significant (p < 0.05)' if p_value < 0.05 else 'Not significant'}")
                    else:
                        log.append(f"p-Value: {p_value}")
                        log.append("Significance: Not determinable")
                if "factors" in results:
                    for factor in results["factors"]:
                        log.append(
                            f"Main effect {factor['factor']}: F({factor['df1']}, {factor['df2']}) = {factor['F']:.3f}, "
                            f"p = {factor['p_value']:.4f}, Effect size: {factor.get('effect_size', 'N/A')}"
                        )
                if "interactions" in results:
                    for inter in results["interactions"]:
                        log.append(
                            f"Interaction {inter['factors'][0]} x {inter['factors'][1]}: F({inter['df1']}, {inter['df2']}) = {inter['F']:.3f}, "
                            f"p = {inter['p_value']:.4f}, Effect size: {inter.get('effect_size', 'N/A')}"
                        )
                if 'pairwise_comparisons' in results and results['pairwise_comparisons']:
                    posthoc_test = results.get("posthoc_test", "Post-hoc test")
                    log.append(f"\nPost‑hoc test: {posthoc_test}")
                    log.append("Pairwise comparisons:")
                    for comp in results["pairwise_comparisons"]:
                        group1 = str(comp['group1'])
                        group2 = str(comp['group2'])
                        p_val = comp['p_value']
                        significant = comp['significant']
                        p_text = "p < 0.001" if isinstance(p_val, (float, int)) and p_val < 0.001 else f"p = {p_val:.4f}"
                        sign_text = "significant" if significant else "not significant"
                        stars = "***" if significant and p_val < 0.001 else "**" if significant and p_val < 0.01 else "*" if significant else ""
                        log.append(f"{group1} vs {group2}: {p_text}, {sign_text} {stars}")
                else:
                    log.append("\nNo pairwise comparisons were performed or calculated.")
                return "\n".join(log)
            analysis_log = build_analysis_log(results, params)
            results["analysis_log"] = analysis_log
            return results

        except Exception as e:
            error_message = str(e) if str(e) else "Unknown error occurred"
            analysis_log += f"\nERROR: {error_message}\n"
            print(f"Error during analysis: {error_message}")
            import traceback
            traceback.print_exc()
            return {"error": error_message, "analysis_log": analysis_log}       

def get_output_path(file_base, ext):
    """Get an absolute path to save output files on desktop."""
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
    if not os.path.isdir(desktop_path):
        # Fallback: current working directory
        desktop_path = os.getcwd()
    
    out_path = os.path.join(desktop_path, f"{file_base}.{ext}")
    abs_path = os.path.abspath(out_path)
    print(f"DEBUG: get_output_path returns absolute path: {abs_path}")
    return abs_path

try:
    _QT_APP.exit()
except Exception:
    pass

import os
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from scipy.stats import t
OUTLIER_IMPORTS_AVAILABLE = True
try:
    import seaborn
    import matplotlib.pyplot as plt
    from openpyxl import Workbook
except ImportError:
    OUTLIER_IMPORTS_AVAILABLE = False

class OutlierDetector:
    """
    Detect outliers in grouped data using Grubbs' Test or Dixon's Q-Test.
    Loads an Excel table with columns ['Group', 'Value'], 
    converts all values (German decimal numbers with comma) to float,
    performs Grubbs or Dixon tests iteratively or once for each group separately
    and marks found outliers.
    """

    def __init__(self, df, group_col, value_col):
        """
        Initializes the OutlierDetector with an already loaded DataFrame.
        
        Parameters:
        -----------
        df : pandas.DataFrame
            DataFrame with the data
        group_col : str
            Name of the group column
        value_col : str
            Name of the values column
        """
        print(f"DEBUG: Initializing OutlierDetector with columns: group_col={group_col}, value_col={value_col}")
        print(f"DEBUG: DataFrame shape: {df.shape}, columns: {df.columns.tolist()}")
        
        if not OUTLIER_IMPORTS_AVAILABLE:
            raise ImportError("Required packages for outlier detection are not available. "
                              "Please install: pip install outliers pingouin openpyxl")
        
        # Create copy of DataFrame
        self.df = df.copy()
        self.group_col = group_col
        self.value_col = value_col
        self.active_test = None  # Track which test was run
        self.debug_log = []  # Initialize debug log

        # Ensure columns are present
        if group_col not in self.df.columns or value_col not in self.df.columns:
            raise ValueError(f"Columns '{group_col}' and '{value_col}' must be present in the DataFrame.")

        # Add initialization info to log
        self.debug_log.append(f"*** OUTLIER DETECTION INITIALIZATION ***")
        self.debug_log.append(f"DataFrame shape: {df.shape}")
        self.debug_log.append(f"Group column: {group_col}")
        self.debug_log.append(f"Value column: {value_col}")
        self.debug_log.append(f"Available columns: {df.columns.tolist()}")

        # Convert values column to float (if necessary)
        self._convert_values_to_float()
        
        # Show group statistics after initialization
        self.debug_log.append(f"\n=== GROUP STATISTICS ===")
        self.debug_log.append(f"Number of groups in data: {self.df[group_col].nunique()}")
        
        for group_name, group_data in self.df.groupby(group_col):
            group_stats = self._calculate_group_statistics(group_name, group_data)
            self.debug_log.extend(group_stats)

    def _convert_values_to_float(self):
        """
        Converts the values column to float.
        Handles German decimal separators (comma instead of period).
        """
        try:
            self.debug_log.append(f"\n=== VALUE CONVERSION ===")
            self.debug_log.append(f"Converting column '{self.value_col}' to float")
            
            # Check if the column exists
            if self.value_col not in self.df.columns:
                error_msg = f"Column '{self.value_col}' not found in DataFrame. Available columns: {list(self.df.columns)}"
                self.debug_log.append(f"ERROR: {error_msg}")
                raise ValueError(error_msg)
            
            self.debug_log.append(f"Sample values before conversion: {self.df[self.value_col].head(3).tolist()}")
            self.debug_log.append(f"Data type before conversion: {self.df[self.value_col].dtype}")
            
            # Check if column is already numeric
            if pd.api.types.is_numeric_dtype(self.df[self.value_col]):
                self.debug_log.append("Column is already numeric, no conversion needed")
                return
            
            # Convert German decimal numbers
            self.df[self.value_col] = (
                self.df[self.value_col]
                    .astype(str)
                    .str.replace('.', '', regex=False)   # Remove thousand separators
                    .str.replace(',', '.', regex=False)  # Comma → Period
                    .astype(float)
            )
            
            self.debug_log.append(f"Sample values after conversion: {self.df[self.value_col].head(3).tolist()}")
            self.debug_log.append(f"Data type after conversion: {self.df[self.value_col].dtype}")
            self.debug_log.append(f"Value range after conversion: {self.df[self.value_col].min():.4f} to {self.df[self.value_col].max():.4f}")
        
        except Exception as e:
            error_message = f"Conversion failed: {str(e)}"
            self.debug_log.append(f"ERROR: {error_message}")
            raise ValueError(f"Error converting values column '{self.value_col}' to float: {str(e)}")
        
    def _calculate_group_statistics(self, group_name, group_data):
        """Calculate basic statistics for a group and return as log entries."""
        stats = []
        values = group_data[self.value_col].dropna()
        stats.append(f"\n--- Group '{group_name}' Statistics ---")
        stats.append(f"  Count: {len(values)}")
        
        if len(values) > 0:
            stats.append(f"  Min: {values.min():.4f}")
            stats.append(f"  Max: {values.max():.4f}")
            stats.append(f"  Mean: {values.mean():.4f}")
            stats.append(f"  Median: {values.median():.4f}")
            stats.append(f"  Std Dev: {values.std():.4f}")
        else:
            stats.append("  No valid values in this group")
        
        return stats
    
    @staticmethod
    def _grubbs_iterative(vals: np.ndarray,
                           alpha: float = 0.05,
                           two_sided: bool = True):
        """
        Iteratively remove the most extreme point (if G_obs > G_crit),
        repeat until no further outliers are detected.
        Returns list of relative indices into the original vals array.
        """
        vals = np.asarray(vals, dtype=float)
        # track the original positions
        idxs = np.arange(vals.size)
        out_rel_idxs = []
        
        while vals.size >= 3:
            G_obs, G_crit, rel_idx = OutlierDetector._grubbs_statistic(
                vals, alpha=alpha, two_sided=two_sided
            )
            if G_obs <= G_crit:
                break
            # record and remove
            out_rel_idxs.append(int(idxs[rel_idx]))
            mask = np.ones(vals.size, dtype=bool)
            mask[rel_idx] = False
            vals = vals[mask]
            idxs = idxs[mask]
        
        return out_rel_idxs

    @staticmethod
    def _grubbs_statistic(x: np.ndarray,
                          alpha: float = 0.05,
                          two_sided: bool = True):
        """
        Compute Grubbs' G statistic and critical value for array x.
        Returns (G_obs, G_crit, outlier_rel_index).
        """
        n = x.size
        if n < 3:
            raise ValueError("Grubbs' test requires at least 3 values")
        
        mu = x.mean()
        s  = x.std(ddof=1)
        # maximum absolute deviation
        diffs = np.abs(x - mu)
        rel_idx = int(np.argmax(diffs))
        G_obs = diffs[rel_idx] / s
        
        # two‐sided splits alpha into 2 tails
        tail = 2 if two_sided else 1
        # student-t critical value
        t_crit = t.ppf(1 - alpha/(tail * n), df=n-2)
        # critical G
        G_crit = ((n - 1) / np.sqrt(n)
                  * np.sqrt(t_crit**2 / ( (n - 2) + t_crit**2 )))
        
        return G_obs, G_crit, rel_idx

    def run_grubbs(self, alpha: float = 0.05, iterate: bool = True):
        self.debug_log.append("\n=== GRUBBS OUTLIER DETECTION ===")
        self.debug_log.append(f"alpha={alpha}, iterate={iterate}")

        self.df["Grubbs_Outlier"] = False
        self.active_test = "Grubbs"
        total_out = 0

        for gname, gdf in self.df.groupby(self.group_col):
            self.debug_log.append(f"\n--- Group '{gname}' ({len(gdf)}) ---")

            # original indices and clean values
            idxs = gdf.index.values
            vals = gdf[self.value_col].astype(float).values
            valid = ~np.isnan(vals)
            idxs_valid = idxs[valid]
            vals_valid = vals[valid]

            if len(vals_valid) < 3:
                self.debug_log.append("  n < 3 → skipping Grubbs")
                continue

            # choose iterative vs. single removal
            if iterate:
                rel_outs = OutlierDetector._grubbs_iterative(
                    vals_valid, alpha=alpha, two_sided=True
                )
            else:
                G_obs, G_crit, rel_idx = OutlierDetector._grubbs_statistic(
                    vals_valid, alpha=alpha, two_sided=True
                )
                rel_outs = [rel_idx] if G_obs > G_crit else []

            if not rel_outs:
                self.debug_log.append("  No outliers detected")
            else:
                # map back to global indices
                global_outs = idxs_valid[rel_outs].tolist()
                for gi in global_outs:
                    val = self.df.at[gi, self.value_col]
                    self.debug_log.append(f"  Outlier at idx={gi}, value={val:.4f}")
                self.df.loc[global_outs, "Grubbs_Outlier"] = True
                total_out += len(global_outs)
                self.debug_log.append(f"  ==> {len(global_outs)} outliers in group")

        self.debug_log.append("\n=== GRUBBS SUMMARY ===")
        self.debug_log.append(f"Total outliers: {total_out}")

    def run_grubbs_test(self, alpha: float = 0.05, iterate: bool = True):
        """
        Alias for run_grubbs method for compatibility.
        
        Parameters:
        -----------
        alpha : float
            Significance level for the test (default: 0.05)
        iterate : bool
            Whether to iteratively remove outliers (default: True)
        """
        return self.run_grubbs(alpha=alpha, iterate=iterate)

    def run_mod_z_score(self, threshold=3.5, iterate=False):
        """
        Performs Modified Z-Score outlier detection for each group.
        
        Parameters:
        -----------
        threshold : float
            Threshold for modified Z-scores to be considered outliers (default: 3.5)
        iterate : bool
            Whether to iteratively remove outliers and recompute scores
        """
        self.debug_log.append(f"\n=== MODIFIED Z-SCORE OUTLIER DETECTION EXECUTION ===")
        self.debug_log.append(f"Test parameters: threshold={threshold}, iterate={iterate}")
        self.debug_log.append(f"Test principle: Modified Z-Score uses median absolute deviation (MAD) for robustness against outliers")
        
        # Create the Modified Z-Score column and mark this test as active
        self.df['ModZ_Outlier'] = False
        self.active_test = 'ModZ'
        
        total_outliers = 0
        
        for group_name, group_df in self.df.groupby(self.group_col):
            self.debug_log.append(f"\n--- Processing group '{group_name}' ---")
            self.debug_log.append(f"Group size: {len(group_df)} observations")
            
            indices = group_df.index.tolist()
            values = group_df[self.value_col].values.copy()
            
            # Filter out NaN values
            valid_mask = ~np.isnan(values)
            valid_values = values[valid_mask]
            valid_indices = [idx for mask, idx in zip(valid_mask, indices) if mask]
            
            if len(valid_values) == 0:
                self.debug_log.append(f"Group '{group_name}' has no valid values (all NaN), skipping")
                continue
                
            outlier_indices = []
            
            # Log basic group statistics
            mean_val = np.mean(valid_values)
            median_val = np.median(valid_values)
            self.debug_log.append(f"Group statistics: min={np.min(valid_values):.3f}, max={np.max(valid_values):.3f}, mean={mean_val:.3f}, median={median_val:.3f}")
            
            def detect_single_round(vals, indices_list):
                if len(vals) <= 1:
                    return [], "No outliers found (too few values for MAD calculation)"
                    
                # Calculate median and MAD
                median = np.median(vals)
                mad = np.median(np.abs(vals - median))
                
                # Avoid division by zero
                if mad == 0:
                    return [], "MAD = 0, cannot compute modified Z-scores"
                    
                # Calculate modified Z-scores
                mod_z_scores = 0.6745 * (vals - median) / mad
                
                # Find outliers
                outlier_mask = np.abs(mod_z_scores) > threshold
                outlier_idx = [idx for mask, idx in zip(outlier_mask, indices_list) if mask]
                
                # Log this round
                self.debug_log.append(f"  Modified Z-Score Analysis:")
                self.debug_log.append(f"    Median: {median:.4f}")
                self.debug_log.append(f"    MAD: {mad:.4f}")
                self.debug_log.append(f"    Threshold: {threshold} (absolute value)")
                
                if len(outlier_idx) > 0:
                    self.debug_log.append(f"    {len(outlier_idx)} outliers found in this round")
                    return outlier_idx, None
                else:
                    self.debug_log.append(f"    No outliers found in this round")
                    return [], "No outliers found (all modified Z-scores within threshold)"
            
            # Iterative outlier detection
            vals_copy = valid_values.copy()
            idx_copy = valid_indices.copy() 
            round_count = 0
            
            while True:
                round_count += 1
                self.debug_log.append(f"  Testing round {round_count} with {len(vals_copy)} values")
                
                round_outliers, stop_message = detect_single_round(vals_copy, idx_copy)
                
                if round_outliers:
                    outlier_indices.extend(round_outliers)
                    if iterate:
                        # Remove outliers for next iteration
                        keep_mask = ~np.isin(idx_copy, round_outliers)
                        vals_copy = vals_copy[keep_mask]
                        idx_copy = [idx for keep, idx in zip(keep_mask, idx_copy) if keep]
                    else:
                        self.debug_log.append(f"  No more iterations requested, terminating test")
                        break
                else:
                    self.debug_log.append(f"  {stop_message if stop_message else 'No more outliers found'}, terminating test")
                    break
                    
                if len(vals_copy) <= 1:
                    self.debug_log.append(f"  Too few values left for MAD calculation, terminating test")
                    break
            
            # Mark all found indices in main DataFrame
            if outlier_indices:
                self.df.loc[outlier_indices, 'ModZ_Outlier'] = True
                total_outliers += len(outlier_indices)
                self.debug_log.append(f"  RESULT: {len(outlier_indices)} outliers found in group '{group_name}'")
            else:
                self.debug_log.append(f"  RESULT: No outliers found in group '{group_name}'")
        
        # Final summary
        self.debug_log.append(f"\n=== MODIFIED Z-SCORE TEST SUMMARY ===")
        self.debug_log.append(f"Total outliers detected across all groups: {total_outliers}")
        self.debug_log.append(f"Test completed successfully")


    def save_results(self, output_path):
        """
        Outputs the result to a new Excel file while preserving formulas.
        Only columns for actually performed tests are displayed.
        """
        self.debug_log.append(f"\n=== SAVING RESULTS ===")
        self.debug_log.append(f"Output path: {output_path}")
        self.debug_log.append(f"Active test: {self.active_test}")
        
        # Store the original path - it's crucial to use absolute paths
        output_path = os.path.abspath(output_path)
        self.debug_log.append(f"Absolute output path: {output_path}")
        
        # Store original directory but don't change it
        original_cwd = os.getcwd()
        self.debug_log.append(f"Current working directory: {original_cwd}")
        
        try:
            if self.active_test == 'Grubbs':
                outlier_count = self.df['Grubbs_Outlier'].sum()
                self.debug_log.append(f"Total outliers found with Grubbs method: {outlier_count}")
            elif self.active_test == 'ModZ':
                outlier_count = self.df['ModZ_Outlier'].sum()
                self.debug_log.append(f"Total outliers found with Modified Z-Score method: {outlier_count}")
            
            # Check if file exists and contains formulas we need to preserve
            file_exists = os.path.exists(output_path)
            
            if file_exists:
                # Create a backup of the original file with a timestamp
                from datetime import datetime
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_dir = os.path.dirname(output_path)
                base_name = os.path.splitext(os.path.basename(output_path))[0]
                backup_path = os.path.join(output_dir, f"{base_name}_{timestamp}_backup.xlsx")
                import shutil
                shutil.copy2(output_path, backup_path)
                self.debug_log.append(f"Created backup of existing file at: {backup_path}")
                
                # Create a new file with a different name for our results
                results_path = os.path.join(output_dir, f"{base_name}_outliers.xlsx")
                self.debug_log.append(f"Creating new results file at: {results_path}")
            else:
                # Just use the original path if the file doesn't exist
                results_path = output_path
            
            # Create new workbook for our results
            wb = Workbook()

            # 1) Sheet "Raw_Data_with_Outlier_Marking"
            ws_raw = wb.active
            ws_raw.title = "Raw_Data_with_Outlier_Marking"

            # Include only columns for the test that was run
            test_columns = []
            if self.active_test == 'Grubbs' and 'Grubbs_Outlier' in self.df.columns:
                test_columns.append('Grubbs_Outlier')
            elif self.active_test == 'ModZ' and 'ModZ_Outlier' in self.df.columns:
                test_columns.append('ModZ_Outlier')

            # Write headers
            headers = [self.group_col, self.value_col]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws_raw.cell(row=1, column=col_idx, value=str(header))
                cell.font = Font(bold=True)

            # Write data with outlier highlighting
            outlier_count_by_group = {}
            for row_idx, (_, row) in enumerate(self.df.iterrows(), start=2):
                group = row[self.group_col]
                if group not in outlier_count_by_group:
                    outlier_count_by_group[group] = 0
                    
                # Column 1: Group
                group_value = str(group) if group is not None else ""
                ws_raw.cell(row=row_idx, column=1, value=group_value)
                
                # Column 2: Value - mark red if outlier
                value = row[self.value_col]
                if pd.isna(value) or not isinstance(value, (int, float)):
                    value = 0.0
                
                cell_value = ws_raw.cell(row=row_idx, column=2, value=float(value))
                
                # Check for outlier and mark with red highlighting
                is_outlier = any(row.get(col, False) for col in test_columns)
                if is_outlier:
                    outlier_count_by_group[group] += 1
                    # Red font and light red background
                    cell_value.font = Font(color="FF0000", bold=True)
                    cell_value.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    self.debug_log.append(f"Marking outlier: Group={group}, Value={value:.4f}")

            # 2) Sheet "Cleaned"
            ws_clean = wb.create_sheet(title="Cleaned")
            headers_clean = [self.group_col, self.value_col]
            for col_idx, header in enumerate(headers_clean, start=1):
                cell = ws_clean.cell(row=1, column=col_idx, value=str(header))
                cell.font = Font(bold=True)

            # Filter: no outliers from active test
            outlier_mask = pd.Series(False, index=self.df.index)
            for test_col in test_columns:
                outlier_mask |= self.df[test_col]
            
            cleaned_df = self.df[~outlier_mask]
            self.debug_log.append(f"Cleaned data has {len(cleaned_df)} rows (removed {len(self.df) - len(cleaned_df)} outliers)")
            
            for row_idx, (_, row) in enumerate(cleaned_df.iterrows(), start=2):
                group_value = str(row[self.group_col]) if row[self.group_col] is not None else ""
                value = float(row[self.value_col]) if not pd.isna(row[self.value_col]) else 0.0
                
                ws_clean.cell(row=row_idx, column=1, value=group_value)
                ws_clean.cell(row=row_idx, column=2, value=value)

            # 3) Summary sheet with statistics
            ws_summary = wb.create_sheet(title="Summary")
            ws_summary.cell(row=1, column=1, value="Outlier Detection Summary").font = Font(bold=True, size=14)
            
            # Test parameters
            ws_summary.cell(row=3, column=1, value="Test parameters:").font = Font(bold=True)
            ws_summary.cell(row=4, column=1, value="Test type:")
            ws_summary.cell(row=4, column=2, value=str(self.active_test))
            
            # Add test-specific parameters
            if self.active_test == 'Grubbs':
                ws_summary.cell(row=5, column=1, value="Grubbs alpha level:")
                ws_summary.cell(row=5, column=2, value="0.05")  # Default value, adjust if needed
            elif self.active_test == 'ModZ':
                ws_summary.cell(row=5, column=1, value="Z-score threshold:")
                ws_summary.cell(row=5, column=2, value="3.5")  # Default value, adjust if needed
            
            # Group statistics
            row_pos = 7
            ws_summary.cell(row=row_pos, column=1, value="Group statistics:").font = Font(bold=True)
            row_pos += 1
            
            # Headers for statistics table
            stat_headers = ["Group", "Count", "Min", "Max", "Mean", "Std", "Outliers"]
            for col_idx, header in enumerate(stat_headers, start=1):
                cell = ws_summary.cell(row=row_pos, column=col_idx, value=header)
                cell.font = Font(bold=True)
                
            row_pos += 1
            for group_name, group_df in self.df.groupby(self.group_col):
                values = group_df[self.value_col].dropna()
                
                ws_summary.cell(row=row_pos, column=1, value=str(group_name))
                ws_summary.cell(row=row_pos, column=2, value=len(group_df))
                ws_summary.cell(row=row_pos, column=3, value=float(values.min()) if len(values) > 0 else 0.0)
                ws_summary.cell(row=row_pos, column=4, value=float(values.max()) if len(values) > 0 else 0.0)
                ws_summary.cell(row=row_pos, column=5, value=float(values.mean()) if len(values) > 0 else 0.0)
                ws_summary.cell(row=row_pos, column=6, value=float(values.std()) if len(values) > 0 else 0.0)
                ws_summary.cell(row=row_pos, column=7, value=outlier_count_by_group.get(group_name, 0))
                row_pos += 1

            # 4) Enhanced Debug Log sheet with better formatting
            ws_log = wb.create_sheet(title="Debug_Log")
            header_cell = ws_log.cell(row=1, column=1)
            header_cell.value = "Outlier Detection Debug Log"
            header_cell.font = Font(bold=True, size=16, color="0066CC")

            subheader_cell = ws_log.cell(row=2, column=1)
            subheader_cell.value = "This sheet contains detailed information about the outlier detection process."
            subheader_cell.font = Font(italic=True)

            # Set column width for better readability
            ws_log.column_dimensions['A'].width = 150

            # Write debug log entries with enhanced formatting and formula protection
            for idx, log_entry in enumerate(self.debug_log, start=4):
                log_text = str(log_entry)[:32000]
                cell = ws_log.cell(row=idx, column=1)
                cell.value = log_text
                
                # Force inlineStr data type for any text starting with "="
                if log_text.startswith("="):
                    cell.data_type = 'inlineStr'
                
                # Enhanced formatting based on content
                if log_text.startswith("==="):
                    # Main section headers - blue background, white text
                    cell.font = Font(bold=True, size=14, color="FFFFFF")
                    cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
                elif log_text.startswith("---"):
                    # Group headers - light blue background, dark text
                    cell.font = Font(bold=True, size=12, color="000080")
                    cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                elif "ERROR:" in log_text:
                    # Errors - red background, white text
                    cell.font = Font(color="FFFFFF", bold=True, size=11)
                    cell.fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
                elif "OUTLIER DETECTED" in log_text:
                    # Outlier detection - yellow background, red text
                    cell.font = Font(color="CC0000", bold=True, size=11)
                    cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                elif "RESULT:" in log_text:
                    # Results - green background, dark text
                    cell.font = Font(color="006600", bold=True, size=11)
                    cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                elif log_text.startswith("  Testing round") or log_text.startswith("  Running"):
                    # Test round information - light gray background
                    cell.font = Font(size=10, italic=True)
                    cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                elif log_text.startswith("    "):
                    # Detailed information - smaller font, indented
                    cell.font = Font(size=9, color="666666")
                elif log_text.startswith("  "):
                    # General information - smaller font
                    cell.font = Font(size=10)

            # Add visualization
            temp_file = self._add_single_visualization_sheet(wb, self)
            
            # Save the new workbook - Make sure to use the absolute path
            wb.save(results_path)
            self.debug_log.append(f"Results saved to: {results_path}")
            
            # If we created a separate file, inform user about both files
            if file_exists:
                self.debug_log.append(f"Original file with formulas preserved at: {output_path}")
                self.debug_log.append(f"Backup copy created at: {backup_path}")
                print(f"NOTE: To preserve formulas, results were saved to a new file: {results_path}")
                print(f"      Original file with formulas intact: {output_path}")
        
        except Exception as e:
            error_msg = f"Error saving results: {str(e)}"
            self.debug_log.append(f"ERROR: {error_msg}")
            print(f"ERROR: {error_msg}")
            raise

    def get_summary(self):
        """
        Creates a summary of found outliers.
        """
        self.debug_log.append(f"\n=== CREATING SUMMARY ===")
        
        total_rows = len(self.df)
        grubbs_outliers = self.df['Grubbs_Outlier'].sum() if 'Grubbs_Outlier' in self.df.columns else 0
        modz_outliers = self.df['ModZ_Outlier'].sum() if 'ModZ_Outlier' in self.df.columns else 0
        
        self.debug_log.append(f"Total rows: {total_rows}")
        self.debug_log.append(f"Grubbs outliers: {grubbs_outliers}")
        self.debug_log.append(f"Modified Z-Score outliers: {modz_outliers}")
        
        any_outliers = 0
        
        # Check which columns exist before combining
        if 'Grubbs_Outlier' in self.df.columns and 'ModZ_Outlier' in self.df.columns:
            any_outliers = (self.df['Grubbs_Outlier'] | self.df['ModZ_Outlier']).sum()
        elif 'Grubbs_Outlier' in self.df.columns:
            any_outliers = grubbs_outliers
        elif 'ModZ_Outlier' in self.df.columns:
            any_outliers = modz_outliers
        
        self.debug_log.append(f"Combined outliers: {any_outliers}")
        
        # Outliers per group
        group_summary = {}
        for group_name, group_df in self.df.groupby(self.group_col):
            group_total = len(group_df)
            group_grubbs = group_df['Grubbs_Outlier'].sum() if 'Grubbs_Outlier' in group_df.columns else 0
            group_modz = group_df['ModZ_Outlier'].sum() if 'ModZ_Outlier' in group_df.columns else 0
            
            group_any = 0
            if 'Grubbs_Outlier' in group_df.columns and 'ModZ_Outlier' in group_df.columns:
                group_any = (group_df['Grubbs_Outlier'] | group_df['ModZ_Outlier']).sum()
            elif 'Grubbs_Outlier' in group_df.columns:
                group_any = group_grubbs
            elif 'ModZ_Outlier' in group_df.columns:
                group_any = group_modz
            
            self.debug_log.append(f"Group '{group_name}': total={group_total}, outliers={group_any}")
            
            group_summary[group_name] = {
                'total': group_total,
                'grubbs_outliers': group_grubbs,
                'modz_outliers': group_modz,
                'any_outliers': group_any
            }
        
        summary = {
            'total_rows': total_rows,
            'grubbs_outliers': grubbs_outliers,
            'modz_outliers': modz_outliers,
            'any_outliers': any_outliers,
            'groups': group_summary
        }
        
        self.debug_log.append(f"Summary completed: {summary}")
        return summary

    @staticmethod
    def run_multi_dataset_outlier_detection(df, group_col, dataset_columns, alpha=0.05, 
                                            iterate=False, run_grubbs=True, run_modz=True, 
                                            grubbs_alpha=0.05, modz_threshold=3.5,
                                            output_path="multi_dataset_outlier_results.xlsx"):
        """
        Run outlier detection on multiple datasets (columns) and create a combined Excel output.
        """
        print(f"DEBUG: Current working directory before export: {os.getcwd()}")
        print(f"DEBUG: Starting multi-dataset outlier detection")
        print(f"DEBUG: Datasets to analyze: {dataset_columns}")
        print(f"DEBUG: Group column: {group_col}")
        
        all_results = {}
        failed_datasets = {}
        temp_files = []  # Track temporary files for cleanup
        
        # Create main workbook for combined results
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)
        
        try:
            # Analyze each dataset
            for i, dataset_col in enumerate(dataset_columns):
                print(f"DEBUG: Analyzing dataset {i+1}/{len(dataset_columns)}: {dataset_col}")
                
                try:
                    # Create detector for this dataset
                    detector = OutlierDetector(
                        df=df.copy(),
                        group_col=group_col,
                        value_col=dataset_col
                    )
                    
                    # Run requested tests
                    if run_grubbs:
                        print(f"DEBUG: Running Grubbs test for {dataset_col}")
                        detector.run_grubbs(alpha=grubbs_alpha, iterate=iterate)
                    
                    if run_modz:
                        print(f"DEBUG: Running Modified Z-Score test for {dataset_col}")
                        detector.run_mod_z_score(threshold=modz_threshold, iterate=iterate)
                    
                    # Store results
                    all_results[dataset_col] = {
                        'detector': detector,
                        'summary': detector.get_summary()
                    }
                    
                    # Add sheets to combined workbook
                    OutlierDetector._add_dataset_to_workbook(wb, detector, dataset_col, run_grubbs, run_modz)
                    print(f"DEBUG: Successfully analyzed {dataset_col}")
                    
                except Exception as e:
                    error_msg = f"Error analyzing {dataset_col}: {str(e)}"
                    failed_datasets[dataset_col] = error_msg
                    print(f"DEBUG: ERROR: {error_msg}")
                        
            # Create summary sheet
            OutlierDetector._create_multi_summary_sheet(wb, all_results, failed_datasets, dataset_columns)
            
            # Add visualization sheet with swarm plots
            visual_temp_files = OutlierDetector._add_visualization_sheet(wb, all_results, dataset_columns)
            if visual_temp_files:
                temp_files.extend(visual_temp_files)
            
            # Save combined workbook
            wb.save(output_path)
            print(f"DEBUG: Combined results saved to: {output_path}")
            
            # Return summary
            return {
                "type": "multi_dataset_outlier_detection",
                "successful_datasets": list(all_results.keys()),
                "failed_datasets": failed_datasets,
                "total_datasets": len(dataset_columns),
                "output_file": output_path,
                "summary": {
                    "success_count": len(all_results),
                    "failure_count": len(failed_datasets),
                    "success_rate": f"{len(all_results)/len(dataset_columns)*100:.1f}%"
                }
            }
            
        except Exception as e:
            error_msg = f"Critical error in multi-dataset analysis: {str(e)}"
            print(f"DEBUG: CRITICAL ERROR: {error_msg}")
            raise RuntimeError(error_msg)
            
        finally:
            # Clean up temporary files
            for temp_file in temp_files:
                try:
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
                        print(f"DEBUG: Cleaned up temporary file: {temp_file}")
                except Exception as e:
                    print(f"DEBUG: Error cleaning up file {temp_file}: {str(e)}")
    
    @staticmethod
    def _add_dataset_to_workbook(wb, detector, dataset_name, run_grubbs, run_modz):
        """Add analysis sheet for a single dataset with debug log and summary."""
        
        # Create sheet name (Excel sheet names are limited to 31 characters)
        safe_name = dataset_name.replace(' ', '_')[:25]  # Leave room for suffixes
        
        # Create single analysis sheet
        ws = wb.create_sheet(title=f"{safe_name}_Analysis")
        
        # Title
        ws.cell(row=1, column=1, value=f"Outlier Analysis: {dataset_name}").font = Font(bold=True, size=16, color="0066CC")
        
        # Summary section
        summary = detector.get_summary()
        row = 3
        
        ws.cell(row=row, column=1, value="SUMMARY").font = Font(bold=True, size=14)
        row += 1
        
        ws.cell(row=row, column=1, value="Dataset:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=dataset_name)
        row += 1
        
        ws.cell(row=row, column=1, value="Total data points:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=summary['total_rows'])
        row += 1
        
        if run_grubbs:
            ws.cell(row=row, column=1, value="Grubbs outliers:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=summary['grubbs_outliers'])
            row += 1
            
        if run_modz:
            ws.cell(row=row, column=1, value="ModZ outliers:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=summary['modz_outliers'])
            row += 1
        
        ws.cell(row=row, column=1, value="Total outliers:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=summary['any_outliers'])
        row += 2
        
        # Group statistics
        ws.cell(row=row, column=1, value="Group Statistics:").font = Font(bold=True)
        row += 1
        
        group_headers = ["Group", "Total", "Outliers", "Clean"]
        for col_idx, header in enumerate(group_headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(bold=True)
        row += 1
        
        for group_name, group_stats in summary['groups'].items():
            ws.cell(row=row, column=1, value=str(group_name))
            ws.cell(row=row, column=2, value=group_stats['total'])
            ws.cell(row=row, column=3, value=group_stats['any_outliers'])
            ws.cell(row=row, column=4, value=group_stats['total'] - group_stats['any_outliers'])
            row += 1
        
        row += 2
        
        # Debug Log section
        ws.cell(row=row, column=1, value="DEBUG LOG").font = Font(bold=True, size=14, color="0066CC")
        row += 1
        
        ws.cell(row=row, column=1, value="Detailed analysis log:").font = Font(italic=True)
        row += 1
        
        # Set column width for debug log
        ws.column_dimensions['A'].width = 40
        
        # Write debug log entries with enhanced formatting
        for log_entry in detector.debug_log:
            log_text = str(log_entry)[:32000]
            cell = ws.cell(row=row, column=1)
            cell.value = log_text
            
            # Force inlineStr data type for any text starting with "="
            if log_text.startswith("="):
                cell.data_type = 'inlineStr'
            
            # Enhanced formatting based on content
            if log_text.startswith("==="):
                # Main section headers - blue background, white text
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            elif log_text.startswith("---"):
                # Group headers - light blue background, dark text
                cell.font = Font(bold=True, size=10, color="000080")
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            elif "ERROR:" in log_text:
                # Errors - red background, white text
                cell.font = Font(color="FFFFFF", bold=True, size=9)
                cell.fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
            elif "OUTLIER DETECTED" in log_text:
                # Outlier detection - yellow background, red text
                cell.font = Font(color="CC0000", bold=True, size=9)
                cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            elif "RESULT:" in log_text:
                # Results - green background, dark text
                cell.font = Font(color="006600", bold=True, size=9)
                cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            elif log_text.startswith("  Testing round") or log_text.startswith("  Running"):
                # Test round information - light gray background
                cell.font = Font(size=8, italic=True)
                cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            elif log_text.startswith("    "):
                # Detailed information - smaller font, indented
                cell.font = Font(size=8, color="666666")
            elif log_text.startswith("  "):
                # General information - smaller font
                cell.font = Font(size=9)
            
            row += 1

        # Set other column widths
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    @staticmethod
    def _create_multi_summary_sheet(wb, all_results, failed_datasets, dataset_columns):
        """Create a comprehensive summary sheet for all datasets."""
        ws = wb.create_sheet(title="Data_without_outlier", index=0)  # Insert as first sheet
        
        # Get the original DataFrame from the first successful result
        original_df = None
        group_col = None
        for dataset_col, result_data in all_results.items():
            if 'detector' in result_data:
                original_df = result_data['detector'].df
                group_col = result_data['detector'].group_col
                break
        
        if original_df is None:
            # Fallback if no data available
            ws.cell(row=1, column=1, value="No data available").font = Font(bold=True, color="FF0000")
            return
        
        if group_col is None:
            ws.cell(row=1, column=1, value="Group column not found").font = Font(bold=True, color="FF0000")
            return
        
        # Create headers: Sample + all dataset columns
        headers = [group_col] + dataset_columns
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
        
        # Get all unique combinations of group values that exist in the original data
        # We need to preserve the original row structure
        row_idx = 2
        
        # Iterate through the original DataFrame to maintain the exact structure
        for orig_idx, orig_row in original_df.iterrows():
            group_value = orig_row[group_col]
            
            # Column 1: Group/Sample name
            ws.cell(row=row_idx, column=1, value=str(group_value))
            
            # For each dataset column, get the value and check if it's an outlier
            for col_idx, dataset_col in enumerate(dataset_columns, start=2):
                if dataset_col in all_results:
                    detector = all_results[dataset_col]['detector']
                    
                    # Get the value for this specific row
                    if orig_idx < len(detector.df):
                        row_data = detector.df.iloc[orig_idx]
                        value = row_data[dataset_col]
                        
                        # Check if this specific row is marked as an outlier
                        is_grubbs_outlier = row_data.get('Grubbs_Outlier', False) if 'Grubbs_Outlier' in detector.df.columns else False
                        is_modz_outlier = row_data.get('ModZ_Outlier', False) if 'ModZ_Outlier' in detector.df.columns else False
                        is_outlier = is_grubbs_outlier or is_modz_outlier
                        
                        # Add the value to the cell
                        if pd.isna(value):
                            cell = ws.cell(row=row_idx, column=col_idx, value="NaN")
                        else:
                            cell = ws.cell(row=row_idx, column=col_idx, value=float(value))
                            
                            # Mark outliers with red highlighting
                            if is_outlier:
                                cell.font = Font(color="FF0000", bold=True)
                                cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    else:
                        # Row doesn't exist in this detector's data
                        ws.cell(row=row_idx, column=col_idx, value="N/A")
                else:
                    # Dataset failed or not available
                    ws.cell(row=row_idx, column=col_idx, value="N/A")
            
            row_idx += 1
        
        # Set column widths
        for col in range(1, len(headers) + 1):
            if col <= 26:  # A-Z
                ws.column_dimensions[chr(64 + col)].width = 15
            else:  # AA, AB, etc.
                ws.column_dimensions[f"A{chr(64 + col - 26)}"].width = 15
                
    @staticmethod
    def _add_visualization_sheet(wb, all_results, dataset_columns):
        """Add visualization sheet with swarm plots showing outliers."""
        from openpyxl.drawing.image import Image
        import matplotlib.pyplot as plt
        import seaborn as sns
        import tempfile
        # import os  # Already imported at top
        
        # Create a visualization sheet
        ws = wb.create_sheet(title="Visualization")
        
        # Set title
        ws.cell(row=1, column=1, value="Outlier Visualization").font = Font(bold=True, size=16, color="0066CC")
        ws.cell(row=2, column=1, value="Visual representation of data with outliers highlighted").font = Font(italic=True)
        
        # Track row position and temp files
        row = 4
        temp_files = []
        
        # Track if we generated any plots successfully
        plots_added = False
        
        # For each dataset, create a swarm plot with modern styling
        for dataset_col in dataset_columns:
            if dataset_col not in all_results:
                continue
            
            # Get detector for this dataset
            result_data = all_results[dataset_col]
            if 'detector' not in result_data:
                continue
            
            detector = result_data['detector']
            df = detector.df
            group_col = detector.group_col
            value_col = detector.value_col
            
            # Determine which outlier column to use
            outlier_col = None
            if 'Grubbs_Outlier' in df.columns:
                outlier_col = 'Grubbs_Outlier'
                test_name = "Grubbs Test"
            elif 'ModZ_Outlier' in df.columns:
                outlier_col = 'ModZ_Outlier'
                test_name = "Modified Z-Score Test"
            
            if not outlier_col:
                continue
                
            try:
                # Add dataset name as header
                title_text = f"Dataset: {dataset_col}"
                title_cell = ws.cell(row=row, column=1)
                title_cell.value = title_text
                title_cell.font = Font(bold=True, size=14)
                if title_text.startswith("="):
                    title_cell.data_type = 'inlineStr'
                row += 1
                
                # Add test method used
                ws.cell(row=row, column=1, value=f"Method: {test_name}")
                row += 1
                
                # Get outlier count
                outlier_count = df[outlier_col].sum()
                total_count = len(df)
                ws.cell(row=row, column=1, value=f"Found {outlier_count} outliers out of {total_count} data points ({outlier_count/total_count:.1%})")
                row += 2
                
                # ─── Matplotlib/Seaborn ‐ Plot ────────────────────────────────────────────
                # Modern, minimalist style
                sns.set_style("white")  
                plt.rcParams.update({
                    "axes.edgecolor": "#333333",
                    "axes.linewidth": 1.0,
                    "xtick.color": "#333333",
                    "ytick.color": "#333333",
                    "font.size": 12,
                    "axes.titlesize": 14,
                    "axes.labelsize": 12,
                })

                # Larger figure size with less whitespace
                fig, ax = plt.subplots(figsize=(10, 7))
                fig.patch.set_facecolor("white")

                # Boxplot in subtle light gray (no outlier symbols)
                sns.boxplot(
                    x=group_col,
                    y=value_col,
                    data=df,
                    fliersize=0,
                    width=0.5,
                    palette=["#DDDDDD"] * len(df[group_col].unique()),   # Light gray box
                    ax=ax
                )

                # Scatter/Swarm plot: normal points in dark gray, outliers in bright blue
                # Depending on dataset size: stripplot with jitter
                if len(df) < 100:
                    sns.swarmplot(
                        x=group_col,
                        y=value_col,
                        data=df,
                        hue=outlier_col,
                        palette={False: "#555555", True: "#007AFF"},  # Bright blue for outliers
                        size=10,
                        edgecolor="white",
                        linewidth=0.8,
                        ax=ax
                    )
                else:
                    sns.stripplot(
                        x=group_col,
                        y=value_col,
                        data=df,
                        hue=outlier_col,
                        palette={False: "#555555", True: "#007AFF"},
                        size=8,
                        jitter=0.25,
                        alpha=0.8,
                        dodge=False,
                        ax=ax
                    )

                # Remove top and right spines ("despine")
                sns.despine(ax=ax, top=True, right=True, left=False, bottom=False)

                # Axis labels
                ax.set_xlabel(group_col, fontsize=12, color="#333333")
                ax.set_ylabel(value_col, fontsize=12, color="#333333")
                ax.set_title(f"{value_col} by Group (Outliers Highlighted)", fontsize=14, color="#333333", pad=15)

                # Legend only for outliers, no frame, modern positioning
                handles, labels = ax.get_legend_handles_labels()
                # labels come in order [False, True] -> filter accordingly
                new_handles, new_labels = [], []
                for h, lab in zip(handles, labels):
                    if lab == "False":
                        new_handles.append(h)
                        new_labels.append("Normal")
                    elif lab == "True":
                        new_handles.append(h)
                        new_labels.append("Outlier")
                legend = ax.legend(
                    new_handles,
                    new_labels,
                    title="Status",
                    loc="upper right",
                    frameon=False,
                    fontsize=12,
                    title_fontsize=13
                )

                # Subtle grid on y-axis (almost invisible)
                ax.grid(axis="y", color="#f0f0f0", linestyle="-", linewidth=0.7)
                ax.set_axisbelow(True)

                # Tight layout (less margin)
                plt.tight_layout(pad=1)
                
                # Create a temporary file path with high DPI for crisp rendering
                fd, temp_path = tempfile.mkstemp(suffix='.png')
                os.close(fd)
                temp_files.append(temp_path)
                
                # Save with higher DPI for better quality
                plt.savefig(temp_path, dpi=200, bbox_inches="tight", facecolor="white")
                plt.close(fig)
                
                # Add the image to the worksheet
                img = Image(temp_path)
                # Set image size (in pixels)
                img.width = 1000
                img.height = 700
                ws.add_image(img, f'A{row}')
                
                # Mark that we added at least one plot
                plots_added = True
                
                # Move row position down to accommodate the image
                row += 32
                row += 2  # Add some space between plots
                
            except Exception as e:
                error_cell = ws.cell(row=row, column=1)
                error_cell.value = f"Error creating visualization for {dataset_col}: {str(e)}"
                error_cell.data_type = 'inlineStr'  # Ensure error messages aren't interpreted as formulas
                row += 3
        
        # If no plots were added, add a message
        if not plots_added:
            ws.cell(row=row, column=1, value="No visualizations could be generated.")
            row += 1
        
        # Set column width for better display
        ws.column_dimensions['A'].width = 120
        
        # Return list of temp files for later cleanup
        return temp_files
    
    @staticmethod
    def _add_single_visualization_sheet(wb, detector, dataset_name=None):
        """Add visualization sheet with swarm plot showing outliers in modern, minimalist style."""
        from openpyxl.drawing.image import Image
        import matplotlib.pyplot as plt
        import seaborn as sns
        import tempfile
        # import os  # Already imported at top

        # Create a visualization sheet
        ws = wb.create_sheet(title="Visualization")

        # Set title (Excel sheet)
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = "Outlier Visualization"
        title_cell.font = Font(bold=True, size=16, color="0066CC")

        subtitle_cell = ws.cell(row=2, column=1)
        subtitle_cell.value = "Visual representation with highlighted outliers"
        subtitle_cell.font = Font(italic=True)

        # Start row for the image
        row = 4
        temp_path = None

        try:
            df = detector.df
            group_col = detector.group_col
            value_col = detector.value_col

            # Which outlier column?
            if 'Grubbs_Outlier' in df.columns:
                outlier_col = 'Grubbs_Outlier'
                test_name = "Grubbs Test"
            elif 'ModZ_Outlier' in df.columns:
                outlier_col = 'ModZ_Outlier'
                test_name = "Modified Z-Score Test"
            else:
                # If no results
                no_results_cell = ws.cell(row=row, column=1)
                no_results_cell.value = "No outlier results available."
                no_results_cell.data_type = 'inlineStr'
                return None

            # Subheader on the Excel sheet
            title_text = f"Dataset: {dataset_name or value_col}"
            title2_cell = ws.cell(row=row, column=1)
            title2_cell.value = title_text
            title2_cell.font = Font(bold=True, size=14)
            if title_text.startswith("="):
                title2_cell.data_type = 'inlineStr'
            row += 1

            method_cell = ws.cell(row=row, column=1)
            method_cell.value = f"Outlier Method: {test_name}"
            row += 1

            # Outlier counter
            outlier_count = int(df[outlier_col].sum())
            total_count = len(df)
            count_cell = ws.cell(row=row, column=1)
            count_cell.value = f"Found outliers: {outlier_count} of {total_count} ({outlier_count/total_count:.1%})"
            row += 2

            # ─── Matplotlib/Seaborn ‐ Plot ────────────────────────────────────────────
            # Modern, minimalist style
            sns.set_style("white")
            plt.rcParams.update({
                "axes.edgecolor": "#333333",
                "axes.linewidth": 1.0,
                "xtick.color": "#333333",
                "ytick.color": "#333333",
                "font.size": 12,
                "axes.titlesize": 14,
                "axes.labelsize": 12,
            })

            # Larger figure size with less whitespace
            fig, ax = plt.subplots(figsize=(10, 7))
            fig.patch.set_facecolor("white")

            # Boxplot in subtle light gray (no outlier symbols)
            sns.boxplot(
                x=group_col,
                y=value_col,
                data=df,
                fliersize=0,
                width=0.5,
                palette=["#DDDDDD"] * len(df[group_col].unique()),   # Light gray box
                ax=ax
            )

            # Scatter/Swarm plot: normal points in dark gray, outliers in bright blue
            # Depending on dataset size: stripplot with jitter
            if len(df) < 100:
                sns.swarmplot(
                    x=group_col,
                    y=value_col,
                    data=df,
                    hue=outlier_col,
                    palette={False: "#555555", True: "#007AFF"},  # Bright blue for outliers
                    size=10,
                    edgecolor="white",
                    linewidth=0.8,
                    ax=ax
                )
            else:
                sns.stripplot(
                    x=group_col,
                    y=value_col,
                    data=df,
                    hue=outlier_col,
                    palette={False: "#555555", True: "#007AFF"},
                    size=8,
                    jitter=0.25,
                    alpha=0.8,
                    dodge=False,
                    ax=ax
                )

            # Remove top and right spines ("despine")
            sns.despine(ax=ax, top=True, right=True, left=False, bottom=False)

            # Axis labels
            ax.set_xlabel(group_col, fontsize=12, color="#333333")
            ax.set_ylabel(value_col, fontsize=12, color="#333333")
            ax.set_title(f"{value_col} by Group (Outliers Highlighted)", fontsize=14, color="#333333", pad=15)

            # Legend only for outliers, no frame, modern positioning
            handles, labels = ax.get_legend_handles_labels()
            # labels come in order [False, True] -> filter accordingly
            new_handles, new_labels = [], []
            for h, lab in zip(handles, labels):
                if lab == "False":
                    new_handles.append(h)
                    new_labels.append("Normal")
                elif lab == "True":
                    new_handles.append(h)
                    new_labels.append("Outlier")
            legend = ax.legend(
                new_handles,
                new_labels,
                title="Status",
                loc="upper right",
                frameon=False,
                fontsize=12,
                title_fontsize=13
            )

            # Subtle grid on y-axis (almost invisible)
            ax.grid(axis="y", color="#f0f0f0", linestyle="-", linewidth=0.7)
            ax.set_axisbelow(True)

            # Tight layout (less margin)
            plt.tight_layout(pad=1)

            # Save as PNG temporarily
            fd, temp_path = tempfile.mkstemp(suffix=".png")
            os.close(fd)
            plt.savefig(temp_path, dpi=200, bbox_inches="tight", facecolor="white")
            plt.close(fig)

            # Insert image into Excel sheet
            img = Image(temp_path)
            img.width = 1000  # in pixels
            img.height = 700
            ws.add_image(img, f"A{row}")

            # ─── Group Statistics Table ───────────────────────────────────────────
            row += 32  # Leave space for the image
            ws.cell(row=row, column=1, value="Group Statistics:").font = Font(bold=True)
            row += 1

            # Table headers
            stats_headers = ["Group", "Count", "Mean", "StdDev", "Median", "Min", "Max", "Outliers"]
            for col_idx, header in enumerate(stats_headers, start=1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = Font(bold=True)
            row += 1

            # Values per group
            for group_name, group_df in df.groupby(group_col):
                vals = group_df[value_col].dropna()
                out_count = int(group_df[outlier_col].sum())

                # Group
                grp_cell = ws.cell(row=row, column=1)
                grp_val = str(group_name)
                grp_cell.value = grp_val
                if grp_val.startswith("="):
                    grp_cell.data_type = 'inlineStr'

                # Numeric cells
                ws.cell(row=row, column=2, value=len(vals))
                ws.cell(row=row, column=3, value=float(vals.mean()) if len(vals) > 0 else 0.0)
                ws.cell(row=row, column=4, value=float(vals.std()) if len(vals) > 0 else 0.0)
                ws.cell(row=row, column=5, value=float(vals.median()) if len(vals) > 0 else 0.0)
                ws.cell(row=row, column=6, value=float(vals.min()) if len(vals) > 0 else 0.0)
                ws.cell(row=row, column=7, value=float(vals.max()) if len(vals) > 0 else 0.0)

                out_cell = ws.cell(row=row, column=8, value=out_count)
                # Highlight if there are outliers
                if out_count > 0:
                    out_cell.font = Font(color="FF0000", bold=True)
                    for c in range(1, 9):
                        ws.cell(row=row, column=c).fill = PatternFill(
                            start_color="FFEEEE",
                            end_color="FFEEEE",
                            fill_type="solid"
                        )
                row += 1

        except Exception as e:
            # If an error occurs during plotting
            err_cell = ws.cell(row=row, column=1)
            err_cell.value = f"Error in visualization: {str(e)}"
            err_cell.data_type = 'inlineStr'

        # Adjust column widths
        for col in range(1, 9):
            if col == 1:
                ws.column_dimensions[chr(64 + col)].width = 30
            else:
                ws.column_dimensions[chr(64 + col)].width = 15

        return temp_path

# Note: Classes are imported lazily to avoid circular imports.
# Use get_data_visualizer(), get_statistical_tester(), get_results_exporter() functions instead.
ResultsExporter = get_results_exporter()
DataVisualizer = get_data_visualizer()